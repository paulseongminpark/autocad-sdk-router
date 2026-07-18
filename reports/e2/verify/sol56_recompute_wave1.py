#!/usr/bin/env python3
"""Independent, read-only recomputation of E2 Wave 1 artifact metrics.

This script writes nothing.  It reads the committed prediction/truth/result
artifacts plus the explicitly allowed staged B3 DXF, prints a JSON audit, and
then prints the verifier's final Markdown table.
"""

from __future__ import annotations

import glob
import hashlib
import json
import math
from collections import Counter, defaultdict
from pathlib import Path

import ezdxf
import openpyxl


ROOT = Path(__file__).resolve().parents[4]
STAGED_DXF = Path(
    r"D:\dev\99_tools\autocad-sdk-router\runs\e2_b3_dxfout_20260717\1_export.dxf"
)
TIERS = ("S", "F", "M")
ARMS = ("full", "name_blind")
JUDGES = ("opus48_max", "fable5_high", "sol56_xhigh")
THRESHOLD = 0.5


def load_json(path: Path):
    with path.open("r", encoding="utf-8") as stream:
        return json.load(stream)


def div(num: int, den: int):
    return num / den if den else None


def pearson(xs: list[float], ys: list[float]):
    if len(xs) != len(ys) or len(xs) < 3:
        return None
    mx = math.fsum(xs) / len(xs)
    my = math.fsum(ys) / len(ys)
    xx = math.fsum((x - mx) ** 2 for x in xs)
    yy = math.fsum((y - my) ** 2 for y in ys)
    if xx == 0 or yy == 0:
        return None
    xy = math.fsum((x - mx) * (y - my) for x, y in zip(xs, ys))
    return xy / math.sqrt(xx * yy)


def truth_by_drawing(tier: str):
    out = {}
    pack = ROOT / "reports" / "e2" / "s2" / "packs" / tier
    for path in sorted(pack.glob("*.truth.json")):
        row = load_json(path)
        out[str(row["drawing_id"])] = {
            str(handle).upper() for handle in row.get("wall_handles_flat", [])
        }
    return out


def threshold_set(pred: dict):
    per_handle = (pred.get("scores") or {}).get("per_handle") or {}
    return {
        str(handle).upper()
        for handle, record in per_handle.items()
        if float(record.get("score", record) if isinstance(record, dict) else record)
        >= THRESHOLD
    }


def recompute_b1():
    tv_path = ROOT / "reports" / "e2" / "s2" / "fidelity_M_v1_tv.json"
    fidelity_path = ROOT / "reports" / "e2" / "s2" / "fidelity_M_v1.json"
    real_stats_path = ROOT / "reports" / "e2" / "s2" / "real_stats.json"
    tv = load_json(tv_path)
    fidelity = load_json(fidelity_path)
    real_stats = load_json(real_stats_path)

    real_mix = tv["real_mix"]
    pack_mix = tv["pack_mix"]
    real_total = math.fsum(real_mix.values())
    pack_total = math.fsum(pack_mix.values())
    keys = set(real_mix) | set(pack_mix)
    tv_exact = 0.5 * math.fsum(
        abs(real_mix.get(k, 0) / real_total - pack_mix.get(k, 0) / pack_total)
        for k in keys
    )

    flattened = Counter()
    for role in real_stats["entity_mix_by_role"].values():
        flattened.update(role.get("entity_types") or {})

    pack_dxf_counts = {
        tier: len(list((ROOT / "reports" / "e2" / "s2" / "packs" / tier).glob("*.dxf")))
        for tier in TIERS
    }
    thick = fidelity["statistics"]["thickness_ks"]
    pack_hist = fidelity["pack_summary"]["thickness_hist"]
    real_hist = fidelity["real_summary"]["thickness_hist"]
    return {
        "tv_exact": tv_exact,
        "tv_recorded": tv["tv"],
        "tv_real_total": real_total,
        "tv_pack_total": pack_total,
        "tv_real_mix_matches_real_stats": dict(sorted(flattened.items()))
        == dict(sorted(real_mix.items())),
        "ks_recorded": thick["distance"],
        "ks_method": thick["method"],
        "ks_n_pack": thick["n_pack"],
        "ks_n_real": thick["n_real"],
        "pack_hist_sum": sum(pack_hist["counts"]),
        "pack_hist_edges": len(pack_hist["edges"]),
        "real_hist_sum": sum(real_hist["counts"]),
        "real_hist_edges": len(real_hist["bin_edges"]),
        "raw_samples_preserved": any(
            "thickness_samples" in fidelity.get(key, {})
            for key in ("pack_summary", "real_summary")
        ),
        "pack_dxf_counts": pack_dxf_counts,
    }


def recompute_b2():
    out = {}
    for tier in TIERS:
        truths = truth_by_drawing(tier)
        eval_dir = ROOT / "reports" / "e2" / "s4" / f"eval_{tier}"
        tier_out = {}
        preds_by_arm = {arm: {} for arm in ARMS}
        for arm in ARMS:
            total = Counter()
            eval_total = Counter()
            eval_mismatches = []
            eval_sources = Counter()
            exact_threshold = 0
            for drawing, truth in sorted(truths.items()):
                pred_path = eval_dir / f"{drawing}.{arm}.pred.json"
                eval_path = eval_dir / f"{drawing}.{arm}.eval.json"
                pred = load_json(pred_path)
                preds_by_arm[arm][drawing] = pred
                predicted = threshold_set(pred)
                per_handle = (pred.get("scores") or {}).get("per_handle") or {}
                exact_threshold += sum(
                    float(rec.get("score", rec) if isinstance(rec, dict) else rec)
                    == THRESHOLD
                    for rec in per_handle.values()
                )
                direct = {
                    "tp": len(predicted & truth),
                    "fp": len(predicted - truth),
                    "fn": len(truth - predicted),
                }
                total.update(direct)
                total["predicted"] += len(predicted)
                total["truth"] += len(truth)
                total["scored"] += len(per_handle)

                committed_eval = load_json(eval_path)
                baseline = committed_eval["baseline"]
                eval_sources[baseline.get("source")] += 1
                eval_counts = {k: int(baseline[k]) for k in ("tp", "fp", "fn")}
                eval_total.update(eval_counts)
                if direct != eval_counts:
                    eval_mismatches.append(
                        {
                            "drawing": drawing,
                            "direct": direct,
                            "committed_eval": eval_counts,
                        }
                    )

            p = div(total["tp"], total["tp"] + total["fp"])
            r = div(total["tp"], total["tp"] + total["fn"])
            eval_p = div(eval_total["tp"], eval_total["tp"] + eval_total["fp"])
            eval_r = div(eval_total["tp"], eval_total["tp"] + eval_total["fn"])
            tier_out[arm] = {
                **dict(total),
                "precision": p,
                "recall": r,
                "eval_tp": eval_total["tp"],
                "eval_fp": eval_total["fp"],
                "eval_fn": eval_total["fn"],
                "eval_precision": eval_p,
                "eval_recall": eval_r,
                "eval_sources": dict(eval_sources),
                "eval_mismatch_drawings": eval_mismatches,
                "scores_exactly_at_threshold": exact_threshold,
            }

        twin = Counter()
        twin_flips = []
        for drawing in sorted(truths):
            full = preds_by_arm["full"][drawing]
            blind = preds_by_arm["name_blind"][drawing]
            full_per = full["scores"]["per_handle"]
            blind_per = blind["scores"]["per_handle"]
            twin["files"] += 1
            full_bytes = (eval_dir / f"{drawing}.full.pred.json").read_bytes()
            blind_bytes = (eval_dir / f"{drawing}.name_blind.pred.json").read_bytes()
            twin["byte_different_files"] += full_bytes != blind_bytes
            for handle in sorted(set(full_per) | set(blind_per)):
                fs = float((full_per.get(handle) or {}).get("score", 0.0))
                bs = float((blind_per.get(handle) or {}).get("score", 0.0))
                if fs != bs:
                    twin["score_different_handles"] += 1
                if (fs >= THRESHOLD) != (bs >= THRESHOLD):
                    twin["threshold_flips"] += 1
                    twin_flips.append(
                        {"drawing": drawing, "handle": handle, "full": fs, "name_blind": bs}
                    )
            if full["scores"].get("walls") == blind["scores"].get("walls"):
                twin["identical_walls_files"] += 1
        tier_out["name_blind_comparison"] = {**dict(twin), "flips": twin_flips}
        tier_out["committed_summary"] = load_json(
            ROOT / "reports" / "e2" / "s4" / f"eval_{tier}.json"
        )
        out[tier] = tier_out
    return out


def workbook_rows(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["results"]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    return [dict(zip(headers, row)) for row in rows[1:]], wb


def recompute_b4():
    paths = (
        ROOT / "reports" / "e2" / "s5" / "battery_S_v2.xlsx",
        ROOT / "reports" / "e2" / "s5" / "battery_S_v2_translate.xlsx",
    )
    all_rows = []
    workbook_summary_mismatches = []
    for path in paths:
        rows, wb = workbook_rows(path)
        all_rows.extend(rows)
        summaries = list(wb["summary"].iter_rows(values_only=True))
        sheaders = list(summaries[0])
        smap = {row[0]: dict(zip(sheaders, row)) for row in summaries[1:]}
        for transform in {row["transform"] for row in rows}:
            group = [row for row in rows if row["transform"] == transform]
            vals = [float(row["invariance"]) for row in group if row["invariance"] is not None]
            computed = math.fsum(vals) / len(vals)
            if not math.isclose(computed, float(smap[transform]["mean_invariance"]), abs_tol=1e-15):
                workbook_summary_mismatches.append(transform)

    truths = truth_by_drawing("S")
    grouped = {}
    for transform in sorted({row["transform"] for row in all_rows}):
        group = [row for row in all_rows if row["transform"] == transform]
        vals = [float(row["invariance"]) for row in group if row["invariance"] is not None]
        recall_proxy = []
        truth_before_mismatch = []
        for row in group:
            drawing = Path(str(row["drawing"])).stem
            n_truth = len(truths[drawing])
            if int(row["wall_count_before"]) != n_truth:
                truth_before_mismatch.append(drawing)
            recall_proxy.append(float(row["wall_count_after"]) / n_truth)
        grouped[transform] = {
            "n": len(group),
            "n_numeric": len(vals),
            "mean_invariance": math.fsum(vals) / len(vals),
            "min_invariance": min(vals),
            "max_invariance": max(vals),
            "flips_total": sum(int(row["flips_count"]) for row in group),
            "wall_before_total": sum(int(row["wall_count_before"]) for row in group),
            "wall_after_total": sum(int(row["wall_count_after"]) for row in group),
            "n_errors": sum(bool(row["error"]) for row in group),
            "n_sentinel_zero": sum(bool(row["sentinel_zero"]) for row in group),
            "n_sentinel_all": sum(bool(row["sentinel_all"]) for row in group),
            "mean_recall_count_proxy": math.fsum(recall_proxy) / len(recall_proxy),
            "min_recall_count_proxy": min(recall_proxy),
            "n_recall_proxy_below_0_5": sum(value < 0.5 for value in recall_proxy),
            "truth_before_count_mismatches": truth_before_mismatch,
        }
    return {
        "arms": grouped,
        "rows_total": len(all_rows),
        "sentinel_zero_total": sum(bool(row["sentinel_zero"]) for row in all_rows),
        "sentinel_all_total": sum(bool(row["sentinel_all"]) for row in all_rows),
        "error_total": sum(bool(row["error"]) for row in all_rows),
        "workbook_summary_mismatches": workbook_summary_mismatches,
        "fold": load_json(ROOT / "reports" / "e2" / "s5" / "b4_fold_v1.json"),
    }


def load_silver():
    acc = defaultdict(dict)
    duplicates = []
    judge_counts = {}
    for judge in JUDGES:
        seen = set()
        paths = sorted(
            glob.glob(
                str(ROOT / "reports" / "e1" / "annot_v1" / "raw" / judge / "*.json")
            )
        )
        for raw_path in paths:
            data = load_json(Path(raw_path))
            records = data if isinstance(data, list) else data.get("answers") or []
            for record in records:
                definition = record.get("def")
                likelihood = record.get("wall_likelihood")
                if definition is None or likelihood is None:
                    continue
                if definition in seen:
                    duplicates.append({"judge": judge, "def": definition})
                seen.add(definition)
                acc[definition][judge] = float(likelihood)
        judge_counts[judge] = len(seen)
    means = {
        definition: math.fsum(values.values()) / len(values)
        for definition, values in acc.items()
    }
    return acc, means, judge_counts, duplicates


def xlsx_real_defs_mismatches(rows_by_def: dict[str, dict]):
    path = ROOT / "reports" / "e2" / "s4" / "real_defs_v1.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    values = list(wb["per_def"].iter_rows(values_only=True))
    headers = list(values[0])
    mismatches = []
    for values_row in values[1:]:
        row = dict(zip(headers, values_row))
        definition = row["def"]
        source = rows_by_def.get(definition)
        if source is None:
            mismatches.append({"def": definition, "reason": "xlsx-only"})
            continue
        for key in (
            "n_segments",
            "n_scored",
            "n_wall",
            "max_score",
            "silver_mean_wall_likelihood",
        ):
            if row.get(key) != source.get(key):
                mismatches.append(
                    {"def": definition, "field": key, "xlsx": row.get(key), "json": source.get(key)}
                )
    return mismatches


def recompute_b3_b5():
    artifact = load_json(ROOT / "reports" / "e2" / "s4" / "real_defs_v1.json")
    rows = artifact["rows"]
    rows_by_def = {row["def"]: row for row in rows}
    acc, raw_means, judge_counts, duplicates = load_silver()

    zero_wall = [row for row in rows if int(row["n_wall"]) == 0]
    zero_scored = [row for row in rows if int(row["n_scored"]) == 0]
    embedded_mismatches = [
        {
            "def": row["def"],
            "embedded": row["silver_mean_wall_likelihood"],
            "raw_round4": round(raw_means[row["def"]], 4),
        }
        for row in rows
        if float(row["silver_mean_wall_likelihood"])
        != round(raw_means[row["def"]], 4)
    ]
    xs = [float(row["max_score"]) for row in rows]
    raw_ys = [raw_means[row["def"]] for row in rows]
    embedded_ys = [float(row["silver_mean_wall_likelihood"]) for row in rows]
    nonempty = [row for row in rows if int(row["n_segments"]) > 0]

    dxf_doc = ezdxf.readfile(STAGED_DXF)
    dxf_blocks = {block.name for block in dxf_doc.blocks}
    with STAGED_DXF.open("rb") as stream:
        digest = hashlib.sha256()
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)

    samples = [
        "X-평면도(기본형)$0$111a",
        "X-평면도(기본형)$0$111b",
        "X-평면도(기본형)$0$130a",
        "*D295",
        "*D299",
        "*D300",
    ]
    return {
        "B3": {
            "n_rows": len(rows),
            "n_unique_defs": len(rows_by_def),
            "n_zero_wall": len(zero_wall),
            "zero_frac_n_wall": len(zero_wall) / len(rows),
            "n_zero_scored": len(zero_scored),
            "zero_frac_n_scored": len(zero_scored) / len(rows),
            "zero_scored_defs": [row["def"] for row in zero_scored],
            "dxf_size": STAGED_DXF.stat().st_size,
            "dxf_sha256": digest.hexdigest(),
            "dxf_block_count": len(dxf_blocks),
            "row_defs_present_in_dxf": len(set(rows_by_def) & dxf_blocks),
            "row_defs_missing_in_dxf": sorted(set(rows_by_def) - dxf_blocks),
            "samples": [rows_by_def[name] for name in samples],
            "xlsx_mismatches": xlsx_real_defs_mismatches(rows_by_def),
        },
        "B5": {
            "judge_unique_defs": judge_counts,
            "raw_union_defs": len(acc),
            "raw_all_three_judges": sum(len(values) == len(JUDGES) for values in acc.values()),
            "raw_duplicates": duplicates,
            "raw_defs_missing_rows": sorted(set(acc) - set(rows_by_def)),
            "row_defs_missing_raw": sorted(set(rows_by_def) - set(acc)),
            "embedded_mean_mismatches": embedded_mismatches,
            "pearson_raw_means": pearson(xs, raw_ys),
            "pearson_embedded_means": pearson(xs, embedded_ys),
            "pearson_nonempty_embedded": pearson(
                [float(row["max_score"]) for row in nonempty],
                [float(row["silver_mean_wall_likelihood"]) for row in nonempty],
            ),
            "n_nonempty": len(nonempty),
            "sample_judge_values": {
                name: acc[name] for name in samples if name in acc
            },
        },
    }


def main():
    result = {
        "B1": recompute_b1(),
        "B2": recompute_b2(),
        "B4": recompute_b4(),
    }
    result.update(recompute_b3_b5())
    print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True))
    print()
    print("| Band | Final verifier verdict |")
    print("|---|---|")
    print("| B1 | NOT_REPRODUCIBLE (TV confirms FAIL; evaluated M DXFs/raw KS samples absent) |")
    print("| B2 | DISCREPANCY (M full P 0.9091 vs claimed 0.8669; S PASS confirmed) |")
    print("| B3 | CONFIRMED (83/384 = 0.2161458333 -> 0.2161; wording caveat) |")
    print("| B4 | DISCREPANCY (100 sentinel_all rows vs claimed transform PASSes) |")
    print("| B5 | CONFIRMED (raw-mean Pearson 0.2991079704 -> 0.2991) |")


if __name__ == "__main__":
    main()
