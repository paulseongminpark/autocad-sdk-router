#!/usr/bin/env python3
"""G1 family-diverse numeric requalification harness for verifier.py.

This cell-local program is intentionally read-constrained.  It reads the
sealed gen2 family pack and the unchanged verifier, writes only to its own
cell directory, evaluates reward families before opening hidden families,
and emits measurements without a qualification PASS/FAIL decision.
"""

from __future__ import annotations

import argparse
import copy
import csv
import hashlib
import importlib.util
import json
import math
import os
import re
import sys
import time
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence


SCHEMA = "e2.g1-verifier-families.results.v1"
SELFTEST_SCHEMA = "e2.g1-verifier-families.selftest.v1"
SEED = 20260719
CPU_CAP_SECONDS = 86_400.0

CELL_DIR = Path(__file__).resolve().parent
SOURCE_ROOT = Path(r"D:\runs\e2_program\cells\gen2_families").resolve()
PACKS_ROOT = (SOURCE_ROOT / "packs").resolve()
VERIFIER_PATH = Path(
    r"D:\dev\99_tools\autocad-sdk-router\tools\e2\instruments\verifier.py"
).resolve()
PREREG_PATH = (CELL_DIR / "prereg.json").resolve()
PREREG_CSV_PATH = (CELL_DIR / "PREREG.csv").resolve()
RESULTS_PATH = (CELL_DIR / "results.json").resolve()
EVIDENCE_PATH = (CELL_DIR / "evidence.xlsx").resolve()
REPORT_PATH = (CELL_DIR / "REPORT.md").resolve()

PREREG_SHA256 = "2fb296255a2430a106dfed306090dacb5f372fd9217a24528e874b59510d3ce1"
PREREG_CSV_SHA256 = "2a43583d34e582914f8a67c7874be92d58f17f013a1eb0bbdb57102dfc7482e1"
VERIFIER_SHA256 = "72e33ab0e87e96defd00f74c5a22ae6c5cb001c69b740e627edeedaf2a80b690"
PACK_INVENTORY_SHA256 = "14ffd7bc07f52552824d59f723c22cce7ff1c0ad54cf58bbdb6cf72771dc3aea"
PACK_INVENTORY_FILE_COUNT = 404
CONFIG_SHA256 = "bf9e4d5c9e7facba3bebf7865bd188b7566f4b319bf1007ab13932bced7d3a07"
SPLIT_SHA256 = "0c06b83b44bdf8b3aa5123b910a6a47f4d9ddfbd725523eff5fea132ed814b6d"

CONFIG = {
    "angle_tol_deg": 5.0,
    "overlap_min": 0.65,
    "thickness_band_mm": [50.0, 400.0],
    "selection": "current_verifier_defaults_no_tuning",
}
REWARD_FAMILIES = ("F03", "F04", "F06", "F07", "F08")
HIDDEN_FAMILIES = ("F01", "F02", "F05")
ALL_FAMILIES = tuple(sorted(REWARD_FAMILIES + HIDDEN_FAMILIES))
ENTITY_TYPES = (
    "3DFACE",
    "ARC",
    "CIRCLE",
    "ELLIPSE",
    "HATCH",
    "INSERT",
    "LINE",
    "LWPOLYLINE",
    "MTEXT",
    "POINT",
    "SPLINE",
    "TEXT",
    "WIPEOUT",
)
ORIGINAL_PERTURBATIONS = (
    "wall_remove_single",
    "wall_remove_pair",
    "lure_add",
    "neighbor_swap",
    "pair_swap",
    "orphan_add",
)
SENTINELS = (
    "empty_claim",
    "whole_universe_claim",
    "duplicate_true_handle",
    "unknown_handle_add",
)
EXPECTED_POSITIVE = 2800
EXPECTED_FALSE = 4600
EXPECTED_TOTAL = 7400


class BoundaryViolation(RuntimeError):
    """Raised before a path outside the sealed read/write boundary is opened."""


class BudgetKill(RuntimeError):
    """Raised before the next measurement when the sealed CPU cap is exhausted."""


FIREWALL_AUDIT: dict[str, Any] = {
    "configuration_frozen": False,
    "hidden_hash_only_reads_before_freeze": 0,
    "hidden_content_reads_before_freeze": 0,
    "hidden_content_reads_after_freeze": 0,
    "freeze_event_utc": None,
}


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def canonical_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def sha256_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def sha256_text(payload: str) -> str:
    return sha256_bytes(payload.encode("utf-8"))


def _norm(path: str | os.PathLike[str]) -> Path:
    return Path(path).resolve(strict=False)


def _under(path: Path, root: Path) -> bool:
    try:
        path.relative_to(root)
        return True
    except ValueError:
        return False


def _family_from_pack_filename(path: Path) -> str | None:
    match = re.match(r"^f(0[1-8])_\d{3}(?:\.truth)?\.json$", path.name, re.IGNORECASE)
    if match:
        return "F" + match.group(1)
    match = re.match(r"^f(0[1-8])_\d{3}\.dxf$", path.name, re.IGNORECASE)
    return "F" + match.group(1) if match else None


def assert_read_allowed(path: str | os.PathLike[str], *, hash_only: bool = False) -> Path:
    resolved = _norm(path)
    allowed = resolved == VERIFIER_PATH or _under(resolved, SOURCE_ROOT) or _under(resolved, CELL_DIR)
    if not allowed:
        raise BoundaryViolation(f"read denied outside sealed allowlist: {resolved}")
    family = _family_from_pack_filename(resolved)
    if family in HIDDEN_FAMILIES:
        if not FIREWALL_AUDIT["configuration_frozen"]:
            if hash_only:
                FIREWALL_AUDIT["hidden_hash_only_reads_before_freeze"] += 1
            else:
                FIREWALL_AUDIT["hidden_content_reads_before_freeze"] += 1
                raise BoundaryViolation(
                    f"hidden-family content read denied before configuration freeze: {resolved}"
                )
        elif not hash_only:
            FIREWALL_AUDIT["hidden_content_reads_after_freeze"] += 1
    return resolved


def assert_write_allowed(path: str | os.PathLike[str]) -> Path:
    resolved = _norm(path)
    if not _under(resolved, CELL_DIR):
        raise BoundaryViolation(f"write denied outside cell directory: {resolved}")
    return resolved


def read_bytes(path: str | os.PathLike[str], *, hash_only: bool = False) -> bytes:
    resolved = assert_read_allowed(path, hash_only=hash_only)
    return resolved.read_bytes()


def read_text(path: str | os.PathLike[str]) -> str:
    return read_bytes(path).decode("utf-8-sig")


def load_json(path: str | os.PathLike[str]) -> Any:
    return json.loads(read_text(path))


def write_text(path: str | os.PathLike[str], payload: str) -> None:
    resolved = assert_write_allowed(path)
    resolved.write_text(payload, encoding="utf-8", newline="\n")


def file_sha256(path: str | os.PathLike[str], *, hash_only: bool = True) -> str:
    return sha256_bytes(read_bytes(path, hash_only=hash_only))


def pack_inventory() -> dict[str, Any]:
    files = sorted(
        (path for path in PACKS_ROOT.rglob("*") if path.is_file()),
        key=lambda path: str(path).casefold(),
    )
    lines: list[str] = []
    for path in files:
        relative = path.relative_to(PACKS_ROOT).as_posix()
        lines.append(f"{relative}|{file_sha256(path, hash_only=True)}")
    canonical = "\n".join(lines) + "\n"
    return {
        "file_count": len(files),
        "sha256": sha256_text(canonical),
        "canonical_line_count": len(lines),
    }


def validate_seals_and_inputs() -> dict[str, Any]:
    actual = {
        "prereg_json_sha256": file_sha256(PREREG_PATH),
        "prereg_csv_sha256": file_sha256(PREREG_CSV_PATH),
        "verifier_sha256": file_sha256(VERIFIER_PATH),
        "pack_inventory": pack_inventory(),
    }
    expected = {
        "prereg_json_sha256": PREREG_SHA256,
        "prereg_csv_sha256": PREREG_CSV_SHA256,
        "verifier_sha256": VERIFIER_SHA256,
        "pack_inventory": {
            "file_count": PACK_INVENTORY_FILE_COUNT,
            "sha256": PACK_INVENTORY_SHA256,
        },
    }
    mismatches: list[str] = []
    for key in ("prereg_json_sha256", "prereg_csv_sha256", "verifier_sha256"):
        if actual[key] != expected[key]:
            mismatches.append(f"{key}: expected {expected[key]}, observed {actual[key]}")
    for key in ("file_count", "sha256"):
        if actual["pack_inventory"][key] != expected["pack_inventory"][key]:
            mismatches.append(
                f"pack_inventory.{key}: expected {expected['pack_inventory'][key]}, "
                f"observed {actual['pack_inventory'][key]}"
            )
    if canonical_json(CONFIG) != (
        '{"angle_tol_deg":5.0,"overlap_min":0.65,'
        '"selection":"current_verifier_defaults_no_tuning",'
        '"thickness_band_mm":[50.0,400.0]}'
    ):
        mismatches.append("configuration canonicalization drift")
    if sha256_text(
        '{"angle_tol_deg":5.0,"overlap_min":0.65,'
        '"thickness_band_mm":[50.0,400.0],'
        '"selection":"current_verifier_defaults_no_tuning"}'
    ) != CONFIG_SHA256:
        mismatches.append("sealed configuration SHA drift")
    if mismatches:
        raise RuntimeError("sealed input mismatch: " + "; ".join(mismatches))
    return actual


def import_verifier():
    before = file_sha256(VERIFIER_PATH)
    previous = sys.dont_write_bytecode
    sys.dont_write_bytecode = True
    try:
        spec = importlib.util.spec_from_file_location("g1_sealed_verifier", VERIFIER_PATH)
        if spec is None or spec.loader is None:
            raise RuntimeError("cannot create import specification for verifier")
        module = importlib.util.module_from_spec(spec)
        sys.modules[spec.name] = module
        spec.loader.exec_module(module)
    finally:
        sys.dont_write_bytecode = previous
    after = file_sha256(VERIFIER_PATH)
    if before != VERIFIER_SHA256 or after != VERIFIER_SHA256:
        raise RuntimeError("verifier hash changed during read-only import")
    return module


@dataclass(frozen=True)
class DrawingRef:
    tier: str
    family_id: str
    drawing_id: str
    dxf_path: Path
    truth_path: Path
    dxf_sha256: str
    truth_sha256: str


@dataclass(frozen=True)
class DxfEntity:
    entity_type: str
    handle: str
    layer: str
    tags: tuple[tuple[int, str], ...]


class CpuBudget:
    def __init__(self, prior_cpu_seconds: float = 0.0) -> None:
        self.prior_cpu_seconds = max(0.0, float(prior_cpu_seconds))
        self.started = time.process_time()

    @property
    def run_cpu_seconds(self) -> float:
        return time.process_time() - self.started

    @property
    def total_cpu_seconds(self) -> float:
        return self.prior_cpu_seconds + self.run_cpu_seconds

    def ensure(self, next_case: str) -> None:
        if self.total_cpu_seconds >= CPU_CAP_SECONDS:
            raise BudgetKill(
                f"CPU cap reached before case {next_case}: {self.total_cpu_seconds:.6f}s"
            )


def load_drawing_refs() -> list[DrawingRef]:
    refs: list[DrawingRef] = []
    for tier in ("S", "F", "M"):
        manifest_path = PACKS_ROOT / tier / "manifest.json"
        manifest = load_json(manifest_path)
        for entry in manifest.get("files", []):
            refs.append(
                DrawingRef(
                    tier=tier,
                    family_id=str(entry["family_id"]),
                    drawing_id=str(entry["drawing_id"]),
                    dxf_path=(manifest_path.parent / str(entry["dxf"])).resolve(),
                    truth_path=(manifest_path.parent / str(entry["truth"])).resolve(),
                    dxf_sha256=str(entry["dxf_sha256"]),
                    truth_sha256=str(entry["truth_sha256"]),
                )
            )
    refs.sort(key=lambda item: (item.family_id, item.drawing_id, item.tier))
    counts = Counter(ref.family_id for ref in refs)
    if len(refs) != 200 or set(counts) != set(ALL_FAMILIES) or any(value != 25 for value in counts.values()):
        raise RuntimeError(f"unexpected drawing/family cardinality: n={len(refs)}, counts={dict(counts)}")
    if len({ref.drawing_id for ref in refs}) != 200:
        raise RuntimeError("drawing IDs are not unique")
    return refs


def parse_dxf_entities(path: Path) -> list[DxfEntity]:
    lines = read_text(path).splitlines()
    if len(lines) % 2:
        raise ValueError(f"odd ASCII DXF line count: {path} ({len(lines)})")
    pairs: list[tuple[int, str]] = []
    for index in range(0, len(lines), 2):
        try:
            code = int(lines[index].strip())
        except ValueError as exc:
            raise ValueError(f"invalid DXF group code at line {index + 1}: {path}") from exc
        pairs.append((code, lines[index + 1].strip()))

    section: str | None = None
    awaiting_section_name = False
    current_type: str | None = None
    current_tags: list[tuple[int, str]] = []
    entities: list[DxfEntity] = []

    def flush() -> None:
        nonlocal current_type, current_tags
        if current_type is None:
            return
        handle = next((value for code, value in current_tags if code == 5), "")
        layer = next((value for code, value in current_tags if code == 8), "")
        if handle:
            entities.append(
                DxfEntity(current_type.upper(), handle, layer, tuple(current_tags))
            )
        current_type = None
        current_tags = []

    for code, value in pairs:
        if code == 0 and value == "SECTION":
            flush()
            awaiting_section_name = True
            section = None
            continue
        if awaiting_section_name and code == 2:
            section = value.upper()
            awaiting_section_name = False
            continue
        if code == 0 and value == "ENDSEC":
            flush()
            section = None
            continue
        if section != "ENTITIES":
            continue
        if code == 0:
            flush()
            current_type = value.upper()
            current_tags = []
        elif current_type is not None:
            current_tags.append((code, value))
    flush()
    return entities


def _first(entity: DxfEntity, code: int, default: str | None = None) -> str | None:
    return next((value for item_code, value in entity.tags if item_code == code), default)


def _float(entity: DxfEntity, code: int, default: float | None = None) -> float | None:
    raw = _first(entity, code)
    if raw is None:
        return default
    try:
        value = float(raw)
    except ValueError:
        return default
    return value if math.isfinite(value) else default


def _points(entity: DxfEntity, x_code: int, y_code: int) -> list[tuple[float, float]]:
    points: list[tuple[float, float]] = []
    pending_x: float | None = None
    for code, raw in entity.tags:
        if code == x_code:
            try:
                pending_x = float(raw)
            except ValueError:
                pending_x = None
        elif code == y_code and pending_x is not None:
            try:
                y_value = float(raw)
            except ValueError:
                pending_x = None
                continue
            if math.isfinite(pending_x) and math.isfinite(y_value):
                points.append((pending_x, y_value))
            pending_x = None
    return points


def _point(entity: DxfEntity, x_code: int, y_code: int) -> tuple[float, float] | None:
    points = _points(entity, x_code, y_code)
    return points[0] if points else None


def _distance(a: tuple[float, float], b: tuple[float, float]) -> float:
    return math.hypot(b[0] - a[0], b[1] - a[1])


def _marker(entity: DxfEntity, length: float = 1.0) -> list[tuple[tuple[float, float], tuple[float, float], str]]:
    origin = None
    for x_code, y_code in ((10, 20), (11, 21), (12, 22), (13, 23), (14, 24)):
        origin = _point(entity, x_code, y_code)
        if origin is not None:
            break
    digest = hashlib.sha256(f"{entity.entity_type}|{entity.handle}".encode("utf-8")).digest()
    if origin is None:
        origin = (10_000_000.0 + int.from_bytes(digest[:2], "big") * 10.0,
                  10_000_000.0 + int.from_bytes(digest[2:4], "big") * 10.0)
    angle = 2.0 * math.pi * int.from_bytes(digest[4:6], "big") / 65536.0
    safe_length = min(1000.0, max(1.0, abs(float(length))))
    end = (origin[0] + safe_length * math.cos(angle),
           origin[1] + safe_length * math.sin(angle))
    return [(origin, end, f"{entity.entity_type.lower()}-proxy")]


def entity_segments(entity: DxfEntity) -> list[tuple[tuple[float, float], tuple[float, float], str]]:
    kind = entity.entity_type
    segments: list[tuple[tuple[float, float], tuple[float, float], str]] = []

    def add(a: tuple[float, float] | None, b: tuple[float, float] | None, label: str) -> None:
        if a is not None and b is not None and _distance(a, b) > 1e-9:
            segments.append((a, b, label))

    if kind == "LINE":
        add(_point(entity, 10, 20), _point(entity, 11, 21), "line")
    elif kind == "LWPOLYLINE":
        points = _points(entity, 10, 20)
        for left, right in zip(points, points[1:]):
            add(left, right, "poly-edge")
        flags = int(_float(entity, 70, 0.0) or 0)
        if flags & 1 and len(points) > 2:
            add(points[-1], points[0], "poly-edge")
    elif kind == "ARC":
        center = _point(entity, 10, 20)
        radius = _float(entity, 40)
        start = _float(entity, 50)
        end = _float(entity, 51)
        if center is not None and radius is not None and radius > 0 and start is not None and end is not None:
            sweep = (end - start) % 360.0
            if sweep <= 1e-12:
                sweep = 360.0
            chord_count = max(4, int(math.ceil(sweep / 7.5)))
            points = [
                (
                    center[0] + radius * math.cos(math.radians(start + sweep * index / chord_count)),
                    center[1] + radius * math.sin(math.radians(start + sweep * index / chord_count)),
                )
                for index in range(chord_count + 1)
            ]
            for left, right in zip(points, points[1:]):
                add(left, right, "arc-chord")
    elif kind == "CIRCLE":
        center = _point(entity, 10, 20)
        radius = _float(entity, 40)
        if center is not None and radius is not None and radius > 0:
            points = [
                (
                    center[0] + radius * math.cos(2.0 * math.pi * index / 32.0),
                    center[1] + radius * math.sin(2.0 * math.pi * index / 32.0),
                )
                for index in range(33)
            ]
            for left, right in zip(points, points[1:]):
                add(left, right, "circle-chord")
    elif kind == "3DFACE":
        points = [
            point for point in (
                _point(entity, 10, 20),
                _point(entity, 11, 21),
                _point(entity, 12, 22),
                _point(entity, 13, 23),
            ) if point is not None
        ]
        compact: list[tuple[float, float]] = []
        for point in points:
            if not compact or _distance(compact[-1], point) > 1e-9:
                compact.append(point)
        for left, right in zip(compact, compact[1:]):
            add(left, right, "3dface-edge")
        if len(compact) > 2:
            add(compact[-1], compact[0], "3dface-edge")
    elif kind == "ELLIPSE":
        center = _point(entity, 10, 20)
        major = _point(entity, 11, 21)
        ratio = _float(entity, 40, 1.0)
        start = _float(entity, 41, 0.0)
        end = _float(entity, 42, 2.0 * math.pi)
        if center is not None and major is not None and ratio is not None and start is not None and end is not None:
            major_length = math.hypot(major[0], major[1])
            if major_length > 1e-9 and ratio > 0:
                sweep = (end - start) % (2.0 * math.pi)
                if sweep <= 1e-12:
                    sweep = 2.0 * math.pi
                count = max(8, int(math.ceil(48.0 * sweep / (2.0 * math.pi))))
                ux, uy = major[0] / major_length, major[1] / major_length
                vx, vy = -uy * major_length * ratio, ux * major_length * ratio
                points = []
                for index in range(count + 1):
                    param = start + sweep * index / count
                    points.append((
                        center[0] + major[0] * math.cos(param) + vx * math.sin(param),
                        center[1] + major[1] * math.cos(param) + vy * math.sin(param),
                    ))
                for left, right in zip(points, points[1:]):
                    add(left, right, "ellipse-chord")
    elif kind == "HATCH":
        points = _points(entity, 10, 20)
        for left, right in zip(points[:65], points[1:65]):
            add(left, right, "hatch-boundary")
        if len(points) > 2:
            add(points[min(len(points), 65) - 1], points[0], "hatch-boundary")
    elif kind == "INSERT":
        origin = _point(entity, 10, 20)
        if origin is not None:
            rotation = math.radians(_float(entity, 50, 0.0) or 0.0)
            scale = max(abs(_float(entity, 41, 1.0) or 1.0), abs(_float(entity, 42, 1.0) or 1.0))
            add(origin, (origin[0] + scale * math.cos(rotation), origin[1] + scale * math.sin(rotation)), "insert-marker")
    elif kind in {"TEXT", "MTEXT"}:
        origin = _point(entity, 10, 20)
        height = _float(entity, 40, 1.0) or 1.0
        direction = _point(entity, 11, 21) if kind == "MTEXT" else None
        if origin is not None:
            if direction is not None and math.hypot(direction[0], direction[1]) > 1e-9:
                norm = math.hypot(direction[0], direction[1])
                ux, uy = direction[0] / norm, direction[1] / norm
            else:
                angle = math.radians(_float(entity, 50, 0.0) or 0.0)
                ux, uy = math.cos(angle), math.sin(angle)
            length = min(1000.0, max(1.0, abs(height)))
            add(origin, (origin[0] + length * ux, origin[1] + length * uy), f"{kind.lower()}-baseline")
    elif kind == "POINT":
        origin = _point(entity, 10, 20)
        if origin is not None:
            add(origin, (origin[0] + 1.0, origin[1]), "point-marker")
    elif kind == "SPLINE":
        points = _points(entity, 10, 20)
        if len(points) < 2:
            points = _points(entity, 11, 21)
        for left, right in zip(points[:65], points[1:65]):
            add(left, right, "spline-control-chain")
    elif kind == "WIPEOUT":
        points = _points(entity, 14, 24)
        for left, right in zip(points[:65], points[1:65]):
            add(left, right, "wipeout-boundary")
        if len(points) > 2:
            add(points[min(len(points), 65) - 1], points[0], "wipeout-boundary")

    return segments or _marker(entity, _float(entity, 40, 1.0) or 1.0)


def wallish_layer(layer: str) -> bool:
    normalized = "".join(character if character.isalnum() else " " for character in (layer or "").upper())
    return bool(set(normalized.split()) & {"WALL", "WALLS", "WA", "BEARING", "벽"})


def rank_hash(token: str) -> str:
    return sha256_text(f"g1_verifier_families|{SEED}|{token}")


def rank_choice(values: Iterable[Any], token: str, stringify=lambda item: str(item)) -> Any:
    materialized = list(values)
    if not materialized:
        raise ValueError(f"no candidates for deterministic choice: {token}")
    return min(materialized, key=lambda item: (rank_hash(f"{token}|{stringify(item)}"), stringify(item)))


def natural_handle_key(value: str) -> tuple[int, int | str]:
    try:
        return (0, int(value, 16))
    except ValueError:
        return (1, value)


def canonical_handles(values: Iterable[Any]) -> list[str]:
    return sorted({str(value) for value in values}, key=natural_handle_key)


def select_representatives(
    entities: Sequence[DxfEntity], truth_wall_handles: set[str], drawing_id: str
) -> tuple[dict[str, DxfEntity], list[dict[str, Any]]]:
    representatives: dict[str, DxfEntity] = {}
    records: list[dict[str, Any]] = []
    for entity_type in ENTITY_TYPES:
        all_candidates = [
            entity for entity in entities
            if entity.entity_type == entity_type and entity.handle not in truth_wall_handles
        ]
        preferred = [entity for entity in all_candidates if not wallish_layer(entity.layer)]
        pool = preferred or all_candidates
        chosen = rank_choice(
            pool,
            f"{drawing_id}|representative|{entity_type}",
            stringify=lambda item: f"{item.handle}|{item.layer}|{item.entity_type}",
        )
        representatives[entity_type] = chosen
        records.append({
            "entity_type": entity_type,
            "handle": chosen.handle,
            "layer": chosen.layer,
            "nonwallish_layer": not wallish_layer(chosen.layer),
            "preferred_pool_used": bool(preferred),
            "segment_count": len(entity_segments(chosen)),
        })
    return representatives, records


def build_compact_ir(
    entities: Sequence[DxfEntity], truth: Mapping[str, Any]
) -> tuple[dict[str, Any], dict[str, DxfEntity], list[dict[str, Any]], list[str]]:
    drawing_id = str(truth["drawing_id"])
    truth_wall_handles = set(map(str, truth.get("wall_handles_flat", [])))
    class_handles = set(map(str, truth.get("class_of_handle", {}).keys()))
    representatives, representative_records = select_representatives(
        entities, truth_wall_handles, drawing_id
    )
    wanted = class_handles | truth_wall_handles | {entity.handle for entity in representatives.values()}
    by_handle: dict[str, list[DxfEntity]] = defaultdict(list)
    for entity in entities:
        by_handle[entity.handle].append(entity)
    missing_wanted = sorted(wanted - set(by_handle), key=natural_handle_key)
    missing_walls = sorted(truth_wall_handles - set(by_handle), key=natural_handle_key)
    if missing_walls:
        raise RuntimeError(f"truth wall handles absent from DXF {drawing_id}: {missing_walls}")

    segments: list[dict[str, Any]] = []
    for handle in sorted(wanted & set(by_handle), key=natural_handle_key):
        for entity in by_handle[handle]:
            for p0, p1, kind in entity_segments(entity):
                if _distance(p0, p1) <= 1e-9:
                    continue
                segments.append({
                    "sid": f"s{len(segments) + 1:07d}",
                    "handle": handle,
                    "pts": [[float(p0[0]), float(p0[1])], [float(p1[0]), float(p1[1])]],
                    "layer": entity.layer,
                    "kind": entity.entity_type,
                    "geometry_adapter_kind": kind,
                    "label": "unknown",
                    "source": "gen2_family_dxf",
                })
    if not segments:
        raise RuntimeError(f"compact SEG-IR is empty: {drawing_id}")
    return ({
        "ir": "seg.v1",
        "drawing_id": drawing_id,
        "units": "mm",
        "scale_mm_per_unit": 1.0,
        "segments": segments,
    }, representatives, representative_records, missing_wanted)


def per_drawing_case_plan(drawing_id: str) -> list[dict[str, Any]]:
    plan: list[dict[str, Any]] = []
    for entity_type in ENTITY_TYPES:
        plan.append({
            "case_id": f"{drawing_id}::true_exact::{entity_type}",
            "truth_class": "true",
            "case_kind": "true_exact",
            "entity_type": entity_type,
        })
    plan.append({
        "case_id": f"{drawing_id}::label_poisoned_true",
        "truth_class": "true",
        "case_kind": "label_poisoned_true",
        "entity_type": None,
    })
    for entity_type in ENTITY_TYPES:
        plan.append({
            "case_id": f"{drawing_id}::entity_add::{entity_type}",
            "truth_class": "false",
            "case_kind": "entity_add",
            "entity_type": entity_type,
        })
    for name in ORIGINAL_PERTURBATIONS + SENTINELS:
        plan.append({
            "case_id": f"{drawing_id}::{name}",
            "truth_class": "false",
            "case_kind": name,
            "entity_type": None,
        })
    return plan


def run_selftests(verifier_module=None, *, include_inventory: bool = True) -> dict[str, Any]:
    started = time.process_time()
    checks: dict[str, bool] = {}
    details: dict[str, Any] = {}

    first_plan = per_drawing_case_plan("selftest_001")
    second_plan = per_drawing_case_plan("selftest_001")
    first_hash = sha256_text(canonical_json(first_plan))
    second_hash = sha256_text(canonical_json(second_plan))
    checks["battery_generation_deterministic"] = first_hash == second_hash
    checks["battery_cases_per_drawing"] = len(first_plan) == 37
    details["case_plan_sha256"] = first_hash
    details["case_count"] = len(first_plan)

    forbidden_examples = {
        "original_cad": Path(r"D:\original_cad\forbidden_sample.dwg"),
        "repository_test": Path(r"D:\dev\99_tools\autocad-sdk-router\tests\forbidden.py"),
        "cubic_base_test": Path(r"D:\datasets\CubiCasa5k\test\forbidden.json"),
        "val_b": Path(r"D:\runs\e2_program\cells\w2_09_valb\forbidden.json"),
    }
    denied: dict[str, bool] = {}
    for name, path in forbidden_examples.items():
        try:
            assert_read_allowed(path)
        except BoundaryViolation:
            denied[name] = True
        else:
            denied[name] = False
    checks["forbidden_read_paths_denied"] = all(denied.values())
    details["forbidden_path_denials"] = denied
    checks["allowed_pack_path_accepted"] = (
        assert_read_allowed(PACKS_ROOT / "S" / "manifest.json")
        == (PACKS_ROOT / "S" / "manifest.json").resolve()
    )
    checks["allowed_verifier_path_accepted"] = assert_read_allowed(VERIFIER_PATH) == VERIFIER_PATH
    checks["verifier_hash_sealed"] = file_sha256(VERIFIER_PATH) == VERIFIER_SHA256
    checks["prereg_hash_sealed"] = file_sha256(PREREG_PATH) == PREREG_SHA256
    checks["prereg_csv_hash_sealed"] = file_sha256(PREREG_CSV_PATH) == PREREG_CSV_SHA256
    if include_inventory:
        inventory = pack_inventory()
        checks["pack_inventory_sealed"] = (
            inventory["file_count"] == PACK_INVENTORY_FILE_COUNT
            and inventory["sha256"] == PACK_INVENTORY_SHA256
        )
        details["pack_inventory"] = inventory

    module = verifier_module or import_verifier()
    verifier_ok, verifier_lines = module.run_selftest(emit=False)
    checks["unchanged_verifier_selftest"] = bool(verifier_ok)
    details["unchanged_verifier_selftest_lines"] = verifier_lines
    checks["hidden_content_reads_before_freeze_zero"] = (
        FIREWALL_AUDIT["hidden_content_reads_before_freeze"] == 0
    )
    return {
        "schema": SELFTEST_SCHEMA,
        "selftest_ok": all(checks.values()),
        "checks": checks,
        "details": details,
        "cpu_seconds": time.process_time() - started,
    }


def case_record(
    *, ref: DrawingRef, phase: str, case_id: str, truth_class: str,
    case_kind: str, entity_type: str | None, result: Mapping[str, Any],
    fallback: str | None = None, source_handle: str | None = None,
) -> dict[str, Any]:
    accepted = bool(result["accepted"])
    is_error = (truth_class == "true" and not accepted) or (
        truth_class == "false" and accepted
    )
    reasons = list(result.get("reason_codes", []))
    if is_error:
        if truth_class == "true":
            harness_cause = "true claim rejected: " + (", ".join(reasons) or "no reason code")
        else:
            harness_cause = "false claim accepted without a rejecting check"
    else:
        harness_cause = None
    return {
        "case_id": case_id,
        "phase": phase,
        "family_id": ref.family_id,
        "tier": ref.tier,
        "drawing_id": ref.drawing_id,
        "truth_class": truth_class,
        "case_kind": case_kind,
        "entity_type": entity_type,
        "source_handle": source_handle,
        "accepted": accepted,
        "rejected": not accepted,
        "is_error": is_error,
        "reason_codes": reasons,
        "claim_raw_count": int(result["claim"]["raw_count"]),
        "claim_unique_count": int(result["claim"]["unique_count"]),
        "fallback": fallback,
        "harness_cause": harness_cause,
    }


def build_perturbation_claims(
    *, truth: Mapping[str, Any], model: Mapping[str, Any], drawing_id: str,
) -> tuple[dict[str, list[str]], dict[str, str | None], dict[str, str | None]]:
    true_handles = set(map(str, truth["wall_handles_flat"]))
    negatives = set(model["universe_handles"]) - true_handles
    if not true_handles or len(negatives) < 2:
        raise RuntimeError(f"insufficient true/negative handles: {drawing_id}")
    removed = rank_choice(true_handles, f"{drawing_id}|removed_handle")
    truth_walls = [
        [str(value) for value in wall.get("handles", []) if str(value) in true_handles]
        for wall in truth.get("walls", [])
        if wall.get("handles")
    ]
    truth_walls = [wall for wall in truth_walls if wall]
    selected_wall = rank_choice(
        truth_walls,
        f"{drawing_id}|selected_wall",
        stringify=lambda values: ",".join(canonical_handles(values)),
    )
    support_counts = {
        handle: len(set(model["support"].get(handle, set())) & negatives)
        for handle in negatives
    }
    max_support = max(support_counts.values())
    lure = rank_choice(
        [handle for handle, count in support_counts.items() if count == max_support],
        f"{drawing_id}|lure",
    )
    removed_centroid = model["centroids"][removed]
    nearest_distance = min(
        _distance(removed_centroid, model["centroids"][handle]) for handle in negatives
    )
    nearest = [
        handle for handle in negatives
        if abs(_distance(removed_centroid, model["centroids"][handle]) - nearest_distance) <= 1e-9
    ]
    neighbor = rank_choice(nearest, f"{drawing_id}|neighbor")
    negative_pairs = [
        tuple(record["handles"])
        for record in model["all_pairs"]
        if set(record["handles"]) <= negatives
    ]
    fallbacks: dict[str, str | None] = {name: None for name in ORIGINAL_PERTURBATIONS}
    if negative_pairs:
        negative_pair = rank_choice(
            negative_pairs,
            f"{drawing_id}|negative_pair",
            stringify=lambda values: ",".join(canonical_handles(values)),
        )
    else:
        negative_pair = tuple(
            sorted(negatives, key=lambda handle: (rank_hash(f"{drawing_id}|pair_fallback|{handle}"), handle))[:2]
        )
        fallbacks["pair_swap"] = "no qualifying negative pair; used seed-ranked two-negative fallback"
    orphan_candidates = [
        handle for handle in negatives
        if not (set(model["support"].get(handle, set())) & true_handles)
    ]
    if orphan_candidates:
        orphan = rank_choice(orphan_candidates, f"{drawing_id}|orphan")
    else:
        min_support = min(
            len(set(model["support"].get(handle, set())) & true_handles) for handle in negatives
        )
        orphan = rank_choice(
            [
                handle for handle in negatives
                if len(set(model["support"].get(handle, set())) & true_handles) == min_support
            ],
            f"{drawing_id}|orphan_fallback",
        )
        fallbacks["orphan_add"] = f"no zero-support negative; used minimum-support fallback ({min_support})"

    universe = canonical_handles(model["universe_handles"])
    ordered_true = canonical_handles(true_handles)
    claims = {
        "wall_remove_single": canonical_handles(true_handles - {removed}),
        "wall_remove_pair": canonical_handles(true_handles - set(selected_wall)),
        "lure_add": canonical_handles(true_handles | {lure}),
        "neighbor_swap": canonical_handles((true_handles - {removed}) | {neighbor}),
        "pair_swap": canonical_handles((true_handles - set(selected_wall)) | set(negative_pair)),
        "orphan_add": canonical_handles(true_handles | {orphan}),
        "empty_claim": [],
        "whole_universe_claim": universe,
        "duplicate_true_handle": ordered_true + [ordered_true[0]],
        "unknown_handle_add": canonical_handles(true_handles | {f"G1_UNKNOWN_{drawing_id}"}),
    }
    source_handles = {
        "wall_remove_single": removed,
        "wall_remove_pair": ";".join(canonical_handles(selected_wall)),
        "lure_add": lure,
        "neighbor_swap": f"{removed}->{neighbor}",
        "pair_swap": ";".join(canonical_handles(negative_pair)),
        "orphan_add": orphan,
        "empty_claim": None,
        "whole_universe_claim": None,
        "duplicate_true_handle": ordered_true[0],
        "unknown_handle_add": f"G1_UNKNOWN_{drawing_id}",
    }
    return claims, fallbacks, source_handles


def evaluate_drawing(ref: DrawingRef, phase: str, verifier_module, budget: CpuBudget) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    budget.ensure(f"{ref.drawing_id}::load")
    if file_sha256(ref.truth_path) != ref.truth_sha256:
        raise RuntimeError(f"truth SHA mismatch: {ref.truth_path}")
    if file_sha256(ref.dxf_path) != ref.dxf_sha256:
        raise RuntimeError(f"DXF SHA mismatch: {ref.dxf_path}")
    truth = load_json(ref.truth_path)
    if str(truth.get("family_id")) != ref.family_id or str(truth.get("drawing_id")) != ref.drawing_id:
        raise RuntimeError(f"manifest/truth identity mismatch: {ref.drawing_id}")
    entities = parse_dxf_entities(ref.dxf_path)
    raw_mix = Counter(entity.entity_type for entity in entities)
    truth_mix = {str(key): int(value) for key, value in truth.get("entity_mix", {}).items()}
    if set(raw_mix) != set(ENTITY_TYPES) or dict(raw_mix) != truth_mix:
        raise RuntimeError(
            f"raw entity mix mismatch: {ref.drawing_id}; raw={dict(raw_mix)}, truth={truth_mix}"
        )
    ir, representatives, representative_records, missing_wanted = build_compact_ir(entities, truth)
    true_handles = canonical_handles(truth["wall_handles_flat"])
    model = verifier_module.analyze_seg_ir(
        ir,
        angle_tol_deg=CONFIG["angle_tol_deg"],
        overlap_min=CONFIG["overlap_min"],
        thickness_band_mm=tuple(CONFIG["thickness_band_mm"]),
    )

    cases: list[dict[str, Any]] = []
    budget.ensure(f"{ref.drawing_id}::true_exact")
    clean_result = verifier_module.verify_claim(ir, true_handles, analysis=model)
    for entity_type in ENTITY_TYPES:
        cases.append(case_record(
            ref=ref,
            phase=phase,
            case_id=f"{ref.drawing_id}::true_exact::{entity_type}",
            truth_class="true",
            case_kind="true_exact",
            entity_type=entity_type,
            result=clean_result,
            source_handle=representatives[entity_type].handle,
        ))

    budget.ensure(f"{ref.drawing_id}::label_poisoned_true")
    poisoned = copy.deepcopy(ir)
    true_set = set(true_handles)
    for segment in poisoned["segments"]:
        segment["label"] = "other" if str(segment["handle"]) in true_set else "wall"
    poisoned_model = verifier_module.analyze_seg_ir(
        poisoned,
        angle_tol_deg=CONFIG["angle_tol_deg"],
        overlap_min=CONFIG["overlap_min"],
        thickness_band_mm=tuple(CONFIG["thickness_band_mm"]),
    )
    poisoned_result = verifier_module.verify_claim(poisoned, true_handles, analysis=poisoned_model)
    cases.append(case_record(
        ref=ref,
        phase=phase,
        case_id=f"{ref.drawing_id}::label_poisoned_true",
        truth_class="true",
        case_kind="label_poisoned_true",
        entity_type=None,
        result=poisoned_result,
    ))

    for entity_type in ENTITY_TYPES:
        budget.ensure(f"{ref.drawing_id}::entity_add::{entity_type}")
        representative = representatives[entity_type]
        claim = canonical_handles(set(true_handles) | {representative.handle})
        result = verifier_module.verify_claim(ir, claim, analysis=model)
        cases.append(case_record(
            ref=ref,
            phase=phase,
            case_id=f"{ref.drawing_id}::entity_add::{entity_type}",
            truth_class="false",
            case_kind="entity_add",
            entity_type=entity_type,
            result=result,
            source_handle=representative.handle,
        ))

    perturbations, fallbacks, source_handles = build_perturbation_claims(
        truth=truth, model=model, drawing_id=ref.drawing_id
    )
    for case_kind in ORIGINAL_PERTURBATIONS + SENTINELS:
        budget.ensure(f"{ref.drawing_id}::{case_kind}")
        result = verifier_module.verify_claim(ir, perturbations[case_kind], analysis=model)
        cases.append(case_record(
            ref=ref,
            phase=phase,
            case_id=f"{ref.drawing_id}::{case_kind}",
            truth_class="false",
            case_kind=case_kind,
            entity_type=None,
            result=result,
            fallback=fallbacks.get(case_kind),
            source_handle=source_handles.get(case_kind),
        ))

    if len(cases) != 37:
        raise RuntimeError(f"unexpected per-drawing case count: {ref.drawing_id} -> {len(cases)}")
    coverage = {
        "phase": phase,
        "family_id": ref.family_id,
        "tier": ref.tier,
        "drawing_id": ref.drawing_id,
        "raw_entity_mix": dict(sorted(raw_mix.items())),
        "raw_entity_count": sum(raw_mix.values()),
        "compact_ir_segment_count": len(ir["segments"]),
        "compact_ir_handle_count": len(model["universe_handles"]),
        "truth_wall_handle_count": len(true_handles),
        "verifier_expected_handle_count": len(model["expected_handles"]),
        "representatives": representative_records,
        "missing_requested_handles": missing_wanted,
        "perturbation_fallbacks": {
            key: value for key, value in fallbacks.items() if value is not None
        },
        "family_template_sha256": str(truth.get("family_template_sha256", "")),
        "family_core_signature": str(truth.get("family_core_signature", "")),
    }
    return cases, coverage


def rate_stats(rows: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    true_rows = [row for row in rows if row["truth_class"] == "true"]
    false_rows = [row for row in rows if row["truth_class"] == "false"]
    true_rejected = sum(bool(row["rejected"]) for row in true_rows)
    false_accepted = sum(bool(row["accepted"]) for row in false_rows)
    return {
        "n": len(rows),
        "accepted": sum(bool(row["accepted"]) for row in rows),
        "rejected": sum(bool(row["rejected"]) for row in rows),
        "true_n": len(true_rows),
        "true_accepted": len(true_rows) - true_rejected,
        "true_rejected": true_rejected,
        "frr": true_rejected / len(true_rows) if true_rows else None,
        "false_n": len(false_rows),
        "false_accepted": false_accepted,
        "false_rejected": len(false_rows) - false_accepted,
        "far": false_accepted / len(false_rows) if false_rows else None,
        "error_count": true_rejected + false_accepted,
    }


def grouped_stats(rows: Sequence[Mapping[str, Any]], key: str) -> list[dict[str, Any]]:
    groups: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    for row in rows:
        group_value = row.get(key)
        if group_value is not None:
            groups[str(group_value)].append(row)
    return [
        {key: group_value, **rate_stats(groups[group_value])}
        for group_value in sorted(groups)
    ]


def family_entity_stats(rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[tuple[str, str], list[Mapping[str, Any]]] = defaultdict(list)
    for row in rows:
        if row.get("entity_type") is not None:
            groups[(str(row["family_id"]), str(row["entity_type"]))].append(row)
    return [
        {"family_id": family, "entity_type": entity_type, **rate_stats(groups[(family, entity_type)])}
        for family, entity_type in sorted(groups)
    ]


def case_kind_stats(rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[tuple[str, str], list[Mapping[str, Any]]] = defaultdict(list)
    for row in rows:
        groups[(str(row["case_kind"]), str(row["truth_class"]))].append(row)
    return [
        {"case_kind": case_kind, "truth_class": truth_class, **rate_stats(groups[(case_kind, truth_class)])}
        for case_kind, truth_class in sorted(groups)
    ]


def coverage_stats(coverage_rows: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    raw_total: Counter[str] = Counter()
    raw_by_family: dict[str, Counter[str]] = defaultdict(Counter)
    fallbacks: list[dict[str, Any]] = []
    missing: list[dict[str, Any]] = []
    nonpreferred_representatives: list[dict[str, Any]] = []
    family_hashes: dict[str, dict[str, str]] = {}
    for row in coverage_rows:
        family = str(row["family_id"])
        raw_total.update(row["raw_entity_mix"])
        raw_by_family[family].update(row["raw_entity_mix"])
        family_hashes[family] = {
            "family_template_sha256": str(row["family_template_sha256"]),
            "family_core_signature": str(row["family_core_signature"]),
        }
        if row["missing_requested_handles"]:
            missing.append({
                "drawing_id": row["drawing_id"],
                "handles": row["missing_requested_handles"],
            })
        if row["perturbation_fallbacks"]:
            fallbacks.append({
                "drawing_id": row["drawing_id"],
                "fallbacks": row["perturbation_fallbacks"],
            })
        for representative in row["representatives"]:
            if not representative["preferred_pool_used"]:
                nonpreferred_representatives.append({
                    "drawing_id": row["drawing_id"],
                    **representative,
                })
    return {
        "raw_entity_totals": dict(sorted(raw_total.items())),
        "raw_entity_totals_by_family": {
            family: dict(sorted(counter.items()))
            for family, counter in sorted(raw_by_family.items())
        },
        "family_hashes": dict(sorted(family_hashes.items())),
        "missing_requested_handles": missing,
        "perturbation_fallbacks": fallbacks,
        "nonpreferred_representatives": nonpreferred_representatives,
        "compact_ir_segments": {
            "min": min(int(row["compact_ir_segment_count"]) for row in coverage_rows),
            "max": max(int(row["compact_ir_segment_count"]) for row in coverage_rows),
            "sum": sum(int(row["compact_ir_segment_count"]) for row in coverage_rows),
        },
    }


def _table_rows(stats_rows: Sequence[Mapping[str, Any]], leading: Sequence[str]) -> tuple[list[str], list[list[Any]]]:
    columns = list(leading) + [
        "true_n", "true_accepted", "true_rejected", "frr",
        "false_n", "false_accepted", "false_rejected", "far", "n", "error_count",
    ]
    return columns, [[row.get(column) for column in columns] for row in stats_rows]


def build_evidence_workbook(results: Mapping[str, Any]) -> dict[str, Any]:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.workbook.properties import CalcProperties

    output = assert_write_allowed(EVIDENCE_PATH)
    temp_output = assert_write_allowed(CELL_DIR / "_evidence_build.tmp.xlsx")
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)

    navy = "17324D"
    teal = "0F6B78"
    light = "E8F1F5"
    pale = "F5F8FA"
    white = "FFFFFF"
    red = "FDE9E7"
    thin_gray = Side(style="thin", color="B8C4CC")
    section_border = Border(bottom=Side(style="medium", color=teal))
    table_counter = 0

    def title(sheet, text: str, width: int) -> None:
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, width))
        cell = sheet.cell(1, 1, text)
        cell.font = Font(name="Aptos Display", size=16, bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(vertical="center")
        sheet.row_dimensions[1].height = 26
        sheet.sheet_view.showGridLines = False

    def add_table_sheet(name: str, sheet_title: str, columns: Sequence[str], rows: Sequence[Sequence[Any]], freeze: str = "A4"):
        nonlocal table_counter
        sheet = workbook.create_sheet(name)
        title(sheet, sheet_title, len(columns))
        for index, column in enumerate(columns, 1):
            cell = sheet.cell(3, index, column)
            cell.font = Font(name="Aptos", bold=True, color=white)
            cell.fill = PatternFill("solid", fgColor=teal)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = section_border
        for row_index, row in enumerate(rows, 4):
            for column_index, value in enumerate(row, 1):
                cell = sheet.cell(row_index, column_index, value)
                cell.font = Font(name="Aptos", size=10)
                cell.alignment = Alignment(
                    horizontal="right" if isinstance(value, (int, float)) and not isinstance(value, bool) else "left",
                    vertical="top",
                    wrap_text=isinstance(value, str) and len(value) > 30,
                )
                if row_index % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=pale)
        if rows:
            table_counter += 1
            end_column = sheet.cell(3, len(columns)).column_letter
            table = Table(displayName=f"G1Table{table_counter}", ref=f"A3:{end_column}{3 + len(rows)}")
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
                showRowStripes=True, showColumnStripes=False,
            )
            sheet.add_table(table)
        sheet.freeze_panes = freeze
        for column_index, column in enumerate(columns, 1):
            # Excel table filter buttons consume part of the header cell.  Give
            # every generated header explicit breathing room so labels remain
            # visible in the native workbook render.
            width = min(48, max(14, len(str(column)) + 5))
            for row in list(rows)[:200]:
                value = row[column_index - 1]
                if value is not None:
                    width = min(48, max(width, min(48, len(str(value)) + 2)))
            sheet.column_dimensions[sheet.cell(3, column_index).column_letter].width = width
        return sheet

    with PREREG_CSV_PATH.open("r", encoding="utf-8-sig", newline="") as stream:
        prereg_rows = list(csv.reader(stream))
    prereg_sheet = add_table_sheet(
        "PREREG", "G1 Sealed Preregistration", prereg_rows[0], prereg_rows[1:]
    )
    prereg_sheet.column_dimensions["A"].width = 22
    prereg_sheet.column_dimensions["B"].width = 34
    prereg_sheet.column_dimensions["C"].width = 90

    summary = workbook.create_sheet("Summary")
    title(summary, "G1 Verifier Family-Diversity Measurements", 8)
    summary.sheet_view.showGridLines = False
    seal_rows = [
        ("prereg.json SHA-256", PREREG_SHA256),
        ("PREREG.csv SHA-256", PREREG_CSV_SHA256),
        ("verifier.py SHA-256", VERIFIER_SHA256),
        ("family split SHA-256", SPLIT_SHA256),
        ("configuration SHA-256", CONFIG_SHA256),
        ("CPU-hours", results["cpu"]["total_cpu_hours"]),
        ("numeric signature SHA-256", results["reproducibility"]["numeric_signature_sha256"]),
    ]
    for row_index, (label, value) in enumerate(seal_rows, 3):
        summary.cell(row_index, 1, label).font = Font(name="Aptos", bold=True, color=navy)
        summary.cell(row_index, 2, value).font = Font(name="Aptos")
        summary.cell(row_index, 1).fill = PatternFill("solid", fgColor=light)
    start = 12
    summary_headers = ["scope", "true_n", "true_rejected", "FRR", "false_n", "false_accepted", "FAR", "total_n"]
    for col, header in enumerate(summary_headers, 1):
        cell = summary.cell(start, col, header)
        cell.font = Font(name="Aptos", bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=teal)
        cell.alignment = Alignment(horizontal="center")
    summary_scopes = [
        ("overall", results["measurements"]["overall"]),
        ("reward", next(row for row in results["measurements"]["by_phase"] if row["phase"] == "reward")),
        ("hidden", next(row for row in results["measurements"]["by_phase"] if row["phase"] == "hidden")),
    ]
    for offset, (scope, row) in enumerate(summary_scopes, 1):
        excel_row = start + offset
        values = [scope, row["true_n"], row["true_rejected"], None,
                  row["false_n"], row["false_accepted"], None, None]
        for col, value in enumerate(values, 1):
            summary.cell(excel_row, col, value)
        summary.cell(excel_row, 4, f"=IF(B{excel_row}=0,0,C{excel_row}/B{excel_row})")
        summary.cell(excel_row, 7, f"=IF(E{excel_row}=0,0,F{excel_row}/E{excel_row})")
        summary.cell(excel_row, 8, f"=B{excel_row}+E{excel_row}")
        summary.cell(excel_row, 4).number_format = "0.000000000"
        summary.cell(excel_row, 7).number_format = "0.000000000"
    summary.freeze_panes = "A12"
    summary.column_dimensions["A"].width = 31
    summary.column_dimensions["B"].width = 68
    for letter in "CDEFGH":
        summary.column_dimensions[letter].width = 16

    family_columns, family_rows = _table_rows(results["measurements"]["by_family"], ["family_id"])
    family_sheet = add_table_sheet("Family", "Per-Family FAR and FRR", family_columns, family_rows)
    entity_columns, entity_rows = _table_rows(results["measurements"]["by_entity_type"], ["entity_type"])
    entity_sheet = add_table_sheet("Entity", "Per-Entity-Type FAR and FRR", entity_columns, entity_rows)
    family_entity_columns, family_entity_rows = _table_rows(
        results["measurements"]["by_family_entity_type"], ["family_id", "entity_type"]
    )
    family_entity_sheet = add_table_sheet(
        "FamilyEntity", "Family by Entity-Type Decomposition", family_entity_columns, family_entity_rows
    )
    perturb_columns, perturb_rows = _table_rows(
        results["measurements"]["by_case_kind"], ["case_kind", "truth_class"]
    )
    perturb_sheet = add_table_sheet(
        "Perturbation", "Claim-Family Measurements", perturb_columns, perturb_rows
    )

    rate_column_names = {"frr", "far"}
    for sheet in (family_sheet, entity_sheet, family_entity_sheet, perturb_sheet):
        headers = {str(cell.value).lower(): cell.column for cell in sheet[3]}
        for header in rate_column_names:
            if header in headers:
                for row_index in range(4, sheet.max_row + 1):
                    sheet.cell(row_index, headers[header]).number_format = "0.000000000"

    failure_columns = [
        "case_id", "phase", "family_id", "tier", "drawing_id", "truth_class",
        "case_kind", "entity_type", "source_handle", "accepted", "reason_codes",
        "fallback", "harness_cause",
    ]
    failure_rows = [
        [
            row.get(column) if column != "reason_codes" else ";".join(row.get("reason_codes", []))
            for column in failure_columns
        ]
        for row in results["failures"]
    ]
    failures_sheet = add_table_sheet(
        "Failures", "Exhaustive False-Accept / True-Reject Ledger", failure_columns, failure_rows
    )
    for row_index in range(4, failures_sheet.max_row + 1):
        for cell in failures_sheet[row_index]:
            cell.fill = PatternFill("solid", fgColor=red)

    case_columns = [
        "case_id", "phase", "family_id", "tier", "drawing_id", "truth_class",
        "case_kind", "entity_type", "source_handle", "accepted", "rejected",
        "is_error", "reason_codes", "claim_raw_count", "claim_unique_count", "fallback",
    ]
    case_rows = [
        [
            row.get(column) if column != "reason_codes" else ";".join(row.get("reason_codes", []))
            for column in case_columns
        ]
        for row in results["cases"]
    ]
    add_table_sheet("Cases", "Complete 7,400-Case Evidence Ledger", case_columns, case_rows)

    coverage_columns = ["family_id", "entity_type", "raw_entity_count"]
    coverage_rows: list[list[Any]] = []
    for family, entity_counts in results["coverage"]["raw_entity_totals_by_family"].items():
        for entity_type, count in entity_counts.items():
            coverage_rows.append([family, entity_type, count])
    add_table_sheet("SourceCoverage", "Raw Source Entity Coverage", coverage_columns, coverage_rows)

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.000000000" if abs(cell.value) < 1 else "#,##0.000000"
                elif isinstance(cell.value, int):
                    cell.number_format = "#,##0"
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.auto_filter.ref = sheet.auto_filter.ref

    workbook.save(temp_output)
    os.replace(temp_output, output)

    loaded = load_workbook(output, data_only=False, read_only=False)
    formula_count = 0
    formula_errors: list[str] = []
    dimensions: dict[str, str] = {}
    for sheet in loaded.worksheets:
        dimensions[sheet.title] = sheet.calculate_dimension()
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
                    if any(token in cell.value for token in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")):
                        formula_errors.append(f"{sheet.title}!{cell.coordinate}:{cell.value}")
    loaded.close()
    return {
        "path": str(output),
        "sha256": file_sha256(output),
        "sheet_names": workbook.sheetnames,
        "dimensions": dimensions,
        "formula_count": formula_count,
        "formula_error_tokens": formula_errors,
        "file_size_bytes": output.stat().st_size,
    }


def fmt_rate(value: Any) -> str:
    return "n/a" if value is None else f"{float(value):.9f}"


def markdown_stats_table(rows: Sequence[Mapping[str, Any]], leading: Sequence[str]) -> list[str]:
    headers = list(leading) + ["true n", "true reject", "FRR", "false n", "false accept", "FAR"]
    output = ["| " + " | ".join(headers) + " |", "|" + "|".join("---" for _ in headers) + "|"]
    for row in rows:
        values = [str(row[key]) for key in leading]
        values += [
            str(row["true_n"]), str(row["true_rejected"]), fmt_rate(row["frr"]),
            str(row["false_n"]), str(row["false_accepted"]), fmt_rate(row["far"]),
        ]
        output.append("| " + " | ".join(values) + " |")
    return output


def build_report(results: Mapping[str, Any]) -> str:
    lines: list[str] = [
        "# G1 verifier family-diversity numeric report",
        "",
        "| sealed artifact | SHA-256 |",
        "|---|---|",
        f"| `prereg.json` | `{PREREG_SHA256}` |",
        f"| `PREREG.csv` | `{PREREG_CSV_SHA256}` |",
        "",
        "## Scope and immutable inputs",
        "",
        "- domain: metadata-bearing synthetic packs only; no domain-expansion claim",
        f"- unchanged verifier SHA-256: `{VERIFIER_SHA256}`",
        f"- pack inventory: {PACK_INVENTORY_FILE_COUNT} files; SHA-256 `{PACK_INVENTORY_SHA256}`",
        f"- family split SHA-256: `{SPLIT_SHA256}`",
        f"- verifier configuration SHA-256: `{CONFIG_SHA256}`",
        "- configuration: angle tolerance 5.0 degrees; overlap minimum 0.65; thickness 50.0-400.0 mm; no setting search",
        "- qualification threshold comparison and qualification verdict are intentionally omitted for the orchestrator",
        "",
        "## Battery composition",
        "",
        f"- drawings: {results['battery']['drawings']}; families: {results['battery']['families']}; raw entity types: {len(ENTITY_TYPES)}",
        f"- true claims: {results['battery']['actual_positive']}; false claims: {results['battery']['actual_false']}; total: {results['battery']['actual_total']}",
        "- per drawing: 13 exact true entity-context probes + 1 label-poisoned true probe + 13 entity-add false probes + 6 original perturbations + 4 sentinels",
        "- true claim rule: exact `wall_handles_flat` from the sealed synthetic truth ledger",
        "- false claim rules: entity add, wall remove single/pair, lure add, neighbor swap, pair swap, orphan add, empty, whole universe, duplicate, unknown",
        "- raw entity types: " + ", ".join(f"`{value}`" for value in ENTITY_TYPES),
        "",
        "## Family firewall evidence",
        "",
        "- reward families: " + ", ".join(REWARD_FAMILIES),
        "- hidden families: " + ", ".join(HIDDEN_FAMILIES),
        f"- reward cases completed before freeze: {results['family_firewall']['reward_case_count']}",
        f"- configuration freeze UTC: {results['family_firewall']['freeze_event_utc']}",
        f"- hidden content reads before freeze: {results['family_firewall']['hidden_content_reads_before_freeze']}",
        f"- hidden content reads after freeze: {results['family_firewall']['hidden_content_reads_after_freeze']}",
        f"- pre-freeze hidden hash-only reads: {results['family_firewall']['hidden_hash_only_reads_before_freeze']}",
        "",
        "## Overall and phase measurements",
        "",
    ]
    overall_rows = [{"scope": "overall", **results["measurements"]["overall"]}]
    overall_rows += [{"scope": row["phase"], **row} for row in results["measurements"]["by_phase"]]
    lines.extend(markdown_stats_table(overall_rows, ["scope"]))
    lines += ["", "## Per-family measurements", ""]
    lines.extend(markdown_stats_table(results["measurements"]["by_family"], ["family_id"]))
    lines += ["", "## Per-entity-type measurements", ""]
    lines.extend(markdown_stats_table(results["measurements"]["by_entity_type"], ["entity_type"]))
    lines += ["", "## Claim-family measurements", ""]
    lines.extend(markdown_stats_table(results["measurements"]["by_case_kind"], ["case_kind", "truth_class"]))
    lines += ["", "## Exhaustive measurement-error ledger", ""]
    lines.append(f"- count: {len(results['failures'])}")
    for failure in results["failures"]:
        lines.append(
            "- `{case_id}` family={family_id} tier={tier} truth={truth_class} "
            "kind={case_kind} entity={entity_type} accepted={accepted} reasons={reasons} "
            "cause={cause}".format(
                case_id=failure["case_id"], family_id=failure["family_id"], tier=failure["tier"],
                truth_class=failure["truth_class"], case_kind=failure["case_kind"],
                entity_type=failure.get("entity_type"), accepted=str(failure["accepted"]).lower(),
                reasons=",".join(failure["reason_codes"]) or "none",
                cause=failure["harness_cause"],
            )
        )
    lines += [
        "",
        "## CPU accounting",
        "",
        f"- prior invocation CPU-s: {results['cpu']['prior_cpu_seconds']:.6f}",
        f"- measurement invocation CPU-s: {results['cpu']['run_cpu_seconds']:.6f}",
        f"- cumulative CPU-s: {results['cpu']['total_cpu_seconds']:.6f}",
        f"- cumulative CPU-h: {results['cpu']['total_cpu_hours']:.9f}",
        f"- sealed cap CPU-h: {results['cpu']['cap_cpu_hours']:.3f}",
        "",
        "## Selftest evidence",
        "",
        f"- pre-measurement selftest_ok: {str(results['selftests']['pre']['selftest_ok']).lower()}",
        f"- post-measurement selftest_ok: {str(results['selftests']['post']['selftest_ok']).lower()}",
        f"- deterministic case-plan SHA-256: `{results['selftests']['pre']['details']['case_plan_sha256']}`",
        f"- forbidden-path denials: {canonical_json(results['selftests']['pre']['details']['forbidden_path_denials'])}",
        "",
        "## Evidence workbook",
        "",
        f"- SHA-256: `{results['evidence']['sha256']}`",
        f"- worksheets: {', '.join(results['evidence']['sheet_names'])}",
        f"- formula cells: {results['evidence']['formula_count']}; formula error-token count: {len(results['evidence']['formula_error_tokens'])}",
        "",
        "## Unresolved and interpretation boundaries",
        "",
        f"- missing requested handle records: {len(results['coverage']['missing_requested_handles'])}",
        f"- perturbation fallback drawings: {len(results['coverage']['perturbation_fallbacks'])}",
        f"- nonpreferred entity representative records: {len(results['coverage']['nonpreferred_representatives'])}",
        f"- process exception records: {len(results.get('process_exceptions', []))}",
        "- verifier-reward domain remains metadata-bearing synthetic packs only; name-blind training tracks remain closed",
        "- no original CAD, repository test, CubiCasa test, or val-B content was opened",
        f"- reproducibility wording: {results['reproducibility']['claim_wording']}",
        f"- numeric signature SHA-256: `{results['reproducibility']['numeric_signature_sha256']}`",
    ]
    for exception in results.get("process_exceptions", []):
        lines.append(
            f"- process exception `{exception['id']}`: {exception['description']} "
            f"(mutation={str(exception['mutation']).lower()})"
        )
    lines += ["", "CELL_COMPLETE: g1_verifier_families"]
    return "\n".join(lines) + "\n"


def run_measurement(prior_cpu_seconds: float) -> dict[str, Any]:
    budget = CpuBudget(prior_cpu_seconds)
    inputs_before = validate_seals_and_inputs()
    verifier_module = import_verifier()
    pre_selftest = run_selftests(verifier_module, include_inventory=False)
    if not pre_selftest["selftest_ok"]:
        raise RuntimeError("pre-measurement selftest boolean conjunction is false")
    refs = load_drawing_refs()
    cases: list[dict[str, Any]] = []
    coverage_rows: list[dict[str, Any]] = []
    reward_refs = [ref for ref in refs if ref.family_id in REWARD_FAMILIES]
    hidden_refs = [ref for ref in refs if ref.family_id in HIDDEN_FAMILIES]

    for index, ref in enumerate(reward_refs, 1):
        drawing_cases, drawing_coverage = evaluate_drawing(ref, "reward", verifier_module, budget)
        cases.extend(drawing_cases)
        coverage_rows.append(drawing_coverage)
        if index % 25 == 0:
            print(f"reward drawings {index}/{len(reward_refs)}; cases={len(cases)}; cpu_s={budget.total_cpu_seconds:.3f}")

    reward_case_count = len(cases)
    if reward_case_count != 125 * 37:
        raise RuntimeError(f"reward case count mismatch before freeze: {reward_case_count}")
    FIREWALL_AUDIT["configuration_frozen"] = True
    FIREWALL_AUDIT["freeze_event_utc"] = utc_now()

    for index, ref in enumerate(hidden_refs, 1):
        drawing_cases, drawing_coverage = evaluate_drawing(ref, "hidden", verifier_module, budget)
        cases.extend(drawing_cases)
        coverage_rows.append(drawing_coverage)
        if index % 25 == 0:
            print(f"hidden drawings {index}/{len(hidden_refs)}; cases={len(cases)}; cpu_s={budget.total_cpu_seconds:.3f}")

    post_selftest = run_selftests(verifier_module, include_inventory=False)
    if not post_selftest["selftest_ok"]:
        raise RuntimeError("post-measurement selftest boolean conjunction is false")
    inputs_after = validate_seals_and_inputs()

    positive_count = sum(row["truth_class"] == "true" for row in cases)
    false_count = sum(row["truth_class"] == "false" for row in cases)
    if (positive_count, false_count, len(cases)) != (EXPECTED_POSITIVE, EXPECTED_FALSE, EXPECTED_TOTAL):
        raise RuntimeError(
            f"battery cardinality mismatch: true={positive_count}, false={false_count}, total={len(cases)}"
        )
    family_counts = Counter(row["family_id"] for row in cases)
    if set(family_counts) != set(ALL_FAMILIES) or any(value != 925 for value in family_counts.values()):
        raise RuntimeError(f"family case count mismatch: {dict(family_counts)}")

    failures = [row for row in cases if row["is_error"]]
    coverage = coverage_stats(coverage_rows)
    measurements = {
        "overall": rate_stats(cases),
        "by_phase": grouped_stats(cases, "phase"),
        "by_family": grouped_stats(cases, "family_id"),
        "by_entity_type": grouped_stats(cases, "entity_type"),
        "by_family_entity_type": family_entity_stats(cases),
        "by_case_kind": case_kind_stats(cases),
    }
    numeric_payload = {
        "battery": {"true": positive_count, "false": false_count, "total": len(cases)},
        "measurements": measurements,
        "failure_case_ids": [row["case_id"] for row in failures],
        "coverage": coverage,
    }
    results: dict[str, Any] = {
        "schema": SCHEMA,
        "cell": "g1_verifier_families",
        "measured_utc": utc_now(),
        "scope": "metadata-bearing synthetic pack requalification only; no domain expansion",
        "provenance": {
            "prereg_json_sha256": PREREG_SHA256,
            "prereg_csv_sha256": PREREG_CSV_SHA256,
            "script_sha256": file_sha256(Path(__file__).resolve()),
            "verifier_sha256": VERIFIER_SHA256,
            "pack_inventory_before": inputs_before["pack_inventory"],
            "pack_inventory_after": inputs_after["pack_inventory"],
        },
        "family_firewall": {
            "reward_families": list(REWARD_FAMILIES),
            "hidden_families": list(HIDDEN_FAMILIES),
            "split_sha256": SPLIT_SHA256,
            "configuration_sha256": CONFIG_SHA256,
            "configuration": CONFIG,
            "reward_case_count": reward_case_count,
            **FIREWALL_AUDIT,
        },
        "battery": {
            "drawings": 200,
            "families": 8,
            "entity_types": list(ENTITY_TYPES),
            "expected_positive": EXPECTED_POSITIVE,
            "actual_positive": positive_count,
            "expected_false": EXPECTED_FALSE,
            "actual_false": false_count,
            "expected_total": EXPECTED_TOTAL,
            "actual_total": len(cases),
            "rate_definitions": {
                "far": "false claims accepted / false claims",
                "frr": "true claims rejected / true claims",
            },
        },
        "measurements": measurements,
        "failures": failures,
        "failure_count": len(failures),
        "coverage": coverage,
        "coverage_by_drawing": coverage_rows,
        "cases": cases,
        "selftests": {"pre": pre_selftest, "post": post_selftest},
        "cpu": {
            "prior_cpu_seconds": budget.prior_cpu_seconds,
            "run_cpu_seconds": budget.run_cpu_seconds,
            "total_cpu_seconds": budget.total_cpu_seconds,
            "total_cpu_hours": budget.total_cpu_seconds / 3600.0,
            "cap_cpu_hours": CPU_CAP_SECONDS / 3600.0,
            "budget_kill": False,
        },
        "reproducibility": {
            "claim_wording": "휘발 필드 제외 수치 전 필드 동일",
            "numeric_signature_sha256": sha256_text(canonical_json(numeric_payload)),
        },
        "qualification_verdict": None,
        "qualification_verdict_authority": "orchestrator",
    }
    evidence_audit = build_evidence_workbook(results)
    results["evidence"] = evidence_audit
    write_text(RESULTS_PATH, json.dumps(results, ensure_ascii=False, indent=2, sort_keys=True) + "\n")
    report = build_report(results)
    write_text(REPORT_PATH, report)
    return results


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="G1 verifier family-diversity numeric harness")
    parser.add_argument("--selftest", action="store_true", help="run deterministic and read-boundary checks")
    parser.add_argument("--run", action="store_true", help="execute reward then hidden measurements")
    parser.add_argument("--prior-cpu-seconds", type=float, default=0.0,
                        help="CPU seconds consumed by earlier cell-local debug/selftest invocations")
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    if args.selftest == args.run:
        print("ERROR: choose exactly one of --selftest or --run", file=sys.stderr)
        return 2
    if args.selftest:
        result = run_selftests(include_inventory=True)
        print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True))
        return 0 if result["selftest_ok"] else 1
    try:
        result = run_measurement(args.prior_cpu_seconds)
    except BudgetKill as exc:
        print(f"BUDGET_KILL: {exc}", file=sys.stderr)
        return 3
    except Exception as exc:  # numeric evidence command-line boundary
        print(f"ERROR: {type(exc).__name__}: {exc}", file=sys.stderr)
        return 2
    overall = result["measurements"]["overall"]
    print(json.dumps({
        "true_n": overall["true_n"],
        "true_rejected": overall["true_rejected"],
        "frr": overall["frr"],
        "false_n": overall["false_n"],
        "false_accepted": overall["false_accepted"],
        "far": overall["far"],
        "cpu_hours": result["cpu"]["total_cpu_hours"],
        "qualification_verdict": None,
    }, ensure_ascii=False, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
