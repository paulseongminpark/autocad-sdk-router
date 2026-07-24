#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""G9/A63 zero-learning set-assembly enumeration diagnostic.

The harness reads only the sealed metadata-bearing synthetic family packs and
the frozen G1 verifier.  It fits the W2-02 two-hop HistGradientBoosting scorer
on reward families, freezes it, and only then opens hidden-family content.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import heapq
import importlib.util
import io
import json
import math
import os
import pickle
import platform
import re
import sys
import time
import traceback
import warnings
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence


# Keep CPU accounting interpretable and the sklearn fits deterministic.
for _thread_variable in (
    "OMP_NUM_THREADS",
    "OPENBLAS_NUM_THREADS",
    "MKL_NUM_THREADS",
    "VECLIB_MAXIMUM_THREADS",
    "NUMEXPR_NUM_THREADS",
):
    os.environ.setdefault(_thread_variable, "1")

import numpy as np
from sklearn.ensemble import HistGradientBoostingClassifier


SCHEMA = "e2.g9-rl-diag.results.v1"
PREREG_SCHEMA = "e2.g9-rl-diag.prereg.v1"
SELFTEST_SCHEMA = "e2.g9-rl-diag.selftest.v1"
SEED = 20260719
MODEL_SEEDS = (17, 29, 43)
BOOTSTRAP_SEED = 20260719
BOOTSTRAP_REPLICATES = 10_000
BEAM_WIDTHS = (4, 16, 64)
PRIMARY_BEAM_WIDTH = 64
PROBABILITY_THRESHOLD = 0.5
TERMINAL_F1_WEIGHT = 0.5
TERMINAL_VERIFIER_WEIGHT = 0.5
EXACT_MAX_HANDLES = 18
EXACT_MAX_SUBSETS = 1 << EXACT_MAX_HANDLES
EXACT_CPU_SECONDS_PER_DRAWING = 8.0
BNB_NODE_CAP = 100_000
BNB_CPU_SECONDS_PER_DRAWING = 8.0
CPU_CAP_HOURS = 72.0
CPU_CAP_SECONDS = CPU_CAP_HOURS * 3600.0
NUMERIC_TOLERANCE = 1.0e-12

FEATURE_SCALE_MM_PER_PSEUDO_PX = 12.0
FEATURE_RADIUS_PX = 20.0
FEATURE_ANGLE_BINS = 12
FEATURE_ANGLE_TOL_DEG = 5.0
FEATURE_OVERLAP_MIN = 0.65
FEATURE_THICKNESS_BAND_MM = (50.0, 400.0)
FEATURE_THICKNESS_BAND_PX = tuple(
    value / FEATURE_SCALE_MM_PER_PSEUDO_PX for value in FEATURE_THICKNESS_BAND_MM
)
FEATURE_SNAP_PX = 6.0 / FEATURE_SCALE_MM_PER_PSEUDO_PX

VERIFIER_CONFIG = {
    "angle_tol_deg": 5.0,
    "overlap_min": 0.65,
    "thickness_band_mm": [50.0, 400.0],
    "selection": "current_verifier_defaults_no_tuning",
}
VERIFIER_CONFIG_SHA256 = "bf9e4d5c9e7facba3bebf7865bd188b7566f4b319bf1007ab13932bced7d3a07"
VERIFIER_CONFIG_SEAL_TEXT = (
    '{"angle_tol_deg":5.0,"overlap_min":0.65,'
    '"thickness_band_mm":[50.0,400.0],'
    '"selection":"current_verifier_defaults_no_tuning"}'
)
VERIFIER_SHA256 = "72e33ab0e87e96defd00f74c5a22ae6c5cb001c69b740e627edeedaf2a80b690"
PACK_INVENTORY_SHA256 = "14ffd7bc07f52552824d59f723c22cce7ff1c0ad54cf58bbdb6cf72771dc3aea"
PACK_INVENTORY_FILE_COUNT = 404
SPLIT_SHA256 = "0c06b83b44bdf8b3aa5123b910a6a47f4d9ddfbd725523eff5fea132ed814b6d"
REWARD_FAMILIES = ("F03", "F04", "F06", "F07", "F08")
HIDDEN_FAMILIES = ("F01", "F02", "F05")
ALL_FAMILIES = tuple(sorted(REWARD_FAMILIES + HIDDEN_FAMILIES))
ENTITY_TYPES = (
    "3DFACE",
    "ARC",
    "CIRCLE",
    "ELLIPSE",
    "HATCH",
    "INSERT",
    "LINE",
    "LWPOLYLINE",
    "MTEXT",
    "POINT",
    "SPLINE",
    "TEXT",
    "WIPEOUT",
)

BASE_FEATURE_NAMES = (
    "parallel",
    "thickness",
    "junction",
    "log10_len",
    "sin2t",
    "cos2t",
)
CONTEXT_FEATURE_NAMES = (
    "parallel_band_neighbor_count",
    "nearest_parallel_gap_px",
    "junction_degree",
    "drawing_length_percentile",
    "radius_density_r20_per_px2",
    "neighbor_angle_entropy_r20",
)
TWOHOP_FEATURE_NAMES = (
    "junction_neighbor_degree_sum",
    "junction_neighbor_degree_max",
    "junction_neighbor_degree_variance",
    "twohop_parallel_count",
    "twohop_parallel_gap_min_px",
    "twohop_parallel_gap_mean_px",
    "twohop_parallel_gap_std_px",
    "twohop_parallel_gap_max_px",
    "junction_component_size",
    "junction_component_thickness_rate",
    "twohop_ring_count",
    "twohop_length_mean_px",
    "twohop_length_std_px",
    "twohop_length_max_px",
    "twohop_angle_alignment_mean",
    "twohop_angle_alignment_std",
    "twohop_angle_aligned_fraction_2deg",
    "collinear_extension_chain_length_px_le2",
)
FULL_FEATURE_NAMES = BASE_FEATURE_NAMES + CONTEXT_FEATURE_NAMES + TWOHOP_FEATURE_NAMES

CELL_DIR = Path(__file__).resolve().parent
SOURCE_ROOT = Path(r"D:\runs\e2_program\cells\gen2_families").resolve()
PACKS_ROOT = (SOURCE_ROOT / "packs").resolve()
PACKET_PATH = Path(r"D:\runs\e2_program\build\PACKET_g9_rl_diag.md").resolve()
VERIFIER_PATH = Path(
    r"D:\dev\99_tools\autocad-sdk-router\tools\e2\instruments\verifier.py"
).resolve()
G1_REPORT_PATH = Path(
    r"D:\runs\e2_program\cells\g1_verifier_families\REPORT.md"
).resolve()
G1_PREREG_PATH = Path(
    r"D:\runs\e2_program\cells\g1_verifier_families\prereg.json"
).resolve()
G1_PREREG_CSV_PATH = Path(
    r"D:\runs\e2_program\cells\g1_verifier_families\PREREG.csv"
).resolve()
W2_02_DIR = Path(r"D:\runs\e2_program\cells\w2_02_twohop").resolve()
W2_02_PROTOCOL_PATHS = {
    "w2_02_script": (W2_02_DIR / "w2_02_twohop.py").resolve(),
    "w2_02_prereg": (W2_02_DIR / "prereg.json").resolve(),
    "w2_02_results": (W2_02_DIR / "results.json").resolve(),
    "w2_02_report": (W2_02_DIR / "REPORT.md").resolve(),
}
PREREG_R2_PATH = Path(
    r"D:\dev\99_tools\autocad-sdk-router\reports\e2\prereg_r2_v1.json"
).resolve()
AMENDMENT1_PATH = Path(
    r"D:\dev\99_tools\autocad-sdk-router\reports\e2\prereg_r2_v1_amendment1.json"
).resolve()
AMENDMENT2_PATH = Path(
    r"D:\dev\99_tools\autocad-sdk-router\reports\e2\prereg_r2_v1_amendment2.json"
).resolve()
FINAL_PLAN_PATH = Path(
    r"D:\dev\99_tools\autocad-sdk-router\reports\e2\synthesis\FINAL_PROGRAM_PLAN.md"
).resolve()

PREREG_PATH = (CELL_DIR / "prereg.json").resolve()
PREREG_CSV_PATH = (CELL_DIR / "PREREG.csv").resolve()
RESULTS_PATH = (CELL_DIR / "results.json").resolve()
EVIDENCE_PATH = (CELL_DIR / "evidence.xlsx").resolve()
REPORT_PATH = (CELL_DIR / "REPORT.md").resolve()

EXACT_ALLOWED_READS = {
    PACKET_PATH,
    VERIFIER_PATH,
    G1_REPORT_PATH,
    G1_PREREG_PATH,
    G1_PREREG_CSV_PATH,
    PREREG_R2_PATH,
    AMENDMENT1_PATH,
    AMENDMENT2_PATH,
    FINAL_PLAN_PATH,
    *W2_02_PROTOCOL_PATHS.values(),
}


class BoundaryViolation(RuntimeError):
    """Raised before a path outside the sealed read boundary is opened."""


class BudgetKill(RuntimeError):
    """Raised when the preregistered cumulative CPU cap is reached."""


_hidden_content_open = False
_boundary_counters: Counter[str] = Counter()


@dataclass(frozen=True)
class DrawingRef:
    tier: str
    family_id: str
    drawing_id: str
    dxf_path: Path
    truth_path: Path
    dxf_sha256: str
    truth_sha256: str


@dataclass(frozen=True)
class DxfEntity:
    entity_type: str
    handle: str
    layer: str
    tags: tuple[tuple[int, str], ...]


@dataclass
class DrawingData:
    ref: DrawingRef
    truth: dict[str, Any]
    ir: dict[str, Any]
    feature_ir: dict[str, Any]
    handles: list[str]
    features: np.ndarray
    labels: np.ndarray
    duplicate_segment_records_collapsed: int
    representative_records: list[dict[str, Any]]
    missing_wanted_handles: list[str]


class CpuBudget:
    def __init__(self, prior_cpu_seconds: float = 0.0) -> None:
        self.prior_cpu_seconds = max(0.0, float(prior_cpu_seconds))
        self.started = time.process_time()

    @property
    def run_cpu_seconds(self) -> float:
        return time.process_time() - self.started

    @property
    def total_cpu_seconds(self) -> float:
        return self.prior_cpu_seconds + self.run_cpu_seconds

    def ensure(self, stage: str) -> None:
        if self.total_cpu_seconds >= CPU_CAP_SECONDS:
            raise BudgetKill(
                f"CPU cap reached before {stage}: {self.total_cpu_seconds:.6f}s "
                f">= {CPU_CAP_SECONDS:.6f}s"
            )


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def canonical_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def sha256_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def sha256_text(payload: str) -> str:
    return sha256_bytes(payload.encode("utf-8"))


def _norm(path: str | os.PathLike[str]) -> Path:
    return Path(os.path.abspath(os.fspath(path))).resolve(strict=False)


def _under(path: Path, root: Path) -> bool:
    try:
        return os.path.commonpath((str(path), str(root))).casefold() == str(root).casefold()
    except ValueError:
        return False


def _family_from_pack_filename(path: Path) -> str | None:
    match = re.match(
        r"^f(0[1-8])_\d{3}(?:\.truth\.json|\.dxf)$", path.name, re.IGNORECASE
    )
    return f"F{match.group(1)}" if match else None


def assert_read_allowed(
    path: str | os.PathLike[str], *, hash_only: bool = False
) -> Path:
    resolved = _norm(path)
    if _under(resolved, CELL_DIR):
        _boundary_counters["cell_reads"] += 1
        return resolved
    if _under(resolved, SOURCE_ROOT):
        family = _family_from_pack_filename(resolved)
        if family in HIDDEN_FAMILIES and not hash_only and not _hidden_content_open:
            _boundary_counters["hidden_content_denials_before_freeze"] += 1
            raise BoundaryViolation(
                f"hidden-family content read denied before scorer freeze: {resolved}"
            )
        if hash_only:
            _boundary_counters["pack_hash_only_reads"] += 1
        elif family in HIDDEN_FAMILIES:
            _boundary_counters["hidden_content_reads_after_freeze"] += 1
        else:
            _boundary_counters["reward_or_manifest_content_reads"] += 1
        return resolved
    if resolved in EXACT_ALLOWED_READS:
        _boundary_counters["sealed_exact_reads"] += 1
        return resolved
    _boundary_counters["forbidden_denials"] += 1
    raise BoundaryViolation(f"read outside sealed G9 boundary: {resolved}")


def assert_write_allowed(path: str | os.PathLike[str]) -> Path:
    resolved = _norm(path)
    if not _under(resolved, CELL_DIR):
        raise BoundaryViolation(f"write outside G9 cell: {resolved}")
    return resolved


def read_bytes(path: str | os.PathLike[str], *, hash_only: bool = False) -> bytes:
    return assert_read_allowed(path, hash_only=hash_only).read_bytes()


def read_text(path: str | os.PathLike[str]) -> str:
    return read_bytes(path).decode("utf-8-sig")


def load_json(path: str | os.PathLike[str]) -> Any:
    return json.loads(read_text(path))


def file_sha256(path: str | os.PathLike[str], *, hash_only: bool = True) -> str:
    return sha256_bytes(read_bytes(path, hash_only=hash_only))


def json_ready(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): json_ready(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [json_ready(item) for item in value]
    if isinstance(value, np.ndarray):
        return json_ready(value.tolist())
    if isinstance(value, np.generic):
        return value.item()
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, float) and not math.isfinite(value):
        return None
    return value


def write_text(path: str | os.PathLike[str], payload: str) -> None:
    resolved = assert_write_allowed(path)
    resolved.write_text(payload, encoding="utf-8", newline="\n")


def write_json(path: str | os.PathLike[str], value: Any) -> None:
    write_text(
        path,
        json.dumps(json_ready(value), ensure_ascii=False, indent=2, sort_keys=True) + "\n",
    )


def pack_inventory() -> dict[str, Any]:
    files = sorted(
        (path for path in PACKS_ROOT.rglob("*") if path.is_file()),
        key=lambda path: str(path).casefold(),
    )
    lines = [
        f"{path.relative_to(PACKS_ROOT).as_posix()}|{file_sha256(path, hash_only=True)}"
        for path in files
    ]
    return {
        "file_count": len(files),
        "sha256": sha256_text("\n".join(lines) + "\n"),
        "canonical_line_count": len(lines),
    }


def current_input_hashes() -> dict[str, Any]:
    fixed = {
        "packet": PACKET_PATH,
        "verifier": VERIFIER_PATH,
        "g1_report": G1_REPORT_PATH,
        "g1_prereg": G1_PREREG_PATH,
        "g1_prereg_csv": G1_PREREG_CSV_PATH,
        "prereg_r2_v1": PREREG_R2_PATH,
        "amendment1": AMENDMENT1_PATH,
        "amendment2": AMENDMENT2_PATH,
        "final_program_plan": FINAL_PLAN_PATH,
        "g9_script": Path(__file__).resolve(),
        **W2_02_PROTOCOL_PATHS,
    }
    return {
        "files": {name: file_sha256(path) for name, path in fixed.items()},
        "pack_inventory": pack_inventory(),
    }


def frozen_prereg() -> dict[str, Any]:
    inputs = current_input_hashes()
    return {
        "schema": PREREG_SCHEMA,
        "cell": "g9_rl_diag",
        "scope": (
            "zero-learning set-assembly enumeration diagnostic within the sealed "
            "metadata-bearing gen2 synthetic family pack only"
        ),
        "universe": {
            "root": str(SOURCE_ROOT),
            "drawings": 200,
            "families": list(ALL_FAMILIES),
            "pack_inventory_file_count": PACK_INVENTORY_FILE_COUNT,
            "pack_inventory_sha256": PACK_INVENTORY_SHA256,
            "candidate_universe": (
                "every unique valid handle in the G1 compact SEG-IR adapter: all "
                "truth-ledger class handles plus one deterministic non-wall "
                "representative for each of 13 raw entity types"
            ),
            "feature_row_adapter": (
                "one first-valid segment per candidate handle in natural-handle "
                "order; duplicate chord/edge records are collapsed only for scorer "
                "features while the full G1 compact IR is retained for verifier calls"
            ),
        },
        "family_split": {
            "reward_families": list(REWARD_FAMILIES),
            "hidden_families": list(HIDDEN_FAMILIES),
            "split_sha256": SPLIT_SHA256,
            "firewall": (
                "reward DXF/truth content only until scorer fit and fingerprint freeze; "
                "hidden files permit hash-only reads before that event"
            ),
        },
        "scorer": {
            "protocol": "W2-02 full 30-column two-hop classical feature protocol",
            "feature_names": list(FULL_FEATURE_NAMES),
            "coordinate_mapping": (
                "synthetic millimetres divided by 12.0 mm/pseudo-px so frozen W2-02 "
                "20px radius, 6mm snap, and 50-400mm band retain physical meaning"
            ),
            "fit_rows": "reward-family candidate-handle rows only",
            "labels": "synthetic wall_handles_flat membership only",
            "estimator": (
                "sklearn HistGradientBoostingClassifier defaults except random_state"
            ),
            "seeds": list(MODEL_SEEDS),
            "ensemble": "arithmetic mean of three predict_proba wall probabilities",
            "threshold": PROBABILITY_THRESHOLD,
            "freeze_fingerprint": (
                "SHA-256 of protocol-5 estimator pickle for each seed plus train "
                "matrix/label hashes and ordered feature schema; ensemble fingerprint "
                "is canonical SHA-256 over those fields"
            ),
            "post_fit_mutation": "forbidden",
            "GNN": "forbidden",
        },
        "verifier": {
            "path": str(VERIFIER_PATH),
            "sha256": VERIFIER_SHA256,
            "configuration": VERIFIER_CONFIG,
            "configuration_sha256": VERIFIER_CONFIG_SHA256,
            "truth_label_usage": "none inside verifier",
        },
        "terminal_objective": {
            "formula": "0.5 * set_F1(candidate_subset, synthetic_truth) + 0.5 * verifier_accepted",
            "range": [0.0, 1.0],
            "set_F1_empty_rule": "0.0, including empty-empty (not present in this pack)",
            "use": (
                "completed-state comparison only; partial beam states are ranked by "
                "frozen scorer Bernoulli log-likelihood"
            ),
        },
        "policies": {
            "greedy": (
                "sort handles by descending frozen ensemble probability with natural "
                "handle tie-break; include the prefix while p>=0.5 and stop at the "
                "first p<0.5"
            ),
            "beam": (
                "widths 4,16,64; traverse the same sorted include/exclude decision "
                "tree; at each equal-depth frontier retain highest cumulative frozen "
                "Bernoulli log-likelihood, then canonical bitmask; at completion choose "
                "highest terminal objective, then log-likelihood, then canonical set"
            ),
            "beam_widths": list(BEAM_WIDTHS),
            "primary_beam_width": PRIMARY_BEAM_WIDTH,
            "exact": (
                "if handle_count<=18, exhaustively enumerate all 2^n terminal subsets "
                "subject to 262144 subsets and 8 CPU-seconds per drawing"
            ),
            "exact_max_handles": EXACT_MAX_HANDLES,
            "exact_max_subsets": EXACT_MAX_SUBSETS,
            "exact_cpu_seconds_per_drawing": EXACT_CPU_SECONDS_PER_DRAWING,
            "upper_bound_replacement": (
                "otherwise use deterministic best-bound branch-and-bound; admissible "
                "F1 bound includes every remaining true handle and no remaining false "
                "handle, verifier bonus remains possible only while processed decisions "
                "match the verifier-expected set; branch order follows frozen p>=0.5; "
                "cap 100000 nodes and 8 CPU-seconds; report incumbent, bound, gap, and "
                "certification per drawing"
            ),
            "bnb_node_cap": BNB_NODE_CAP,
            "bnb_cpu_seconds_per_drawing": BNB_CPU_SECONDS_PER_DRAWING,
        },
        "metrics": {
            "hidden_only_primary": True,
            "per_policy": [
                "drawing set-F1",
                "family pooled set-F1",
                "hidden pooled set-F1",
                "verifier acceptance rate",
            ],
            "primary_delta_estimand": (
                "arithmetic mean of drawing-level set-F1 differences across 75 hidden drawings"
            ),
            "beam_minus_greedy": "beam64 drawing F1 minus greedy drawing F1",
            "exact_or_bound_minus_beam": (
                "certified exact incumbent drawing F1, otherwise admissible F1 upper "
                "bound 1.0, minus beam64 drawing F1"
            ),
            "greedy_exact_match": (
                "canonical selected-set equality only on drawings with certified optimum"
            ),
            "cpu": "time.process_time cumulative across selftest and measurement invocation",
            "judgment_output": "forbidden; numeric measurements only",
        },
        "bootstrap": {
            "unit": "hidden family cluster",
            "replicates": BOOTSTRAP_REPLICATES,
            "rng_seed": BOOTSTRAP_SEED,
            "resampling": (
                "sample three hidden family IDs with replacement; retain all 25 drawings "
                "for each sampled family and average drawing-level paired deltas"
            ),
            "ci": "percentile 2.5 and 97.5",
        },
        "budget": {
            "resource": "CPU-h",
            "cap_h": CPU_CAP_HOURS,
            "cap_seconds": CPU_CAP_SECONDS,
            "overrun_status": "BUDGET_KILL",
        },
        "selftest": {
            "search_determinism": "repeat fixture greedy/beam/B&B and compare canonical signatures",
            "scorer_freeze": (
                "fit the actual reward matrix twice and require identical estimator and prediction fingerprints"
            ),
            "boundary_denials": [
                "original CAD",
                "repository tests",
                "val-B",
                "CubiCasa",
            ],
        },
        "read_boundary": {
            "allowed": [
                str(SOURCE_ROOT),
                str(VERIFIER_PATH),
                "sealed G1/W2-02/prereg/plan reference files",
                str(CELL_DIR),
            ],
            "forbidden": ["original CAD", "repository tests", "test", "val-B", "CubiCasa"],
        },
        "dependencies": ["Python standard library", "numpy", "sklearn", "openpyxl"],
        "output_policy": {
            "root": str(CELL_DIR),
            "required": [
                "prereg.json",
                "PREREG.csv",
                "g9_rl_diag.py",
                "results.json",
                "evidence.xlsx",
                "REPORT.md",
            ],
            "report_last_line": "CELL_COMPLETE: g9_rl_diag",
        },
        "reproducibility_claim_wording": "휘발 필드 제외 수치 전 필드 동일",
        "input_hashes": inputs,
    }


def _flatten(value: Any, prefix: str = "") -> list[tuple[str, str]]:
    rows: list[tuple[str, str]] = []
    if isinstance(value, dict):
        for key in sorted(value):
            child = f"{prefix}.{key}" if prefix else str(key)
            rows.extend(_flatten(value[key], child))
    elif isinstance(value, list):
        rows.append((prefix, canonical_json(value)))
    else:
        rows.append((prefix, canonical_json(value)))
    return rows


def seal() -> dict[str, Any]:
    if PREREG_PATH.exists() or PREREG_CSV_PATH.exists():
        raise FileExistsError("refusing to replace an existing G9 preregistration seal")
    frozen = frozen_prereg()
    content_hash = sha256_text(canonical_json(frozen))
    payload = {
        "schema": PREREG_SCHEMA,
        "sealed_at_utc": utc_now(),
        "frozen": frozen,
        "content_hash": content_hash,
    }
    write_json(PREREG_PATH, payload)
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerow(["path", "value"])
    writer.writerow(["content_hash", content_hash])
    for key, value in _flatten(frozen):
        writer.writerow([key, value])
    write_text(PREREG_CSV_PATH, buffer.getvalue())
    return {
        "content_hash": content_hash,
        "prereg_json_sha256": file_sha256(PREREG_PATH),
        "prereg_csv_sha256": file_sha256(PREREG_CSV_PATH),
    }


def validate_seals_and_inputs() -> dict[str, Any]:
    prereg = load_json(PREREG_PATH)
    frozen = prereg.get("frozen")
    observed_content_hash = sha256_text(canonical_json(frozen))
    if prereg.get("content_hash") != observed_content_hash:
        raise RuntimeError("prereg.json content hash mismatch")
    csv_text = read_text(PREREG_CSV_PATH)
    csv_rows = list(csv.DictReader(io.StringIO(csv_text)))
    csv_hashes = [row["value"] for row in csv_rows if row["path"] == "content_hash"]
    if csv_hashes != [observed_content_hash]:
        raise RuntimeError("PREREG.csv content hash mismatch")
    current = current_input_hashes()
    if frozen.get("input_hashes") != current:
        raise RuntimeError("sealed input hash drift detected")
    inventory = current["pack_inventory"]
    if inventory != {
        "file_count": PACK_INVENTORY_FILE_COUNT,
        "sha256": PACK_INVENTORY_SHA256,
        "canonical_line_count": PACK_INVENTORY_FILE_COUNT,
    }:
        raise RuntimeError(f"pack inventory drift: {inventory}")
    if current["files"]["verifier"] != VERIFIER_SHA256:
        raise RuntimeError("verifier hash drift")
    if sha256_text(VERIFIER_CONFIG_SEAL_TEXT) != VERIFIER_CONFIG_SHA256:
        raise RuntimeError("verifier configuration hash drift")
    return {
        "content_hash": observed_content_hash,
        "prereg_json_sha256": file_sha256(PREREG_PATH),
        "prereg_csv_sha256": file_sha256(PREREG_CSV_PATH),
        "inputs": current,
    }


def import_verifier() -> Any:
    before = file_sha256(VERIFIER_PATH)
    previous = sys.dont_write_bytecode
    sys.dont_write_bytecode = True
    try:
        spec = importlib.util.spec_from_file_location("g9_sealed_verifier", VERIFIER_PATH)
        if spec is None or spec.loader is None:
            raise ImportError("cannot create verifier import specification")
        module = importlib.util.module_from_spec(spec)
        sys.modules[spec.name] = module
        spec.loader.exec_module(module)
    finally:
        sys.dont_write_bytecode = previous
    after = file_sha256(VERIFIER_PATH)
    if before != VERIFIER_SHA256 or after != VERIFIER_SHA256:
        raise RuntimeError("verifier changed during read-only import")
    return module


def load_drawing_refs() -> list[DrawingRef]:
    refs: list[DrawingRef] = []
    for tier in ("S", "F", "M"):
        manifest_path = PACKS_ROOT / tier / "manifest.json"
        manifest = load_json(manifest_path)
        for entry in manifest.get("files", []):
            refs.append(
                DrawingRef(
                    tier=tier,
                    family_id=str(entry["family_id"]),
                    drawing_id=str(entry["drawing_id"]),
                    dxf_path=(manifest_path.parent / str(entry["dxf"])).resolve(),
                    truth_path=(manifest_path.parent / str(entry["truth"])).resolve(),
                    dxf_sha256=str(entry["dxf_sha256"]),
                    truth_sha256=str(entry["truth_sha256"]),
                )
            )
    refs.sort(key=lambda item: (item.family_id, item.drawing_id, item.tier))
    counts = Counter(ref.family_id for ref in refs)
    if len(refs) != 200 or set(counts) != set(ALL_FAMILIES):
        raise RuntimeError(f"unexpected universe cardinality: {len(refs)}, {dict(counts)}")
    if any(count != 25 for count in counts.values()):
        raise RuntimeError(f"unexpected family cardinality: {dict(counts)}")
    if len({ref.drawing_id for ref in refs}) != 200:
        raise RuntimeError("drawing IDs are not unique")
    return refs


def parse_dxf_entities(path: Path) -> list[DxfEntity]:
    lines = read_text(path).splitlines()
    if len(lines) % 2:
        raise ValueError(f"odd ASCII DXF line count: {path} ({len(lines)})")
    pairs: list[tuple[int, str]] = []
    for index in range(0, len(lines), 2):
        try:
            code = int(lines[index].strip())
        except ValueError as exc:
            raise ValueError(f"invalid DXF group code at line {index + 1}: {path}") from exc
        pairs.append((code, lines[index + 1].strip()))
    section: str | None = None
    awaiting_section_name = False
    current_type: str | None = None
    current_tags: list[tuple[int, str]] = []
    entities: list[DxfEntity] = []

    def flush() -> None:
        nonlocal current_type, current_tags
        if current_type is None:
            return
        handle = next((value for code, value in current_tags if code == 5), "")
        layer = next((value for code, value in current_tags if code == 8), "")
        if handle:
            entities.append(DxfEntity(current_type.upper(), handle, layer, tuple(current_tags)))
        current_type = None
        current_tags = []

    for code, value in pairs:
        if code == 0 and value == "SECTION":
            flush()
            awaiting_section_name = True
            section = None
            continue
        if awaiting_section_name and code == 2:
            section = value.upper()
            awaiting_section_name = False
            continue
        if code == 0 and value == "ENDSEC":
            flush()
            section = None
            continue
        if section != "ENTITIES":
            continue
        if code == 0:
            flush()
            current_type = value.upper()
            current_tags = []
        elif current_type is not None:
            current_tags.append((code, value))
    flush()
    return entities


def _first(entity: DxfEntity, code: int, default: str | None = None) -> str | None:
    return next((value for item_code, value in entity.tags if item_code == code), default)


def _float(entity: DxfEntity, code: int, default: float | None = None) -> float | None:
    raw = _first(entity, code)
    if raw is None:
        return default
    try:
        value = float(raw)
    except ValueError:
        return default
    return value if math.isfinite(value) else default


def _points(entity: DxfEntity, x_code: int, y_code: int) -> list[tuple[float, float]]:
    points: list[tuple[float, float]] = []
    pending_x: float | None = None
    for code, raw in entity.tags:
        if code == x_code:
            try:
                pending_x = float(raw)
            except ValueError:
                pending_x = None
        elif code == y_code and pending_x is not None:
            try:
                y_value = float(raw)
            except ValueError:
                pending_x = None
                continue
            if math.isfinite(pending_x) and math.isfinite(y_value):
                points.append((pending_x, y_value))
            pending_x = None
    return points


def _point(entity: DxfEntity, x_code: int, y_code: int) -> tuple[float, float] | None:
    points = _points(entity, x_code, y_code)
    return points[0] if points else None


def _distance(a: tuple[float, float], b: tuple[float, float]) -> float:
    return math.hypot(b[0] - a[0], b[1] - a[1])


def _marker(
    entity: DxfEntity, length: float = 1.0
) -> list[tuple[tuple[float, float], tuple[float, float], str]]:
    origin = None
    for x_code, y_code in ((10, 20), (11, 21), (12, 22), (13, 23), (14, 24)):
        origin = _point(entity, x_code, y_code)
        if origin is not None:
            break
    digest = hashlib.sha256(f"{entity.entity_type}|{entity.handle}".encode("utf-8")).digest()
    if origin is None:
        origin = (
            10_000_000.0 + int.from_bytes(digest[:2], "big") * 10.0,
            10_000_000.0 + int.from_bytes(digest[2:4], "big") * 10.0,
        )
    angle = 2.0 * math.pi * int.from_bytes(digest[4:6], "big") / 65536.0
    safe_length = min(1000.0, max(1.0, abs(float(length))))
    end = (
        origin[0] + safe_length * math.cos(angle),
        origin[1] + safe_length * math.sin(angle),
    )
    return [(origin, end, f"{entity.entity_type.lower()}-proxy")]


def entity_segments(
    entity: DxfEntity,
) -> list[tuple[tuple[float, float], tuple[float, float], str]]:
    kind = entity.entity_type
    segments: list[tuple[tuple[float, float], tuple[float, float], str]] = []

    def add(
        a: tuple[float, float] | None, b: tuple[float, float] | None, label: str
    ) -> None:
        if a is not None and b is not None and _distance(a, b) > 1e-9:
            segments.append((a, b, label))

    if kind == "LINE":
        add(_point(entity, 10, 20), _point(entity, 11, 21), "line")
    elif kind == "LWPOLYLINE":
        points = _points(entity, 10, 20)
        for left, right in zip(points, points[1:]):
            add(left, right, "poly-edge")
        flags = int(_float(entity, 70, 0.0) or 0)
        if flags & 1 and len(points) > 2:
            add(points[-1], points[0], "poly-edge")
    elif kind == "ARC":
        center = _point(entity, 10, 20)
        radius = _float(entity, 40)
        start = _float(entity, 50)
        end = _float(entity, 51)
        if center is not None and radius and start is not None and end is not None:
            sweep = (end - start) % 360.0
            if sweep <= 1e-12:
                sweep = 360.0
            count = max(4, int(math.ceil(sweep / 7.5)))
            points = [
                (
                    center[0] + radius * math.cos(math.radians(start + sweep * i / count)),
                    center[1] + radius * math.sin(math.radians(start + sweep * i / count)),
                )
                for i in range(count + 1)
            ]
            for left, right in zip(points, points[1:]):
                add(left, right, "arc-chord")
    elif kind == "CIRCLE":
        center = _point(entity, 10, 20)
        radius = _float(entity, 40)
        if center is not None and radius and radius > 0:
            points = [
                (
                    center[0] + radius * math.cos(2.0 * math.pi * i / 32.0),
                    center[1] + radius * math.sin(2.0 * math.pi * i / 32.0),
                )
                for i in range(33)
            ]
            for left, right in zip(points, points[1:]):
                add(left, right, "circle-chord")
    elif kind == "3DFACE":
        points = [
            point
            for point in (
                _point(entity, 10, 20),
                _point(entity, 11, 21),
                _point(entity, 12, 22),
                _point(entity, 13, 23),
            )
            if point is not None
        ]
        compact: list[tuple[float, float]] = []
        for point in points:
            if not compact or _distance(compact[-1], point) > 1e-9:
                compact.append(point)
        for left, right in zip(compact, compact[1:]):
            add(left, right, "3dface-edge")
        if len(compact) > 2:
            add(compact[-1], compact[0], "3dface-edge")
    elif kind == "ELLIPSE":
        center = _point(entity, 10, 20)
        major = _point(entity, 11, 21)
        ratio = _float(entity, 40, 1.0)
        start = _float(entity, 41, 0.0)
        end = _float(entity, 42, 2.0 * math.pi)
        if center is not None and major is not None and ratio and start is not None and end is not None:
            major_length = math.hypot(major[0], major[1])
            if major_length > 1e-9 and ratio > 0:
                sweep = (end - start) % (2.0 * math.pi)
                if sweep <= 1e-12:
                    sweep = 2.0 * math.pi
                count = max(8, int(math.ceil(48.0 * sweep / (2.0 * math.pi))))
                ux, uy = major[0] / major_length, major[1] / major_length
                vx, vy = -uy * major_length * ratio, ux * major_length * ratio
                points = []
                for i in range(count + 1):
                    param = start + sweep * i / count
                    points.append(
                        (
                            center[0] + major[0] * math.cos(param) + vx * math.sin(param),
                            center[1] + major[1] * math.cos(param) + vy * math.sin(param),
                        )
                    )
                for left, right in zip(points, points[1:]):
                    add(left, right, "ellipse-chord")
    elif kind == "HATCH":
        points = _points(entity, 10, 20)
        for left, right in zip(points[:65], points[1:65]):
            add(left, right, "hatch-boundary")
        if len(points) > 2:
            add(points[min(len(points), 65) - 1], points[0], "hatch-boundary")
    elif kind == "INSERT":
        origin = _point(entity, 10, 20)
        if origin is not None:
            rotation = math.radians(_float(entity, 50, 0.0) or 0.0)
            scale = max(
                abs(_float(entity, 41, 1.0) or 1.0),
                abs(_float(entity, 42, 1.0) or 1.0),
            )
            add(
                origin,
                (
                    origin[0] + scale * math.cos(rotation),
                    origin[1] + scale * math.sin(rotation),
                ),
                "insert-marker",
            )
    elif kind in {"TEXT", "MTEXT"}:
        origin = _point(entity, 10, 20)
        height = _float(entity, 40, 1.0) or 1.0
        direction = _point(entity, 11, 21) if kind == "MTEXT" else None
        if origin is not None:
            if direction is not None and math.hypot(direction[0], direction[1]) > 1e-9:
                norm = math.hypot(direction[0], direction[1])
                ux, uy = direction[0] / norm, direction[1] / norm
            else:
                angle = math.radians(_float(entity, 50, 0.0) or 0.0)
                ux, uy = math.cos(angle), math.sin(angle)
            length = min(1000.0, max(1.0, abs(height)))
            add(origin, (origin[0] + length * ux, origin[1] + length * uy), f"{kind.lower()}-baseline")
    elif kind == "POINT":
        origin = _point(entity, 10, 20)
        if origin is not None:
            add(origin, (origin[0] + 1.0, origin[1]), "point-marker")
    elif kind == "SPLINE":
        points = _points(entity, 10, 20)
        if len(points) < 2:
            points = _points(entity, 11, 21)
        for left, right in zip(points[:65], points[1:65]):
            add(left, right, "spline-control-chain")
    elif kind == "WIPEOUT":
        points = _points(entity, 14, 24)
        for left, right in zip(points[:65], points[1:65]):
            add(left, right, "wipeout-boundary")
        if len(points) > 2:
            add(points[min(len(points), 65) - 1], points[0], "wipeout-boundary")
    return segments or _marker(entity, _float(entity, 40, 1.0) or 1.0)


def wallish_layer(layer: str) -> bool:
    normalized = "".join(
        character if character.isalnum() else " " for character in (layer or "").upper()
    )
    return bool(set(normalized.split()) & {"WALL", "WALLS", "WA", "BEARING", "벽"})


def natural_handle_key(value: str) -> tuple[int, int | str]:
    try:
        return (0, int(value, 16))
    except ValueError:
        return (1, value)


def canonical_handles(values: Iterable[Any]) -> list[str]:
    return sorted({str(value) for value in values}, key=natural_handle_key)


def rank_hash(token: str) -> str:
    return sha256_text(f"g1_verifier_families|{SEED}|{token}")


def rank_choice(values: Iterable[Any], token: str, stringify=lambda item: str(item)) -> Any:
    materialized = list(values)
    if not materialized:
        raise ValueError(f"no candidates for deterministic choice: {token}")
    return min(
        materialized,
        key=lambda item: (rank_hash(f"{token}|{stringify(item)}"), stringify(item)),
    )


def select_representatives(
    entities: Sequence[DxfEntity], truth_wall_handles: set[str], drawing_id: str
) -> tuple[dict[str, DxfEntity], list[dict[str, Any]]]:
    representatives: dict[str, DxfEntity] = {}
    records: list[dict[str, Any]] = []
    for entity_type in ENTITY_TYPES:
        candidates = [
            entity
            for entity in entities
            if entity.entity_type == entity_type and entity.handle not in truth_wall_handles
        ]
        preferred = [entity for entity in candidates if not wallish_layer(entity.layer)]
        chosen = rank_choice(
            preferred or candidates,
            f"{drawing_id}|representative|{entity_type}",
            stringify=lambda item: f"{item.handle}|{item.layer}|{item.entity_type}",
        )
        representatives[entity_type] = chosen
        records.append(
            {
                "entity_type": entity_type,
                "handle": chosen.handle,
                "layer": chosen.layer,
                "nonwallish_layer": not wallish_layer(chosen.layer),
                "preferred_pool_used": bool(preferred),
                "segment_count": len(entity_segments(chosen)),
            }
        )
    return representatives, records


def build_compact_ir(
    entities: Sequence[DxfEntity], truth: Mapping[str, Any]
) -> tuple[dict[str, Any], list[dict[str, Any]], list[str]]:
    drawing_id = str(truth["drawing_id"])
    truth_wall_handles = set(map(str, truth.get("wall_handles_flat", [])))
    class_handles = set(map(str, truth.get("class_of_handle", {}).keys()))
    representatives, representative_records = select_representatives(
        entities, truth_wall_handles, drawing_id
    )
    wanted = class_handles | truth_wall_handles | {
        entity.handle for entity in representatives.values()
    }
    by_handle: dict[str, list[DxfEntity]] = defaultdict(list)
    for entity in entities:
        by_handle[entity.handle].append(entity)
    missing_wanted = sorted(wanted - set(by_handle), key=natural_handle_key)
    missing_walls = sorted(truth_wall_handles - set(by_handle), key=natural_handle_key)
    if missing_walls:
        raise RuntimeError(f"truth wall handles absent from DXF {drawing_id}: {missing_walls}")
    segments: list[dict[str, Any]] = []
    for handle in sorted(wanted & set(by_handle), key=natural_handle_key):
        for entity in by_handle[handle]:
            for p0, p1, adapter_kind in entity_segments(entity):
                if _distance(p0, p1) <= 1e-9:
                    continue
                segments.append(
                    {
                        "sid": f"s{len(segments) + 1:07d}",
                        "handle": handle,
                        "pts": [
                            [float(p0[0]), float(p0[1])],
                            [float(p1[0]), float(p1[1])],
                        ],
                        "layer": entity.layer,
                        "kind": entity.entity_type,
                        "geometry_adapter_kind": adapter_kind,
                        "label": "unknown",
                        "source": "gen2_family_dxf",
                    }
                )
    if not segments:
        raise RuntimeError(f"compact SEG-IR is empty: {drawing_id}")
    return (
        {
            "ir": "seg.v1",
            "drawing_id": drawing_id,
            "units": "mm",
            "scale_mm_per_unit": 1.0,
            "segments": segments,
        },
        representative_records,
        missing_wanted,
    )


def collapse_feature_ir(ir: Mapping[str, Any]) -> tuple[dict[str, Any], list[str], int]:
    by_handle: dict[str, dict[str, Any]] = {}
    valid_records = 0
    for segment in ir.get("segments", []) or []:
        points = segment.get("pts") or []
        if len(points) < 2:
            continue
        p0 = (float(points[0][0]), float(points[0][1]))
        p1 = (float(points[-1][0]), float(points[-1][1]))
        if _distance(p0, p1) <= 1e-9:
            continue
        valid_records += 1
        handle = str(segment.get("handle") or segment.get("sid"))
        by_handle.setdefault(handle, dict(segment))
    handles = sorted(by_handle, key=natural_handle_key)
    segments = []
    for index, handle in enumerate(handles, 1):
        record = dict(by_handle[handle])
        record["sid"] = f"f{index:06d}"
        segments.append(record)
    feature_ir = {
        "ir": "seg.v1",
        "drawing_id": str(ir.get("drawing_id") or ""),
        "units": "pseudo-px",
        "scale_mm_per_unit": FEATURE_SCALE_MM_PER_PSEUDO_PX,
        "segments": segments,
    }
    return feature_ir, handles, valid_records - len(handles)


def _angle_difference(left: float | np.ndarray, right: float | np.ndarray) -> np.ndarray:
    difference = np.abs(np.asarray(left) - np.asarray(right)) % 180.0
    return np.minimum(difference, 180.0 - difference)


def _band_membership(distance: float, lo: float, hi: float) -> float:
    if lo <= distance <= hi:
        return 1.0
    margin = max(1.0e-9, 0.5 * (hi - lo))
    if distance < lo:
        return max(0.0, 1.0 - (lo - distance) / margin)
    return max(0.0, 1.0 - (distance - hi) / margin)


def _proper_intersection(
    a0: np.ndarray, a1: np.ndarray, b0: np.ndarray, b1: np.ndarray, eps: float = 1e-9
) -> bool:
    r = a1 - a0
    s = b1 - b0
    cross = float(r[0] * s[1] - r[1] * s[0])
    if abs(cross) < eps:
        return False
    q = b0 - a0
    t = float((q[0] * s[1] - q[1] * s[0]) / cross)
    u = float((q[0] * r[1] - q[1] * r[0]) / cross)
    return eps < t < 1.0 - eps and eps < u < 1.0 - eps


def _perpendicular(point: np.ndarray, origin: np.ndarray, unit: np.ndarray) -> float:
    delta = point - origin
    return abs(float(delta[0] * (-unit[1]) + delta[1] * unit[0]))


def _junction_pair(
    a0: np.ndarray,
    a1: np.ndarray,
    b0: np.ndarray,
    b1: np.ndarray,
    unit_a: np.ndarray,
    unit_b: np.ndarray,
    length_a: float,
    length_b: float,
) -> bool:
    if _proper_intersection(a0, a1, b0, b1):
        return True
    for pa in (a0, a1):
        for pb in (b0, b1):
            if float(np.linalg.norm(pa - pb)) <= FEATURE_SNAP_PX:
                return True
    for endpoint in (a0, a1):
        projection = float((endpoint - b0) @ unit_b)
        if (
            FEATURE_SNAP_PX < projection < length_b - FEATURE_SNAP_PX
            and _perpendicular(endpoint, b0, unit_b) <= FEATURE_SNAP_PX
        ):
            return True
    for endpoint in (b0, b1):
        projection = float((endpoint - a0) @ unit_a)
        if (
            FEATURE_SNAP_PX < projection < length_a - FEATURE_SNAP_PX
            and _perpendicular(endpoint, a0, unit_a) <= FEATURE_SNAP_PX
        ):
            return True
    return False


def _geometry_arrays(feature_ir: Mapping[str, Any]) -> dict[str, Any]:
    handles: list[str] = []
    p1_rows: list[list[float]] = []
    p2_rows: list[list[float]] = []
    for segment in feature_ir.get("segments", []) or []:
        points = segment.get("pts") or []
        if len(points) < 2:
            continue
        p0 = np.asarray(points[0][:2], dtype=np.float64) / FEATURE_SCALE_MM_PER_PSEUDO_PX
        p1 = np.asarray(points[-1][:2], dtype=np.float64) / FEATURE_SCALE_MM_PER_PSEUDO_PX
        if float(np.linalg.norm(p1 - p0)) <= 1e-12:
            continue
        handles.append(str(segment.get("handle") or segment.get("sid")))
        p1_rows.append([float(p0[0]), float(p0[1])])
        p2_rows.append([float(p1[0]), float(p1[1])])
    if len(handles) != len(set(handles)):
        raise RuntimeError("feature IR contains duplicate handles after collapse")
    p1 = np.asarray(p1_rows, dtype=np.float64).reshape((-1, 2))
    p2 = np.asarray(p2_rows, dtype=np.float64).reshape((-1, 2))
    delta = p2 - p1
    lengths = np.hypot(delta[:, 0], delta[:, 1])
    if np.any(lengths <= 0.0):
        raise RuntimeError("degenerate feature segment")
    units = delta / lengths[:, None]
    angles = np.degrees(np.arctan2(delta[:, 1], delta[:, 0])) % 180.0
    midpoints = (p1 + p2) / 2.0
    return {
        "handles": handles,
        "p1": p1,
        "p2": p2,
        "delta": delta,
        "lengths": lengths,
        "units": units,
        "angles": angles,
        "midpoints": midpoints,
    }


def extract_twohop_features(
    feature_ir: Mapping[str, Any]
) -> tuple[list[str], np.ndarray, dict[str, Any]]:
    geometry = _geometry_arrays(feature_ir)
    handles = geometry["handles"]
    p1 = geometry["p1"]
    p2 = geometry["p2"]
    lengths = geometry["lengths"]
    units = geometry["units"]
    angles = geometry["angles"]
    midpoints = geometry["midpoints"]
    n = len(handles)
    if n == 0:
        raise RuntimeError("feature extraction received zero handles")

    angle_diff = np.zeros((n, n), dtype=np.float64)
    overlap = np.zeros((n, n), dtype=np.float64)
    gap = np.full((n, n), np.inf, dtype=np.float64)
    junction = np.zeros((n, n), dtype=bool)
    for i in range(n):
        angle_diff[i] = _angle_difference(angles[i], angles)
        dx1 = p1[:, 0] - p1[i, 0]
        dy1 = p1[:, 1] - p1[i, 1]
        dx2 = p2[:, 0] - p1[i, 0]
        dy2 = p2[:, 1] - p1[i, 1]
        ta = dx1 * units[i, 0] + dy1 * units[i, 1]
        tb = dx2 * units[i, 0] + dy2 * units[i, 1]
        lower = np.minimum(ta, tb)
        upper = np.maximum(ta, tb)
        overlap_length = np.maximum(
            np.minimum(lengths[i], upper) - np.maximum(0.0, lower), 0.0
        )
        overlap[i] = overlap_length / np.minimum(lengths[i], lengths)
        mx = midpoints[:, 0] - p1[i, 0]
        my = midpoints[:, 1] - p1[i, 1]
        gap[i] = np.abs(mx * (-units[i, 1]) + my * units[i, 0])
        overlap[i, i] = 0.0
        gap[i, i] = np.inf
    for i in range(n):
        for j in range(i + 1, n):
            joined = _junction_pair(
                p1[i], p2[i], p1[j], p2[j], units[i], units[j], lengths[i], lengths[j]
            )
            junction[i, j] = joined
            junction[j, i] = joined

    lo, hi = FEATURE_THICKNESS_BAND_PX
    parallel_gate = (
        (angle_diff <= FEATURE_ANGLE_TOL_DEG)
        & (overlap >= FEATURE_OVERLAP_MIN)
        & np.isfinite(gap)
    )
    parallel_band = parallel_gate & (gap >= lo) & (gap <= hi)
    parallel_score = np.zeros(n, dtype=np.float64)
    thickness_score = np.zeros(n, dtype=np.float64)
    for i in range(n):
        if np.any(parallel_band[i]):
            parallel_score[i] = float(np.max(overlap[i, parallel_band[i]]))
        near_candidates = np.flatnonzero(
            (angle_diff[i] <= FEATURE_ANGLE_TOL_DEG)
            & (overlap[i] > 0.0)
            & np.isfinite(gap[i])
            & (np.arange(n) != i)
        )
        if len(near_candidates):
            chosen = min(
                near_candidates,
                key=lambda j: (
                    0 if overlap[i, j] >= FEATURE_OVERLAP_MIN else 1,
                    gap[i, j],
                    int(j),
                ),
            )
            thickness_score[i] = _band_membership(float(gap[i, chosen]), lo, hi)
    degrees = junction.sum(axis=1).astype(np.int64)
    junction_score = np.minimum(1.0, degrees.astype(np.float64) / 2.0)

    angles_rad = np.arctan2(
        p2[:, 1] - p1[:, 1], p2[:, 0] - p1[:, 0]
    )
    base = np.column_stack(
        (
            parallel_score,
            thickness_score,
            junction_score,
            np.log10(lengths + 1.0),
            np.sin(2.0 * angles_rad),
            np.cos(2.0 * angles_rad),
        )
    )

    nearest_gap = np.full(n, np.nan, dtype=np.float64)
    parallel_count = parallel_band.sum(axis=1).astype(np.float64)
    for i in range(n):
        candidates = np.flatnonzero(parallel_gate[i])
        if len(candidates):
            nearest_gap[i] = float(np.min(gap[i, candidates]))
    length_percentile = (
        np.searchsorted(np.sort(lengths), lengths, side="right").astype(np.float64) / n
    )
    midpoint_delta = midpoints[:, None, :] - midpoints[None, :, :]
    midpoint_distance = np.sqrt(np.sum(midpoint_delta * midpoint_delta, axis=2))
    radius_adjacency = (
        (midpoint_distance <= FEATURE_RADIUS_PX)
        & (midpoint_distance > 0.0)
        & ~np.eye(n, dtype=bool)
    )
    radius_degree = radius_adjacency.sum(axis=1).astype(np.float64)
    radius_density = radius_degree / (math.pi * FEATURE_RADIUS_PX**2)
    angle_bins = np.minimum(
        (angles / (180.0 / FEATURE_ANGLE_BINS)).astype(np.int64),
        FEATURE_ANGLE_BINS - 1,
    )
    angle_entropy = np.zeros(n, dtype=np.float64)
    for i in range(n):
        neighbors = np.flatnonzero(radius_adjacency[i])
        if len(neighbors):
            hist = np.bincount(angle_bins[neighbors], minlength=FEATURE_ANGLE_BINS)
            probabilities = hist[hist > 0].astype(np.float64) / len(neighbors)
            angle_entropy[i] = float(
                -np.sum(probabilities * np.log(probabilities)) / math.log(FEATURE_ANGLE_BINS)
            )
    context = np.column_stack(
        (
            parallel_count,
            nearest_gap,
            degrees.astype(np.float64),
            length_percentile,
            radius_density,
            angle_entropy,
        )
    )

    twohop = np.full((n, len(TWOHOP_FEATURE_NAMES)), np.nan, dtype=np.float64)
    for i in range(n):
        neighbors = np.flatnonzero(junction[i])
        if len(neighbors):
            neighbor_degrees = degrees[neighbors].astype(np.float64)
            twohop[i, 0] = float(np.sum(neighbor_degrees))
            twohop[i, 1] = float(np.max(neighbor_degrees))
            twohop[i, 2] = float(np.var(neighbor_degrees, ddof=0))
        else:
            twohop[i, 0:3] = 0.0

    component_labels = np.full(n, -1, dtype=np.int64)
    components: list[list[int]] = []
    for start in range(n):
        if component_labels[start] >= 0:
            continue
        label = len(components)
        stack = [start]
        members: list[int] = []
        component_labels[start] = label
        while stack:
            current = stack.pop()
            members.append(current)
            for neighbor in np.flatnonzero(junction[current]):
                if component_labels[neighbor] < 0:
                    component_labels[neighbor] = label
                    stack.append(int(neighbor))
        components.append(members)
    thickness_flags = thickness_score > 0.0
    for members in components:
        rate = float(np.mean(thickness_flags[members]))
        for index in members:
            twohop[index, 8] = float(len(members))
            twohop[index, 9] = rate

    second_ring_memberships = 0
    for i in range(n):
        direct = set(map(int, np.flatnonzero(radius_adjacency[i])))
        ring: set[int] = set()
        for neighbor in direct:
            ring.update(map(int, np.flatnonzero(radius_adjacency[neighbor])))
        ring.difference_update(direct)
        ring.discard(i)
        ordered_ring = np.asarray(sorted(ring), dtype=np.int64)
        second_ring_memberships += len(ordered_ring)
        twohop[i, 10] = float(len(ordered_ring))
        if len(ordered_ring) == 0:
            twohop[i, 3] = 0.0
            continue
        ring_lengths = lengths[ordered_ring]
        differences = _angle_difference(angles[i], angles[ordered_ring])
        alignment = np.cos(np.radians(differences))
        twohop[i, 11] = float(np.mean(ring_lengths))
        twohop[i, 12] = float(np.std(ring_lengths, ddof=0))
        twohop[i, 13] = float(np.max(ring_lengths))
        twohop[i, 14] = float(np.mean(alignment))
        twohop[i, 15] = float(np.std(alignment, ddof=0))
        twohop[i, 16] = float(np.mean(differences <= 2.0))
        qualifying = (
            (differences <= FEATURE_ANGLE_TOL_DEG)
            & (overlap[i, ordered_ring] >= FEATURE_OVERLAP_MIN)
            & (gap[i, ordered_ring] >= lo)
            & (gap[i, ordered_ring] <= hi)
        )
        selected_gaps = gap[i, ordered_ring][qualifying]
        twohop[i, 3] = float(len(selected_gaps))
        if len(selected_gaps):
            twohop[i, 4] = float(np.min(selected_gaps))
            twohop[i, 5] = float(np.mean(selected_gaps))
            twohop[i, 6] = float(np.std(selected_gaps, ddof=0))
            twohop[i, 7] = float(np.max(selected_gaps))

    extension = np.zeros((n, n), dtype=bool)
    for i in range(n):
        for j in range(i + 1, n):
            endpoint_distance = min(
                float(np.linalg.norm(left - right))
                for left in (p1[i], p2[i])
                for right in (p1[j], p2[j])
            )
            if endpoint_distance > FEATURE_RADIUS_PX:
                continue
            if float(_angle_difference(angles[i], angles[j])) > 2.0:
                continue
            offset = _perpendicular(midpoints[j], p1[i], units[i])
            if offset > FEATURE_SNAP_PX:
                continue
            extension[i, j] = True
            extension[j, i] = True
    chain_spans = lengths.copy()
    for i in range(n):
        origin = p1[i]
        unit = units[i]
        for neighbor in np.flatnonzero(extension[i]):
            paths = [[i, int(neighbor)]]
            for third in np.flatnonzero(extension[neighbor]):
                if int(third) != i:
                    paths.append([i, int(neighbor), int(third)])
            for path in paths:
                indices = np.asarray(path, dtype=np.int64)
                projections = np.concatenate(
                    ((p1[indices] - origin) @ unit, (p2[indices] - origin) @ unit)
                )
                chain_spans[i] = max(
                    chain_spans[i], float(np.max(projections) - np.min(projections))
                )
    twohop[:, 17] = chain_spans
    matrix = np.column_stack((base, context, twohop)).astype(np.float32, copy=False)
    if matrix.shape != (n, len(FULL_FEATURE_NAMES)):
        raise RuntimeError(f"feature shape drift: {matrix.shape}")
    return handles, matrix, {
        "handle_count": n,
        "junction_edges": int(np.sum(junction) // 2),
        "junction_components": len(components),
        "radius_edges": int(np.sum(radius_adjacency) // 2),
        "second_ring_memberships": int(second_ring_memberships),
        "extension_edges": int(np.sum(extension) // 2),
        "feature_count": len(FULL_FEATURE_NAMES),
    }


def load_drawing(ref: DrawingRef) -> DrawingData:
    if file_sha256(ref.dxf_path) != ref.dxf_sha256:
        raise RuntimeError(f"DXF hash mismatch: {ref.drawing_id}")
    if file_sha256(ref.truth_path) != ref.truth_sha256:
        raise RuntimeError(f"truth hash mismatch: {ref.drawing_id}")
    truth = load_json(ref.truth_path)
    if str(truth.get("family_id")) != ref.family_id:
        raise RuntimeError(f"family identity mismatch: {ref.drawing_id}")
    if str(truth.get("drawing_id")) != ref.drawing_id:
        raise RuntimeError(f"drawing identity mismatch: {ref.drawing_id}")
    entities = parse_dxf_entities(ref.dxf_path)
    raw_mix = Counter(entity.entity_type for entity in entities)
    truth_mix = {str(key): int(value) for key, value in truth.get("entity_mix", {}).items()}
    if dict(raw_mix) != truth_mix:
        raise RuntimeError(f"raw entity mix mismatch: {ref.drawing_id}")
    ir, representative_records, missing_wanted = build_compact_ir(entities, truth)
    feature_ir, expected_handles, duplicate_count = collapse_feature_ir(ir)
    handles, features, _meta = extract_twohop_features(feature_ir)
    if handles != expected_handles:
        raise RuntimeError(f"feature handle alignment mismatch: {ref.drawing_id}")
    wall = set(map(str, truth.get("wall_handles_flat", [])))
    if not wall or not wall <= set(handles):
        raise RuntimeError(f"truth/candidate universe mismatch: {ref.drawing_id}")
    labels = np.asarray([1 if handle in wall else 0 for handle in handles], dtype=np.int8)
    return DrawingData(
        ref=ref,
        truth=truth,
        ir=ir,
        feature_ir=feature_ir,
        handles=handles,
        features=features,
        labels=labels,
        duplicate_segment_records_collapsed=duplicate_count,
        representative_records=representative_records,
        missing_wanted_handles=missing_wanted,
    )


def load_drawings(
    refs: Sequence[DrawingRef], budget: CpuBudget, phase: str
) -> list[DrawingData]:
    output: list[DrawingData] = []
    for index, ref in enumerate(refs, 1):
        budget.ensure(f"{phase} drawing {ref.drawing_id}")
        output.append(load_drawing(ref))
        if index % 25 == 0 or index == len(refs):
            print(f"[{phase}] loaded {index}/{len(refs)} drawings", flush=True)
    return output


def matrix_sha256(matrix: np.ndarray) -> str:
    contiguous = np.ascontiguousarray(matrix)
    digest = hashlib.sha256()
    digest.update(str(contiguous.dtype).encode("ascii"))
    digest.update(canonical_json(list(contiguous.shape)).encode("utf-8"))
    digest.update(contiguous.tobytes(order="C"))
    return digest.hexdigest()


def estimator_sha256(model: HistGradientBoostingClassifier) -> str:
    return sha256_bytes(pickle.dumps(model, protocol=5))


def fit_scorer(
    reward_drawings: Sequence[DrawingData], budget: CpuBudget
) -> tuple[list[HistGradientBoostingClassifier], dict[str, Any]]:
    X = np.vstack([drawing.features for drawing in reward_drawings]).astype(
        np.float32, copy=False
    )
    y = np.concatenate([drawing.labels for drawing in reward_drawings]).astype(
        np.int8, copy=False
    )
    if X.shape[1] != len(FULL_FEATURE_NAMES) or set(np.unique(y)) != {0, 1}:
        raise RuntimeError("scorer training matrix contract failure")
    model_records: list[dict[str, Any]] = []
    models: list[HistGradientBoostingClassifier] = []
    for seed in MODEL_SEEDS:
        budget.ensure(f"scorer fit seed {seed}")
        model = HistGradientBoostingClassifier(random_state=seed)
        started_wall = time.perf_counter()
        started_cpu = time.process_time()
        with warnings.catch_warnings(record=True) as caught:
            warnings.simplefilter("always")
            model.fit(X, y)
        probability = model.predict_proba(X)[:, 1].astype(np.float64)
        model_records.append(
            {
                "seed": seed,
                "estimator_sha256": estimator_sha256(model),
                "training_prediction_sha256": matrix_sha256(probability),
                "n_iter": int(model.n_iter_),
                "fit_wall_seconds": time.perf_counter() - started_wall,
                "fit_cpu_seconds": time.process_time() - started_cpu,
                "warning_count": len(caught),
                "warnings": [str(item.message)[:300] for item in caught],
                "parameters": json_ready(model.get_params(deep=False)),
            }
        )
        models.append(model)
    fingerprint_payload = {
        "feature_names": list(FULL_FEATURE_NAMES),
        "train_X_sha256": matrix_sha256(X),
        "train_y_sha256": matrix_sha256(y),
        "model_seeds": list(MODEL_SEEDS),
        "model_sha256": [row["estimator_sha256"] for row in model_records],
        "prediction_sha256": [row["training_prediction_sha256"] for row in model_records],
    }
    fingerprint = sha256_text(canonical_json(fingerprint_payload))
    meta = {
        "training_drawings": len(reward_drawings),
        "training_rows": int(len(y)),
        "positive_rows": int(np.sum(y == 1)),
        "negative_rows": int(np.sum(y == 0)),
        "positive_fraction": float(np.mean(y == 1)),
        "train_X_sha256": fingerprint_payload["train_X_sha256"],
        "train_y_sha256": fingerprint_payload["train_y_sha256"],
        "model_records": model_records,
        "fingerprint_payload": fingerprint_payload,
        "ensemble_fingerprint_sha256": fingerprint,
    }
    return models, meta


def score_drawing(
    models: Sequence[HistGradientBoostingClassifier], drawing: DrawingData
) -> np.ndarray:
    predictions = np.vstack(
        [model.predict_proba(drawing.features)[:, 1] for model in models]
    ).astype(np.float64)
    probability = np.mean(predictions, axis=0)
    if probability.shape != (len(drawing.handles),) or not np.all(np.isfinite(probability)):
        raise RuntimeError(f"invalid scorer output: {drawing.ref.drawing_id}")
    return probability


def ordered_indices(handles: Sequence[str], probabilities: np.ndarray) -> list[int]:
    return sorted(
        range(len(handles)),
        key=lambda index: (-float(probabilities[index]), natural_handle_key(handles[index])),
    )


def reorder_problem(
    handles: Sequence[str], probabilities: np.ndarray, truth: set[str], expected: set[str]
) -> dict[str, Any]:
    indices = ordered_indices(handles, probabilities)
    ordered_handles = [handles[index] for index in indices]
    ordered_probability = np.asarray([probabilities[index] for index in indices], dtype=np.float64)
    truth_mask = 0
    expected_mask = 0
    for index, handle in enumerate(ordered_handles):
        if handle in truth:
            truth_mask |= 1 << index
        if handle in expected:
            expected_mask |= 1 << index
    return {
        "handles": ordered_handles,
        "probabilities": ordered_probability,
        "truth_mask": truth_mask,
        "expected_mask": expected_mask,
        "all_mask": (1 << len(ordered_handles)) - 1,
    }


def set_metrics(mask: int, truth_mask: int, n: int) -> dict[str, Any]:
    all_mask = (1 << n) - 1
    tp = (mask & truth_mask).bit_count()
    fp = (mask & (~truth_mask & all_mask)).bit_count()
    fn = ((~mask & all_mask) & truth_mask).bit_count()
    precision = tp / (tp + fp) if tp + fp else 0.0
    recall = tp / (tp + fn) if tp + fn else 0.0
    f1 = 2.0 * tp / (2 * tp + fp + fn) if 2 * tp + fp + fn else 0.0
    return {
        "tp": tp,
        "fp": fp,
        "fn": fn,
        "precision": precision,
        "recall": recall,
        "f1": f1,
    }


def terminal_fast(mask: int, truth_mask: int, expected_mask: int, n: int) -> dict[str, Any]:
    metrics = set_metrics(mask, truth_mask, n)
    accepted = mask == expected_mask
    objective = (
        TERMINAL_F1_WEIGHT * metrics["f1"]
        + TERMINAL_VERIFIER_WEIGHT * float(accepted)
    )
    return {**metrics, "accepted_fast": accepted, "objective": objective}


def decision_log_likelihood(mask: int, probabilities: np.ndarray) -> float:
    result = 0.0
    for index, raw_probability in enumerate(probabilities):
        probability = min(1.0 - 1e-12, max(1e-12, float(raw_probability)))
        result += math.log(probability if mask & (1 << index) else 1.0 - probability)
    return result


def greedy_mask(probabilities: np.ndarray) -> tuple[int, int]:
    mask = 0
    stop_index = len(probabilities)
    for index, probability in enumerate(probabilities):
        if float(probability) < PROBABILITY_THRESHOLD:
            stop_index = index
            break
        mask |= 1 << index
    return mask, stop_index


def beam_search(
    probabilities: np.ndarray, truth_mask: int, expected_mask: int, width: int
) -> dict[str, Any]:
    states: list[tuple[float, int]] = [(0.0, 0)]
    frontier_peak = 1
    expanded = 0
    for index, raw_probability in enumerate(probabilities):
        probability = min(1.0 - 1e-12, max(1e-12, float(raw_probability)))
        include_log = math.log(probability)
        exclude_log = math.log(1.0 - probability)
        candidates: list[tuple[float, int]] = []
        for log_score, mask in states:
            candidates.append((log_score + exclude_log, mask))
            candidates.append((log_score + include_log, mask | (1 << index)))
        expanded += len(candidates)
        candidates.sort(key=lambda state: (-state[0], state[1]))
        states = candidates[:width]
        frontier_peak = max(frontier_peak, len(states))
    n = len(probabilities)
    ranked = []
    for log_score, mask in states:
        terminal = terminal_fast(mask, truth_mask, expected_mask, n)
        ranked.append((terminal["objective"], log_score, mask, terminal))
    ranked.sort(key=lambda row: (-row[0], -row[1], row[2]))
    objective, log_score, mask, terminal = ranked[0]
    return {
        "mask": mask,
        "terminal": terminal,
        "log_likelihood": log_score,
        "width": width,
        "expanded_states": expanded,
        "frontier_peak": frontier_peak,
        "completed_candidates": len(states),
    }


def exhaustive_search(
    probabilities: np.ndarray, truth_mask: int, expected_mask: int
) -> dict[str, Any]:
    n = len(probabilities)
    total = 1 << n
    started = time.process_time()
    best: tuple[float, float, int, dict[str, Any]] | None = None
    evaluated = 0
    for mask in range(total):
        if time.process_time() - started > EXACT_CPU_SECONDS_PER_DRAWING:
            return {
                "completed": False,
                "reason": "exact_cpu_cap",
                "evaluated_subsets": evaluated,
                "cpu_seconds": time.process_time() - started,
            }
        terminal = terminal_fast(mask, truth_mask, expected_mask, n)
        log_score = decision_log_likelihood(mask, probabilities)
        candidate = (terminal["objective"], log_score, -mask, terminal)
        if best is None or candidate[:3] > best[:3]:
            best = candidate
        evaluated += 1
    assert best is not None
    return {
        "completed": True,
        "mask": -best[2],
        "terminal": best[3],
        "log_likelihood": best[1],
        "evaluated_subsets": evaluated,
        "cpu_seconds": time.process_time() - started,
        "certified": True,
        "objective_upper_bound": best[0],
        "objective_gap": 0.0,
        "method": "exhaustive_exact",
    }


def partial_upper_bound(
    depth: int, mask: int, truth_mask: int, expected_mask: int, n: int
) -> tuple[float, float, bool]:
    all_mask = (1 << n) - 1
    processed = (1 << depth) - 1
    tp = (mask & truth_mask).bit_count()
    fp = (mask & (~truth_mask & all_mask)).bit_count()
    remaining_true = (truth_mask & (~processed & all_mask)).bit_count()
    possible_tp = tp + remaining_true
    f1_upper = (
        2.0 * possible_tp / (2 * possible_tp + fp)
        if 2 * possible_tp + fp
        else 0.0
    )
    verifier_possible = ((mask ^ expected_mask) & processed) == 0
    objective_upper = (
        TERMINAL_F1_WEIGHT * f1_upper
        + TERMINAL_VERIFIER_WEIGHT * float(verifier_possible)
    )
    return objective_upper, f1_upper, verifier_possible


def branch_and_bound(
    probabilities: np.ndarray,
    truth_mask: int,
    expected_mask: int,
    initial_mask: int,
) -> dict[str, Any]:
    n = len(probabilities)
    started = time.process_time()
    initial_terminal = terminal_fast(initial_mask, truth_mask, expected_mask, n)
    incumbent_mask = initial_mask
    incumbent_terminal = initial_terminal
    incumbent_log = decision_log_likelihood(initial_mask, probabilities)
    root_upper, root_f1_upper, root_possible = partial_upper_bound(
        0, 0, truth_mask, expected_mask, n
    )
    heap: list[tuple[float, float, int, int, float, bool, int]] = []
    sequence = 0
    heapq.heappush(heap, (-root_upper, 0.0, 0, 0, root_f1_upper, root_possible, sequence))
    nodes = 0
    pruned = 0
    stop_reason = "heap_exhausted"
    while heap:
        if nodes >= BNB_NODE_CAP:
            stop_reason = "node_cap"
            break
        if time.process_time() - started > BNB_CPU_SECONDS_PER_DRAWING:
            stop_reason = "cpu_cap"
            break
        neg_upper, neg_prefix_log, depth, mask, _f1_upper, _possible, _seq = heapq.heappop(heap)
        upper = -neg_upper
        prefix_log = -neg_prefix_log
        if upper <= incumbent_terminal["objective"] + NUMERIC_TOLERANCE:
            pruned += 1
            continue
        nodes += 1
        if depth == n:
            terminal = terminal_fast(mask, truth_mask, expected_mask, n)
            log_score = decision_log_likelihood(mask, probabilities)
            ranking = (terminal["objective"], log_score, -mask)
            incumbent_ranking = (
                incumbent_terminal["objective"], incumbent_log, -incumbent_mask
            )
            if ranking > incumbent_ranking:
                incumbent_mask = mask
                incumbent_terminal = terminal
                incumbent_log = log_score
            continue
        raw_probability = float(probabilities[depth])
        probability = min(1.0 - 1e-12, max(1e-12, raw_probability))
        preferred_include = probability >= PROBABILITY_THRESHOLD
        decisions = (True, False) if preferred_include else (False, True)
        for include in decisions:
            child_mask = mask | (1 << depth) if include else mask
            child_log = prefix_log + math.log(probability if include else 1.0 - probability)
            child_depth = depth + 1
            child_upper, child_f1_upper, child_possible = partial_upper_bound(
                child_depth, child_mask, truth_mask, expected_mask, n
            )
            if child_upper <= incumbent_terminal["objective"] + NUMERIC_TOLERANCE:
                pruned += 1
                continue
            sequence += 1
            heapq.heappush(
                heap,
                (
                    -child_upper,
                    -child_log,
                    child_depth,
                    child_mask,
                    child_f1_upper,
                    child_possible,
                    sequence,
                ),
            )
    remaining_upper = max(
        incumbent_terminal["objective"],
        -heap[0][0] if heap else incumbent_terminal["objective"],
    )
    certified = remaining_upper <= incumbent_terminal["objective"] + NUMERIC_TOLERANCE
    return {
        "completed": True,
        "mask": incumbent_mask,
        "terminal": incumbent_terminal,
        "log_likelihood": incumbent_log,
        "evaluated_nodes": nodes,
        "pruned_nodes": pruned,
        "remaining_frontier": len(heap),
        "cpu_seconds": time.process_time() - started,
        "certified": certified,
        "objective_upper_bound": remaining_upper,
        "objective_gap": max(0.0, remaining_upper - incumbent_terminal["objective"]),
        "f1_upper_bound": incumbent_terminal["f1"] if certified else 1.0,
        "stop_reason": stop_reason,
        "method": "branch_and_bound_upper_bound_replacement",
    }


def exact_or_bound(
    probabilities: np.ndarray,
    truth_mask: int,
    expected_mask: int,
    initial_mask: int,
) -> dict[str, Any]:
    n = len(probabilities)
    if n <= EXACT_MAX_HANDLES and (1 << n) <= EXACT_MAX_SUBSETS:
        exact = exhaustive_search(probabilities, truth_mask, expected_mask)
        if exact.get("completed"):
            exact["f1_upper_bound"] = exact["terminal"]["f1"]
            return exact
    result = branch_and_bound(probabilities, truth_mask, expected_mask, initial_mask)
    result["replacement_reason"] = (
        "handle_count_above_exact_cap"
        if n > EXACT_MAX_HANDLES
        else "exhaustive_time_cap"
    )
    return result


def mask_handles(mask: int, handles: Sequence[str]) -> list[str]:
    return canonical_handles(
        handle for index, handle in enumerate(handles) if mask & (1 << index)
    )


def policy_record(
    name: str,
    mask: int,
    problem: Mapping[str, Any],
    drawing: DrawingData,
    verifier_module: Any,
    verifier_analysis: Mapping[str, Any],
    extra: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    handles = mask_handles(mask, problem["handles"])
    verifier_result = verifier_module.verify_claim(
        drawing.ir, handles, analysis=verifier_analysis
    )
    terminal = terminal_fast(
        mask,
        int(problem["truth_mask"]),
        int(problem["expected_mask"]),
        len(problem["handles"]),
    )
    if bool(verifier_result["accepted"]) != bool(terminal["accepted_fast"]):
        raise RuntimeError(
            f"fast/full verifier disagreement {drawing.ref.drawing_id} {name}"
        )
    return {
        "policy": name,
        "selected_count": len(handles),
        "selected_handles": handles,
        "set_metrics": {
            key: terminal[key]
            for key in ("tp", "fp", "fn", "precision", "recall", "f1")
        },
        "verifier_accepted": bool(verifier_result["accepted"]),
        "verifier_reason_codes": list(verifier_result["reason_codes"]),
        "terminal_objective": float(terminal["objective"]),
        **dict(extra or {}),
    }


def _feature_fixture() -> dict[str, Any]:
    segments = []
    records = (
        ("A1", (0.0, 0.0), (2400.0, 0.0)),
        ("A2", (0.0, 200.0), (2400.0, 200.0)),
        ("B1", (0.0, 1800.0), (2400.0, 1800.0)),
        ("B2", (0.0, 2000.0), (2400.0, 2000.0)),
        ("N1", (3000.0, 3000.0), (3500.0, 3400.0)),
    )
    for index, (handle, p0, p1) in enumerate(records, 1):
        segments.append(
            {
                "sid": f"fixture_{index}",
                "handle": handle,
                "pts": [list(p0), list(p1)],
                "layer": "WALL" if handle[0] in {"A", "B"} else "ANNO",
                "kind": "LINE",
            }
        )
    return {
        "ir": "seg.v1",
        "drawing_id": "g9_feature_fixture",
        "units": "pseudo-px",
        "scale_mm_per_unit": FEATURE_SCALE_MM_PER_PSEUDO_PX,
        "segments": segments,
    }


def boundary_selftest() -> dict[str, Any]:
    probes = {
        "original_cad": Path(r"D:\forbidden\original_cad\1.dwg"),
        "repository_test": Path(
            r"D:\dev\99_tools\autocad-sdk-router\tests\unit\forbidden.py"
        ),
        "val_b": Path(r"D:\runs\e2_program\cells\w2_09_valb\forbidden.json"),
        "cubicasa": Path(
            r"D:\dev\99_tools\autocad-sdk-router\runs\e2_ext_cubicasa\ir\test\forbidden.segir.json"
        ),
    }
    results: dict[str, bool] = {}
    for name, path in probes.items():
        try:
            assert_read_allowed(path)
        except BoundaryViolation:
            results[name] = True
        else:
            results[name] = False
    return {"probes": {name: str(path) for name, path in probes.items()}, "denied": results}


def search_fixture_signature() -> dict[str, Any]:
    probabilities = np.asarray([0.93, 0.78, 0.61, 0.42, 0.17], dtype=np.float64)
    truth_mask = (1 << 0) | (1 << 3)
    expected_mask = truth_mask
    greedy, stop_index = greedy_mask(probabilities)
    beams = {
        str(width): beam_search(probabilities, truth_mask, expected_mask, width)
        for width in BEAM_WIDTHS
    }
    bound = branch_and_bound(
        probabilities,
        truth_mask,
        expected_mask,
        int(beams[str(PRIMARY_BEAM_WIDTH)]["mask"]),
    )
    payload = {
        "greedy_mask": greedy,
        "greedy_stop_index": stop_index,
        "beam_masks": {key: int(value["mask"]) for key, value in beams.items()},
        "beam_objectives": {
            key: float(value["terminal"]["objective"]) for key, value in beams.items()
        },
        "bound_mask": int(bound["mask"]),
        "bound_certified": bool(bound["certified"]),
        "bound_gap": float(bound["objective_gap"]),
    }
    return {"payload": payload, "sha256": sha256_text(canonical_json(payload))}


def run_selftests_with_reward(
    reward_drawings: Sequence[DrawingData], budget: CpuBudget
) -> tuple[dict[str, Any], list[HistGradientBoostingClassifier], dict[str, Any]]:
    fixture_handles_a, fixture_matrix_a, fixture_meta_a = extract_twohop_features(
        _feature_fixture()
    )
    fixture_handles_b, fixture_matrix_b, fixture_meta_b = extract_twohop_features(
        _feature_fixture()
    )
    feature_deterministic = (
        fixture_handles_a == fixture_handles_b
        and np.array_equal(fixture_matrix_a, fixture_matrix_b, equal_nan=True)
        and fixture_meta_a == fixture_meta_b
    )
    first_models, first_meta = fit_scorer(reward_drawings, budget)
    second_models, second_meta = fit_scorer(reward_drawings, budget)
    scorer_hash_equal = (
        first_meta["ensemble_fingerprint_sha256"]
        == second_meta["ensemble_fingerprint_sha256"]
    )
    prediction_hashes_a: list[str] = []
    prediction_hashes_b: list[str] = []
    for drawing in reward_drawings:
        prediction_hashes_a.append(matrix_sha256(score_drawing(first_models, drawing)))
        prediction_hashes_b.append(matrix_sha256(score_drawing(second_models, drawing)))
    scorer_predictions_equal = prediction_hashes_a == prediction_hashes_b
    search_a = search_fixture_signature()
    search_b = search_fixture_signature()
    boundary = boundary_selftest()
    boundary_ok = all(boundary["denied"].values())
    result = {
        "schema": SELFTEST_SCHEMA,
        "selftest_ok": bool(
            feature_deterministic
            and scorer_hash_equal
            and scorer_predictions_equal
            and search_a == search_b
            and boundary_ok
        ),
        "feature_determinism": {
            "ok": feature_deterministic,
            "fixture_handles": fixture_handles_a,
            "fixture_matrix_sha256": matrix_sha256(fixture_matrix_a),
            "fixture_meta": fixture_meta_a,
        },
        "scorer_freeze": {
            "ok": scorer_hash_equal and scorer_predictions_equal,
            "first_ensemble_fingerprint_sha256": first_meta[
                "ensemble_fingerprint_sha256"
            ],
            "second_ensemble_fingerprint_sha256": second_meta[
                "ensemble_fingerprint_sha256"
            ],
            "prediction_hash_list_sha256_first": sha256_text(
                canonical_json(prediction_hashes_a)
            ),
            "prediction_hash_list_sha256_second": sha256_text(
                canonical_json(prediction_hashes_b)
            ),
        },
        "search_determinism": {
            "ok": search_a == search_b,
            "first": search_a,
            "second": search_b,
        },
        "boundary": {**boundary, "ok": boundary_ok},
        "hidden_content_reads_during_selftest": int(
            _boundary_counters["hidden_content_reads_after_freeze"]
        ),
        "reproducibility_claim": "휘발 필드 제외 수치 전 필드 동일",
    }
    if not result["selftest_ok"]:
        raise RuntimeError(f"G9 selftest failed: {json_ready(result)}")
    return result, first_models, first_meta


def evaluate_hidden_drawing(
    drawing: DrawingData,
    models: Sequence[HistGradientBoostingClassifier],
    verifier_module: Any,
    budget: CpuBudget,
) -> dict[str, Any]:
    budget.ensure(f"hidden policy evaluation {drawing.ref.drawing_id}")
    probabilities = score_drawing(models, drawing)
    truth = set(map(str, drawing.truth.get("wall_handles_flat", [])))
    verifier_analysis = verifier_module.analyze_seg_ir(
        drawing.ir,
        angle_tol_deg=VERIFIER_CONFIG["angle_tol_deg"],
        overlap_min=VERIFIER_CONFIG["overlap_min"],
        thickness_band_mm=tuple(VERIFIER_CONFIG["thickness_band_mm"]),
    )
    expected = set(map(str, verifier_analysis["expected_handles"]))
    true_verification = verifier_module.verify_claim(
        drawing.ir, canonical_handles(truth), analysis=verifier_analysis
    )
    if expected != truth or not true_verification["accepted"]:
        raise RuntimeError(
            f"G1 oracle alignment failure on {drawing.ref.drawing_id}: "
            f"expected={len(expected)} truth={len(truth)} accepted={true_verification['accepted']}"
        )
    problem = reorder_problem(drawing.handles, probabilities, truth, expected)
    greedy, stop_index = greedy_mask(problem["probabilities"])
    beam_results = {
        width: beam_search(
            problem["probabilities"],
            int(problem["truth_mask"]),
            int(problem["expected_mask"]),
            width,
        )
        for width in BEAM_WIDTHS
    }
    primary_beam = beam_results[PRIMARY_BEAM_WIDTH]
    exact = exact_or_bound(
        problem["probabilities"],
        int(problem["truth_mask"]),
        int(problem["expected_mask"]),
        int(primary_beam["mask"]),
    )
    policies: dict[str, Any] = {}
    policies["greedy"] = policy_record(
        "greedy",
        greedy,
        problem,
        drawing,
        verifier_module,
        verifier_analysis,
        {
            "stop_index": stop_index,
            "log_likelihood": decision_log_likelihood(
                greedy, problem["probabilities"]
            ),
        },
    )
    for width, search in beam_results.items():
        policies[f"beam_{width}"] = policy_record(
            f"beam_{width}",
            int(search["mask"]),
            problem,
            drawing,
            verifier_module,
            verifier_analysis,
            {
                key: value
                for key, value in search.items()
                if key not in {"mask", "terminal"}
            },
        )
    policies["exact_or_bound_incumbent"] = policy_record(
        "exact_or_bound_incumbent",
        int(exact["mask"]),
        problem,
        drawing,
        verifier_module,
        verifier_analysis,
        {
            key: value
            for key, value in exact.items()
            if key not in {"mask", "terminal"}
        },
    )
    exact_or_upper_f1 = (
        float(exact["terminal"]["f1"])
        if exact.get("certified")
        else float(exact.get("f1_upper_bound", 1.0))
    )
    ordered_scores = [
        {"handle": handle, "probability": float(probability)}
        for handle, probability in zip(problem["handles"], problem["probabilities"])
    ]
    return {
        "drawing_id": drawing.ref.drawing_id,
        "family_id": drawing.ref.family_id,
        "tier": drawing.ref.tier,
        "candidate_handle_count": len(drawing.handles),
        "truth_handle_count": len(truth),
        "verifier_expected_handle_count": len(expected),
        "verifier_truth_alignment": True,
        "duplicate_segment_records_collapsed_for_features": drawing.duplicate_segment_records_collapsed,
        "missing_wanted_handles": drawing.missing_wanted_handles,
        "ordered_frozen_scores": ordered_scores,
        "policies": policies,
        "exact_or_upper_f1": exact_or_upper_f1,
        "beam64_minus_greedy_f1": float(
            policies["beam_64"]["set_metrics"]["f1"]
            - policies["greedy"]["set_metrics"]["f1"]
        ),
        "exact_or_upper_minus_beam64_f1": float(
            exact_or_upper_f1 - policies["beam_64"]["set_metrics"]["f1"]
        ),
        "greedy_matches_certified_exact_set": bool(
            exact.get("certified") and greedy == int(exact["mask"])
        ),
        "exact_or_bound": {
            key: json_ready(value)
            for key, value in exact.items()
            if key not in {"mask", "terminal"}
        }
        | {
            "incumbent_set_f1": float(exact["terminal"]["f1"]),
            "incumbent_objective": float(exact["terminal"]["objective"]),
            "incumbent_selected_handles": mask_handles(
                int(exact["mask"]), problem["handles"]
            ),
        },
    }


def pooled_policy_stats(rows: Sequence[Mapping[str, Any]], policy: str) -> dict[str, Any]:
    records = [row["policies"][policy] for row in rows]
    tp = sum(int(record["set_metrics"]["tp"]) for record in records)
    fp = sum(int(record["set_metrics"]["fp"]) for record in records)
    fn = sum(int(record["set_metrics"]["fn"]) for record in records)
    precision = tp / (tp + fp) if tp + fp else 0.0
    recall = tp / (tp + fn) if tp + fn else 0.0
    pooled_f1 = 2.0 * tp / (2 * tp + fp + fn) if 2 * tp + fp + fn else 0.0
    drawing_f1 = np.asarray(
        [record["set_metrics"]["f1"] for record in records], dtype=np.float64
    )
    acceptance = np.asarray(
        [bool(record["verifier_accepted"]) for record in records], dtype=np.float64
    )
    objectives = np.asarray(
        [record["terminal_objective"] for record in records], dtype=np.float64
    )
    return {
        "drawings": len(records),
        "tp": tp,
        "fp": fp,
        "fn": fn,
        "precision": precision,
        "recall": recall,
        "pooled_set_f1": pooled_f1,
        "drawing_set_f1_mean": float(np.mean(drawing_f1)),
        "drawing_set_f1_std_population": float(np.std(drawing_f1, ddof=0)),
        "drawing_set_f1_min": float(np.min(drawing_f1)),
        "drawing_set_f1_max": float(np.max(drawing_f1)),
        "verifier_accept_count": int(np.sum(acceptance)),
        "verifier_accept_rate": float(np.mean(acceptance)),
        "terminal_objective_mean": float(np.mean(objectives)),
    }


def bootstrap_deltas(rows: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    by_family = {
        family: [row for row in rows if row["family_id"] == family]
        for family in HIDDEN_FAMILIES
    }
    if any(len(values) != 25 for values in by_family.values()):
        raise RuntimeError("hidden family bootstrap cardinality failure")
    rng = np.random.default_rng(BOOTSTRAP_SEED)
    beam_values = np.empty(BOOTSTRAP_REPLICATES, dtype=np.float64)
    upper_values = np.empty(BOOTSTRAP_REPLICATES, dtype=np.float64)
    families = list(HIDDEN_FAMILIES)
    for replicate in range(BOOTSTRAP_REPLICATES):
        sampled = rng.integers(0, len(families), size=len(families))
        selected_rows: list[Mapping[str, Any]] = []
        for index in sampled:
            selected_rows.extend(by_family[families[int(index)]])
        beam_values[replicate] = float(
            np.mean([row["beam64_minus_greedy_f1"] for row in selected_rows])
        )
        upper_values[replicate] = float(
            np.mean(
                [row["exact_or_upper_minus_beam64_f1"] for row in selected_rows]
            )
        )

    def summary(values: np.ndarray, point: float) -> dict[str, Any]:
        return {
            "point_delta": point,
            "bootstrap_mean": float(np.mean(values)),
            "bootstrap_std_population": float(np.std(values, ddof=0)),
            "ci95_percentile": [
                float(np.percentile(values, 2.5)),
                float(np.percentile(values, 97.5)),
            ],
        }

    beam_point = float(np.mean([row["beam64_minus_greedy_f1"] for row in rows]))
    upper_point = float(
        np.mean([row["exact_or_upper_minus_beam64_f1"] for row in rows])
    )
    return {
        "unit": "hidden family cluster",
        "replicates": BOOTSTRAP_REPLICATES,
        "seed": BOOTSTRAP_SEED,
        "beam64_minus_greedy": summary(beam_values, beam_point),
        "exact_or_upper_minus_beam64": summary(upper_values, upper_point),
        "replicate_values": {
            "beam64_minus_greedy": beam_values.tolist(),
            "exact_or_upper_minus_beam64": upper_values.tolist(),
        },
    }


def aggregate_measurements(rows: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    policies = (
        "greedy",
        "beam_4",
        "beam_16",
        "beam_64",
        "exact_or_bound_incumbent",
    )
    by_family: dict[str, Any] = {}
    for family in HIDDEN_FAMILIES:
        family_rows = [row for row in rows if row["family_id"] == family]
        by_family[family] = {
            policy: pooled_policy_stats(family_rows, policy) for policy in policies
        }
    pooled = {policy: pooled_policy_stats(rows, policy) for policy in policies}
    bootstrap = bootstrap_deltas(rows)
    upper_rows = [row for row in rows if row["exact_or_bound"]["method"] != "exhaustive_exact"]
    return {
        "hidden_drawings": len(rows),
        "hidden_families": list(HIDDEN_FAMILIES),
        "by_family": by_family,
        "pooled": pooled,
        "bootstrap": bootstrap,
        "greedy_matches_certified_exact_count": sum(
            bool(row["greedy_matches_certified_exact_set"]) for row in rows
        ),
        "certified_optimum_count": sum(
            bool(row["exact_or_bound"]["certified"]) for row in rows
        ),
        "upper_bound_replacement_count": len(upper_rows),
        "upper_bound_replacement_drawings": [
            {
                "drawing_id": row["drawing_id"],
                "family_id": row["family_id"],
                "candidate_handle_count": row["candidate_handle_count"],
                **row["exact_or_bound"],
            }
            for row in upper_rows
        ],
    }


def _style_header(row: Iterable[Any]) -> None:
    from openpyxl.styles import Font, PatternFill

    for cell in row:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")


def build_evidence_workbook(result: Mapping[str, Any]) -> dict[str, Any]:
    from openpyxl import Workbook, load_workbook

    path = assert_write_allowed(EVIDENCE_PATH)
    workbook = Workbook()
    prereg_sheet = workbook.active
    prereg_sheet.title = "PREREG"
    prereg_sheet.append(["field", "value"])
    _style_header(prereg_sheet[1])
    prereg_sheet.append(["prereg_json_sha256", result["seals"]["prereg_json_sha256"]])
    prereg_sheet.append(["prereg_csv_sha256", result["seals"]["prereg_csv_sha256"]])
    prereg_sheet.append(["prereg_content_hash", result["seals"]["content_hash"]])
    prereg_sheet.append(["split_sha256", SPLIT_SHA256])
    prereg_sheet.append(["verifier_sha256", VERIFIER_SHA256])
    prereg_sheet.append(["scorer_fingerprint", result["scorer"]["ensemble_fingerprint_sha256"]])

    summary_sheet = workbook.create_sheet("Summary")
    summary_sheet.append(
        [
            "policy",
            "drawings",
            "tp",
            "fp",
            "fn",
            "precision",
            "recall",
            "pooled_set_f1",
            "drawing_set_f1_mean",
            "verifier_accept_count",
            "verifier_accept_rate",
            "terminal_objective_mean",
        ]
    )
    _style_header(summary_sheet[1])
    for policy, stats in result["measurements"]["pooled"].items():
        summary_sheet.append(
            [
                policy,
                stats["drawings"],
                stats["tp"],
                stats["fp"],
                stats["fn"],
                stats["precision"],
                stats["recall"],
                stats["pooled_set_f1"],
                stats["drawing_set_f1_mean"],
                stats["verifier_accept_count"],
                stats["verifier_accept_rate"],
                stats["terminal_objective_mean"],
            ]
        )

    family_sheet = workbook.create_sheet("Family")
    family_sheet.append(
        [
            "family_id",
            "policy",
            "drawings",
            "tp",
            "fp",
            "fn",
            "pooled_set_f1",
            "drawing_set_f1_mean",
            "verifier_accept_rate",
        ]
    )
    _style_header(family_sheet[1])
    for family, policies in result["measurements"]["by_family"].items():
        for policy, stats in policies.items():
            family_sheet.append(
                [
                    family,
                    policy,
                    stats["drawings"],
                    stats["tp"],
                    stats["fp"],
                    stats["fn"],
                    stats["pooled_set_f1"],
                    stats["drawing_set_f1_mean"],
                    stats["verifier_accept_rate"],
                ]
            )

    drawing_sheet = workbook.create_sheet("Drawing")
    drawing_sheet.append(
        [
            "drawing_id",
            "family_id",
            "tier",
            "candidate_handles",
            "truth_handles",
            "greedy_f1",
            "beam4_f1",
            "beam16_f1",
            "beam64_f1",
            "exact_incumbent_f1",
            "exact_or_upper_f1",
            "beam64_minus_greedy",
            "exact_or_upper_minus_beam64",
            "greedy_matches_certified_exact",
            "exact_method",
            "exact_certified",
            "objective_upper_bound",
            "objective_gap",
            "greedy_accept",
            "beam64_accept",
            "exact_incumbent_accept",
        ]
    )
    _style_header(drawing_sheet[1])
    for row in result["drawings"]:
        drawing_sheet.append(
            [
                row["drawing_id"],
                row["family_id"],
                row["tier"],
                row["candidate_handle_count"],
                row["truth_handle_count"],
                row["policies"]["greedy"]["set_metrics"]["f1"],
                row["policies"]["beam_4"]["set_metrics"]["f1"],
                row["policies"]["beam_16"]["set_metrics"]["f1"],
                row["policies"]["beam_64"]["set_metrics"]["f1"],
                row["policies"]["exact_or_bound_incumbent"]["set_metrics"]["f1"],
                row["exact_or_upper_f1"],
                row["beam64_minus_greedy_f1"],
                row["exact_or_upper_minus_beam64_f1"],
                row["greedy_matches_certified_exact_set"],
                row["exact_or_bound"]["method"],
                row["exact_or_bound"]["certified"],
                row["exact_or_bound"]["objective_upper_bound"],
                row["exact_or_bound"]["objective_gap"],
                row["policies"]["greedy"]["verifier_accepted"],
                row["policies"]["beam_64"]["verifier_accepted"],
                row["policies"]["exact_or_bound_incumbent"]["verifier_accepted"],
            ]
        )

    bootstrap_sheet = workbook.create_sheet("Bootstrap")
    bootstrap_sheet.append(
        ["replicate", "beam64_minus_greedy", "exact_or_upper_minus_beam64"]
    )
    _style_header(bootstrap_sheet[1])
    bootstrap = result["measurements"]["bootstrap"]["replicate_values"]
    for index, (beam, upper) in enumerate(
        zip(
            bootstrap["beam64_minus_greedy"],
            bootstrap["exact_or_upper_minus_beam64"],
        ),
        1,
    ):
        bootstrap_sheet.append([index, beam, upper])

    upper_sheet = workbook.create_sheet("UpperBounds")
    upper_sheet.append(
        [
            "drawing_id",
            "family_id",
            "candidate_handles",
            "method",
            "replacement_reason",
            "certified",
            "evaluated_nodes",
            "pruned_nodes",
            "remaining_frontier",
            "objective_upper_bound",
            "objective_gap",
            "f1_upper_bound",
            "stop_reason",
            "cpu_seconds",
        ]
    )
    _style_header(upper_sheet[1])
    for row in result["measurements"]["upper_bound_replacement_drawings"]:
        upper_sheet.append(
            [
                row.get("drawing_id"),
                row.get("family_id"),
                row.get("candidate_handle_count"),
                row.get("method"),
                row.get("replacement_reason"),
                row.get("certified"),
                row.get("evaluated_nodes"),
                row.get("pruned_nodes"),
                row.get("remaining_frontier"),
                row.get("objective_upper_bound"),
                row.get("objective_gap"),
                row.get("f1_upper_bound"),
                row.get("stop_reason"),
                row.get("cpu_seconds"),
            ]
        )

    model_sheet = workbook.create_sheet("Scorer")
    model_sheet.append(
        [
            "seed",
            "estimator_sha256",
            "training_prediction_sha256",
            "n_iter",
            "fit_wall_seconds",
            "fit_cpu_seconds",
            "warning_count",
        ]
    )
    _style_header(model_sheet[1])
    for row in result["scorer"]["model_records"]:
        model_sheet.append(
            [
                row["seed"],
                row["estimator_sha256"],
                row["training_prediction_sha256"],
                row["n_iter"],
                row["fit_wall_seconds"],
                row["fit_cpu_seconds"],
                row["warning_count"],
            ]
        )

    selftest_sheet = workbook.create_sheet("Selftest")
    selftest_sheet.append(["field", "value"])
    _style_header(selftest_sheet[1])
    selftest_sheet.append(["selftest_ok", result["selftest"]["selftest_ok"]])
    selftest_sheet.append(
        [
            "search_signature",
            result["selftest"]["search_determinism"]["first"]["sha256"],
        ]
    )
    selftest_sheet.append(
        [
            "scorer_fingerprint",
            result["selftest"]["scorer_freeze"][
                "first_ensemble_fingerprint_sha256"
            ],
        ]
    )
    for name, denied in result["selftest"]["boundary"]["denied"].items():
        selftest_sheet.append([f"boundary_denied_{name}", denied])

    input_sheet = workbook.create_sheet("Inputs")
    input_sheet.append(["name", "sha256"])
    _style_header(input_sheet[1])
    for name, digest in result["seals"]["inputs"]["files"].items():
        input_sheet.append([name, digest])
    input_sheet.append(
        ["pack_inventory", result["seals"]["inputs"]["pack_inventory"]["sha256"]]
    )

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    workbook.save(path)
    reloaded = load_workbook(path, read_only=True, data_only=False)
    names = reloaded.sheetnames
    formula_cells = 0
    for sheet in reloaded.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_cells += 1
    reloaded.close()
    return {
        "path": str(path),
        "worksheets": names,
        "formula_cells": formula_cells,
        "sha256": file_sha256(path),
    }


def fmt(value: float) -> str:
    return f"{float(value):.9f}"


def render_report(
    result: Mapping[str, Any], results_sha256: str, evidence_sha256: str
) -> str:
    prereg = load_json(PREREG_PATH)["frozen"]
    lines = [
        "# G9/A63 set-assembly enumeration diagnostic",
        "",
        f"PREREG_JSON_SHA256: {result['seals']['prereg_json_sha256']}",
        f"PREREG_CSV_SHA256: {result['seals']['prereg_csv_sha256']}",
        f"PREREG_CONTENT_HASH: {result['seals']['content_hash']}",
        "",
        "This report contains numeric measurement and search-accounting output only. "
        "It emits no RL kill or survival judgment.",
        "",
        "## Sealed scope and scorer freeze",
        "",
        f"- synthetic drawings: 200 across 8 families; reward fit 125, hidden evaluation 75",
        f"- family split SHA-256: {SPLIT_SHA256}",
        f"- pack inventory: {PACK_INVENTORY_FILE_COUNT} files; SHA-256 {PACK_INVENTORY_SHA256}",
        f"- verifier SHA-256: {VERIFIER_SHA256}",
        f"- verifier configuration: angle 5.0 degrees; overlap 0.65; thickness 50-400 mm",
        f"- scorer features: {len(FULL_FEATURE_NAMES)} W2-02 full two-hop columns",
        f"- reward handle rows: {result['scorer']['training_rows']}",
        f"- reward positive rows: {result['scorer']['positive_rows']}",
        f"- scorer ensemble fingerprint: {result['scorer']['ensemble_fingerprint_sha256']}",
        f"- hidden content opened only after freeze: {result['family_firewall']['hidden_opened_after_scorer_freeze']}",
        "- original CAD reads: 0; repository test reads: 0; val-B reads: 0; CubiCasa reads: 0",
        "- training updates after scorer freeze: 0; GNN use: 0; subagents used: 0",
        "",
        "## Sealed policy definitions",
        "",
        f"- terminal objective: {prereg['terminal_objective']['formula']}",
        f"- greedy: {prereg['policies']['greedy']}",
        f"- beam: {prereg['policies']['beam']}",
        f"- exact: {prereg['policies']['exact']}",
        f"- deterministic upper-bound replacement: {prereg['policies']['upper_bound_replacement']}",
        "",
        "## Hidden pooled measurements",
        "",
        "| policy | TP | FP | FN | precision | recall | pooled set-F1 | drawing mean set-F1 | verifier accepted | acceptance rate | terminal objective mean |",
        "|---|---:|---:|---:|---:|---:|---:|---:|---:|---:|---:|",
    ]
    for policy, stats in result["measurements"]["pooled"].items():
        lines.append(
            f"| {policy} | {stats['tp']} | {stats['fp']} | {stats['fn']} | "
            f"{fmt(stats['precision'])} | {fmt(stats['recall'])} | "
            f"{fmt(stats['pooled_set_f1'])} | {fmt(stats['drawing_set_f1_mean'])} | "
            f"{stats['verifier_accept_count']}/{stats['drawings']} | "
            f"{fmt(stats['verifier_accept_rate'])} | "
            f"{fmt(stats['terminal_objective_mean'])} |"
        )
    lines += ["", "## Hidden family measurements", ""]
    lines += [
        "| family | policy | TP | FP | FN | pooled set-F1 | drawing mean set-F1 | verifier acceptance rate |",
        "|---|---|---:|---:|---:|---:|---:|---:|",
    ]
    for family, policies in result["measurements"]["by_family"].items():
        for policy, stats in policies.items():
            lines.append(
                f"| {family} | {policy} | {stats['tp']} | {stats['fp']} | {stats['fn']} | "
                f"{fmt(stats['pooled_set_f1'])} | {fmt(stats['drawing_set_f1_mean'])} | "
                f"{fmt(stats['verifier_accept_rate'])} |"
            )
    bootstrap = result["measurements"]["bootstrap"]
    lines += [
        "",
        "## Family-cluster bootstrap deltas",
        "",
        f"- replicates: {bootstrap['replicates']}; seed: {bootstrap['seed']}; hidden family clusters: 3",
        "",
        "| delta | point | bootstrap mean | bootstrap population SD | 95% percentile CI |",
        "|---|---:|---:|---:|---:|",
    ]
    for label, key in (
        ("beam64 minus greedy", "beam64_minus_greedy"),
        ("exact or upper minus beam64", "exact_or_upper_minus_beam64"),
    ):
        stats = bootstrap[key]
        lines.append(
            f"| {label} | {fmt(stats['point_delta'])} | {fmt(stats['bootstrap_mean'])} | "
            f"{fmt(stats['bootstrap_std_population'])} | "
            f"[{fmt(stats['ci95_percentile'][0])}, {fmt(stats['ci95_percentile'][1])}] |"
        )
    lines += [
        "",
        "## Exact and upper-bound accounting",
        "",
        f"- exhaustive-exact drawing count: {75 - result['measurements']['upper_bound_replacement_count']}",
        f"- deterministic B&B replacement drawing count: {result['measurements']['upper_bound_replacement_count']}",
        f"- certified optimum count: {result['measurements']['certified_optimum_count']}",
        f"- greedy selected set equals certified optimum: {result['measurements']['greedy_matches_certified_exact_count']}/75",
        "",
        "| drawing | family | handles | method | certified | objective upper | gap | stop | CPU-s |",
        "|---|---|---:|---|---|---:|---:|---|---:|",
    ]
    for row in result["measurements"]["upper_bound_replacement_drawings"]:
        lines.append(
            f"| {row['drawing_id']} | {row['family_id']} | {row['candidate_handle_count']} | "
            f"{row['method']} | {str(row['certified']).lower()} | "
            f"{fmt(row['objective_upper_bound'])} | {fmt(row['objective_gap'])} | "
            f"{row['stop_reason']} | {float(row['cpu_seconds']):.6f} |"
        )
    lines += [
        "",
        "## CPU and deterministic checks",
        "",
        f"- cumulative CPU-h: {result['cpu']['total_cpu_hours']:.9f} / {CPU_CAP_HOURS:.1f}",
        f"- measurement wall-s: {result['cpu']['wall_seconds']:.6f}",
        f"- selftest_ok: {str(result['selftest']['selftest_ok']).lower()}",
        f"- search signature SHA-256: {result['selftest']['search_determinism']['first']['sha256']}",
        f"- scorer repeat fingerprint equal: {str(result['selftest']['scorer_freeze']['ok']).lower()}",
        f"- boundary denials: {canonical_json(result['selftest']['boundary']['denied'])}",
        "- reproducibility wording: 휘발 필드 제외 수치 전 필드 동일",
        "",
        "## Artifact hashes",
        "",
        "| artifact | SHA-256 |",
        "|---|---|",
        f"| g9_rl_diag.py | {result['artifacts']['g9_rl_diag_py_sha256']} |",
        f"| prereg.json | {result['seals']['prereg_json_sha256']} |",
        f"| PREREG.csv | {result['seals']['prereg_csv_sha256']} |",
        f"| results.json | {results_sha256} |",
        f"| evidence.xlsx | {evidence_sha256} |",
        "",
        "## Unresolved and disclosures",
        "",
        "- All hidden drawings exceeded the exhaustive handle cap if listed above; their deterministic B&B incumbent/bound records are not relabeled as exhaustive enumeration.",
        "- Feature rows use the sealed first-valid-segment handle adapter; duplicate curved/polyline segment records remain present in full verifier IR and are counted per drawing in results.json.",
        "- Process exception initial_git_probe: repository grounding invoked read-only Git status probes before/while the packet text was first loaded. No Git mutation occurred; no later Git command was used.",
        "- Process exception superseded_pre_measurement_seal: the first dual seal was preserved and superseded before any pack-content read because its validator used a different JSON key order from G1's published configuration seal text.",
        "- No subagent was created or used.",
        "- No RL track judgment is emitted here; the orchestrator owns the sealed band comparison.",
        "",
        "CELL_COMPLETE: g9_rl_diag",
    ]
    return "\n".join(lines) + "\n"


def execute(prior_cpu_seconds: float = 0.0) -> dict[str, Any]:
    global _hidden_content_open
    wall_started = time.perf_counter()
    budget = CpuBudget(prior_cpu_seconds)
    seals = validate_seals_and_inputs()
    refs = load_drawing_refs()
    reward_refs = [ref for ref in refs if ref.family_id in REWARD_FAMILIES]
    hidden_refs = [ref for ref in refs if ref.family_id in HIDDEN_FAMILIES]
    if len(reward_refs) != 125 or len(hidden_refs) != 75:
        raise RuntimeError("family split drawing count failure")
    reward_drawings = load_drawings(reward_refs, budget, "reward")
    selftest, models, scorer = run_selftests_with_reward(reward_drawings, budget)
    freeze_event = {
        "event": "SCORER_FROZEN_BEFORE_HIDDEN_OPEN",
        "at_utc": utc_now(),
        "ensemble_fingerprint_sha256": scorer["ensemble_fingerprint_sha256"],
        "input_hashes_after_fit": current_input_hashes(),
        "training_updates_after_event": 0,
    }
    if freeze_event["input_hashes_after_fit"] != seals["inputs"]:
        raise RuntimeError("input drift at scorer freeze event")
    _hidden_content_open = True
    hidden_drawings = load_drawings(hidden_refs, budget, "hidden")
    verifier_module = import_verifier()
    drawing_rows: list[dict[str, Any]] = []
    for index, drawing in enumerate(hidden_drawings, 1):
        drawing_rows.append(
            evaluate_hidden_drawing(drawing, models, verifier_module, budget)
        )
        if index % 10 == 0 or index == len(hidden_drawings):
            print(f"[search] evaluated {index}/{len(hidden_drawings)} hidden drawings", flush=True)
    if current_input_hashes() != seals["inputs"]:
        raise RuntimeError("sealed input drift after measurement")
    measurements = aggregate_measurements(drawing_rows)
    result: dict[str, Any] = {
        "schema": SCHEMA,
        "cell": "g9_rl_diag",
        "status": "COMPLETE_NUMERIC",
        "completed_at_utc": utc_now(),
        "seals": seals,
        "scope": {
            "synthetic_only": True,
            "reward_drawings": len(reward_drawings),
            "hidden_drawings": len(hidden_drawings),
            "original_cad_reads": 0,
            "repository_test_reads": 0,
            "val_b_reads": 0,
            "cubicasa_reads": 0,
            "GNN_used": False,
            "training_after_scorer_freeze": False,
            "subagents_used": 0,
            "repository_modified": False,
            "git_mutation": False,
        },
        "family_firewall": {
            "reward_families": list(REWARD_FAMILIES),
            "hidden_families": list(HIDDEN_FAMILIES),
            "split_sha256": SPLIT_SHA256,
            "freeze_event": freeze_event,
            "hidden_opened_after_scorer_freeze": True,
            "boundary_counters": dict(_boundary_counters),
        },
        "scorer": scorer,
        "selftest": selftest,
        "drawings": drawing_rows,
        "measurements": measurements,
        "cpu": {
            "prior_cpu_seconds": prior_cpu_seconds,
            "run_cpu_seconds": budget.run_cpu_seconds,
            "total_cpu_seconds": budget.total_cpu_seconds,
            "total_cpu_hours": budget.total_cpu_seconds / 3600.0,
            "cap_cpu_hours": CPU_CAP_HOURS,
            "wall_seconds": time.perf_counter() - wall_started,
        },
        "runtime": {
            "python": sys.version,
            "platform": platform.platform(),
            "numpy": np.__version__,
            "sklearn_estimator": "HistGradientBoostingClassifier",
            "thread_environment": {
                name: os.environ.get(name)
                for name in (
                    "OMP_NUM_THREADS",
                    "OPENBLAS_NUM_THREADS",
                    "MKL_NUM_THREADS",
                    "NUMEXPR_NUM_THREADS",
                )
            },
        },
        "artifacts": {
            "g9_rl_diag_py_sha256": file_sha256(Path(__file__).resolve()),
        },
        "process_exceptions": [
            {
                "id": "initial_git_probe",
                "mutation": False,
                "description": (
                    "read-only Git status probes occurred during initial grounding before/while "
                    "the packet prohibition was first loaded; no later Git command was used"
                ),
            }
            ,
            {
                "id": "superseded_pre_measurement_seal",
                "mutation": False,
                "pack_content_reads_before_supersession": 0,
                "hidden_content_reads_before_supersession": 0,
                "superseded_prereg_json_sha256": "f41346a7774fe04d722ea52ecf90d1b987631194682c0fd0df68f2385bc027ce",
                "superseded_prereg_csv_sha256": "b2732d2d749405d349dee366c7629043011f8a9f9f98b98e1187f4475004c794",
                "superseded_content_hash": "d09980cdefc891c779866f1ae002a62f91d021c1348d22aa921f0edfe49916ba",
                "description": (
                    "superseded before measurement because the local validator canonicalized "
                    "the G1 configuration in a different key order from its published seal"
                ),
            }
        ],
        "reproducibility_claim": "휘발 필드 제외 수치 전 필드 동일",
        "judgment": None,
    }
    workbook = build_evidence_workbook(result)
    result["artifacts"]["evidence"] = workbook
    result["cpu"].update(
        {
            "run_cpu_seconds": budget.run_cpu_seconds,
            "total_cpu_seconds": budget.total_cpu_seconds,
            "total_cpu_hours": budget.total_cpu_seconds / 3600.0,
            "wall_seconds": time.perf_counter() - wall_started,
        }
    )
    write_json(RESULTS_PATH, result)
    results_sha = file_sha256(RESULTS_PATH)
    report = render_report(result, results_sha, workbook["sha256"])
    write_text(REPORT_PATH, report)
    if not report.rstrip().endswith("CELL_COMPLETE: g9_rl_diag"):
        raise RuntimeError("report terminator failure")
    return {
        "status": result["status"],
        "results": str(RESULTS_PATH),
        "results_sha256": results_sha,
        "evidence": str(EVIDENCE_PATH),
        "evidence_sha256": workbook["sha256"],
        "report": str(REPORT_PATH),
        "report_sha256": file_sha256(REPORT_PATH),
        "cpu_hours": result["cpu"]["total_cpu_hours"],
        "hidden_drawings": len(drawing_rows),
        "scorer_fingerprint": scorer["ensemble_fingerprint_sha256"],
    }


def standalone_selftest(prior_cpu_seconds: float = 0.0) -> dict[str, Any]:
    budget = CpuBudget(prior_cpu_seconds)
    seals = validate_seals_and_inputs()
    refs = load_drawing_refs()
    reward_refs = [ref for ref in refs if ref.family_id in REWARD_FAMILIES]
    reward_drawings = load_drawings(reward_refs, budget, "selftest-reward")
    selftest, _models, scorer = run_selftests_with_reward(reward_drawings, budget)
    if _boundary_counters["hidden_content_reads_after_freeze"] != 0:
        raise RuntimeError("standalone selftest opened hidden content")
    return {
        **selftest,
        "seals": {
            "prereg_json_sha256": seals["prereg_json_sha256"],
            "prereg_csv_sha256": seals["prereg_csv_sha256"],
        },
        "actual_scorer_fingerprint": scorer["ensemble_fingerprint_sha256"],
        "cpu_seconds": budget.total_cpu_seconds,
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="G9/A63 zero-learning set-assembly enumeration diagnostic"
    )
    modes = parser.add_mutually_exclusive_group(required=True)
    modes.add_argument("--seal", action="store_true", help="write dual preregistration seals")
    modes.add_argument("--selftest", action="store_true", help="run sealed deterministic/boundary tests")
    modes.add_argument("--run", action="store_true", help="execute full reward-fit and hidden diagnostic")
    parser.add_argument(
        "--prior-cpu-seconds",
        type=float,
        default=0.0,
        help="cumulative CPU seconds from a prior failed/resumed invocation",
    )
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    try:
        if args.seal:
            output = seal()
        elif args.selftest:
            output = standalone_selftest(args.prior_cpu_seconds)
        else:
            output = execute(args.prior_cpu_seconds)
    except BudgetKill as exc:
        print(f"BUDGET_KILL: {exc}", file=sys.stderr)
        return 3
    except Exception as exc:
        print(f"ERROR: {type(exc).__name__}: {exc}", file=sys.stderr)
        traceback.print_exc()
        return 2
    print(json.dumps(json_ready(output), ensure_ascii=False, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
