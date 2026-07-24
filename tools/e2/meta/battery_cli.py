#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""CARD S5-D — metamorphic battery runner CLI.

run --drawings DIR --staging DIR --detector-cmd "CMD {dxf} {out}"
    --transforms rotate,mirror,scale,units,explode,rename
    --budget-drawings N --timeout-s T --out-xlsx R.xlsx

Imports sibling modules by file path (no package __init__.py):
  transforms_rigid.py, transforms_struct.py, invariance.py
Missing siblings -> print names and exit 3 (parallel-build contract).

Originals are READ-ONLY: transforms always write under --staging.
"""
from __future__ import annotations

import argparse
import importlib.util
import json
import os
import shutil
import subprocess
import sys
import tempfile
import time
import traceback
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

# ---------------------------------------------------------------------------
# Sibling path imports
# ---------------------------------------------------------------------------

_META_DIR = Path(__file__).resolve().parent
_SIBLING_NAMES = ("transforms_rigid.py", "transforms_struct.py", "invariance.py")

RIGID_KINDS = frozenset({"rotate", "translate", "mirror", "scale", "units"})
STRUCT_KINDS = frozenset({"explode", "rename"})
KNOWN_TRANSFORMS = frozenset(
    {"rotate", "mirror", "scale", "units", "explode", "rename", "translate"}
)

DEFAULT_PARAMS: Dict[str, Dict[str, Any]] = {
    "rotate": {"angle": 90.0},
    "mirror": {"axis": "x"},
    "scale": {"factor": 2.0},
    "units": {},
    "translate": {"dx": 100.0, "dy": 50.0},
    "explode": {},
    "rename": {"scheme": "anonymize"},
}


def _load_module(name: str, path: Path):
    """Load a plain .py sibling via importlib (no package)."""
    spec = importlib.util.spec_from_file_location(name, str(path))
    if spec is None or spec.loader is None:
        raise ImportError(f"cannot create spec for {path}")
    mod = importlib.util.module_from_spec(spec)
    sys.modules[name] = mod
    spec.loader.exec_module(mod)
    return mod


def load_siblings(
    meta_dir: Optional[Path] = None,
) -> Tuple[Any, Any, Any, List[str]]:
    """Return (rigid, struct, inv, missing_names).

    missing_names non-empty means callers of `run` must exit 3.
    """
    base = meta_dir or _META_DIR
    missing: List[str] = []
    paths = {n: base / n for n in _SIBLING_NAMES}
    for n, p in paths.items():
        if not p.is_file():
            missing.append(n)
    if missing:
        return None, None, None, missing
    rigid = _load_module("e2_meta_transforms_rigid", paths["transforms_rigid.py"])
    struct = _load_module("e2_meta_transforms_struct", paths["transforms_struct.py"])
    inv = _load_module("e2_meta_invariance", paths["invariance.py"])
    return rigid, struct, inv, []


# ---------------------------------------------------------------------------
# Detector subprocess
# ---------------------------------------------------------------------------

def _format_detector_cmd(template: str, dxf: Path, out: Path) -> str:
    return template.format(dxf=str(dxf), out=str(out))


def run_detector(cmd_template: str, dxf: Path, out: Path, timeout_s: float) -> Dict[str, Any]:
    """Run detector-cmd once; write prediction JSON to out; return parsed dict."""
    out.parent.mkdir(parents=True, exist_ok=True)
    cmd_str = _format_detector_cmd(cmd_template, dxf, out)
    # Use shell=True so user templates with quotes/redirects work as documented.
    proc = subprocess.run(
        cmd_str,
        shell=True,
        capture_output=True,
        text=True,
        timeout=timeout_s,
    )
    if proc.returncode != 0:
        raise RuntimeError(
            f"detector exit {proc.returncode}: stderr={proc.stderr[:500]!r} "
            f"stdout={proc.stdout[:500]!r}"
        )
    if not out.is_file():
        raise RuntimeError(f"detector did not write prediction file: {out}")
    with out.open("r", encoding="utf-8") as f:
        return json.load(f)


def _wall_count(pred: Dict[str, Any], threshold: float = 0.5) -> int:
    per = (pred.get("scores") or {}).get("per_handle") or {}
    n = 0
    for _h, rec in per.items():
        if isinstance(rec, dict):
            score = float(rec.get("score", 0.0))
        else:
            score = float(rec)
        if score >= threshold:
            n += 1
    return n


def _n_handles(pred: Dict[str, Any]) -> int:
    per = (pred.get("scores") or {}).get("per_handle") or {}
    return len(per)


def _normalize_handle_map(meta: Any) -> Any:
    """Pass through dict maps; treat 'identity'/None as identity for compare."""
    if meta is None:
        return "identity"
    if isinstance(meta, dict):
        if meta.get("handle_map") is not None:
            return meta["handle_map"]
        # explode/rename may return the map directly
        return meta
    return meta


# ---------------------------------------------------------------------------
# Apply one transform via siblings
# ---------------------------------------------------------------------------

def apply_transform(
    rigid: Any,
    struct: Any,
    kind: str,
    dxf_in: Path,
    dxf_out: Path,
    seed: int = 0,
) -> Any:
    """Write transformed DXF to dxf_out (never in-place). Return handle_map-ish."""
    dxf_out.parent.mkdir(parents=True, exist_ok=True)
    params = dict(DEFAULT_PARAMS.get(kind, {}))
    if kind in RIGID_KINDS:
        meta = rigid.transform(str(dxf_in), str(dxf_out), kind, params, seed)
        return _normalize_handle_map(meta)
    if kind == "explode":
        hmap = struct.explode(str(dxf_in), str(dxf_out))
        return hmap if hmap is not None else "identity"
    if kind == "rename":
        scheme = params.get("scheme", "anonymize")
        # rename_layers returns layer_map; handles unchanged -> identity for compare
        _layer_map = struct.rename_layers(str(dxf_in), str(dxf_out), scheme, seed)
        return "identity"
    raise ValueError(f"unknown transform kind: {kind}")


# ---------------------------------------------------------------------------
# Battery run
# ---------------------------------------------------------------------------

def collect_drawings(drawings_dir: Path, budget: int) -> List[Path]:
    files = sorted(drawings_dir.glob("*.dxf")) + sorted(drawings_dir.glob("*.DXF"))
    # de-dupe case on Windows
    seen = set()
    out: List[Path] = []
    for p in files:
        key = str(p.resolve()).lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(p)
        if len(out) >= budget:
            break
    return out


def run_battery(
    *,
    drawings_dir: Path,
    staging_dir: Path,
    detector_cmd: str,
    transforms: Sequence[str],
    budget_drawings: int,
    timeout_s: float,
    out_xlsx: Path,
    rigid: Any,
    struct: Any,
    inv: Any,
    seed: int = 0,
) -> Dict[str, Any]:
    drawings = collect_drawings(drawings_dir, budget_drawings)
    staging_dir.mkdir(parents=True, exist_ok=True)
    rows: List[Dict[str, Any]] = []

    for di, dxf in enumerate(drawings):
        stem = dxf.stem
        for ti, kind in enumerate(transforms):
            if kind not in KNOWN_TRANSFORMS:
                rows.append(
                    {
                        "drawing": dxf.name,
                        "transform": kind,
                        "invariance": None,
                        "flips_count": None,
                        "sentinel_zero": None,
                        "sentinel_all": None,
                        "wall_count_before": None,
                        "wall_count_after": None,
                        "timing_s": None,
                        "error": f"unknown transform: {kind}",
                    }
                )
                continue

            t0 = time.perf_counter()
            err: Optional[str] = None
            inv_score: Optional[float] = None
            flips_count: Optional[int] = None
            sz_flag: Optional[bool] = None
            sa_flag: Optional[bool] = None
            wc_b: Optional[int] = None
            wc_a: Optional[int] = None
            try:
                staged = staging_dir / f"{stem}__{kind}__{di}_{ti}.dxf"
                handle_map = apply_transform(rigid, struct, kind, dxf, staged, seed=seed)

                pred_dir = staging_dir / "_preds" / stem / kind
                pred_dir.mkdir(parents=True, exist_ok=True)
                pred_before_path = pred_dir / "before.json"
                pred_after_path = pred_dir / "after.json"

                pred_before = run_detector(detector_cmd, dxf, pred_before_path, timeout_s)
                pred_after = run_detector(detector_cmd, staged, pred_after_path, timeout_s)

                cmp = inv.compare(pred_before, pred_after, handle_map)
                inv_score = float(cmp.get("invariance", 0.0))
                flips = cmp.get("flips") or []
                flips_count = len(flips)

                wc_b = _wall_count(pred_before)
                wc_a = _wall_count(pred_after)
                n_h = max(_n_handles(pred_before), _n_handles(pred_after), 1)

                sz_flag = bool(inv.sentinel_zero(pred_before)) or bool(
                    inv.sentinel_zero(pred_after)
                )
                sa_flag = bool(inv.sentinel_all(pred_before, n_h)) or bool(
                    inv.sentinel_all(pred_after, n_h)
                )
            except subprocess.TimeoutExpired as e:
                err = f"timeout: {e}"
            except Exception as e:
                err = f"{type(e).__name__}: {e}"
            timing = time.perf_counter() - t0
            rows.append(
                {
                    "drawing": dxf.name,
                    "transform": kind,
                    "invariance": inv_score,
                    "flips_count": flips_count,
                    "sentinel_zero": sz_flag,
                    "sentinel_all": sa_flag,
                    "wall_count_before": wc_b,
                    "wall_count_after": wc_a,
                    "timing_s": round(timing, 6),
                    "error": err,
                }
            )

    write_xlsx(rows, out_xlsx, list(transforms))
    return {"n_drawings": len(drawings), "n_rows": len(rows), "rows": rows, "out_xlsx": str(out_xlsx)}


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------

def write_xlsx(rows: List[Dict[str, Any]], out_xlsx: Path, transforms: List[str]) -> None:
    import openpyxl
    from openpyxl import Workbook

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "results"
    headers = [
        "drawing",
        "transform",
        "invariance",
        "flips_count",
        "sentinel_zero",
        "sentinel_all",
        "wall_count_before",
        "wall_count_after",
        "timing_s",
        "error",
    ]
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h) for h in headers])

    # summary: per-transform means
    ws2 = wb.create_sheet("summary")
    ws2.append(
        [
            "transform",
            "n_rows",
            "mean_invariance",
            "mean_flips_count",
            "mean_timing_s",
            "n_sentinel_zero",
            "n_sentinel_all",
            "n_errors",
        ]
    )
    by_t: Dict[str, List[Dict[str, Any]]] = {}
    for r in rows:
        by_t.setdefault(r["transform"], []).append(r)
    order = list(transforms) + [t for t in by_t if t not in transforms]
    for t in order:
        group = by_t.get(t, [])
        invs = [float(g["invariance"]) for g in group if g.get("invariance") is not None]
        flips = [float(g["flips_count"]) for g in group if g.get("flips_count") is not None]
        times = [float(g["timing_s"]) for g in group if g.get("timing_s") is not None]
        n_sz = sum(1 for g in group if g.get("sentinel_zero"))
        n_sa = sum(1 for g in group if g.get("sentinel_all"))
        n_err = sum(1 for g in group if g.get("error"))
        ws2.append(
            [
                t,
                len(group),
                (sum(invs) / len(invs)) if invs else None,
                (sum(flips) / len(flips)) if flips else None,
                (sum(times) / len(times)) if times else None,
                n_sz,
                n_sa,
                n_err,
            ]
        )

    wb.save(str(out_xlsx))


# ---------------------------------------------------------------------------
# Selftest stubs (only used when siblings absent — parallel-build isolation)
# ---------------------------------------------------------------------------

_STUB_RIGID = r'''
import shutil
from pathlib import Path

def transform(dxf_in, dxf_out, kind, params, seed):
    Path(dxf_out).parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(dxf_in, dxf_out)
    return {"kind": kind, "params": params or {}, "handle_map": "identity"}
'''

_STUB_STRUCT = r'''
import shutil
from pathlib import Path

def explode(dxf_in, dxf_out):
    Path(dxf_out).parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(dxf_in, dxf_out)
    return "identity"

def rename_layers(dxf_in, dxf_out, scheme, seed):
    Path(dxf_out).parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(dxf_in, dxf_out)
    return {}
'''

_STUB_INV = r'''
def compare(pred_before, pred_after, handle_map):
    # Identity mock: same fixed predictions -> perfect invariance, no flips.
    return {"invariance": 1.0, "flips": []}

def sentinel_zero(pred):
    per = (pred.get("scores") or {}).get("per_handle") or {}
    walls = 0
    for rec in per.values():
        s = float(rec.get("score", 0.0) if isinstance(rec, dict) else rec)
        if s >= 0.5:
            walls += 1
    return walls == 0

def sentinel_all(pred, n_handles):
    per = (pred.get("scores") or {}).get("per_handle") or {}
    n = max(int(n_handles), 1)
    walls = 0
    for rec in per.values():
        s = float(rec.get("score", 0.0) if isinstance(rec, dict) else rec)
        if s >= 0.5:
            walls += 1
    return (walls / n) > 0.9

def recall_floor(pred, truth_ledger):
    truth_h = set(truth_ledger.get("wall_handles_flat") or [])
    if not truth_h:
        return 1.0
    per = (pred.get("scores") or {}).get("per_handle") or {}
    hit = 0
    for h in truth_h:
        rec = per.get(h)
        if rec is None:
            continue
        s = float(rec.get("score", 0.0) if isinstance(rec, dict) else rec)
        if s >= 0.5:
            hit += 1
    return hit / len(truth_h)

def verdict(inv, zero_flag, all_flag, recall, inv_band=0.9, recall_floor_band=None):
    if zero_flag or all_flag:
        return "FAIL"
    if inv < inv_band:
        return "FAIL"
    if recall_floor_band is not None and recall is not None and recall < recall_floor_band:
        return "FAIL"
    return "PASS"
'''


def _install_selftest_stubs(tmpdir: Path) -> Tuple[Any, Any, Any]:
    """Write stub siblings into OS temp and load them (never into the repo)."""
    (tmpdir / "transforms_rigid.py").write_text(_STUB_RIGID, encoding="utf-8")
    (tmpdir / "transforms_struct.py").write_text(_STUB_STRUCT, encoding="utf-8")
    (tmpdir / "invariance.py").write_text(_STUB_INV, encoding="utf-8")
    rigid, struct, inv, missing = load_siblings(tmpdir)
    if missing:
        raise RuntimeError(f"stub install failed, missing: {missing}")
    return rigid, struct, inv


def _make_fixture_dxf(path: Path) -> None:
    import ezdxf

    doc = ezdxf.new("R2010")
    msp = doc.modelspace()
    msp.add_line((0, 0), (3000, 0), dxfattribs={"layer": "WALL"})
    msp.add_line((0, 0), (0, 2400), dxfattribs={"layer": "WALL"})
    msp.add_line((100, 100), (200, 100), dxfattribs={"layer": "MISC"})
    path.parent.mkdir(parents=True, exist_ok=True)
    doc.saveas(str(path))


def _mock_detector_cmd(
    python_exe: str, fixed_pred: Dict[str, Any], work_dir: Path
) -> str:
    """Tiny inline-ish python script that echoes a fixed prediction JSON to {out}.

    Written to OS temp (not the repo) so Windows shell quoting cannot mangle the
    payload. The returned template still uses {dxf} and {out} placeholders.
    """
    script = work_dir / "_mock_detector.py"
    pred_path = work_dir / "_fixed_pred.json"
    pred_path.write_text(json.dumps(fixed_pred), encoding="utf-8")
    script.write_text(
        "import json, sys\n"
        f"PRED = r'''{pred_path.as_posix()}'''\n"
        "dxf, out = sys.argv[1], sys.argv[2]\n"
        "data = json.load(open(PRED, encoding='utf-8'))\n"
        "open(out, 'w', encoding='utf-8').write(json.dumps(data))\n",
        encoding="utf-8",
    )
    # Quote paths for shell=True; placeholders remain unexpanded until format().
    return f'"{python_exe}" "{script}" "{{dxf}}" "{{out}}"'


def selftest() -> int:
    print("SELFTEST S5-D battery_cli starting")
    python_exe = sys.executable
    used_stubs = False
    rigid, struct, inv, missing = load_siblings()
    stub_dir_ctx = None
    try:
        if missing:
            print(f"NOTE: siblings missing at import ({', '.join(missing)}); "
                  f"installing temp stubs for selftest wiring check")
            stub_dir_ctx = tempfile.TemporaryDirectory(prefix="s5d_stubs_")
            rigid, struct, inv = _install_selftest_stubs(Path(stub_dir_ctx.name))
            used_stubs = True
        else:
            print("siblings loaded from meta dir")

        fixed_pred = {
            "scores": {
                "per_handle": {
                    "A1": {"score": 0.9},
                    "A2": {"score": 0.85},
                    "B1": {"score": 0.1},
                }
            }
        }

        with tempfile.TemporaryDirectory(prefix="s5d_battery_") as td:
            td_path = Path(td)
            drawings = td_path / "drawings"
            staging = td_path / "staging"
            drawings.mkdir()
            staging.mkdir()
            dxf_path = drawings / "fixture01.dxf"
            _make_fixture_dxf(dxf_path)

            out_xlsx = td_path / "battery_result.xlsx"
            transforms = ["rotate", "mirror"]  # exactly 2
            detector_cmd = _mock_detector_cmd(python_exe, fixed_pred, td_path)
            print(f"detector-cmd template: {detector_cmd[:160]}...")

            result = run_battery(
                drawings_dir=drawings,
                staging_dir=staging,
                detector_cmd=detector_cmd,
                transforms=transforms,
                budget_drawings=1,
                timeout_s=30.0,
                out_xlsx=out_xlsx,
                rigid=rigid,
                struct=struct,
                inv=inv,
                seed=0,
            )

            expected_rows = 1 * len(transforms)  # 2
            assert out_xlsx.is_file(), f"xlsx not written: {out_xlsx}"
            assert result["n_rows"] == expected_rows, (
                f"row count {result['n_rows']} != expected {expected_rows}"
            )

            import openpyxl

            wb = openpyxl.load_workbook(str(out_xlsx))
            assert "results" in wb.sheetnames, wb.sheetnames
            assert "summary" in wb.sheetnames, wb.sheetnames
            ws = wb["results"]
            # header + data rows
            data_rows = ws.max_row - 1
            assert data_rows == expected_rows, (
                f"xlsx data rows {data_rows} != {expected_rows}"
            )
            # all rows should have no error for identity-stub / fixed-pred path
            for row in result["rows"]:
                assert row.get("error") is None, f"unexpected error: {row}"
                assert row.get("invariance") is not None

            print(f"PASS selftest: xlsx={out_xlsx} rows={expected_rows} "
                  f"used_stubs={used_stubs}")
            print(f"row transforms: {[r['transform'] for r in result['rows']]}")
            print(f"invariance values: {[r['invariance'] for r in result['rows']]}")
            return 0
    except Exception:
        traceback.print_exc()
        print("FAIL selftest")
        return 1
    finally:
        if stub_dir_ctx is not None:
            stub_dir_ctx.cleanup()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="S5-D metamorphic battery runner")
    p.add_argument("--selftest", action="store_true", help="run self-contained selftest")
    sub = p.add_subparsers(dest="command")

    run_p = sub.add_parser("run", help="run metamorphic battery")
    run_p.add_argument("--drawings", required=True, type=Path, help="dir of source DXFs (read-only)")
    run_p.add_argument("--staging", required=True, type=Path, help="staging dir for transformed copies")
    run_p.add_argument(
        "--detector-cmd",
        required=True,
        help='command template with {dxf} and {out} placeholders',
    )
    run_p.add_argument(
        "--transforms",
        required=True,
        help="comma-separated kinds: rotate,mirror,scale,units,explode,rename",
    )
    run_p.add_argument(
        "--budget-drawings",
        required=True,
        type=int,
        help="MANDATORY cap on number of drawings (runaway defense)",
    )
    run_p.add_argument(
        "--timeout-s",
        required=True,
        type=float,
        help="MANDATORY per-detector-invocation timeout seconds",
    )
    run_p.add_argument("--out-xlsx", required=True, type=Path, help="output workbook path")
    run_p.add_argument("--seed", type=int, default=0)
    return p


def main(argv: Optional[Sequence[str]] = None) -> int:
    argv = list(sys.argv[1:] if argv is None else argv)
    # Allow `python battery_cli.py --selftest` without a subcommand.
    if "--selftest" in argv and "run" not in argv:
        return selftest()

    parser = build_parser()
    args = parser.parse_args(argv)

    if getattr(args, "selftest", False) and args.command is None:
        return selftest()

    if args.command != "run":
        parser.print_help()
        return 2

    if args.budget_drawings < 1:
        print("ERROR: --budget-drawings must be >= 1", file=sys.stderr)
        return 2
    if args.timeout_s <= 0:
        print("ERROR: --timeout-s must be > 0", file=sys.stderr)
        return 2

    rigid, struct, inv, missing = load_siblings()
    if missing:
        print(
            "MISSING sibling modules (exit 3): " + ", ".join(missing),
            file=sys.stderr,
        )
        return 3

    transforms = [t.strip() for t in args.transforms.split(",") if t.strip()]
    if not transforms:
        print("ERROR: --transforms empty", file=sys.stderr)
        return 2

    result = run_battery(
        drawings_dir=args.drawings,
        staging_dir=args.staging,
        detector_cmd=args.detector_cmd,
        transforms=transforms,
        budget_drawings=args.budget_drawings,
        timeout_s=args.timeout_s,
        out_xlsx=args.out_xlsx,
        rigid=rigid,
        struct=struct,
        inv=inv,
        seed=args.seed,
    )
    print(
        json.dumps(
            {
                "ok": True,
                "n_drawings": result["n_drawings"],
                "n_rows": result["n_rows"],
                "out_xlsx": result["out_xlsx"],
            },
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
