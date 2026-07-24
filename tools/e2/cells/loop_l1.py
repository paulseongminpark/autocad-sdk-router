#!/usr/bin/env python3
"""E2 loop L1: regenerate anchor-rich C0 scenes and rerun sealed C1.

All writes are confined to this file's directory.  The repository C0/C1
implementations and preregistration dossier are loaded read-only.  No CAD source
drawing or held-out/test split is accessed, and no gate/theory verdict is emitted.
"""

from __future__ import annotations

import argparse
import contextlib
import copy
import hashlib
import importlib.util
import io
import json
import math
import os
import platform
import re
import statistics
import sys
import time
from collections import Counter
from pathlib import Path
from typing import Any, Iterable, Iterator, Mapping, Sequence

import numpy as np


# Import repository modules without creating or refreshing repository __pycache__ files.
sys.dont_write_bytecode = True


CELL_DIR = Path(__file__).resolve().parent
SCENES_V2_DIR = CELL_DIR / "scenes_v2"
C0V2_NUMBERS_PATH = CELL_DIR / "c0v2_numbers.json"
C1V2_RESULTS_PATH = CELL_DIR / "c1v2_results.json"
EVIDENCE_PATH = CELL_DIR / "evidence.xlsx"
EVIDENCE_FALLBACK_PATH = CELL_DIR / "evidence_rows.json"
REPORT_PATH = CELL_DIR / "REPORT.md"

PACKET_PATH = Path(r"D:\runs\e2_program\build\PACKET_loop_L1_anchor_rich.md")
REPO_ROOT = Path(r"D:\dev\99_tools\autocad-sdk-router")
DOSSIER_PATH = REPO_ROOT / "reports" / "e2" / "dossiers" / "feyerabend_P2.md"
C0_PATH = REPO_ROOT / "tools" / "e2" / "cells" / "feyerabend_c0.py"
C1_PATH = REPO_ROOT / "tools" / "e2" / "cells" / "feyerabend_c1.py"
V1_C0_NUMBERS_PATH = (
    REPO_ROOT / "reports" / "e2" / "cells" / "feyerabend_c0" / "coverage_numbers.json"
)
V1_C1_REPORT_PATH = (
    REPO_ROOT / "reports" / "e2" / "cells" / "feyerabend_c1" / "REPORT.md"
)

SOURCE_PATHS = (
    PACKET_PATH,
    DOSSIER_PATH,
    C0_PATH,
    C1_PATH,
    V1_C0_NUMBERS_PATH,
    V1_C1_REPORT_PATH,
)

LOOP_SCHEMA = "ariadne.e2.loop_l1.anchor_rich.v1"
C0V2_SCHEMA = "ariadne.e2.loop_l1.c0v2_numbers.v1"
C1V2_SCHEMA = "ariadne.e2.loop_l1.c1v2_results.v1"
ANCHOR_COUNT_MIN = 5
ANCHOR_COUNT_MAX = 8


def load_module(name: str, path: Path) -> Any:
    if not path.is_file():
        raise FileNotFoundError(path)
    spec = importlib.util.spec_from_file_location(name, path)
    if spec is None or spec.loader is None:
        raise ImportError(f"cannot load module from {path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[name] = module
    spec.loader.exec_module(module)
    return module


C0 = load_module("loop_l1_feyerabend_c0", C0_PATH)
C1 = load_module("loop_l1_feyerabend_c1", C1_PATH)


def canonical_bytes(value: Any) -> bytes:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8")


def canonical_sha256(value: Any) -> str:
    return hashlib.sha256(canonical_bytes(value)).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def file_manifest(paths: Sequence[Path]) -> dict[str, Any]:
    rows = []
    for path in paths:
        if not path.is_file():
            raise FileNotFoundError(path)
        rows.append(
            {
                "path": str(path),
                "bytes": path.stat().st_size,
                "sha256": sha256_file(path),
            }
        )
    return {"files": rows, "digest": canonical_sha256(rows)}


def write_json_atomic(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    with temporary.open("w", encoding="utf-8", newline="\n") as stream:
        json.dump(payload, stream, ensure_ascii=False, indent=2, sort_keys=True, allow_nan=False)
        stream.write("\n")
    temporary.replace(path)


def write_text_atomic(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    with temporary.open("w", encoding="utf-8", newline="\n") as stream:
        stream.write(text)
    temporary.replace(path)


def anchor_count_for_seed(seed: int) -> int:
    """Direct deterministic variation from the prescribed C0 seed."""
    return ANCHOR_COUNT_MIN + (int(seed) % (ANCHOR_COUNT_MAX - ANCHOR_COUNT_MIN + 1))


# Midpoints occupy at least three of C1's relative 3x3 bins even at n=5.
# Unit vectors vary direction, and span bands are separated by more than 5%.
ANCHOR_LAYOUT = (
    ((5_000.0, 35_000.0), (1.0, 0.0), "R00"),
    ((30_000.0, 35_000.0), (0.0, 1.0), "R10"),
    ((55_000.0, 35_000.0), (0.6, 0.8), "R20"),
    ((5_000.0, 60_000.0), (-0.8, 0.6), "R01"),
    ((55_000.0, 60_000.0), (1.0, 0.0), "R21"),
    ((30_000.0, 85_000.0), (0.0, 1.0), "R12"),
    ((5_000.0, 85_000.0), (-0.6, 0.8), "R02"),
    ((55_000.0, 85_000.0), (0.8, 0.6), "R22"),
)


class AnchorRichSceneBuilder(C0.SceneBuilder):
    """C0 builder extension that changes only normalized reference anchors."""

    def add_reference_anchors(self) -> None:
        count = anchor_count_for_seed(self.seed)
        for anchor_index, (center, direction, region) in enumerate(ANCHOR_LAYOUT[:count]):
            # Adjacent span bands remain distinct beyond C1's log(1.05) tolerance,
            # including the worst seed-derived offsets at the longest span.
            span = (
                800.0
                + 250.0 * anchor_index
                + float(int(C0.seed_fraction(self.seed, f"anchor_span:{anchor_index}") * 100.0))
            )
            cx, cy = center
            dx, dy = direction
            half = 0.5 * span
            p0 = [cx - half * dx, cy - half * dy]
            p1 = [cx + half * dx, cy + half * dy]
            self.anchors.append(
                {
                    "handle": self.handle(f"anchor_rich_dim_{anchor_index}"),
                    "anchor_type": "DIM",
                    "region": region,
                    "p0": p0,
                    "p1": p1,
                    "raw_span": span,
                    "display_value": span,
                    "display_unit": "MM",
                    "text_height": 50.0,
                    "weight": 1.0,
                    "source_span_id": f"{self.base_scene_id}:SPAN_{anchor_index:02d}",
                    "anchor_factory_revision": "v2_anchor_rich_l1",
                }
            )
        self.mutations.add("multiple_reference_span_regions")
        self.mutations.add("anchor_rich_5_to_8_dim_ratio")


@contextlib.contextmanager
def patched_scene_builder() -> Iterator[None]:
    original = C0.SceneBuilder
    C0.SceneBuilder = AnchorRichSceneBuilder
    try:
        yield
    finally:
        C0.SceneBuilder = original


def enrich_base_scene(scene: dict[str, Any]) -> None:
    scene["loop_schema"] = LOOP_SCHEMA
    scene["scene_factory_revision"] = "v2_anchor_rich_l1"
    scene["anchor_count_rule"] = "5 + (uint32_seed mod 4)"
    scene["anchor_count"] = len(scene["anchors"])
    scene["provenance"] = {
        **scene["provenance"],
        "loop_packet": str(PACKET_PATH),
        "c0_import": str(C0_PATH),
        "c1_import": str(C1_PATH),
        "anchor_builder": "AnchorRichSceneBuilder.add_reference_anchors",
    }
    C0.refresh_scene_digests(scene)


def build_v2_structural_scene(
    index: int, calibration_rows: list[tuple[int, int, float, str]]
) -> Any:
    with patched_scene_builder():
        builder = C0.build_structural_scene(index, calibration_rows)
    return builder


def build_v2_base_corpus() -> tuple[list[dict[str, Any]], dict[str, Any]]:
    with patched_scene_builder():
        scenes, build_numbers = C0.build_base_corpus()
    for scene in scenes:
        enrich_base_scene(scene)
    return scenes, build_numbers


def anchor_richness_numbers(
    base_scenes: Sequence[Mapping[str, Any]], all_scenes: Sequence[Mapping[str, Any]]
) -> dict[str, Any]:
    base_counts = [len(scene["anchors"]) for scene in base_scenes]
    all_counts = [len(scene["anchors"]) for scene in all_scenes]
    spatial_bins = []
    canonical_counts = []
    distinct_spans = []
    unit_confidences = []
    unit_statuses: Counter[str] = Counter()
    reference_statuses: Counter[str] = Counter()
    count_rule_mismatches = 0
    distinct_span_mismatches = 0
    canonical_count_mismatches = 0
    spatial_bin_below_three = 0
    non_ratio_anchor_count = 0
    ratio_mismatch_count = 0
    minimum_log_span_separation: list[float] = []

    for scene in base_scenes:
        anchors = list(scene["anchors"])
        expected_count = anchor_count_for_seed(int(scene["seed"]))
        count_rule_mismatches += len(anchors) != expected_count
        prepared, _audit = C1.prepare_anchors(anchors)
        canonical_counts.append(len(prepared))
        canonical_count_mismatches += len(prepared) != len(anchors)
        spans = sorted(float(anchor["raw_span"]) for anchor in anchors)
        distinct = len({C1.round_key(span) for span in spans})
        distinct_spans.append(distinct)
        distinct_span_mismatches += distinct != len(anchors)
        if len(spans) > 1:
            minimum_log_span_separation.append(
                min(math.log(spans[index + 1] / spans[index]) for index in range(len(spans) - 1))
            )
        bins = C1.spatial_bin_count(prepared)
        spatial_bins.append(bins)
        spatial_bin_below_three += bins < 3
        for anchor in anchors:
            if (
                str(anchor.get("anchor_type", "")).upper() not in ("DIM", "TEXT")
                or anchor.get("display_value") is None
            ):
                non_ratio_anchor_count += 1
        model = C1.fit_anchor_model(anchors)
        unit_confidences.append(float(model["confidence_score"]))
        unit_statuses[str(model["unit_status"])] += 1
        reference_statuses[str(model["reference_status"])] += 1

    for scene in all_scenes:
        truth_scale = float(scene["truth_unit_scale"])
        for anchor in scene["anchors"]:
            raw = float(anchor["raw_span"])
            display = float(anchor["display_value"])
            if not math.isclose(display / raw, truth_scale, rel_tol=1e-12, abs_tol=1e-12):
                ratio_mismatch_count += 1

    return {
        "anchor_count_rule": "5 + (uint32_seed mod 4)",
        "base_anchor_count_distribution": dict(sorted(Counter(base_counts).items())),
        "ir_anchor_count_distribution": dict(sorted(Counter(all_counts).items())),
        "base_anchor_count_min": min(base_counts),
        "base_anchor_count_max": max(base_counts),
        "base_anchor_count_mean": statistics.fmean(base_counts),
        "base_anchor_count_median": statistics.median(base_counts),
        "base_spatial_bin_count_distribution": dict(sorted(Counter(spatial_bins).items())),
        "base_spatial_bin_count_min": min(spatial_bins),
        "base_spatial_bin_below_3_scene_count": spatial_bin_below_three,
        "base_distinct_span_count_distribution": dict(sorted(Counter(distinct_spans).items())),
        "base_distinct_span_mismatch_scene_count": distinct_span_mismatches,
        "base_canonical_anchor_count_distribution": dict(sorted(Counter(canonical_counts).items())),
        "base_canonical_anchor_count_mismatch_scene_count": canonical_count_mismatches,
        "base_anchor_count_rule_mismatch_scene_count": count_rule_mismatches,
        "base_non_dim_text_ratio_anchor_count": non_ratio_anchor_count,
        "ir_display_raw_truth_ratio_mismatch_count": ratio_mismatch_count,
        "base_minimum_adjacent_log_span_separation": min(minimum_log_span_separation),
        "c1_ransac_log_tolerance": C1.RANSAC_LOG_TOLERANCE,
        "base_unit_confidence_score": C1.numeric_distribution(unit_confidences),
        "base_unit_status_counts": dict(sorted(unit_statuses.items())),
        "base_reference_status_counts": dict(sorted(reference_statuses.items())),
    }


def nested_get(payload: Mapping[str, Any], dotted: str) -> Any:
    current: Any = payload
    for token in dotted.split("."):
        current = current[token]
    return current


def numeric_delta(before: Any, after: Any) -> Any:
    if isinstance(before, (int, float)) and not isinstance(before, bool) and isinstance(
        after, (int, float)
    ) and not isinstance(after, bool):
        return float(after) - float(before)
    return None


def c0_v1_v2_comparison_rows(
    v1: Mapping[str, Any], v2: Mapping[str, Any], anchor_numbers: Mapping[str, Any]
) -> list[dict[str, Any]]:
    metrics = (
        "scene_counts.base_scene_count",
        "scene_counts.ir_scene_count",
        "truth_pair_numbers.ir_truth_pair_count",
        "truth_validator.error_count",
        "fidelity_numbers_kappa_1.entity_mix_tv",
        "fidelity_numbers_kappa_1.thickness_histogram_ks",
        "mutation_family_coverage.single_reference_span_region.scene_count",
        "mutation_family_coverage.multiple_reference_span_regions.scene_count",
        "determinism_and_scale_numbers.four_scale_topology_mismatch_base_scene_count",
        "determinism_and_scale_numbers.four_scale_normalized_geometry_mismatch_base_scene_count",
        "determinism_and_scale_numbers.four_scale_source_handle_mismatch_base_scene_count",
    )
    rows = []
    for metric in metrics:
        before = nested_get(v1, metric)
        after = nested_get(v2, metric)
        rows.append(
            {"metric": metric, "v1": before, "v2": after, "delta": numeric_delta(before, after)}
        )
    for suffix, value in (
        ("min", anchor_numbers["base_anchor_count_min"]),
        ("mean", anchor_numbers["base_anchor_count_mean"]),
        ("max", anchor_numbers["base_anchor_count_max"]),
    ):
        rows.append(
            {
                "metric": f"ratio_anchor_count_per_scene.{suffix}",
                "v1": 3,
                "v2": value,
                "delta": numeric_delta(3, value),
            }
        )
    rows.extend(
        (
            {
                "metric": "spatial_bin_count_per_scene.min",
                "v1": 3,
                "v2": anchor_numbers["base_spatial_bin_count_min"],
                "delta": numeric_delta(3, anchor_numbers["base_spatial_bin_count_min"]),
            },
            {
                "metric": "distinct_span_mismatch_scene_count",
                "v1": None,
                "v2": anchor_numbers["base_distinct_span_mismatch_scene_count"],
                "delta": None,
            },
        )
    )
    return rows


def parse_markdown_scalar(report: str, metric: str) -> Any:
    pattern = re.compile(rf"^\|\s*{re.escape(metric)}\s*\|\s*([^|]+?)\s*\|\s*$", re.MULTILINE)
    match = pattern.search(report)
    if match is None:
        raise ValueError(f"metric not found in v1 C1 report: {metric}")
    token = match.group(1).strip()
    try:
        return json.loads(token)
    except json.JSONDecodeError:
        return token


def parse_v1_corruption_row(report: str, kind: str) -> dict[str, Any]:
    for line in report.splitlines():
        if not line.startswith(f"| {kind} |"):
            continue
        cells = [cell.strip() for cell in line.strip().strip("|").split("|")]
        if len(cells) != 9:
            raise ValueError(f"unexpected v1 corruption row shape: {cells}")
        return {
            "corruption": cells[0],
            "scene_count": int(cells[1]),
            "unit_status_transitions": json.loads(cells[2]),
            "status_transitions": json.loads(cells[3]),
            "scale_estimate_unchanged_count": int(cells[4]),
            "scale_estimate_changed_count": int(cells[5]),
            "relative_error_median": json.loads(cells[6]),
            "relative_error_p95": json.loads(cells[7]),
            "confidence_score_after_median": json.loads(cells[8]),
        }
    raise ValueError(f"corruption row not found in v1 C1 report: {kind}")


def v1_c1_baseline(report_text: str) -> dict[str, Any]:
    return {
        "source_path": str(V1_C1_REPORT_PATH),
        "source_sha256": sha256_file(V1_C1_REPORT_PATH),
        "estimate_count": parse_markdown_scalar(report_text, "estimate_count"),
        "estimate_coverage": parse_markdown_scalar(report_text, "estimate_coverage"),
        "accuracy_within_5pct": parse_markdown_scalar(report_text, "accuracy_within_5pct"),
        "high_scene_count": parse_markdown_scalar(report_text, "HIGH_scene_count"),
        "high_coverage": parse_markdown_scalar(report_text, "HIGH_coverage"),
        "high_accuracy_within_5pct": parse_markdown_scalar(
            report_text, "HIGH_accuracy_within_5pct"
        ),
        "single_outlier": parse_v1_corruption_row(report_text, "single_outlier"),
        "pair_label_mismatching_anchor_artifact_scene_count": parse_markdown_scalar(
            report_text, "mismatching_anchor_artifact_scene_count"
        ),
    }


STATUS_RANK = {"NONE": 0, "LOW": 1, "HIGH": 2}


def ticket_single_outlier_numbers(
    rows: Sequence[Mapping[str, Any]], baseline: Mapping[str, Any]
) -> dict[str, Any]:
    diagnostics = [row["corruption_diagnostics"]["single_outlier"] for row in rows]

    def increased(prefix: str) -> int:
        return sum(
            STATUS_RANK[str(row[f"{prefix}_after"])] > STATUS_RANK[str(row[f"{prefix}_before"])]
            for row in diagnostics
        )

    score_increased = sum(
        float(row["confidence_score_after"]) > float(row["confidence_score_before"]) + 1e-15
        for row in diagnostics
    )
    score_decreased = sum(
        float(row["confidence_score_after"]) + 1e-15 < float(row["confidence_score_before"])
        for row in diagnostics
    )
    score_unchanged = len(diagnostics) - score_increased - score_decreased
    v2_status_transitions = C1.transition_counter(rows, "single_outlier", "status")
    v2_unit_transitions = C1.transition_counter(rows, "single_outlier", "unit_status")
    v2_reference_transitions = C1.transition_counter(rows, "single_outlier", "reference_status")
    v1_status_transitions = baseline["single_outlier"]["status_transitions"]
    return {
        "scene_count": len(rows),
        "v1_status_transitions": v1_status_transitions,
        "v2_status_transitions": v2_status_transitions,
        "v1_unit_status_transitions": baseline["single_outlier"]["unit_status_transitions"],
        "v2_unit_status_transitions": v2_unit_transitions,
        "v2_reference_status_transitions": v2_reference_transitions,
        "v1_low_to_high_status_count": int(v1_status_transitions.get("LOW->HIGH", 0)),
        "v2_low_to_high_status_count": int(v2_status_transitions.get("LOW->HIGH", 0)),
        "v2_confidence_score_increased_count": score_increased,
        "v2_confidence_score_unchanged_count": score_unchanged,
        "v2_confidence_score_decreased_count": score_decreased,
        "v2_status_rank_increased_count": increased("status"),
        "v2_unit_status_rank_increased_count": increased("unit_status"),
        "v2_reference_status_rank_increased_count": increased("reference_status"),
        "v2_confidence_or_status_increased_count": sum(
            (
                float(row["confidence_score_after"])
                > float(row["confidence_score_before"]) + 1e-15
            )
            or STATUS_RANK[str(row["status_after"])] > STATUS_RANK[str(row["status_before"])]
            for row in diagnostics
        ),
        "v2_scale_estimate_unchanged_count": sum(
            bool(row["scale_estimate_unchanged"]) for row in diagnostics
        ),
        "v2_scale_estimate_changed_count": sum(
            not bool(row["scale_estimate_unchanged"]) for row in diagnostics
        ),
    }


def c1_v1_v2_comparison_rows(
    baseline: Mapping[str, Any], aggregates: Mapping[str, Any], ticket: Mapping[str, Any]
) -> list[dict[str, Any]]:
    overall = aggregates["overall"]
    pairs = (
        ("estimate_count", baseline["estimate_count"], overall["estimate_count"]),
        ("estimate_coverage", baseline["estimate_coverage"], overall["estimate_coverage"]),
        (
            "accuracy_within_5pct",
            baseline["accuracy_within_5pct"],
            overall["accuracy_within_5pct"],
        ),
        ("HIGH_scene_count", baseline["high_scene_count"], overall["high_scene_count"]),
        ("HIGH_coverage", baseline["high_coverage"], overall["high_coverage"]),
        (
            "HIGH_accuracy_within_5pct",
            baseline["high_accuracy_within_5pct"],
            overall["high_accuracy_within_5pct"],
        ),
        (
            "pair_label_mismatch_scene_count",
            baseline["pair_label_mismatching_anchor_artifact_scene_count"],
            aggregates["pair_label_permutation"]["mismatching_anchor_artifact_scene_count"],
        ),
        (
            "single_outlier_LOW_to_HIGH_status_count",
            ticket["v1_low_to_high_status_count"],
            ticket["v2_low_to_high_status_count"],
        ),
        (
            "single_outlier_scale_estimate_unchanged_count",
            baseline["single_outlier"]["scale_estimate_unchanged_count"],
            ticket["v2_scale_estimate_unchanged_count"],
        ),
    )
    return [
        {"metric": metric, "v1": before, "v2": after, "delta": numeric_delta(before, after)}
        for metric, before, after in pairs
    ]


def run_selftests(stream: io.TextIOBase) -> tuple[int, dict[str, Any]]:
    tests: list[dict[str, Any]] = []

    def check(name: str, condition: bool, detail: str) -> None:
        passed = bool(condition)
        print(f"SELFTEST {name}: {'PASS' if passed else 'FAIL'} | {detail}", file=stream)
        tests.append({"name": name, "passed": passed, "detail": detail})

    print("=== loop_l1 anchor-rich selftest ===", file=stream)
    try:
        j = 7
        expected_seed = int.from_bytes(
            hashlib.sha256(f"{C0.SEED_PREFIX}{j}".encode("utf-8")).digest()[:4], "big"
        )
        check(
            "c0_seed_rule_unchanged",
            C0.seed_for_index(j) == expected_seed,
            f"j={j} seed={C0.seed_for_index(j)} expected={expected_seed}",
        )
        fixture_rows = C0.fidelity_schedule()[j][:3]
        first = build_v2_structural_scene(j, fixture_rows).finalize()
        enrich_base_scene(first)
        second = build_v2_structural_scene(j, fixture_rows).finalize()
        enrich_base_scene(second)
        check(
            "same_j_same_sha",
            canonical_sha256(first) == canonical_sha256(second),
            f"j={j} sha256={canonical_sha256(first)}",
        )
        anchors = first["anchors"]
        expected_count = anchor_count_for_seed(first["seed"])
        check(
            "fixture_anchor_count_rule",
            len(anchors) == expected_count and ANCHOR_COUNT_MIN <= len(anchors) <= ANCHOR_COUNT_MAX,
            f"observed={len(anchors)} expected={expected_count}",
        )
        spans = [C1.geometric_span(anchor) for anchor in anchors]
        check(
            "fixture_independent_distinct_spans",
            None not in spans and len({C1.round_key(float(span)) for span in spans}) == len(anchors),
            f"span_count={len(spans)} distinct={len({C1.round_key(float(span)) for span in spans if span is not None})}",
        )
        prepared, audit = C1.prepare_anchors(anchors)
        bins = C1.spatial_bin_count(prepared)
        check(
            "fixture_canonical_independence",
            len(prepared) == len(anchors) and audit["duplicate_count"] == 0,
            f"input={len(anchors)} canonical={len(prepared)} duplicates={audit['duplicate_count']}",
        )
        check("fixture_spatial_bins", bins >= 3, f"n_spatial_bins={bins}")
        model = C1.fit_anchor_model(anchors)
        check(
            "fixture_ratio_estimate",
            model["display_per_raw"] is not None
            and math.isclose(float(model["display_per_raw"]), 1.0, rel_tol=1e-12),
            f"estimate={model['display_per_raw']} unit_status={model['unit_status']} confidence={model['confidence_score']:.12g}",
        )
        scaled = [C0.scale_scene(first, kappa, token) for kappa, token in C0.SCALES]
        check(
            "same_j_four_scale_topology",
            len({scene["topology_sha256"] for scene in scaled}) == 1,
            f"unique_topology_sha={len({scene['topology_sha256'] for scene in scaled})}",
        )
        check(
            "same_j_four_scale_normalized_geometry",
            len({scene["normalized_geometry_sha256"] for scene in scaled}) == 1,
            f"unique_normalized_geometry_sha={len({scene['normalized_geometry_sha256'] for scene in scaled})}",
        )
        positive_errors = [error for scene in scaled for error in C0.validate_scene(scene)]
        check(
            "truth_validator_positive_cases",
            not positive_errors,
            f"scene_count={len(scaled)} error_count={len(positive_errors)}",
        )
        corrupted = copy.deepcopy(scaled[0])
        corrupted["truth_pairs"][0]["handles"][1] = "MISSING_SOURCE_HANDLE"
        corrupted["truth_pairs"][0]["pair_id"] = C0.pair_id(
            corrupted["truth_pairs"][0]["handles"]
        )
        negative_errors = C0.validate_scene(corrupted)
        check(
            "truth_validator_negative_case_honest_fail",
            bool(negative_errors),
            f"error_count={len(negative_errors)}",
        )
        guarded = C1.GuardedScene(anchors)
        guarded_artifact = C1.anchor_artifact_from_scene(guarded)
        check(
            "truth_key_access_guard",
            guarded.accessed == ["anchors"]
            and guarded_artifact["model"]["display_per_raw"] is not None,
            f"accessed_keys={guarded.accessed!r}",
        )
        distribution = Counter()
        distribution_errors = 0
        spatial_errors = 0
        distinct_errors = 0
        for index in range(C0.BASE_SCENE_COUNT):
            builder = AnchorRichSceneBuilder(index)
            builder.add_reference_anchors()
            count = len(builder.anchors)
            distribution[count] += 1
            distribution_errors += count != anchor_count_for_seed(builder.seed)
            prepared_scene, _ = C1.prepare_anchors(builder.anchors)
            spatial_errors += C1.spatial_bin_count(prepared_scene) < 3
            span_keys = {C1.round_key(float(anchor["raw_span"])) for anchor in builder.anchors}
            distinct_errors += len(span_keys) != count
        check(
            "anchor_count_distribution",
            set(distribution) == {5, 6, 7, 8} and distribution_errors == 0,
            f"distribution={dict(sorted(distribution.items()))} rule_mismatches={distribution_errors}",
        )
        check(
            "all_base_spatial_bins_ge_3",
            spatial_errors == 0,
            f"below_3_scene_count={spatial_errors}",
        )
        check(
            "all_base_spans_distinct",
            distinct_errors == 0,
            f"distinct_span_mismatch_scene_count={distinct_errors}",
        )
        entity_types, wall_count = C0.gen2_library_probe()
        required = {"LINE", "ARC", "SPLINE", "HATCH", "LWPOLYLINE"}
        check(
            "permitted_gen2_synthetic_fixture",
            wall_count >= 2 and required.issubset(entity_types),
            f"wall_records={wall_count} entity_types={sorted(entity_types)}",
        )
    except Exception as exc:
        check("unexpected_exception", False, f"{type(exc).__name__}: {exc}")

    c1_stream = io.StringIO()
    c1_tests: dict[str, Any] = {"passed": 0, "total": 0, "tests": []}
    try:
        c1_tests = C1.run_selftests(c1_stream)
        print("=== imported sealed C1 selftests ===", file=stream)
        print(c1_stream.getvalue().rstrip("\n"), file=stream)
        check(
            "sealed_c1_selftests",
            c1_tests["passed"] == c1_tests["total"],
            f"passed={c1_tests['passed']} total={c1_tests['total']}",
        )
    except Exception as exc:
        print(c1_stream.getvalue().rstrip("\n"), file=stream)
        check("sealed_c1_selftests", False, f"{type(exc).__name__}: {exc}")

    passed = sum(test["passed"] for test in tests)
    print(f"SELFTEST SUMMARY: {passed}/{len(tests)} passed", file=stream)
    print(f"SELFTEST_RESULT: {'PASS' if passed == len(tests) else 'FAIL'}", file=stream)
    details = {
        "passed": passed,
        "total": len(tests),
        "tests": tests,
        "sealed_c1": c1_tests,
    }
    return (0 if passed == len(tests) else 1), details


def write_scenes(all_scenes: Sequence[dict[str, Any]]) -> list[Path]:
    SCENES_V2_DIR.mkdir(parents=True, exist_ok=True)
    expected: dict[str, dict[str, Any]] = {
        C0.scene_filename(scene): scene for scene in all_scenes
    }
    if len(expected) != C0.BASE_SCENE_COUNT * len(C0.SCALES):
        raise RuntimeError(f"non-unique scene filenames: {len(expected)}")
    for name, scene in sorted(expected.items()):
        write_json_atomic(SCENES_V2_DIR / name, scene)
    expected_names = set(expected)
    for stale in SCENES_V2_DIR.glob("scene_*.json"):
        if stale.name not in expected_names:
            stale.unlink()
    paths = sorted(SCENES_V2_DIR.glob("scene_*.json"))
    if len(paths) != len(expected):
        raise RuntimeError(f"scene write count mismatch: {len(paths)} != {len(expected)}")
    return paths


@contextlib.contextmanager
def c1_output_paths() -> Iterator[None]:
    original = (C1.INPUT_DIR, C1.RESULTS_PATH, C1.EVIDENCE_PATH, C1.EVIDENCE_FALLBACK_PATH, C1.REPORT_PATH)
    C1.INPUT_DIR = SCENES_V2_DIR
    C1.RESULTS_PATH = C1V2_RESULTS_PATH
    C1.EVIDENCE_PATH = EVIDENCE_PATH
    C1.EVIDENCE_FALLBACK_PATH = EVIDENCE_FALLBACK_PATH
    C1.REPORT_PATH = REPORT_PATH
    try:
        yield
    finally:
        (
            C1.INPUT_DIR,
            C1.RESULTS_PATH,
            C1.EVIDENCE_PATH,
            C1.EVIDENCE_FALLBACK_PATH,
            C1.REPORT_PATH,
        ) = original


def style_table_sheet(sheet: Any, widths: Mapping[str, float]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.auto_filter.ref = sheet.dimensions
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(cell.value, float):
                cell.number_format = "0.000000000000"
            elif isinstance(cell.value, int):
                cell.number_format = "#,##0"


def augment_evidence_workbook(
    c0_numbers: Mapping[str, Any],
    c0_comparison: Sequence[Mapping[str, Any]],
    c1_comparison: Sequence[Mapping[str, Any]],
    aggregates: Mapping[str, Any],
    ticket: Mapping[str, Any],
    source_manifest_before: Mapping[str, Any],
    source_manifest_after: Mapping[str, Any],
) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = openpyxl.load_workbook(EVIDENCE_PATH)
    for name in (
        "overview",
        "anchor_distribution",
        "c0_v1_v2",
        "c1_v1_v2",
        "single_outlier",
        "source_manifest",
    ):
        if name in workbook.sheetnames:
            del workbook[name]

    overview = workbook.create_sheet("overview", 0)
    overview.sheet_view.showGridLines = False
    overview.merge_cells("A1:D1")
    overview["A1"] = "E2 Loop L1 — Anchor-rich C0v2 / sealed C1 reevaluation"
    overview["A1"].fill = PatternFill("solid", fgColor="17365D")
    overview["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    overview["A1"].alignment = Alignment(horizontal="left", vertical="center")
    overview.row_dimensions[1].height = 26
    overview.append(["metric", "value", "unit/scope", "source"])
    anchor = c0_numbers["anchor_richness"]
    overall = aggregates["overall"]
    overview_rows = (
        ("IR scene count", overall["scene_count"], "scenes", "c1v2 aggregates"),
        ("Base scene count", c0_numbers["scene_counts"]["base_scene_count"], "scenes", "c0v2"),
        ("Anchor count min", anchor["base_anchor_count_min"], "per base scene", "c0v2"),
        ("Anchor count mean", anchor["base_anchor_count_mean"], "per base scene", "c0v2"),
        ("Anchor count max", anchor["base_anchor_count_max"], "per base scene", "c0v2"),
        ("Spatial bin min", anchor["base_spatial_bin_count_min"], "per base scene", "c0v2"),
        ("HIGH scene count", overall["high_scene_count"], "unit_status", "c1v2"),
        ("HIGH coverage", overall["high_coverage"], "fraction", "c1v2"),
        ("HIGH accuracy within 5%", overall["high_accuracy_within_5pct"], "fraction", "c1v2"),
        (
            "Single-outlier confidence increase count",
            ticket["v2_confidence_score_increased_count"],
            "scenes",
            "ticket diagnostic",
        ),
        (
            "Single-outlier status increase count",
            ticket["v2_status_rank_increased_count"],
            "scenes",
            "ticket diagnostic",
        ),
        (
            "Pair-label anchor digest mismatch count",
            aggregates["pair_label_permutation"]["mismatching_anchor_artifact_scene_count"],
            "scenes",
            "c1v2",
        ),
        ("C0 truth validator error count", c0_numbers["truth_validator"]["error_count"], "errors", "c0v2"),
    )
    for row in overview_rows:
        overview.append(row)
    for cell in overview[2]:
        cell.fill = PatternFill("solid", fgColor="5B9BD5")
        cell.font = Font(color="FFFFFF", bold=True)
    overview.freeze_panes = "A3"
    overview.column_dimensions["A"].width = 43
    overview.column_dimensions["B"].width = 20
    overview.column_dimensions["C"].width = 20
    overview.column_dimensions["D"].width = 24
    for row in overview.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if isinstance(row[1].value, float):
            row[1].number_format = "0.000000000000"
        elif isinstance(row[1].value, int):
            row[1].number_format = "#,##0"

    distribution = workbook.create_sheet("anchor_distribution", 1)
    distribution.append(["anchor_count", "base_scene_count", "ir_scene_count"])
    base_distribution = anchor["base_anchor_count_distribution"]
    ir_distribution = anchor["ir_anchor_count_distribution"]
    for count in range(ANCHOR_COUNT_MIN, ANCHOR_COUNT_MAX + 1):
        distribution.append(
            [count, int(base_distribution.get(count, base_distribution.get(str(count), 0))), int(ir_distribution.get(count, ir_distribution.get(str(count), 0)))]
        )
    style_table_sheet(distribution, {"A": 18, "B": 22, "C": 20})

    c0_sheet = workbook.create_sheet("c0_v1_v2", 2)
    c0_sheet.append(["metric", "v1", "v2", "delta"])
    for row in c0_comparison:
        c0_sheet.append([row["metric"], row["v1"], row["v2"], row["delta"]])
    style_table_sheet(c0_sheet, {"A": 72, "B": 20, "C": 20, "D": 20})

    c1_sheet = workbook.create_sheet("c1_v1_v2", 3)
    c1_sheet.append(["metric", "v1", "v2", "delta"])
    for row in c1_comparison:
        c1_sheet.append([row["metric"], row["v1"], row["v2"], row["delta"]])
    style_table_sheet(c1_sheet, {"A": 55, "B": 20, "C": 20, "D": 20})

    ticket_sheet = workbook.create_sheet("single_outlier", 4)
    ticket_sheet.append(["metric", "value"])
    for key, value in ticket.items():
        if isinstance(value, Mapping):
            for transition, count in value.items():
                ticket_sheet.append([f"{key}.{transition}", count])
        else:
            ticket_sheet.append([key, value])
    style_table_sheet(ticket_sheet, {"A": 62, "B": 22})

    source_sheet = workbook.create_sheet("source_manifest")
    source_sheet.append(["path", "bytes_before", "sha256_before", "bytes_after", "sha256_after", "digest_equal"])
    after_by_path = {row["path"]: row for row in source_manifest_after["files"]}
    for before in source_manifest_before["files"]:
        after = after_by_path[before["path"]]
        source_sheet.append(
            [
                before["path"],
                before["bytes"],
                before["sha256"],
                after["bytes"],
                after["sha256"],
                before["sha256"] == after["sha256"],
            ]
        )
    style_table_sheet(
        source_sheet,
        {"A": 92, "B": 16, "C": 68, "D": 16, "E": 68, "F": 14},
    )

    workbook.properties.creator = "OpenAI Codex"
    workbook.properties.title = "E2 Loop L1 numeric evidence"
    workbook.properties.subject = "Anchor-rich C0v2 and sealed C1 reevaluation"
    temporary = EVIDENCE_PATH.with_name(EVIDENCE_PATH.stem + ".augment.tmp.xlsx")
    workbook.save(temporary)
    workbook.close()
    temporary.replace(EVIDENCE_PATH)


def verify_evidence_workbook(expected_scale_rows: int) -> dict[str, Any]:
    import openpyxl

    workbook = openpyxl.load_workbook(EVIDENCE_PATH, read_only=True, data_only=False)
    required = {
        "overview",
        "anchor_distribution",
        "c0_v1_v2",
        "c1_v1_v2",
        "single_outlier",
        "source_manifest",
        "scale_confidence",
        "score_bins",
        "corruption_transitions",
        "pair_label_digest",
        "README",
    }
    missing = sorted(required - set(workbook.sheetnames))
    scale_rows = max(0, workbook["scale_confidence"].max_row - 1)
    formula_error_count = 0
    populated_sheet_count = 0
    for sheet in workbook.worksheets:
        if sheet.max_row > 0 and sheet.max_column > 0:
            populated_sheet_count += 1
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and re.search(
                    r"#REF!|#DIV/0!|#VALUE!|#NAME\?|#N/A", cell.value
                ):
                    formula_error_count += 1
    workbook.close()
    if missing or scale_rows != expected_scale_rows or formula_error_count:
        raise RuntimeError(
            "evidence verification failed: "
            f"missing={missing} scale_rows={scale_rows}/{expected_scale_rows} "
            f"formula_errors={formula_error_count}"
        )
    return {
        "required_sheet_count": len(required),
        "observed_sheet_count": populated_sheet_count,
        "missing_required_sheets": missing,
        "scale_confidence_data_rows": scale_rows,
        "formula_error_count": formula_error_count,
        "backend": f"openpyxl {openpyxl.__version__}",
        "visual_render_count": 0,
    }


def format_number(value: Any) -> str:
    return C1.format_number(value)


def markdown_table(headers: Sequence[str], rows: Sequence[Sequence[Any]]) -> str:
    return C1.markdown_table(headers, rows)


def render_report(
    c0_numbers: Mapping[str, Any],
    c1_results: Mapping[str, Any],
    selftest_transcript: str,
    artifact_rows: Sequence[Mapping[str, Any]],
) -> str:
    aggregates = c1_results["aggregates"]
    overall = aggregates["overall"]
    anchor = c0_numbers["anchor_richness"]
    c0_compare = c0_numbers["v1_comparison_rows"]
    c1_compare = c1_results["v1_comparison_rows"]
    ticket = c1_results["ticket_single_outlier"]
    lines: list[str] = [
        "# E2 Loop L1 — 앵커-풍부 장면 재생성 및 C1 재평가",
        "",
        "## 설계 및 실행 경계",
        "",
        f"- C0의 seed 규칙과 구조 생성·scale 복제·fidelity·truth validator를 read-only import했고, anchor builder만 scene당 seed 결정적 {ANCHOR_COUNT_MIN}~{ANCHOR_COUNT_MAX} DIM ratio로 확장했다.",
        "- 각 DIM ratio는 서로 다른 geometry span과 source handle을 가지며 midpoint는 C1의 상대 3×3 공간 bin 중 최소 3개 이상에 놓였다.",
        "- C1의 estimator, 봉인 confidence 식, 네 corruption, pair-label permutation digest는 수정 없이 import·실행했다.",
        f"- 출력은 `{CELL_DIR}` 아래로 한정했다. 원본 CAD와 test split은 접근하지 않았고 repository 및 v1 산출물은 수정하지 않았다.",
        "- 이 보고서는 수치와 selftest 상태만 기록하며 목표·게이트·이론 판정을 기록하지 않는다.",
        "",
        "## Selftest 전문",
        "",
        "```text",
        selftest_transcript.rstrip("\n"),
        "```",
        "",
        "## 앵커 수·독립성 분포",
        "",
        markdown_table(
            ["metric", "value"],
            [
                ("base_anchor_count_distribution", json.dumps(anchor["base_anchor_count_distribution"], sort_keys=True)),
                ("ir_anchor_count_distribution", json.dumps(anchor["ir_anchor_count_distribution"], sort_keys=True)),
                ("base_anchor_count_min", anchor["base_anchor_count_min"]),
                ("base_anchor_count_mean", anchor["base_anchor_count_mean"]),
                ("base_anchor_count_max", anchor["base_anchor_count_max"]),
                ("base_spatial_bin_count_distribution", json.dumps(anchor["base_spatial_bin_count_distribution"], sort_keys=True)),
                ("base_spatial_bin_count_min", anchor["base_spatial_bin_count_min"]),
                ("base_spatial_bin_below_3_scene_count", anchor["base_spatial_bin_below_3_scene_count"]),
                ("base_distinct_span_mismatch_scene_count", anchor["base_distinct_span_mismatch_scene_count"]),
                ("base_canonical_anchor_count_mismatch_scene_count", anchor["base_canonical_anchor_count_mismatch_scene_count"]),
                ("ir_display_raw_truth_ratio_mismatch_count", anchor["ir_display_raw_truth_ratio_mismatch_count"]),
            ],
        ),
        "",
        "## v1 대비 C0v2 변화 수치",
        "",
        markdown_table(
            ["metric", "v1", "v2", "delta"],
            [(row["metric"], row["v1"], row["v2"], row["delta"]) for row in c0_compare],
        ),
        "",
        "## v1 대비 C1v2 변화 수치",
        "",
        markdown_table(
            ["metric", "v1", "v2", "delta"],
            [(row["metric"], row["v1"], row["v2"], row["delta"]) for row in c1_compare],
        ),
        "",
        "## C0v2 커버리지·fidelity·truth 수치 전문",
        "",
        "```json",
        json.dumps(c0_numbers, ensure_ascii=False, indent=2, sort_keys=True, allow_nan=False),
        "```",
        "",
        "## C1v2 전체 scale 추정 수치",
        "",
        markdown_table(
            ["metric", "value"],
            [
                ("scene_count", overall["scene_count"]),
                ("estimate_count", overall["estimate_count"]),
                ("estimate_coverage", overall["estimate_coverage"]),
                ("accuracy_within_5pct", overall["accuracy_within_5pct"]),
                ("HIGH_scene_count", overall["high_scene_count"]),
                ("HIGH_coverage", overall["high_coverage"]),
                ("HIGH_accuracy_within_5pct", overall["high_accuracy_within_5pct"]),
                ("e_s_min", overall["e_s"]["min"]),
                ("e_s_median", overall["e_s"]["median"]),
                ("e_s_p95", overall["e_s"]["p95"]),
                ("e_s_max", overall["e_s"]["max"]),
                ("relative_error_min", overall["relative_error"]["min"]),
                ("relative_error_median", overall["relative_error"]["median"]),
                ("relative_error_p95", overall["relative_error"]["p95"]),
                ("relative_error_max", overall["relative_error"]["max"]),
            ],
        ),
        "",
        "## Scale × confidence 전 행",
        "",
        markdown_table(
            [
                "kappa",
                "unit_status",
                "n",
                "fraction",
                "n_est",
                "accuracy_5pct",
                "e_s_med",
                "e_s_p95",
                "relerr_med",
                "relerr_p95",
                "conf_med",
            ],
            [
                (
                    row["scale_kappa"],
                    row["confidence"],
                    row["scene_count"],
                    row["fraction_of_scale"],
                    row["estimate_count"],
                    row["accuracy_within_5pct"],
                    row["e_s"]["median"],
                    row["e_s"]["p95"],
                    row["relative_error"]["median"],
                    row["relative_error"]["p95"],
                    row["confidence_score"]["median"],
                )
                for row in aggregates["scale_confidence_rows"]
            ],
        ),
        "",
        "## Numeric confidence-score bin별 accuracy",
        "",
        markdown_table(
            ["scale", "score_bin", "n", "n_est", "accuracy_5pct", "relerr_med", "relerr_p95"],
            [
                (
                    row["scale_kappa"],
                    row["score_bin"],
                    row["scene_count"],
                    row["estimate_count"],
                    row["accuracy_within_5pct"],
                    row["relative_error"]["median"],
                    row["relative_error"]["p95"],
                )
                for row in aggregates["score_calibration_rows"]
            ],
        ),
        "",
        "## Corruption 전이 수치",
        "",
        markdown_table(
            [
                "corruption",
                "n",
                "unit_transition",
                "status_transition",
                "reference_transition",
                "scale_same",
                "scale_changed",
                "relerr_med",
                "relerr_p95",
                "conf_after_med",
            ],
            [
                (
                    kind,
                    summary["scene_count"],
                    json.dumps(summary["unit_status_transitions"], sort_keys=True),
                    json.dumps(summary["status_transitions"], sort_keys=True),
                    json.dumps(summary["reference_status_transitions"], sort_keys=True),
                    summary["scale_estimate_unchanged_count"],
                    summary["scale_estimate_changed_count"],
                    summary["relative_error_after"]["median"],
                    summary["relative_error_after"]["p95"],
                    summary["confidence_score_after"]["median"],
                )
                for kind, summary in aggregates["corruption"]["all_four_applied"].items()
            ],
        ),
        "",
        "## Single-outlier 티켓 수치",
        "",
        markdown_table(
            ["metric", "value"],
            [
                (key, json.dumps(value, sort_keys=True) if isinstance(value, Mapping) else value)
                for key, value in ticket.items()
            ],
        ),
        "",
        "## Pair-label permutation digest 수치",
        "",
        markdown_table(
            ["metric", "value"], list(aggregates["pair_label_permutation"].items())
        ),
        "",
        "## C1v2 aggregate 수치 전문",
        "",
        "```json",
        json.dumps(aggregates, ensure_ascii=False, indent=2, sort_keys=True, allow_nan=False),
        "```",
        "",
        "## 산출물 검증 수치",
        "",
        markdown_table(
            ["artifact", "status", "bytes", "sha256_or_digest"],
            [
                (row["artifact"], row["status"], row["bytes"], row["sha256_or_digest"])
                for row in artifact_rows
            ],
        ),
        "",
        "## 미해결",
        "",
        f"- 서로 다른 geometry span을 사용한 reference model의 v2 status 분포는 `{json.dumps(overall['status_counts'], sort_keys=True)}`이고 primary unit_status 분포는 `{json.dumps(overall['unit_status_counts'], sort_keys=True)}`다.",
        f"- evidence workbook은 openpyxl 구조 검증 {c1_results['evidence']['verification']['required_sheet_count']}개 required sheet, formula error {c1_results['evidence']['verification']['formula_error_count']}개로 기록했다. artifact-tool dependency loader가 제공되지 않아 raster render count는 {c1_results['evidence']['verification']['visual_render_count']}이다.",
        "- 원본 CAD 및 test split 접근 수는 0이며, repository/v1 source manifest mismatch 수는 0이다.",
        "- 목표·게이트·이론 판정은 오케스트레이터 비교 범위로 남겼다.",
        "",
        "LOOP_COMPLETE: L1",
    ]
    report = "\n".join(lines) + "\n"
    if report.rstrip("\n").splitlines()[-1] != "LOOP_COMPLETE: L1":
        raise AssertionError("report completion marker is not the last line")
    return report


def artifact_row(path: Path, status: str = "GENERATED") -> dict[str, Any]:
    return {
        "artifact": path.name,
        "status": status,
        "bytes": path.stat().st_size,
        "sha256_or_digest": sha256_file(path),
    }


def execute_full() -> int:
    wall_start = time.perf_counter()
    cpu_start = time.process_time()
    source_manifest_before = file_manifest(SOURCE_PATHS)

    selftest_stream = io.StringIO()
    selftest_rc, selftest_details = run_selftests(selftest_stream)
    selftest_transcript = selftest_stream.getvalue()
    print(selftest_transcript, end="")
    if selftest_rc != 0:
        return 1

    base_scenes, build_numbers = build_v2_base_corpus()
    all_scenes = [
        C0.scale_scene(base_scene, kappa, token)
        for base_scene in base_scenes
        for kappa, token in C0.SCALES
    ]
    c0_numbers, validation_errors = C0.aggregate_numbers(
        base_scenes, all_scenes, build_numbers
    )
    if validation_errors:
        print(f"C0 validator error_count={len(validation_errors)}", file=sys.stderr)
        for error in validation_errors[:20]:
            print(error, file=sys.stderr)
        return 1

    anchor_numbers = anchor_richness_numbers(base_scenes, all_scenes)
    v1_c0_numbers = json.loads(V1_C0_NUMBERS_PATH.read_text(encoding="utf-8"))
    c0_comparison = c0_v1_v2_comparison_rows(v1_c0_numbers, c0_numbers, anchor_numbers)
    c0_numbers["schema"] = C0V2_SCHEMA
    c0_numbers["contract"] = {
        "source_packet": str(PACKET_PATH),
        "source_dossier": str(DOSSIER_PATH),
        "c0_import": str(C0_PATH),
        "c1_import": str(C1_PATH),
        "write_root": str(CELL_DIR),
        "source_cad_access_count": 0,
        "test_split_access_count": 0,
        "random_search_count": 0,
        "gate_verdict_emitted": False,
    }
    c0_numbers["anchor_richness"] = anchor_numbers
    c0_numbers["v1_baseline"] = {
        "path": str(V1_C0_NUMBERS_PATH),
        "bytes": V1_C0_NUMBERS_PATH.stat().st_size,
        "sha256": sha256_file(V1_C0_NUMBERS_PATH),
        "schema": v1_c0_numbers.get("schema"),
    }
    c0_numbers["v1_comparison_rows"] = c0_comparison
    c0_numbers["selftest"] = {**selftest_details, "transcript": selftest_transcript}

    scene_paths = write_scenes(all_scenes)
    write_json_atomic(C0V2_NUMBERS_PATH, c0_numbers)

    c1_wall_start = time.perf_counter()
    c1_cpu_start = time.process_time()
    with c1_output_paths():
        scenes_from_disk, validated_paths, input_validation = C1.load_and_validate_inputs()
        input_manifest_before = C1.input_manifest(validated_paths)
        rows = [C1.evaluate_scene(scene) for scene in scenes_from_disk]
        aggregates = C1.aggregate_results(rows)

    v1_report_text = V1_C1_REPORT_PATH.read_text(encoding="utf-8")
    baseline_c1 = v1_c1_baseline(v1_report_text)
    ticket = ticket_single_outlier_numbers(rows, baseline_c1)
    c1_comparison = c1_v1_v2_comparison_rows(baseline_c1, aggregates, ticket)

    source_manifest_after = file_manifest(SOURCE_PATHS)
    source_manifest_mismatch_count = int(
        source_manifest_before["digest"] != source_manifest_after["digest"]
    )
    if source_manifest_mismatch_count:
        raise RuntimeError("read-only source manifest changed during execution")

    with c1_output_paths():
        workbook_status = C1.write_workbook(aggregates)
        if workbook_status["status"] == "GENERATED" and EVIDENCE_PATH.is_file():
            augment_evidence_workbook(
                c0_numbers,
                c0_comparison,
                c1_comparison,
                aggregates,
                ticket,
                source_manifest_before,
                source_manifest_after,
            )
            workbook_verification = verify_evidence_workbook(
                len(aggregates["scale_confidence_rows"])
            )
            workbook_status = {
                **workbook_status,
                "path": str(EVIDENCE_PATH),
                "bytes": EVIDENCE_PATH.stat().st_size,
                "sha256": sha256_file(EVIDENCE_PATH),
            }
        else:
            workbook_verification = C1.verify_workbook(
                len(aggregates["scale_confidence_rows"])
            )
            workbook_verification = {
                **workbook_verification,
                "required_sheet_count": 0,
                "observed_sheet_count": 0,
                "formula_error_count": 0,
                "visual_render_count": 0,
            }

    input_manifest_after = C1.input_manifest(validated_paths)
    input_manifest_mismatch_count = int(
        input_manifest_before["digest"] != input_manifest_after["digest"]
    )
    if input_manifest_mismatch_count:
        raise RuntimeError("v2 scene input manifest changed during C1 evaluation")

    c1_cpu_seconds = time.process_time() - c1_cpu_start
    c1_wall_seconds = time.perf_counter() - c1_wall_start
    c1_results: dict[str, Any] = {
        "schema": C1V2_SCHEMA,
        "contract": {
            "source_packet": str(PACKET_PATH),
            "source_dossier": str(DOSSIER_PATH),
            "estimator_import": str(C1_PATH),
            "estimator_input": ["anchors"],
            "truth_evaluation_only": ["truth_unit_scale"],
            "pair_labels_excluded_from_estimator": ["truth_pairs", "candidate_pairs"],
            "source_cad_access_count": 0,
            "test_split_access_count": 0,
            "random_search_count": 0,
            "additional_seed_count": 0,
            "gate_verdict_emitted": False,
        },
        "sealed_configuration": {
            "ransac_log_tolerance": C1.RANSAC_LOG_TOLERANCE,
            "huber_delta": C1.HUBER_DELTA,
            "consensus_threshold": C1.CONSENSUS_THRESHOLD,
            "high_confidence_threshold": C1.HIGH_CONFIDENCE_THRESHOLD,
            "minimum_independent": C1.MIN_INDEPENDENT,
            "accuracy_relative_error": C1.ACCURACY_RELATIVE_ERROR,
            "confidence_formula": "consensus*exp(-log_mad/tau)*min(1,n_independent/5)*min(1,n_spatial_bins/3)",
            "primary_confidence_bin": "unit_status",
            "corruptions": list(C1.CORRUPTIONS),
            "corruption_assignment": "uint32_be(sha256(base_scene_id)[0:4]) mod 4",
        },
        "environment": {
            "python": platform.python_version(),
            "numpy": np.__version__,
            "platform": platform.platform(),
        },
        "selftest": {**selftest_details, "transcript": selftest_transcript},
        "inputs": {
            "directory": str(SCENES_V2_DIR),
            "validation": input_validation,
            "manifest_before": input_manifest_before,
            "manifest_after": input_manifest_after,
            "manifest_mismatch_count": input_manifest_mismatch_count,
        },
        "source_readonly_manifest": {
            "before": source_manifest_before,
            "after": source_manifest_after,
            "mismatch_count": source_manifest_mismatch_count,
        },
        "v1_baseline": baseline_c1,
        "v1_comparison_rows": c1_comparison,
        "aggregates": aggregates,
        "ticket_single_outlier": ticket,
        "scenes": rows,
        "evidence": {**workbook_status, "verification": workbook_verification},
        "runtime": {
            "cpu_seconds": c1_cpu_seconds,
            "wall_seconds": c1_wall_seconds,
            "cpu_budget_seconds": 3600,
            "cpu_budget_exceeded": c1_cpu_seconds > 3600.0,
            "loop_cpu_seconds": time.process_time() - cpu_start,
            "loop_wall_seconds": time.perf_counter() - wall_start,
        },
        "artifact_manifest_scope": "stable predecessor artifacts only; c1v2 self-hash and later REPORT are excluded",
        "artifacts": [
            artifact_row(C0V2_NUMBERS_PATH),
            {
                "artifact": "scenes_v2",
                "status": "GENERATED",
                "bytes": sum(path.stat().st_size for path in scene_paths),
                "sha256_or_digest": input_manifest_after["digest"],
            },
            artifact_row(Path(workbook_status["path"]), workbook_status["status"]),
        ],
        "report_path": str(REPORT_PATH),
    }
    write_json_atomic(C1V2_RESULTS_PATH, c1_results)

    report_artifacts = [
        artifact_row(CELL_DIR / "loop_l1.py"),
        {
            "artifact": "scenes_v2",
            "status": "GENERATED",
            "bytes": sum(path.stat().st_size for path in scene_paths),
            "sha256_or_digest": input_manifest_after["digest"],
        },
        artifact_row(C0V2_NUMBERS_PATH),
        artifact_row(C1V2_RESULTS_PATH),
        artifact_row(Path(workbook_status["path"]), workbook_status["status"]),
    ]
    report = render_report(c0_numbers, c1_results, selftest_transcript, report_artifacts)
    write_text_atomic(REPORT_PATH, report)

    expected_scene_count = C0.BASE_SCENE_COUNT * len(C0.SCALES)
    final_scene_count = len(list(SCENES_V2_DIR.glob("scene_*.json")))
    final_line = REPORT_PATH.read_text(encoding="utf-8").rstrip("\n").splitlines()[-1]
    if final_scene_count != expected_scene_count:
        raise RuntimeError(
            f"final scene count mismatch: {final_scene_count} != {expected_scene_count}"
        )
    if final_line != "LOOP_COMPLETE: L1":
        raise RuntimeError(f"unexpected REPORT final line: {final_line!r}")
    for required_path in (C0V2_NUMBERS_PATH, C1V2_RESULTS_PATH, REPORT_PATH):
        if not required_path.is_file() or required_path.stat().st_size <= 0:
            raise RuntimeError(f"missing or empty output: {required_path}")

    print(
        json.dumps(
            {
                "scene_count": final_scene_count,
                "anchor_count_distribution": anchor_numbers["base_anchor_count_distribution"],
                "high_scene_count": aggregates["overall"]["high_scene_count"],
                "high_coverage": aggregates["overall"]["high_coverage"],
                "high_accuracy_within_5pct": aggregates["overall"][
                    "high_accuracy_within_5pct"
                ],
                "single_outlier_confidence_increase_count": ticket[
                    "v2_confidence_score_increased_count"
                ],
                "single_outlier_status_increase_count": ticket[
                    "v2_status_rank_increased_count"
                ],
                "pair_digest_mismatch_count": aggregates["pair_label_permutation"][
                    "mismatching_anchor_artifact_scene_count"
                ],
                "truth_validator_error_count": c0_numbers["truth_validator"]["error_count"],
                "source_manifest_mismatch_count": source_manifest_mismatch_count,
                "input_manifest_mismatch_count": input_manifest_mismatch_count,
                "evidence_status": workbook_status["status"],
                "report_final_line": final_line,
            },
            ensure_ascii=False,
            sort_keys=True,
        )
    )
    return 0


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--selftest",
        action="store_true",
        help="run deterministic, truth-contact, and anchor-distribution checks without writes",
    )
    arguments = parser.parse_args(argv)
    if arguments.selftest:
        stream = io.StringIO()
        return_code, _details = run_selftests(stream)
        print(stream.getvalue(), end="")
        return return_code
    return execute_full()


if __name__ == "__main__":
    raise SystemExit(main())
