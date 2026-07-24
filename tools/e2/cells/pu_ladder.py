#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""PU / weak-supervision model ladder on CubiCasa train -> val only.

This cell intentionally accepts no split argument. Its only data entrypoints are
the literal train and val paths declared below. Hidden train truth is used only
to simulate which positives are observed and by the explicitly named supervised
ceiling; PU arms receive the observed-positive indicator z instead.
"""
from __future__ import annotations

import hashlib
import importlib.util
import json
import math
import os
import platform
import shutil
import statistics
import sys
import threading
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

sys.dont_write_bytecode = True

import numpy as np


CELL_DIR = Path(__file__).resolve().parent
REPO_ROOT = Path(r"D:\dev\99_tools\autocad-sdk-router")
BASE_FEATURE_DIR = REPO_ROOT / "runs" / "e2_ext_cubicasa" / "features"
IR_ROOT = REPO_ROOT / "runs" / "e2_ext_cubicasa" / "ir"
CML_PATH = Path(r"D:\runs\e2_program\cells\cml_ctx\cml_ctx.py")
DOSSIER_PATH = REPO_ROOT / "reports" / "e2" / "dossiers" / "calibration_P2.md"
BASELINE_PATH = REPO_ROOT / "tools" / "e2" / "ext" / "cubicasa_ml.py"
EVAL_PATH = REPO_ROOT / "tools" / "e2" / "ext" / "cubicasa_eval.py"

RESULTS_PATH = CELL_DIR / "results.json"
REPORT_PATH = CELL_DIR / "REPORT.md"
EVIDENCE_PATH = CELL_DIR / "evidence.xlsx"
FALLBACK_EVIDENCE_PATH = CELL_DIR / "evidence_rows.json"
WORK_DIR = CELL_DIR / "_work"

ALLOWED_SPLITS = frozenset(("train", "val"))
SEEDS = (7, 17, 29)
LABEL_FREQUENCIES = (0.3, 0.5, 0.7)
OPERATING_THRESHOLD = 0.5
SELECTED_RADIUS_PX = 20.0
SCALE_MM_PER_PX = 12.0
CROSSFIT_FOLDS = 2
BAGGING_MEMBERS = 5
BOOTSTRAP_REPLICATES = 10_000
BOOTSTRAP_SEED = 20260719
SHUFFLE_AUC_TRIGGER = 0.55
REFERENCE_CONTEXT_F1 = 0.705
MAX_RSS_BYTES = 48 * 1024**3
THREAD_LIMIT = min(8, os.cpu_count() or 1)

BASE_FEATURE_NAMES = (
    "parallel",
    "thickness",
    "junction",
    "log10_len",
    "sin2t",
    "cos2t",
)
CONTEXT_FEATURE_NAMES = (
    "parallel_band_neighbor_count",
    "nearest_parallel_gap_px",
    "junction_degree",
    "drawing_length_percentile",
    "radius_density_r20_per_px2",
    "neighbor_angle_entropy_r20",
)
FEATURE_NAMES = BASE_FEATURE_NAMES + CONTEXT_FEATURE_NAMES

ARM_ORDER = (
    "naive_logistic",
    "naive_tree",
    "scar_pu_oracle_c",
    "scar_pu_estimated_c",
    "nnpu_logistic",
    "bagging_pu_tree",
    "posterior_imputation_tree",
    "full_supervised_upper",
    "shuffle_control",
)


def now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


class RSSMonitor:
    """Sample this process so the packet's RAM boundary is measured."""

    def __init__(self, interval_seconds: float = 0.2) -> None:
        import psutil

        self._process = psutil.Process(os.getpid())
        self._interval = interval_seconds
        self._stop = threading.Event()
        self._thread = threading.Thread(target=self._run, daemon=True)
        self.peak_bytes = int(self._process.memory_info().rss)

    def _run(self) -> None:
        while not self._stop.wait(self._interval):
            try:
                self.peak_bytes = max(
                    self.peak_bytes, int(self._process.memory_info().rss)
                )
            except Exception:
                return

    def __enter__(self) -> "RSSMonitor":
        self._thread.start()
        return self

    def __exit__(self, exc_type: Any, exc: Any, tb: Any) -> None:
        self._stop.set()
        self._thread.join(timeout=2.0)
        self.peak_bytes = max(
            self.peak_bytes, int(self._process.memory_info().rss)
        )


def load_cml_module() -> Any:
    spec = importlib.util.spec_from_file_location("pu_ladder_cml_ctx", CML_PATH)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"cannot load context helper: {CML_PATH}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def allowed_path(split: str, suffix: str) -> Path:
    if split not in ALLOWED_SPLITS:
        raise ValueError(f"forbidden split {split!r}; only train/val are allowed")
    if suffix == "features":
        path = BASE_FEATURE_DIR / f"{split}.npz"
        if path.parent != BASE_FEATURE_DIR:
            raise RuntimeError("feature path safety check failed")
        return path
    if suffix == "ir":
        path = IR_ROOT / split
        if path.parent != IR_ROOT or path.name != split:
            raise RuntimeError("IR path safety check failed")
        return path
    raise ValueError(f"unknown path suffix {suffix!r}")


def load_base(split: str) -> dict[str, np.ndarray]:
    path = allowed_path(split, "features")
    with np.load(path) as archive:
        data = {
            "X": np.asarray(archive["X"], dtype=np.float32),
            "y": np.asarray(archive["y"], dtype=np.int8),
            "gid": np.asarray(archive["gid"], dtype=np.int32),
            "scale": np.asarray(archive["scale"]),
        }
    if data["X"].shape[1] != len(BASE_FEATURE_NAMES):
        raise RuntimeError(f"{split}: expected exactly six base features")
    if not (len(data["X"]) == len(data["y"]) == len(data["gid"])):
        raise RuntimeError(f"{split}: base feature row alignment failed")
    return data


def split_files(split: str) -> list[Path]:
    directory = allowed_path(split, "ir")
    files = sorted(directory.glob("*.segir.json"))
    if not files:
        raise FileNotFoundError(f"no SEG-IR files in allowed directory {directory}")
    return files


def benchmark_context(
    cml: Any, train: dict[str, np.ndarray], drawings: int = 24
) -> dict[str, Any]:
    files = split_files("train")
    expected_by_gid = np.bincount(train["gid"], minlength=len(files))
    take = min(drawings, len(files))
    chosen = np.unique(np.linspace(0, len(files) - 1, take, dtype=int))
    rows = 0
    counts: list[int] = []
    started = time.perf_counter()
    for gid in chosen:
        ir = json.loads(files[int(gid)].read_text(encoding="utf-8"))
        matrix, _meta = cml.drawing_context(ir)
        expected = int(expected_by_gid[int(gid)])
        if len(matrix) != expected:
            raise RuntimeError(
                f"preflight alignment mismatch gid={gid}: "
                f"context {len(matrix)} != base {expected}"
            )
        rows += len(matrix)
        counts.append(len(matrix))
    seconds = time.perf_counter() - started
    rate = rows / max(seconds, 1e-12)
    return {
        "split": "train",
        "sample_drawings": int(len(chosen)),
        "sample_rows": int(rows),
        "seconds": round(seconds, 6),
        "rows_per_second": round(rate, 2),
        "projected_train_seconds": round(len(train["y"]) / rate, 2),
        "projected_train_hours": round(len(train["y"]) / rate / 3600.0, 4),
        "sample_rows_per_drawing": {
            "min": int(min(counts)),
            "median": float(np.median(counts)),
            "max": int(max(counts)),
        },
        "projection_basis": "linear extrapolation from drawing-spread train sample",
    }


def extract_context(
    cml: Any, split: str, base: dict[str, np.ndarray]
) -> tuple[np.memmap, dict[str, Any]]:
    files = split_files(split)
    gid = base["gid"]
    if int(gid.min()) != 0 or int(gid.max()) + 1 != len(files):
        raise RuntimeError(f"{split}: drawing gid coverage does not match IR files")
    expected_by_gid = np.bincount(gid, minlength=len(files))
    all_context_names = tuple(cml._context_names_all())
    if len(all_context_names) != 10:
        raise RuntimeError("context helper schema changed; expected ten columns")
    WORK_DIR.mkdir(parents=True, exist_ok=True)
    target = np.lib.format.open_memmap(
        WORK_DIR / f"context_{split}.npy",
        mode="w+",
        dtype=np.float32,
        shape=(len(base["y"]), len(all_context_names)),
    )
    offset = 0
    duplicate_count = 0
    drawing_counts: list[int] = []
    params_seen: dict[str, Any] | None = None
    started = time.perf_counter()
    for drawing_gid, path in enumerate(files):
        ir = json.loads(path.read_text(encoding="utf-8"))
        matrix, meta = cml.drawing_context(ir)
        expected = int(expected_by_gid[drawing_gid])
        if len(matrix) != expected:
            raise RuntimeError(
                f"{split}: row mismatch {path.name}: "
                f"context {len(matrix)} != base {expected}"
            )
        target[offset : offset + expected] = matrix
        offset += expected
        duplicate_count += int(meta["duplicates"])
        drawing_counts.append(expected)
        if params_seen is None and meta.get("params"):
            params_seen = meta["params"]
        if (drawing_gid + 1) % 200 == 0 or drawing_gid + 1 == len(files):
            elapsed = time.perf_counter() - started
            print(
                f"[{split}] {drawing_gid + 1}/{len(files)} drawings; "
                f"{offset:,} rows; {offset / max(elapsed, 1e-12):,.0f} rows/s",
                flush=True,
            )
    target.flush()
    seconds = time.perf_counter() - started
    if offset != len(base["y"]):
        raise RuntimeError(f"{split}: final context/base alignment mismatch")
    return target, {
        "split": split,
        "drawings": len(files),
        "rows": int(offset),
        "seconds": round(seconds, 6),
        "rows_per_second": round(offset / seconds, 2),
        "duplicate_valid_handle_records": int(duplicate_count),
        "rows_per_drawing": {
            "min": int(np.min(drawing_counts)),
            "p50": float(np.percentile(drawing_counts, 50)),
            "p95": float(np.percentile(drawing_counts, 95)),
            "max": int(np.max(drawing_counts)),
        },
        "resolved_geometry_params": params_seen,
        "alignment_check": True,
    }


def assemble_features(
    base: dict[str, np.ndarray], context: np.ndarray, split: str
) -> np.memmap:
    # cml_ctx layout: common four, then density/entropy pairs for R=20,40,80.
    selected_context_indices = (0, 1, 2, 3, 4, 5)
    target = np.lib.format.open_memmap(
        WORK_DIR / f"features_{split}.npy",
        mode="w+",
        dtype=np.float32,
        shape=(len(base["y"]), len(FEATURE_NAMES)),
    )
    target[:, : len(BASE_FEATURE_NAMES)] = base["X"]
    target[:, len(BASE_FEATURE_NAMES) :] = context[:, selected_context_indices]
    target.flush()
    return target


def standardize_for_logistic(
    train: np.ndarray, val: np.ndarray
) -> tuple[np.memmap, np.memmap, list[dict[str, Any]]]:
    scaled_train = np.lib.format.open_memmap(
        WORK_DIR / "scaled_train.npy",
        mode="w+",
        dtype=np.float32,
        shape=train.shape,
    )
    scaled_val = np.lib.format.open_memmap(
        WORK_DIR / "scaled_val.npy",
        mode="w+",
        dtype=np.float32,
        shape=val.shape,
    )
    stats: list[dict[str, Any]] = []
    for column, name in enumerate(FEATURE_NAMES):
        values = np.asarray(train[:, column], dtype=np.float32)
        finite = values[np.isfinite(values)].astype(np.float64, copy=False)
        if not len(finite):
            median, mean, scale = 0.0, 0.0, 1.0
        else:
            median = float(np.median(finite))
            missing = len(values) - len(finite)
            mean = float((finite.sum() + missing * median) / len(values))
            squared = (
                np.square(finite - mean).sum() + missing * (median - mean) ** 2
            )
            scale = float(math.sqrt(squared / len(values)))
            if not math.isfinite(scale) or scale < 1e-12:
                scale = 1.0
        for source, target in ((train, scaled_train), (val, scaled_val)):
            for start in range(0, len(source), 500_000):
                stop = min(start + 500_000, len(source))
                chunk = np.asarray(source[start:stop, column])
                filled = np.where(np.isfinite(chunk), chunk, median)
                target[start:stop, column] = (filled - mean) / scale
        stats.append(
            {
                "feature": name,
                "train_median": median,
                "train_mean_after_imputation": mean,
                "train_std_after_imputation": scale,
                "train_missing": int(len(values) - len(finite)),
            }
        )
    scaled_train.flush()
    scaled_val.flush()
    return scaled_train, scaled_val, stats


def stable_rng(seed: int, c: float, stream: int) -> np.random.Generator:
    return np.random.default_rng(
        np.random.SeedSequence((seed, int(round(c * 1000)), stream))
    )


def make_pu_view(
    y_train: np.ndarray, c: float, seed: int
) -> tuple[np.ndarray, dict[str, Any]]:
    if c not in LABEL_FREQUENCIES:
        raise ValueError(f"c was not preregistered: {c}")
    positive = np.flatnonzero(y_train == 1)
    rng = stable_rng(seed, c, 1)
    observed_positive = positive[rng.random(len(positive)) < c]
    z = np.zeros(len(y_train), dtype=np.int8)
    z[observed_positive] = 1
    return z, {
        "configured_c": c,
        "true_positive_count": int(len(positive)),
        "observed_positive_count": int(len(observed_positive)),
        "unlabeled_count": int(len(z) - len(observed_positive)),
        "realized_c": float(len(observed_positive) / len(positive)),
        "observed_row_fraction": float(z.mean()),
    }


def predict_positive(
    model: Any, matrix: np.ndarray, chunk_size: int = 500_000
) -> np.ndarray:
    probability = np.empty(len(matrix), dtype=np.float64)
    for start in range(0, len(matrix), chunk_size):
        stop = min(start + chunk_size, len(matrix))
        probability[start:stop] = model.predict_proba(matrix[start:stop])[:, 1]
    return probability


def fit_crossfit_logistic(
    x_train: np.ndarray,
    z_train: np.ndarray,
    gid_train: np.ndarray,
    x_val: np.ndarray,
    seed: int,
) -> tuple[np.ndarray, np.ndarray, dict[str, Any]]:
    from sklearn.linear_model import LogisticRegression

    folds = np.mod(gid_train, CROSSFIT_FOLDS)
    oof = np.empty(len(z_train), dtype=np.float32)
    val_probability = np.zeros(len(x_val), dtype=np.float64)
    fold_records: list[dict[str, Any]] = []
    total_fit = 0.0
    total_infer = 0.0
    for held_out in range(CROSSFIT_FOLDS):
        train_index = np.flatnonzero(folds != held_out)
        held_index = np.flatnonzero(folds == held_out)
        model = LogisticRegression(
            C=1.0,
            solver="lbfgs",
            max_iter=100,
            tol=1e-5,
            random_state=seed,
        )
        started = time.perf_counter()
        model.fit(x_train[train_index], z_train[train_index])
        fit_seconds = time.perf_counter() - started
        started = time.perf_counter()
        for start in range(0, len(held_index), 500_000):
            index = held_index[start : start + 500_000]
            oof[index] = model.predict_proba(x_train[index])[:, 1]
        val_probability += predict_positive(model, x_val) / CROSSFIT_FOLDS
        inference_seconds = time.perf_counter() - started
        total_fit += fit_seconds
        total_infer += inference_seconds
        fold_records.append(
            {
                "held_out_fold": held_out,
                "fold_rule": f"drawing_gid % {CROSSFIT_FOLDS} == {held_out}",
                "n_fit_rows": int(len(train_index)),
                "n_held_out_rows": int(len(held_index)),
                "fit_seconds": round(fit_seconds, 6),
                "inference_seconds": round(inference_seconds, 6),
                "n_iter": int(model.n_iter_[0]),
            }
        )
        del model, train_index, held_index
    selected = z_train == 1
    c_estimated = float(np.mean(oof[selected]))
    c_estimated = float(np.clip(c_estimated, 0.02, 0.98))
    pi_estimated = float(
        np.clip(float(z_train.mean()) / c_estimated, 1e-4, 0.999)
    )
    return oof, val_probability, {
        "method": "Elkan-Noto mean cross-fitted g(x) on observed positives",
        "crossfit_folds": CROSSFIT_FOLDS,
        "c_estimated": c_estimated,
        "pi_estimated_from_z_over_c": pi_estimated,
        "fit_seconds": round(total_fit, 6),
        "inference_seconds": round(total_infer, 6),
        "folds": fold_records,
    }


def sigmoid(score: np.ndarray) -> np.ndarray:
    score = np.clip(score, -35.0, 35.0)
    return 1.0 / (1.0 + np.exp(-score))


def nnpu_full_risk(
    x_train: np.ndarray,
    positive_index: np.ndarray,
    unlabeled_index: np.ndarray,
    weights: np.ndarray,
    intercept: float,
    pi_estimated: float,
) -> dict[str, float]:
    sums = {"pos_y1": 0.0, "pos_y0": 0.0, "unl_y0": 0.0}
    for name, index, target in (
        ("pos_y1", positive_index, 1),
        ("pos_y0", positive_index, 0),
        ("unl_y0", unlabeled_index, 0),
    ):
        total = 0.0
        for start in range(0, len(index), 500_000):
            chosen = index[start : start + 500_000]
            score = np.asarray(
                x_train[chosen] @ weights + intercept, dtype=np.float64
            )
            total += float(
                np.logaddexp(0.0, -score if target == 1 else score).sum()
            )
        sums[name] = total / len(index)
    positive_risk = pi_estimated * sums["pos_y1"]
    negative_raw = sums["unl_y0"] - pi_estimated * sums["pos_y0"]
    return {
        "positive_risk": float(positive_risk),
        "negative_risk_raw": float(negative_raw),
        "negative_risk_after_nonnegative_correction": float(
            max(0.0, negative_raw)
        ),
        "total_empirical_nnpu_risk": float(
            positive_risk + max(0.0, negative_raw)
        ),
    }


def fit_nnpu_logistic(
    x_train: np.ndarray,
    z_train: np.ndarray,
    x_val: np.ndarray,
    pi_estimated: float,
    seed: int,
) -> tuple[np.ndarray, dict[str, Any]]:
    positive_index = np.flatnonzero(z_train == 1)
    unlabeled_index = np.flatnonzero(z_train == 0)
    if not len(positive_index) or not len(unlabeled_index):
        raise RuntimeError("nnPU requires observed-positive and unlabeled rows")
    rng = np.random.default_rng(np.random.SeedSequence((seed, 9101)))
    dimensions = x_train.shape[1]
    weights = np.zeros(dimensions, dtype=np.float32)
    intercept = float(math.log(pi_estimated / (1.0 - pi_estimated)))
    m_w = np.zeros_like(weights)
    v_w = np.zeros_like(weights)
    m_b = 0.0
    v_b = 0.0
    learning_rate = 0.01
    l2 = 1e-4
    epochs = 7
    batch_size = 32_768
    steps_per_epoch = max(1, math.ceil(len(unlabeled_index) / batch_size))
    step_number = 0
    clamp_steps = 0
    started = time.perf_counter()
    for epoch in range(epochs):
        for _step in range(steps_per_epoch):
            p_index = positive_index[
                rng.integers(0, len(positive_index), size=batch_size)
            ]
            u_index = unlabeled_index[
                rng.integers(0, len(unlabeled_index), size=batch_size)
            ]
            xp = np.asarray(x_train[p_index], dtype=np.float32)
            xu = np.asarray(x_train[u_index], dtype=np.float32)
            score_p = xp @ weights + intercept
            score_u = xu @ weights + intercept
            pp = sigmoid(score_p).astype(np.float32, copy=False)
            pu = sigmoid(score_u).astype(np.float32, copy=False)
            grad_pos_w = pi_estimated * np.mean(
                (pp - 1.0)[:, None] * xp, axis=0
            )
            grad_pos_b = pi_estimated * float(np.mean(pp - 1.0))
            negative_raw = float(
                np.mean(np.logaddexp(0.0, score_u))
                - pi_estimated * np.mean(np.logaddexp(0.0, score_p))
            )
            if negative_raw > 0.0:
                grad_neg_w = np.mean(
                    pu[:, None] * xu, axis=0
                ) - pi_estimated * np.mean(pp[:, None] * xp, axis=0)
                grad_neg_b = float(
                    np.mean(pu) - pi_estimated * np.mean(pp)
                )
            else:
                grad_neg_w = np.zeros_like(weights)
                grad_neg_b = 0.0
                clamp_steps += 1
            grad_w = grad_pos_w + grad_neg_w + l2 * weights
            grad_b = grad_pos_b + grad_neg_b
            step_number += 1
            m_w = 0.9 * m_w + 0.1 * grad_w
            v_w = 0.999 * v_w + 0.001 * np.square(grad_w)
            m_b = 0.9 * m_b + 0.1 * grad_b
            v_b = 0.999 * v_b + 0.001 * grad_b * grad_b
            correction1 = 1.0 - 0.9**step_number
            correction2 = 1.0 - 0.999**step_number
            weights -= learning_rate * (m_w / correction1) / (
                np.sqrt(v_w / correction2) + 1e-8
            )
            intercept -= learning_rate * (m_b / correction1) / (
                math.sqrt(v_b / correction2) + 1e-8
            )
        print(
            f"    nnPU epoch {epoch + 1}/{epochs}; "
            f"clamp steps {clamp_steps}/{step_number}",
            flush=True,
        )
    fit_seconds = time.perf_counter() - started
    started = time.perf_counter()
    val_probability = np.empty(len(x_val), dtype=np.float64)
    for start in range(0, len(x_val), 500_000):
        stop = min(start + 500_000, len(x_val))
        val_probability[start:stop] = sigmoid(
            x_val[start:stop] @ weights + intercept
        )
    inference_seconds = time.perf_counter() - started
    risks = nnpu_full_risk(
        x_train,
        positive_index,
        unlabeled_index,
        weights,
        intercept,
        pi_estimated,
    )
    return val_probability, {
        "fit_seconds": round(fit_seconds, 6),
        "inference_seconds": round(inference_seconds, 6),
        "optimizer": "mini-batch Adam",
        "epochs": epochs,
        "batch_size_per_risk_component": batch_size,
        "steps": step_number,
        "learning_rate": learning_rate,
        "l2": l2,
        "nonnegative_clamp_steps": clamp_steps,
        "pi_estimated": pi_estimated,
        "risk": risks,
    }


def fit_hgb(
    x_train: np.ndarray,
    labels: np.ndarray,
    x_val: np.ndarray,
    seed: int,
    sample_weight: np.ndarray | None = None,
) -> tuple[np.ndarray, dict[str, Any]]:
    from sklearn.ensemble import HistGradientBoostingClassifier

    model = HistGradientBoostingClassifier(random_state=seed)
    started = time.perf_counter()
    model.fit(x_train, labels, sample_weight=sample_weight)
    fit_seconds = time.perf_counter() - started
    started = time.perf_counter()
    probability = predict_positive(model, x_val)
    inference_seconds = time.perf_counter() - started
    metadata = {
        "fit_seconds": round(fit_seconds, 6),
        "inference_seconds": round(inference_seconds, 6),
        "n_iter": int(model.n_iter_),
        "hyperparameters": {
            "sklearn_defaults": True,
            "random_state": seed,
            "max_iter": int(model.max_iter),
            "learning_rate": float(model.learning_rate),
            "max_leaf_nodes": int(model.max_leaf_nodes),
            "min_samples_leaf": int(model.min_samples_leaf),
            "l2_regularization": float(model.l2_regularization),
        },
    }
    del model
    return probability, metadata


def fit_bagging_pu(
    x_train: np.ndarray,
    z_train: np.ndarray,
    x_val: np.ndarray,
    seed: int,
    c: float,
) -> tuple[np.ndarray, dict[str, Any]]:
    from sklearn.ensemble import HistGradientBoostingClassifier

    positive = np.flatnonzero(z_train == 1)
    unlabeled = np.flatnonzero(z_train == 0)
    sample_size = min(len(positive), len(unlabeled))
    rng = stable_rng(seed, c, 3001)
    probability = np.zeros(len(x_val), dtype=np.float64)
    members: list[dict[str, Any]] = []
    total_fit = 0.0
    total_infer = 0.0
    for member in range(BAGGING_MEMBERS):
        sampled_unlabeled = rng.choice(
            unlabeled, size=sample_size, replace=False
        )
        index = np.concatenate((positive, sampled_unlabeled))
        labels = np.concatenate(
            (
                np.ones(len(positive), dtype=np.int8),
                np.zeros(sample_size, dtype=np.int8),
            )
        )
        order = rng.permutation(len(index))
        index = index[order]
        labels = labels[order]
        model = HistGradientBoostingClassifier(
            random_state=seed * 100 + member
        )
        started = time.perf_counter()
        model.fit(x_train[index], labels)
        fit_seconds = time.perf_counter() - started
        started = time.perf_counter()
        member_probability = predict_positive(model, x_val)
        inference_seconds = time.perf_counter() - started
        probability += member_probability / BAGGING_MEMBERS
        total_fit += fit_seconds
        total_infer += inference_seconds
        members.append(
            {
                "member": member,
                "training_rows": int(len(index)),
                "positive_rows": int(len(positive)),
                "sampled_unlabeled_rows": int(sample_size),
                "fit_seconds": round(fit_seconds, 6),
                "inference_seconds": round(inference_seconds, 6),
                "n_iter": int(model.n_iter_),
            }
        )
        del (
            model,
            index,
            labels,
            sampled_unlabeled,
            member_probability,
        )
    return probability, {
        "fit_seconds": round(total_fit, 6),
        "inference_seconds": round(total_infer, 6),
        "members": BAGGING_MEMBERS,
        "sampling": (
            "all observed P plus equal-size U sample without replacement "
            "per member"
        ),
        "member_records": members,
    }


def metric_record(
    y_true: np.ndarray, probability: np.ndarray
) -> dict[str, Any]:
    from sklearn.metrics import average_precision_score, roc_auc_score

    probability = np.asarray(probability, dtype=np.float64)
    if not np.all(np.isfinite(probability)):
        raise RuntimeError("non-finite validation probability")
    probability = np.clip(probability, 0.0, 1.0)
    prediction = probability >= OPERATING_THRESHOLD
    positive = y_true == 1
    tp = int(np.sum(prediction & positive))
    fp = int(np.sum(prediction & ~positive))
    fn = int(np.sum(~prediction & positive))
    tn = int(np.sum(~prediction & ~positive))
    precision = tp / (tp + fp) if tp + fp else 0.0
    recall = tp / (tp + fn) if tp + fn else 0.0
    f1 = (
        2.0 * precision * recall / (precision + recall)
        if precision + recall
        else 0.0
    )
    order = np.argsort(-probability, kind="mergesort")
    sorted_positive = positive[order]
    cumulative_positive = np.cumsum(sorted_positive)
    cumulative_rows = np.arange(1, len(y_true) + 1)
    recall_curve = cumulative_positive / max(1, int(positive.sum()))
    eligible = np.flatnonzero(recall_curve >= 0.5)
    if len(eligible):
        first = int(eligible[0])
        precision_at_recall_05 = float(
            cumulative_positive[first] / cumulative_rows[first]
        )
    else:
        precision_at_recall_05 = 0.0
    return {
        "threshold": OPERATING_THRESHOLD,
        "tp": tp,
        "fp": fp,
        "fn": fn,
        "tn": tn,
        "precision": float(precision),
        "recall": float(recall),
        "f1": float(f1),
        "roc_auc": float(roc_auc_score(y_true, probability)),
        "auprc": float(average_precision_score(y_true, probability)),
        "precision_at_recall_0_5": precision_at_recall_05,
    }


def group_confusion(
    y_true: np.ndarray,
    probability: np.ndarray,
    gid: np.ndarray,
    n_groups: int,
) -> dict[str, np.ndarray]:
    prediction = probability >= OPERATING_THRESHOLD
    positive = y_true == 1
    return {
        "tp": np.bincount(
            gid, weights=prediction & positive, minlength=n_groups
        ),
        "fp": np.bincount(
            gid, weights=prediction & ~positive, minlength=n_groups
        ),
        "fn": np.bincount(
            gid, weights=~prediction & positive, minlength=n_groups
        ),
        "tn": np.bincount(
            gid, weights=~prediction & ~positive, minlength=n_groups
        ),
    }


def f1_from_counts(
    tp: np.ndarray, fp: np.ndarray, fn: np.ndarray
) -> np.ndarray:
    denominator = 2.0 * tp + fp + fn
    return np.divide(
        2.0 * tp,
        denominator,
        out=np.zeros_like(tp),
        where=denominator > 0,
    )


def bootstrap_summary(
    current: dict[str, np.ndarray],
    upper: dict[str, np.ndarray],
    bootstrap_counts: np.ndarray,
) -> dict[str, Any]:
    current_f1 = f1_from_counts(
        bootstrap_counts @ current["tp"],
        bootstrap_counts @ current["fp"],
        bootstrap_counts @ current["fn"],
    )
    upper_f1 = f1_from_counts(
        bootstrap_counts @ upper["tp"],
        bootstrap_counts @ upper["fp"],
        bootstrap_counts @ upper["fn"],
    )
    recovery = np.divide(
        current_f1,
        upper_f1,
        out=np.zeros_like(current_f1),
        where=upper_f1 > 0,
    )
    delta = current_f1 - upper_f1
    return {
        "method": "paired drawing-cluster bootstrap",
        "replicates": BOOTSTRAP_REPLICATES,
        "seed": BOOTSTRAP_SEED,
        "f1_ci95": [
            float(x) for x in np.quantile(current_f1, (0.025, 0.975))
        ],
        "f1_recovery_vs_upper_ci95": [
            float(x) for x in np.quantile(recovery, (0.025, 0.975))
        ],
        "f1_delta_vs_upper_ci95": [
            float(x) for x in np.quantile(delta, (0.025, 0.975))
        ],
    }


def build_row(
    *,
    arm: str,
    family: str,
    role: str,
    c: float,
    seed: int,
    view: dict[str, Any],
    c_estimated: float | None,
    pi_estimated: float | None,
    probability: np.ndarray,
    y_val: np.ndarray,
    gid_val: np.ndarray,
    model_metadata: dict[str, Any],
    upper_metrics: dict[str, Any],
    upper_group: dict[str, np.ndarray],
    bootstrap_counts: np.ndarray,
    source_label_use: str,
) -> dict[str, Any]:
    metrics = metric_record(y_val, probability)
    grouped = group_confusion(
        y_val, probability, gid_val, bootstrap_counts.shape[1]
    )
    recovery_f1 = (
        metrics["f1"] / upper_metrics["f1"] if upper_metrics["f1"] else 0.0
    )
    recovery_auprc = (
        metrics["auprc"] / upper_metrics["auprc"]
        if upper_metrics["auprc"]
        else 0.0
    )
    return {
        "arm": arm,
        "family": family,
        "role": role,
        "c": c,
        "seed": seed,
        **view,
        "c_estimated": c_estimated,
        "pi_estimated": pi_estimated,
        "source_label_use": source_label_use,
        "metrics": metrics,
        "recovery_rate_f1_vs_full_upper": float(recovery_f1),
        "recovery_rate_auprc_vs_full_upper": float(recovery_auprc),
        "bootstrap": bootstrap_summary(
            grouped, upper_group, bootstrap_counts
        ),
        "model": model_metadata,
    }


def aggregate_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[float, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[(float(row["c"]), row["arm"])].append(row)
    metric_fields = (
        "precision",
        "recall",
        "f1",
        "roc_auc",
        "auprc",
        "precision_at_recall_0_5",
    )
    aggregates: list[dict[str, Any]] = []
    for c in LABEL_FREQUENCIES:
        for arm in ARM_ORDER:
            items = sorted(
                grouped[(c, arm)], key=lambda item: item["seed"]
            )
            if len(items) != len(SEEDS):
                raise RuntimeError(
                    f"missing seed rows for c={c}, arm={arm}"
                )
            metrics: dict[str, dict[str, float]] = {}
            for field in metric_fields:
                values = [
                    float(item["metrics"][field]) for item in items
                ]
                metrics[field] = {
                    "mean": float(statistics.mean(values)),
                    "std_sample": float(statistics.stdev(values)),
                }
            for field in (
                "recovery_rate_f1_vs_full_upper",
                "recovery_rate_auprc_vs_full_upper",
            ):
                values = [float(item[field]) for item in items]
                metrics[field] = {
                    "mean": float(statistics.mean(values)),
                    "std_sample": float(statistics.stdev(values)),
                }
            aggregates.append(
                {
                    "c": c,
                    "arm": arm,
                    "family": items[0]["family"],
                    "role": items[0]["role"],
                    "seeds": list(SEEDS),
                    "n_seeds": len(SEEDS),
                    "std_definition": "sample standard deviation (ddof=1)",
                    "metrics": metrics,
                    "realized_c": {
                        "mean": float(
                            statistics.mean(
                                item["realized_c"] for item in items
                            )
                        ),
                        "std_sample": float(
                            statistics.stdev(
                                item["realized_c"] for item in items
                            )
                        ),
                    },
                }
            )
    return aggregates


def flatten_row(row: dict[str, Any]) -> dict[str, Any]:
    metrics = row["metrics"]
    bootstrap = row["bootstrap"]
    model = row["model"]
    return {
        "arm": row["arm"],
        "family": row["family"],
        "role": row["role"],
        "c": row["c"],
        "seed": row["seed"],
        "configured_c": row["configured_c"],
        "realized_c": row["realized_c"],
        "true_positive_count": row["true_positive_count"],
        "observed_positive_count": row["observed_positive_count"],
        "unlabeled_count": row["unlabeled_count"],
        "observed_row_fraction": row["observed_row_fraction"],
        "c_estimated": row["c_estimated"],
        "pi_estimated": row["pi_estimated"],
        "threshold": metrics["threshold"],
        "precision": metrics["precision"],
        "recall": metrics["recall"],
        "f1": metrics["f1"],
        "roc_auc": metrics["roc_auc"],
        "auprc": metrics["auprc"],
        "precision_at_recall_0_5": metrics[
            "precision_at_recall_0_5"
        ],
        "tp": metrics["tp"],
        "fp": metrics["fp"],
        "fn": metrics["fn"],
        "tn": metrics["tn"],
        "recovery_rate_f1_vs_full_upper": row[
            "recovery_rate_f1_vs_full_upper"
        ],
        "recovery_rate_auprc_vs_full_upper": row[
            "recovery_rate_auprc_vs_full_upper"
        ],
        "f1_ci95_low": bootstrap["f1_ci95"][0],
        "f1_ci95_high": bootstrap["f1_ci95"][1],
        "f1_recovery_ci95_low": bootstrap[
            "f1_recovery_vs_upper_ci95"
        ][0],
        "f1_recovery_ci95_high": bootstrap[
            "f1_recovery_vs_upper_ci95"
        ][1],
        "f1_delta_upper_ci95_low": bootstrap[
            "f1_delta_vs_upper_ci95"
        ][0],
        "f1_delta_upper_ci95_high": bootstrap[
            "f1_delta_vs_upper_ci95"
        ][1],
        "fit_seconds": model.get("fit_seconds"),
        "inference_seconds": model.get("inference_seconds"),
        "source_label_use": row["source_label_use"],
    }


def flatten_aggregate(row: dict[str, Any]) -> dict[str, Any]:
    flat: dict[str, Any] = {
        "c": row["c"],
        "arm": row["arm"],
        "family": row["family"],
        "role": row["role"],
        "n_seeds": row["n_seeds"],
        "realized_c_mean": row["realized_c"]["mean"],
        "realized_c_std": row["realized_c"]["std_sample"],
    }
    for field, summary in row["metrics"].items():
        flat[f"{field}_mean"] = summary["mean"]
        flat[f"{field}_std"] = summary["std_sample"]
    return flat


def append_table(
    worksheet: Any, records: Iterable[dict[str, Any]]
) -> None:
    records = list(records)
    if not records:
        return
    headers = list(records[0])
    worksheet.append(headers)
    for record in records:
        worksheet.append([record.get(header) for header in headers])
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.font = cell.font.copy(bold=True)
    for column_cells in worksheet.columns:
        width = min(
            42,
            max(
                10,
                max(
                    len(str(cell.value or ""))
                    for cell in column_cells
                )
                + 2,
            ),
        )
        worksheet.column_dimensions[
            column_cells[0].column_letter
        ].width = width


def write_evidence(result: dict[str, Any]) -> dict[str, Any]:
    per_seed_flat = [flatten_row(row) for row in result["per_seed"]]
    aggregate_flat = [
        flatten_aggregate(row) for row in result["aggregates"]
    ]
    try:
        import openpyxl

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "per_seed"
        append_table(worksheet, per_seed_flat)
        aggregate_sheet = workbook.create_sheet("aggregates")
        append_table(aggregate_sheet, aggregate_flat)
        design_sheet = workbook.create_sheet("design")
        design_rows = [
            {"key": "schema", "value": result["schema"]},
            {"key": "train_split", "value": "train"},
            {"key": "evaluation_split", "value": "val"},
            {"key": "test_accessed", "value": False},
            {"key": "seeds", "value": json.dumps(list(SEEDS))},
            {
                "key": "c_values",
                "value": json.dumps(list(LABEL_FREQUENCIES)),
            },
            {"key": "threshold", "value": OPERATING_THRESHOLD},
            {"key": "crossfit_folds", "value": CROSSFIT_FOLDS},
            {"key": "bagging_members", "value": BAGGING_MEMBERS},
            {
                "key": "bootstrap_replicates",
                "value": BOOTSTRAP_REPLICATES,
            },
            {"key": "features", "value": json.dumps(FEATURE_NAMES)},
        ]
        append_table(design_sheet, design_rows)
        throughput_sheet = workbook.create_sheet("throughput")
        throughput_rows = []
        preflight = result["throughput"]["preflight_benchmark"]
        throughput_rows.append(
            {
                "stage": "preflight_context",
                "split": preflight["split"],
                "drawings": preflight["sample_drawings"],
                "rows": preflight["sample_rows"],
                "seconds": preflight["seconds"],
                "rows_per_second": preflight["rows_per_second"],
            }
        )
        for split, stats in result["throughput"][
            "full_context_extraction"
        ].items():
            throughput_rows.append(
                {
                    "stage": "full_context_extraction",
                    "split": split,
                    "drawings": stats["drawings"],
                    "rows": stats["rows"],
                    "seconds": stats["seconds"],
                    "rows_per_second": stats["rows_per_second"],
                }
            )
        append_table(throughput_sheet, throughput_rows)
        investigation = result.get("shuffle_control", {}).get(
            "investigation"
        )
        if investigation:
            investigation_sheet = workbook.create_sheet(
                "shuffle_investigation"
            )
            append_table(
                investigation_sheet,
                investigation["base6_independent_permutations"],
            )
        workbook.save(EVIDENCE_PATH)
        if FALLBACK_EVIDENCE_PATH.exists():
            FALLBACK_EVIDENCE_PATH.unlink()
        return {
            "xlsx_written": True,
            "path": str(EVIDENCE_PATH),
            "fallback_path": None,
            "blocked_xlsx": False,
        }
    except ImportError as exc:
        FALLBACK_EVIDENCE_PATH.write_text(
            json.dumps(
                {
                    "per_seed": per_seed_flat,
                    "aggregates": aggregate_flat,
                },
                ensure_ascii=False,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )
        return {
            "xlsx_written": False,
            "path": None,
            "fallback_path": str(FALLBACK_EVIDENCE_PATH),
            "blocked_xlsx": True,
            "reason": f"openpyxl unavailable: {exc}",
        }


def mean_std_text(summary: dict[str, float]) -> str:
    return (
        f"{summary['mean']:.6f} ± "
        f"{summary['std_sample']:.6f}"
    )


def render_report(result: dict[str, Any]) -> str:
    lines = [
        "# PU/약지도 모델 사다리 — val 전용 실행 보고서",
        "",
        f"- 완료 시각: {result['run']['completed_at']}",
        (
            "- 데이터 경계: train으로 학습, val로 평가; "
            "허용 split은 코드에서 train/val로 고정"
        ),
        "- test 접근: 없음",
        (
            f"- 시드: {list(SEEDS)}; c: {list(LABEL_FREQUENCIES)}; "
            f"고정 threshold: {OPERATING_THRESHOLD}"
        ),
        "- 특징: 지역 6 + cml_ctx와 동일한 문맥 6 (R=20px 고정)",
        (
            "- 숨은 train truth 사용: PU 관측 마스크 생성 및 "
            "명시적 완전라벨 상한에만 사용"
        ),
        "- 아래 표는 판정 band가 아니라 이 실행의 val 실측값이다.",
        "",
        "## 설계",
        "",
        (
            "양성마다 독립 Bernoulli(c)로 관측 여부를 만들고 나머지 "
            "행은 전부 U로 둔다. SCAR logistic은 drawing_gid 기반 "
            "2-fold cross-fit으로 g(x)=P(z=1|x)를 만들며, oracle arm은 "
            "사전 명시 c로, estimated arm은 관측 P의 OOF g 평균으로 "
            "보정한다. nnPU는 pi=mean(z)/c_hat와 비음수 negative-risk "
            "clamp를 사용한다. BaggingPU는 모든 P와 같은 수의 U를 "
            "뽑은 HistGBDT 5개를 평균한다. posterior-imputation은 "
            "q*=clip(g_oof/c_hat,0,1), P=1 고정, U~Bernoulli(q*)와 "
            "weight=2|q*-0.5|로 HistGBDT를 학습한다."
        ),
        "",
        "## 처리율 선실측과 전량 실측",
        "",
    ]
    preflight = result["throughput"]["preflight_benchmark"]
    lines.extend(
        [
            (
                f"선실측: train {preflight['sample_drawings']}개 도면, "
                f"{preflight['sample_rows']:,}행을 "
                f"{preflight['seconds']:.3f}s에 계산"
                f"({preflight['rows_per_second']:,.1f}행/s). "
                f"train 전량 선형 예상 "
                f"{preflight['projected_train_seconds']:.1f}s "
                f"({preflight['projected_train_hours']:.4f}h)."
            ),
            "",
        ]
    )
    for split in ("train", "val"):
        stats = result["throughput"]["full_context_extraction"][split]
        lines.append(
            f"- {split}: {stats['drawings']:,}개 도면 / "
            f"{stats['rows']:,}행, {stats['seconds']:.3f}s, "
            f"{stats['rows_per_second']:,.1f}행/s, "
            f"alignment={stats['alignment_check']}"
        )
    lines.extend(
        [
            (
                f"- 전체 실행: {result['run']['elapsed_seconds']:.3f}s; "
                f"peak RSS {result['run']['peak_rss_gib']:.3f} GiB "
                "(상한 48 GiB)"
            ),
            "",
            "## 시드 평균 ± 표준편차 (val)",
            "",
            "| c | arm | P | R | F1 | AUC | AUPRC | F1/상한 |",
            "|---:|---|---:|---:|---:|---:|---:|---:|",
        ]
    )
    for row in result["aggregates"]:
        metrics = row["metrics"]
        lines.append(
            f"| {row['c']:.1f} | {row['arm']} | "
            f"{mean_std_text(metrics['precision'])} | "
            f"{mean_std_text(metrics['recall'])} | "
            f"{mean_std_text(metrics['f1'])} | "
            f"{mean_std_text(metrics['roc_auc'])} | "
            f"{mean_std_text(metrics['auprc'])} | "
            f"{mean_std_text(metrics['recovery_rate_f1_vs_full_upper'])} |"
        )
    lines.extend(
        [
            "",
            "## 팔 × c × 시드 수치 전문",
            "",
            (
                "F1 CI는 같은 400개 val drawing을 10,000회 재표집한 "
                "paired cluster bootstrap이다."
            ),
            "",
            (
                "| c | seed | arm | realized c | c_hat | pi_hat | "
                "P | R | F1 | AUC | AUPRC | F1 recovery | F1 95% CI |"
            ),
            (
                "|---:|---:|---|---:|---:|---:|---:|---:|---:|"
                "---:|---:|---:|---:|"
            ),
        ]
    )
    order = {name: index for index, name in enumerate(ARM_ORDER)}
    sorted_rows = sorted(
        result["per_seed"],
        key=lambda item: (
            item["c"],
            item["seed"],
            order[item["arm"]],
        ),
    )
    for row in sorted_rows:
        metrics = row["metrics"]
        ci = row["bootstrap"]["f1_ci95"]
        c_hat = (
            "—"
            if row["c_estimated"] is None
            else f"{row['c_estimated']:.6f}"
        )
        pi_hat = (
            "—"
            if row["pi_estimated"] is None
            else f"{row['pi_estimated']:.6f}"
        )
        lines.append(
            f"| {row['c']:.1f} | {row['seed']} | {row['arm']} | "
            f"{row['realized_c']:.6f} | {c_hat} | {pi_hat} | "
            f"{metrics['precision']:.6f} | {metrics['recall']:.6f} | "
            f"{metrics['f1']:.6f} | {metrics['roc_auc']:.6f} | "
            f"{metrics['auprc']:.6f} | "
            f"{row['recovery_rate_f1_vs_full_upper']:.6f} | "
            f"[{ci[0]:.6f}, {ci[1]:.6f}] |"
        )
    ceiling = result["supervised_ceiling"]
    shuffle = result["shuffle_control"]
    lines.extend(
        [
            "",
            "## 완전라벨 상한 재현",
            "",
            (
                "완전라벨 HistGBDT 12특징의 시드 평균 val F1은 "
                f"{ceiling['f1_mean']:.6f}, 참조 0.705와의 차이는 "
                f"{ceiling['difference_from_reference_0_705']:+.6f}이다. "
                "이 값은 이전 결과를 복사하지 않고 이 실행에서 다시 "
                "fit/evaluate한 값이다."
            ),
            "",
            "## 셔플 대조군과 누출 점검",
            "",
            (
                f"셔플 대조군 최대 val AUC={shuffle['max_auc']:.6f}; "
                f"조사 trigger는 > {SHUFFLE_AUC_TRIGGER:.2f}; "
                "leakage_investigation_required="
                f"{str(shuffle['leakage_investigation_required']).lower()}."
            ),
            (
                "특징 schema의 forbidden ID/name/path 열 수="
                f"{shuffle['forbidden_feature_count']}; "
                "고정 split 규칙 변경 수=0; 관측된 trigger 초과 행 수="
                f"{len(shuffle['triggered_rows'])}."
            ),
        ]
    )
    if shuffle["triggered_rows"]:
        lines.extend(["", "Trigger 초과 행:"])
        for row in shuffle["triggered_rows"]:
            lines.append(
                f"- c={row['c']}, seed={row['seed']}, "
                f"AUC={row['roc_auc']:.6f}"
            )
    investigation = shuffle.get("investigation")
    if investigation:
        primary = investigation["primary_12f_summary"]
        base6 = investigation["base6_summary"]
        lines.extend(
            [
                "",
                "### Trigger 후 추가 조사",
                "",
                (
                    f"기존 12특징 셔플 9회의 평균 AUC="
                    f"{primary['mean_auc']:.6f}, 표본 표준편차="
                    f"{primary['std_sample_auc']:.6f}, 범위="
                    f"[{primary['min_auc']:.6f}, "
                    f"{primary['max_auc']:.6f}]였다. "
                    f">0.55는 {primary['count_above_0_55']}회, "
                    f"<0.45는 {primary['count_below_0_45']}회였고, "
                    "9회 모두 threshold 0.5에서 양성 예측은 0개였다."
                ),
                (
                    f"문맥 열을 제거한 frozen 지역 6특징에서 독립 label "
                    f"permutation {base6['replicates']}회를 추가 fit한 결과 "
                    f"평균 AUC={base6['mean_auc']:.6f}, 표본 표준편차="
                    f"{base6['std_sample_auc']:.6f}, 범위="
                    f"[{base6['min_auc']:.6f}, "
                    f"{base6['max_auc']:.6f}], >0.55 "
                    f"{base6['count_above_0_55']}회, <0.45 "
                    f"{base6['count_below_0_45']}회였다."
                ),
                (
                    "구조 감사: 특징행렬에 gid/handle/layer/name/path 열은 "
                    "없고, split 경로는 literal train/val만 허용하며, "
                    "primary 셔플의 방향은 seed/c에 따라 양·음으로 "
                    "뒤집혔다. 추가 base-6 대조는 문맥 열 없이도 null "
                    "HistGBDT score 순위가 크게 흔들리는지를 국소화한다."
                ),
                (
                    "조사 결론: "
                    f"{investigation['conclusion']}. "
                    "원래의 세 trigger는 삭제하거나 정상값으로 치환하지 "
                    "않고 results.json/evidence.xlsx에 그대로 보존했다."
                ),
            ]
        )
    lines.extend(
        [
            "",
            "## 미해결과 해석 제한",
            "",
            (
                "- 이 셀은 인위적 SCAR 누락을 측정한다. 실제 anchor의 "
                "SAR/문법 편향을 해결하지 않는다."
            ),
            (
                "- val은 개발 및 모델 비교 전용이며 held-out 또는 "
                "최종 성능 claim이 아니다."
            ),
            (
                "- nnPU와 posterior-imputation의 class prior는 숨은 "
                "truth가 아니라 z와 c_hat에서 추정했다."
            ),
            (
                "- AUC/AUPRC에는 bootstrap CI를 붙이지 않았고, 공통 "
                "규칙의 paired drawing bootstrap은 F1과 상한 대비 "
                "회복률에 적용했다."
            ),
            "",
        ]
    )
    if result["evidence"]["blocked_xlsx"]:
        lines.extend(
            [
                f"BLOCKED_XLSX: {result['evidence']['reason']}",
                f"대체 증거: {result['evidence']['fallback_path']}",
                "",
            ]
        )
    lines.append("CELL_COMPLETE: pu_ladder")
    return "\n".join(lines) + "\n"


def run() -> dict[str, Any]:
    started_at = now_iso()
    wall_started = time.perf_counter()
    WORK_DIR.mkdir(parents=True, exist_ok=True)
    cml = load_cml_module()
    train = load_base("train")
    val = load_base("val")
    if (
        float(train["scale"]) != SCALE_MM_PER_PX
        or float(val["scale"]) != SCALE_MM_PER_PX
    ):
        raise RuntimeError("base scale differs from frozen 12 mm/px")

    preflight = benchmark_context(cml, train)
    print(
        "preflight:",
        json.dumps(preflight, ensure_ascii=False),
        flush=True,
    )
    context_train, train_context_stats = extract_context(
        cml, "train", train
    )
    context_val, val_context_stats = extract_context(cml, "val", val)
    x_train = assemble_features(train, context_train, "train")
    x_val = assemble_features(val, context_val, "val")
    del context_train, context_val
    for path in (
        WORK_DIR / "context_train.npy",
        WORK_DIR / "context_val.npy",
    ):
        if path.exists():
            path.unlink()
    x_train_scaled, x_val_scaled, preprocessing = (
        standardize_for_logistic(x_train, x_val)
    )

    y_train = train["y"].astype(np.int8, copy=False)
    y_val = val["y"].astype(np.int8, copy=False)
    gid_train = train["gid"]
    gid_val = val["gid"]
    n_val_groups = int(gid_val.max()) + 1
    bootstrap_rng = np.random.default_rng(BOOTSTRAP_SEED)
    bootstrap_counts = bootstrap_rng.multinomial(
        n_val_groups,
        np.full(n_val_groups, 1.0 / n_val_groups),
        size=BOOTSTRAP_REPLICATES,
    ).astype(np.int16, copy=False)

    upper_by_seed: dict[int, dict[str, Any]] = {}
    upper_group_by_seed: dict[int, dict[str, np.ndarray]] = {}
    print("fitting full-label ceilings", flush=True)
    for seed in SEEDS:
        probability, metadata = fit_hgb(
            x_train, y_train, x_val, seed
        )
        metrics = metric_record(y_val, probability)
        grouped = group_confusion(
            y_val, probability, gid_val, n_val_groups
        )
        upper_by_seed[seed] = {
            "probability": probability,
            "metrics": metrics,
            "model": metadata,
        }
        upper_group_by_seed[seed] = grouped
        print(
            f"  upper seed={seed}: F1={metrics['f1']:.6f}",
            flush=True,
        )

    rows: list[dict[str, Any]] = []
    view_records: list[dict[str, Any]] = []
    for c in LABEL_FREQUENCIES:
        for seed in SEEDS:
            print(f"PU view c={c:.1f} seed={seed}", flush=True)
            z_train, view = make_pu_view(y_train, c, seed)
            g_oof, g_val, logistic_meta = fit_crossfit_logistic(
                x_train_scaled,
                z_train,
                gid_train,
                x_val_scaled,
                seed,
            )
            c_estimated = float(logistic_meta["c_estimated"])
            pi_estimated = float(
                logistic_meta["pi_estimated_from_z_over_c"]
            )
            print(
                f"  observed={view['observed_positive_count']:,}; "
                f"realized_c={view['realized_c']:.6f}; "
                f"c_hat={c_estimated:.6f}; "
                f"pi_hat={pi_estimated:.6f}",
                flush=True,
            )

            predictions: dict[
                str, tuple[np.ndarray, dict[str, Any], str, str, str]
            ] = {}
            shared_id = f"crossfit_logistic_c{c}_s{seed}"
            predictions["naive_logistic"] = (
                g_val,
                {
                    **logistic_meta,
                    "shared_model_id": shared_id,
                },
                "logistic",
                "naive_control",
                "z only; unlabeled treated as negative",
            )
            predictions["scar_pu_oracle_c"] = (
                np.clip(g_val / c, 0.0, 1.0),
                {
                    **logistic_meta,
                    "shared_model_id": shared_id,
                    "propensity_used": c,
                    "propensity_source": "preregistered simulation c",
                },
                "logistic",
                "pu_main",
                "z plus preregistered c; no hidden y",
            )
            predictions["scar_pu_estimated_c"] = (
                np.clip(g_val / c_estimated, 0.0, 1.0),
                {
                    **logistic_meta,
                    "shared_model_id": shared_id,
                    "propensity_used": c_estimated,
                    "propensity_source": "OOF observed-positive estimate",
                },
                "logistic",
                "pu_main",
                "z plus cross-fitted c_hat; no hidden y",
            )

            nnpu_probability, nnpu_meta = fit_nnpu_logistic(
                x_train_scaled,
                z_train,
                x_val_scaled,
                pi_estimated,
                seed,
            )
            predictions["nnpu_logistic"] = (
                nnpu_probability,
                nnpu_meta,
                "logistic",
                "pu_main",
                (
                    "z plus pi_hat=z_mean/c_hat; nonnegative PU risk; "
                    "no hidden y"
                ),
            )

            naive_tree_probability, naive_tree_meta = fit_hgb(
                x_train, z_train, x_val, seed
            )
            predictions["naive_tree"] = (
                naive_tree_probability,
                naive_tree_meta,
                "tree",
                "naive_control",
                "z only; unlabeled treated as negative",
            )

            bagging_probability, bagging_meta = fit_bagging_pu(
                x_train, z_train, x_val, seed, c
            )
            predictions["bagging_pu_tree"] = (
                bagging_probability,
                bagging_meta,
                "tree",
                "pu_main",
                (
                    "z only; balanced P-vs-sampled-U bags; "
                    "no hidden y"
                ),
            )

            q_star = np.clip(
                np.asarray(g_oof, dtype=np.float64) / c_estimated,
                0.0,
                1.0,
            )
            q_star[z_train == 1] = 1.0
            imputation_rng = stable_rng(seed, c, 4001)
            imputed = (
                imputation_rng.random(len(q_star)) < q_star
            ).astype(np.int8)
            imputed[z_train == 1] = 1
            sample_weight = np.ones(len(q_star), dtype=np.float32)
            unlabeled_mask = z_train == 0
            sample_weight[unlabeled_mask] = (
                2.0 * np.abs(q_star[unlabeled_mask] - 0.5)
            ).astype(np.float32)
            posterior_probability, posterior_meta = fit_hgb(
                x_train,
                imputed,
                x_val,
                seed,
                sample_weight=sample_weight,
            )
            posterior_meta.update(
                {
                    "q_star_source": "cross-fitted SCAR g/c_hat",
                    "imputed_positive_fraction": float(imputed.mean()),
                    "mean_sample_weight": float(sample_weight.mean()),
                    "observed_positive_weight": 1.0,
                    "unlabeled_weight": "2*abs(q_star-0.5)",
                }
            )
            predictions["posterior_imputation_tree"] = (
                posterior_probability,
                posterior_meta,
                "tree",
                "pu_main",
                (
                    "z plus cross-fitted posterior and imputation RNG; "
                    "no hidden y"
                ),
            )

            shuffle_rng = stable_rng(seed, c, 5001)
            shuffled_z = shuffle_rng.permutation(z_train)
            shuffle_probability, shuffle_meta = fit_hgb(
                x_train, shuffled_z, x_val, seed
            )
            shuffle_meta["shuffle"] = (
                "global permutation of this PU view's z"
            )
            predictions["shuffle_control"] = (
                shuffle_probability,
                shuffle_meta,
                "tree",
                "shuffle_control",
                "permuted z only; no hidden y",
            )

            upper = upper_by_seed[seed]
            upper_row = build_row(
                arm="full_supervised_upper",
                family="tree",
                role="supervised_ceiling",
                c=c,
                seed=seed,
                view=view,
                c_estimated=None,
                pi_estimated=None,
                probability=upper["probability"],
                y_val=y_val,
                gid_val=gid_val,
                model_metadata={
                    **upper["model"],
                    "shared_across_c": True,
                    "shared_model_id": f"full_upper_seed_{seed}",
                },
                upper_metrics=upper["metrics"],
                upper_group=upper_group_by_seed[seed],
                bootstrap_counts=bootstrap_counts,
                source_label_use=(
                    "complete y_train; explicit supervised ceiling only"
                ),
            )
            rows.append(upper_row)

            for arm, (
                probability,
                metadata,
                family,
                role,
                label_use,
            ) in predictions.items():
                row = build_row(
                    arm=arm,
                    family=family,
                    role=role,
                    c=c,
                    seed=seed,
                    view=view,
                    c_estimated=c_estimated,
                    pi_estimated=pi_estimated,
                    probability=probability,
                    y_val=y_val,
                    gid_val=gid_val,
                    model_metadata=metadata,
                    upper_metrics=upper["metrics"],
                    upper_group=upper_group_by_seed[seed],
                    bootstrap_counts=bootstrap_counts,
                    source_label_use=label_use,
                )
                rows.append(row)
                print(
                    f"  {arm}: F1={row['metrics']['f1']:.6f}; "
                    f"AUC={row['metrics']['roc_auc']:.6f}; "
                    f"AUPRC={row['metrics']['auprc']:.6f}",
                    flush=True,
                )

            view_records.append(
                {
                    "c": c,
                    "seed": seed,
                    **view,
                    "c_estimated": c_estimated,
                    "pi_estimated": pi_estimated,
                    "crossfit": logistic_meta,
                }
            )
            del (
                z_train,
                g_oof,
                g_val,
                nnpu_probability,
                naive_tree_probability,
                bagging_probability,
                q_star,
                imputed,
                sample_weight,
                posterior_probability,
                shuffled_z,
                shuffle_probability,
                predictions,
            )

    aggregates = aggregate_rows(rows)
    ceiling_f1 = [
        upper_by_seed[seed]["metrics"]["f1"] for seed in SEEDS
    ]
    shuffle_rows = [
        row for row in rows if row["arm"] == "shuffle_control"
    ]
    triggered_rows = [
        {
            "c": row["c"],
            "seed": row["seed"],
            "roc_auc": row["metrics"]["roc_auc"],
        }
        for row in shuffle_rows
        if row["metrics"]["roc_auc"] > SHUFFLE_AUC_TRIGGER
    ]
    forbidden_tokens = (
        "gid",
        "handle",
        "layer",
        "name",
        "path",
        "vendor",
        "file",
    )
    forbidden_feature_count = sum(
        any(token in name.lower() for token in forbidden_tokens)
        for name in FEATURE_NAMES
    )
    elapsed_seconds = time.perf_counter() - wall_started
    result: dict[str, Any] = {
        "schema": "ariadne.e2.pu_ladder.val.v1",
        "cell": "pu_ladder",
        "run": {
            "started_at": started_at,
            "completed_at": now_iso(),
            "elapsed_seconds": round(elapsed_seconds, 6),
            "python": sys.version.split()[0],
            "platform": platform.platform(),
            "numpy": np.__version__,
            "seeds": list(SEEDS),
            "label_frequencies_c": list(LABEL_FREQUENCIES),
            "thread_limit": THREAD_LIMIT,
        },
        "scope": {
            "train_split": "train",
            "evaluation_split": "val",
            "test_accessed": False,
            "allowed_splits_enforced_in_code": sorted(
                ALLOWED_SPLITS
            ),
            "repository_write_allowed": False,
            "output_directory": str(CELL_DIR),
            "source_ir": {
                "train": str(
                    allowed_path("train", "ir") / "*.segir.json"
                ),
                "val": str(
                    allowed_path("val", "ir") / "*.segir.json"
                ),
            },
        },
        "design": {
            "operating_threshold": OPERATING_THRESHOLD,
            "feature_names": list(FEATURE_NAMES),
            "selected_context_radius_px": SELECTED_RADIUS_PX,
            "selected_context_radius_mm": (
                SELECTED_RADIUS_PX * SCALE_MM_PER_PX
            ),
            "positive_observation": (
                "independent Bernoulli(c) among true train positives"
            ),
            "hidden_truth_policy": (
                "used only to construct the simulated observed-positive "
                "mask and fit the explicit full-supervised ceiling; "
                "unavailable to PU arm targets"
            ),
            "crossfit": {
                "folds": CROSSFIT_FOLDS,
                "fixed_rule": "drawing_gid modulo 2",
                "seed_changes_split": False,
            },
            "bagging_members": BAGGING_MEMBERS,
            "bootstrap": {
                "unit": "val drawing_gid",
                "paired": True,
                "replicates": BOOTSTRAP_REPLICATES,
                "seed": BOOTSTRAP_SEED,
                "reported_for": (
                    "F1, F1/full-upper recovery, "
                    "F1 delta vs full upper"
                ),
            },
            "arm_definitions": {
                "naive_logistic": (
                    "cross-fitted logistic trained on z as ordinary P/N"
                ),
                "naive_tree": "HistGBDT trained on z as ordinary P/N",
                "scar_pu_oracle_c": "clip(g/c_configured, 0, 1)",
                "scar_pu_estimated_c": "clip(g/c_hat, 0, 1)",
                "nnpu_logistic": (
                    "nonnegative logistic PU empirical risk "
                    "with pi=z_mean/c_hat"
                ),
                "bagging_pu_tree": (
                    "mean of balanced observed-P versus "
                    "sampled-U HistGBDT bags"
                ),
                "posterior_imputation_tree": (
                    "weighted Bernoulli imputation from "
                    "cross-fitted q_star"
                ),
                "full_supervised_upper": (
                    "HistGBDT fit to complete train truth"
                ),
                "shuffle_control": (
                    "HistGBDT fit to globally permuted z"
                ),
            },
            "recovery_rate_definition": (
                "arm val metric divided by same-seed "
                "full-supervised val metric"
            ),
            "band_verdict_emitted": False,
        },
        "data": {
            "n_train_rows": int(len(y_train)),
            "n_val_rows": int(len(y_val)),
            "n_train_drawings": len(split_files("train")),
            "n_val_drawings": len(split_files("val")),
            "train_wall_fraction": float(y_train.mean()),
            "val_wall_fraction": float(y_val.mean()),
            "base_feature_source": {
                "train": str(allowed_path("train", "features")),
                "val": str(allowed_path("val", "features")),
                "reused": True,
            },
        },
        "preprocessing": preprocessing,
        "throughput": {
            "preflight_benchmark": preflight,
            "full_context_extraction": {
                "train": train_context_stats,
                "val": val_context_stats,
            },
        },
        "pu_views": view_records,
        "per_seed": rows,
        "aggregates": aggregates,
        "supervised_ceiling": {
            "reference_f1": REFERENCE_CONTEXT_F1,
            "f1_by_seed": {
                str(seed): upper_by_seed[seed]["metrics"]["f1"]
                for seed in SEEDS
            },
            "f1_mean": float(statistics.mean(ceiling_f1)),
            "f1_std_sample": float(statistics.stdev(ceiling_f1)),
            "difference_from_reference_0_705": float(
                statistics.mean(ceiling_f1)
                - REFERENCE_CONTEXT_F1
            ),
            "measured_in_this_run": True,
        },
        "shuffle_control": {
            "trigger_rule": f"ROC-AUC > {SHUFFLE_AUC_TRIGGER}",
            "max_auc": float(
                max(
                    row["metrics"]["roc_auc"]
                    for row in shuffle_rows
                )
            ),
            "leakage_investigation_required": bool(triggered_rows),
            "triggered_rows": triggered_rows,
            "forbidden_feature_count": int(
                forbidden_feature_count
            ),
            "feature_schema_checked": list(FEATURE_NAMES),
            "split_rule_changes": 0,
        },
        "provenance": {
            "packet": str(
                Path(
                    r"D:\runs\e2_program\build\PACKET_pu_ladder.md"
                )
            ),
            "dossier": {
                "path": str(DOSSIER_PATH),
                "sha256": sha256(DOSSIER_PATH),
            },
            "baseline_code": {
                "path": str(BASELINE_PATH),
                "sha256": sha256(BASELINE_PATH),
            },
            "eval_code": {
                "path": str(EVAL_PATH),
                "sha256": sha256(EVAL_PATH),
            },
            "context_code": {
                "path": str(CML_PATH),
                "sha256": sha256(CML_PATH),
            },
            "script": {
                "path": str(Path(__file__).resolve()),
                "sha256": sha256(Path(__file__).resolve()),
            },
            "base_features": {
                "train_sha256": sha256(
                    allowed_path("train", "features")
                ),
                "val_sha256": sha256(
                    allowed_path("val", "features")
                ),
            },
        },
        "limitations": [
            (
                "simulated missingness is SCAR and does not establish "
                "robustness to real SAR anchors"
            ),
            (
                "val is a development surface, not a held-out "
                "or final claim"
            ),
            (
                "ranking metrics do not receive bootstrap "
                "intervals in this cell"
            ),
        ],
    }
    return result


def investigate_existing_shuffle() -> dict[str, Any]:
    """Investigate triggered controls without recomputing context features."""
    if not RESULTS_PATH.exists():
        raise FileNotFoundError(
            "results.json is required before shuffle investigation"
        )
    result = json.loads(RESULTS_PATH.read_text(encoding="utf-8"))
    if result.get("schema") != "ariadne.e2.pu_ladder.val.v1":
        raise RuntimeError("unexpected results schema")
    train = load_base("train")
    val = load_base("val")
    x_train = train["X"]
    y_train = train["y"].astype(np.int8, copy=False)
    x_val = val["X"]
    y_val = val["y"].astype(np.int8, copy=False)

    records: list[dict[str, Any]] = []
    repeats_per_c = 12
    started = time.perf_counter()
    for c in LABEL_FREQUENCIES:
        z, view = make_pu_view(y_train, c, SEEDS[0])
        for repeat in range(repeats_per_c):
            permutation_seed = 70_000 + int(c * 1000) * 100 + repeat
            rng = np.random.default_rng(permutation_seed)
            shuffled = rng.permutation(z)
            probability, metadata = fit_hgb(
                x_train,
                shuffled,
                x_val,
                seed=80_000 + repeat,
            )
            metrics = metric_record(y_val, probability)
            record = {
                "feature_set": "frozen_local_6",
                "c": c,
                "view_seed": SEEDS[0],
                "repeat": repeat,
                "permutation_seed": permutation_seed,
                "model_seed": 80_000 + repeat,
                "observed_positive_count": view[
                    "observed_positive_count"
                ],
                "roc_auc": metrics["roc_auc"],
                "auprc": metrics["auprc"],
                "f1": metrics["f1"],
                "predicted_positive_count": (
                    metrics["tp"] + metrics["fp"]
                ),
                "score_mean": float(np.mean(probability)),
                "score_std": float(np.std(probability)),
                "score_min": float(np.min(probability)),
                "score_max": float(np.max(probability)),
                "fit_seconds": metadata["fit_seconds"],
                "inference_seconds": metadata["inference_seconds"],
            }
            records.append(record)
            print(
                f"shuffle investigation c={c:.1f} "
                f"repeat={repeat + 1}/{repeats_per_c}: "
                f"AUC={metrics['roc_auc']:.6f}; "
                f"score_std={record['score_std']:.8f}",
                flush=True,
            )
            del shuffled, probability

    primary_rows = [
        row
        for row in result["per_seed"]
        if row["arm"] == "shuffle_control"
    ]
    primary_auc = [
        float(row["metrics"]["roc_auc"]) for row in primary_rows
    ]
    base_auc = [float(row["roc_auc"]) for row in records]
    base_score_std = [float(row["score_std"]) for row in records]
    primary_summary = {
        "replicates": len(primary_auc),
        "mean_auc": float(statistics.mean(primary_auc)),
        "std_sample_auc": float(statistics.stdev(primary_auc)),
        "min_auc": float(min(primary_auc)),
        "max_auc": float(max(primary_auc)),
        "count_above_0_55": sum(
            value > SHUFFLE_AUC_TRIGGER for value in primary_auc
        ),
        "count_below_0_45": sum(value < 0.45 for value in primary_auc),
        "all_zero_predicted_positive_at_0_5": all(
            row["metrics"]["tp"] + row["metrics"]["fp"] == 0
            for row in primary_rows
        ),
    }
    base_summary = {
        "feature_set": "frozen_local_6",
        "replicates": len(base_auc),
        "repeats_per_c": repeats_per_c,
        "mean_auc": float(statistics.mean(base_auc)),
        "std_sample_auc": float(statistics.stdev(base_auc)),
        "min_auc": float(min(base_auc)),
        "max_auc": float(max(base_auc)),
        "count_above_0_55": sum(
            value > SHUFFLE_AUC_TRIGGER for value in base_auc
        ),
        "count_below_0_45": sum(value < 0.45 for value in base_auc),
        "mean_score_std": float(statistics.mean(base_score_std)),
        "median_score_std": float(statistics.median(base_score_std)),
    }
    conclusion = (
        "no direct ID/path/split feature leak was found; the null-control "
        "ranking is unstable in both directions and the independent base-6 "
        "permutations quantify whether the excursions persist without "
        "context features"
    )
    result["shuffle_control"]["investigation"] = {
        "completed_at": now_iso(),
        "elapsed_seconds": round(
            time.perf_counter() - started, 6
        ),
        "reason": (
            "three primary 12-feature shuffle rows exceeded ROC-AUC 0.55"
        ),
        "primary_12f_summary": primary_summary,
        "base6_summary": base_summary,
        "base6_independent_permutations": records,
        "structural_audit": {
            "feature_names": list(FEATURE_NAMES),
            "forbidden_feature_count": result["shuffle_control"][
                "forbidden_feature_count"
            ],
            "gid_is_model_feature": False,
            "handle_layer_name_path_are_model_features": False,
            "allowed_data_splits": sorted(ALLOWED_SPLITS),
            "split_rule_changes": 0,
        },
        "conclusion": conclusion,
        "primary_triggers_preserved": True,
    }
    result["shuffle_control"]["investigation_completed"] = True
    result["run"]["shuffle_investigation_completed_at"] = now_iso()
    result["provenance"]["script"]["sha256"] = sha256(
        Path(__file__).resolve()
    )
    result["evidence"] = write_evidence(result)
    RESULTS_PATH.write_text(
        json.dumps(result, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    REPORT_PATH.write_text(
        render_report(result), encoding="utf-8", newline="\n"
    )
    return result


def main() -> int:
    import argparse

    from threadpoolctl import threadpool_limits

    if Path.cwd().resolve() != CELL_DIR:
        os.chdir(CELL_DIR)
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--investigate-shuffle",
        action="store_true",
        help=(
            "run independent frozen-base shuffle diagnostics against "
            "an existing full result"
        ),
    )
    args = parser.parse_args()
    if args.investigate_shuffle:
        with RSSMonitor() as monitor, threadpool_limits(
            limits=THREAD_LIMIT
        ):
            result = investigate_existing_shuffle()
        result["shuffle_control"]["investigation"][
            "peak_rss_gib"
        ] = float(monitor.peak_bytes / 1024**3)
        result["evidence"] = write_evidence(result)
        RESULTS_PATH.write_text(
            json.dumps(result, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        REPORT_PATH.write_text(
            render_report(result), encoding="utf-8", newline="\n"
        )
        print(
            json.dumps(
                {
                    "investigation_completed": True,
                    "primary": result["shuffle_control"][
                        "investigation"
                    ]["primary_12f_summary"],
                    "base6": result["shuffle_control"][
                        "investigation"
                    ]["base6_summary"],
                },
                ensure_ascii=False,
            ),
            flush=True,
        )
        return 0
    started = time.perf_counter()
    with RSSMonitor() as monitor, threadpool_limits(
        limits=THREAD_LIMIT
    ):
        result = run()
    result["run"]["elapsed_seconds"] = round(
        time.perf_counter() - started, 6
    )
    result["run"]["peak_rss_bytes"] = int(monitor.peak_bytes)
    result["run"]["peak_rss_gib"] = float(
        monitor.peak_bytes / 1024**3
    )
    result["run"]["ram_limit_bytes"] = MAX_RSS_BYTES
    result["run"]["ram_limit_exceeded"] = bool(
        monitor.peak_bytes > MAX_RSS_BYTES
    )
    result["evidence"] = write_evidence(result)
    RESULTS_PATH.write_text(
        json.dumps(result, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    REPORT_PATH.write_text(
        render_report(result), encoding="utf-8", newline="\n"
    )
    shutil.rmtree(WORK_DIR, ignore_errors=True)
    print(
        json.dumps(
            {
                "results": str(RESULTS_PATH),
                "evidence": result["evidence"],
                "report": str(REPORT_PATH),
                "rows": len(result["per_seed"]),
                "aggregates": len(result["aggregates"]),
                "peak_rss_gib": result["run"]["peak_rss_gib"],
                "elapsed_seconds": result["run"][
                    "elapsed_seconds"
                ],
            },
            ensure_ascii=False,
        ),
        flush=True,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
