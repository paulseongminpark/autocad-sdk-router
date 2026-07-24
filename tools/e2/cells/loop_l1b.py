#!/usr/bin/env python3
"""E2 loop L1b: mixed anchor-rich and single-span population reevaluation.

The repository C0/C1 implementations and the completed L1 runner are imported
read-only.  All generated artifacts are confined to this file's directory.  No
CAD source drawing, held-out/test split, repository file, or predecessor output
is modified, and no target/theory verdict is emitted.
"""

from __future__ import annotations

import argparse
import contextlib
import copy
import hashlib
import importlib.util
import io
import json
import math
import re
import statistics
import sys
from collections import Counter
from pathlib import Path
from typing import Any, Iterator, Mapping, Sequence


sys.dont_write_bytecode = True

CELL_DIR = Path(__file__).resolve().parent
SCRIPT_PATH = CELL_DIR / "loop_l1b.py"
SCENES_V3_DIR = CELL_DIR / "scenes_v3"
C0V3_NUMBERS_PATH = CELL_DIR / "c0v3_numbers.json"
C1V3_RESULTS_PATH = CELL_DIR / "c1v3_results.json"
EVIDENCE_PATH = CELL_DIR / "evidence.xlsx"
EVIDENCE_FALLBACK_PATH = CELL_DIR / "evidence_rows.json"
REPORT_PATH = CELL_DIR / "REPORT.md"

PACKET_PATH = Path(r"D:\runs\e2_program\build\PACKET_loop_L1b_mixed_population.md")
L1_DIR = Path(r"D:\runs\e2_program\cells\loop_l1")
L1_PATH = L1_DIR / "loop_l1.py"
L1_C0_NUMBERS_PATH = L1_DIR / "c0v2_numbers.json"
L1_C1_RESULTS_PATH = L1_DIR / "c1v2_results.json"
L1_REPORT_PATH = L1_DIR / "REPORT.md"


def load_module(name: str, path: Path) -> Any:
    if not path.is_file():
        raise FileNotFoundError(path)
    spec = importlib.util.spec_from_file_location(name, path)
    if spec is None or spec.loader is None:
        raise ImportError(f"cannot load module from {path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[name] = module
    spec.loader.exec_module(module)
    return module


L1 = load_module("loop_l1b_l1_reference", L1_PATH)
C0 = L1.C0
C1 = L1.C1

REPO_ROOT = L1.REPO_ROOT
DOSSIER_PATH = L1.DOSSIER_PATH
C0_PATH = L1.C0_PATH
C1_PATH = L1.C1_PATH
V1_C0_NUMBERS_PATH = L1.V1_C0_NUMBERS_PATH
V1_C1_REPORT_PATH = L1.V1_C1_REPORT_PATH

LOOP_SCHEMA = "ariadne.e2.loop_l1b.mixed_population.v1"
C0V3_SCHEMA = "ariadne.e2.loop_l1b.c0v3_numbers.v1"
C1V3_SCHEMA = "ariadne.e2.loop_l1b.c1v3_results.v1"
ANCHOR_RICH_BASE_COUNT = 40
SINGLE_SPAN_BASE_COUNT = 10

# feyerabend_P2.md lines 550-561, expanded so the combined final bullet maps
# one-to-one to C0's two manifest families.
DOSSIER_FAMILY_ROWS: tuple[tuple[str, str], ...] = (
    ("순수 LINE 평행쌍", "pure_line_parallel_pair"),
    ("LWPOLYLINE 분절", "lwpolyline_segmentation"),
    ("ARC/SPLINE 인접 또는 교란", "arc_spline_adjacent_or_distractor"),
    ("HATCH boundary 교란", "hatch_boundary_distractor"),
    (
        "nested INSERT와 non-uniform/누적 transform",
        "nested_insert_nonuniform_accumulated_transform",
    ),
    ("부분 overlap과 거의 평행한 조각", "partial_overlap_near_parallel_fragment"),
    (
        "door/window/dimension-like 긴 평행 distractor",
        "door_window_dimension_long_parallel_distractor",
    ),
    ("zero-wall sentinel", "zero_wall_sentinel"),
    ("all-wall sentinel", "all_wall_sentinel"),
    ("단일 reference span 영역", "single_reference_span_region"),
    ("다중 reference span 영역", "multiple_reference_span_regions"),
)

if tuple(row[1] for row in DOSSIER_FAMILY_ROWS) != tuple(C0.MUTATION_FAMILIES):
    raise RuntimeError("dossier-to-C0 mutation family mapping is not one-to-one")


def canonical_bytes(value: Any) -> bytes:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8")


def canonical_sha256(value: Any) -> str:
    return hashlib.sha256(canonical_bytes(value)).hexdigest()


def assignment_key(index: int) -> tuple[str, int]:
    """Stable seed-derived ordering; the first ten ranks are single-span."""
    seed = C0.seed_for_index(index)
    digest = hashlib.sha256(f"{seed}:loop_l1b_population".encode("utf-8")).hexdigest()
    return digest, index


POPULATION_ORDER = tuple(sorted(range(C0.BASE_SCENE_COUNT), key=assignment_key))
SINGLE_SPAN_INDICES = frozenset(POPULATION_ORDER[:SINGLE_SPAN_BASE_COUNT])
ANCHOR_RICH_INDICES = frozenset(POPULATION_ORDER[SINGLE_SPAN_BASE_COUNT:])
POPULATION_RANK = {index: rank for rank, index in enumerate(POPULATION_ORDER)}

if len(SINGLE_SPAN_INDICES) != SINGLE_SPAN_BASE_COUNT:
    raise AssertionError("single-span assignment count mismatch")
if len(ANCHOR_RICH_INDICES) != ANCHOR_RICH_BASE_COUNT:
    raise AssertionError("anchor-rich assignment count mismatch")
if SINGLE_SPAN_INDICES & ANCHOR_RICH_INDICES:
    raise AssertionError("population assignment overlap")


def population_for_index(index: int) -> str:
    return "single_span" if index in SINGLE_SPAN_INDICES else "anchor_rich"


class MixedPopulationSceneBuilder(C0.SceneBuilder):
    """C0 builder extension with an exact seed-ranked 40/10 anchor mix."""

    def add_reference_anchors(self) -> None:
        if self.index not in SINGLE_SPAN_INDICES:
            # Reuse L1's anchor-rich method verbatim for the 40 rich scenes.
            L1.AnchorRichSceneBuilder.add_reference_anchors(self)
            for anchor in self.anchors:
                anchor["anchor_factory_revision"] = "v3_mixed_population_anchor_rich"
            self.mutations.add("mixed_population_anchor_rich")
            return

        # C0-v1-style single region with <=3 anchors.  Two ratio anchors share
        # one reference span, so adding C1's diagnostic outlier cannot raise
        # reference evidence to the minimum three-independent-span threshold.
        for anchor_index in range(2):
            y = 35_000.0 + anchor_index * 1_200.0
            self.anchors.append(
                {
                    "handle": self.handle(f"single_span_dim_{anchor_index}"),
                    "anchor_type": "DIM",
                    "region": "SINGLE",
                    "p0": [5_000.0, y],
                    "p1": [6_000.0, y],
                    "raw_span": 1_000.0,
                    "display_value": 1_000.0,
                    "display_unit": "MM",
                    "text_height": 100.0,
                    "weight": 1.0,
                    "anchor_factory_revision": "v3_mixed_population_single_span",
                }
            )
        self.mutations.add("single_reference_span_region")
        self.mutations.add("mixed_population_single_span")


@contextlib.contextmanager
def patched_scene_builder() -> Iterator[None]:
    original = C0.SceneBuilder
    C0.SceneBuilder = MixedPopulationSceneBuilder
    try:
        yield
    finally:
        C0.SceneBuilder = original


def enrich_base_scene(scene: dict[str, Any]) -> None:
    population = population_for_index(int(scene["base_scene_index"]))
    scene["loop_schema"] = LOOP_SCHEMA
    scene["scene_factory_revision"] = "v3_mixed_population_l1b"
    scene["population_role"] = population
    scene["population_assignment_rank"] = POPULATION_RANK[int(scene["base_scene_index"])]
    scene["population_assignment_rule"] = (
        "rank base indices by sha256(str(seed)+':loop_l1b_population'); "
        "ranks 0..9 single_span, ranks 10..49 anchor_rich"
    )
    scene["anchor_count_rule"] = (
        "2 repeated-span DIM anchors in one region"
        if population == "single_span"
        else "5 + (uint32_seed mod 4) independent-span DIM anchors"
    )
    scene["anchor_count"] = len(scene["anchors"])
    scene["provenance"] = {
        **scene["provenance"],
        "loop_packet": str(PACKET_PATH),
        "l1_import": str(L1_PATH),
        "c0_import": str(C0_PATH),
        "c1_import": str(C1_PATH),
        "anchor_builder": "MixedPopulationSceneBuilder.add_reference_anchors",
    }
    C0.refresh_scene_digests(scene)


def build_mixed_structural_scene(
    index: int, calibration_rows: list[tuple[int, int, float, str]]
) -> Any:
    with patched_scene_builder():
        builder = C0.build_structural_scene(index, calibration_rows)
    return builder


def build_mixed_base_corpus() -> tuple[list[dict[str, Any]], dict[str, Any]]:
    with patched_scene_builder():
        scenes, build_numbers = C0.build_base_corpus()
    for scene in scenes:
        enrich_base_scene(scene)
    build_numbers = {
        **build_numbers,
        "population_assignment": {
            "anchor_rich_base_scene_count": ANCHOR_RICH_BASE_COUNT,
            "single_span_base_scene_count": SINGLE_SPAN_BASE_COUNT,
            "single_span_indices": sorted(SINGLE_SPAN_INDICES),
            "assignment_digest": canonical_sha256(
                [
                    {
                        "base_scene_index": index,
                        "seed": C0.seed_for_index(index),
                        "population": population_for_index(index),
                        "rank": POPULATION_RANK[index],
                    }
                    for index in range(C0.BASE_SCENE_COUNT)
                ]
            ),
        },
    }
    return scenes, build_numbers


def numeric_distribution(values: Sequence[float]) -> dict[str, Any]:
    return C1.numeric_distribution([float(value) for value in values])


def population_numbers(
    base_scenes: Sequence[Mapping[str, Any]], all_scenes: Sequence[Mapping[str, Any]]
) -> dict[str, Any]:
    base_counts = [len(scene["anchors"]) for scene in base_scenes]
    all_counts = [len(scene["anchors"]) for scene in all_scenes]
    spatial_bins: list[int] = []
    canonical_counts: list[int] = []
    distinct_spans: list[int] = []
    confidence_scores: list[float] = []
    unit_statuses: Counter[str] = Counter()
    reference_statuses: Counter[str] = Counter()
    status_by_population: dict[str, Counter[str]] = {
        "anchor_rich": Counter(),
        "single_span": Counter(),
    }
    unit_status_by_population: dict[str, Counter[str]] = {
        "anchor_rich": Counter(),
        "single_span": Counter(),
    }
    reference_status_by_population: dict[str, Counter[str]] = {
        "anchor_rich": Counter(),
        "single_span": Counter(),
    }
    population_counts: Counter[str] = Counter()
    rule_mismatches = 0
    canonical_mismatches = 0
    spatial_mismatches = 0
    distinctness_mismatches = 0
    mutation_role_mismatches = 0
    non_ratio_anchor_count = 0
    ratio_mismatch_count = 0
    rich_log_separations: list[float] = []
    assignment_rows: list[dict[str, Any]] = []

    for scene in sorted(base_scenes, key=lambda value: int(value["base_scene_index"])):
        index = int(scene["base_scene_index"])
        population = population_for_index(index)
        population_counts[population] += 1
        anchors = list(scene["anchors"])
        expected_count = 2 if population == "single_span" else L1.anchor_count_for_seed(int(scene["seed"]))
        rule_mismatches += len(anchors) != expected_count
        prepared, audit = C1.prepare_anchors(anchors)
        canonical_counts.append(len(prepared))
        canonical_mismatches += len(prepared) != len(anchors)
        spans = sorted(float(anchor["raw_span"]) for anchor in anchors)
        distinct = len({C1.round_key(span) for span in spans})
        distinct_spans.append(distinct)
        expected_distinct = 1 if population == "single_span" else len(anchors)
        distinctness_mismatches += distinct != expected_distinct
        if population == "anchor_rich" and len(spans) > 1:
            rich_log_separations.append(
                min(math.log(spans[offset + 1] / spans[offset]) for offset in range(len(spans) - 1))
            )
        bins = C1.spatial_bin_count(prepared)
        spatial_bins.append(bins)
        if population == "anchor_rich":
            spatial_mismatches += bins < 3
        regions = sorted({str(anchor.get("region")) for anchor in anchors})
        families = set(scene["mutation_families"])
        if population == "single_span":
            mutation_role_mismatches += not (
                len(regions) == 1
                and "single_reference_span_region" in families
                and "multiple_reference_span_regions" not in families
            )
        else:
            mutation_role_mismatches += not (
                len(regions) >= 3
                and "multiple_reference_span_regions" in families
                and "single_reference_span_region" not in families
            )
        non_ratio_anchor_count += sum(
            str(anchor.get("anchor_type", "")).upper() not in ("DIM", "TEXT")
            or anchor.get("display_value") is None
            for anchor in anchors
        )
        model = C1.fit_anchor_model(anchors)
        confidence_scores.append(float(model["confidence_score"]))
        unit_status = str(model["unit_status"])
        reference_status = str(model["reference_status"])
        status = str(model["status"])
        unit_statuses[unit_status] += 1
        reference_statuses[reference_status] += 1
        unit_status_by_population[population][unit_status] += 1
        reference_status_by_population[population][reference_status] += 1
        status_by_population[population][status] += 1
        assignment_rows.append(
            {
                "base_scene_index": index,
                "seed": int(scene["seed"]),
                "assignment_rank": POPULATION_RANK[index],
                "assignment_sha256": assignment_key(index)[0],
                "population": population,
                "anchor_count": len(anchors),
                "canonical_anchor_count": len(prepared),
                "distinct_span_count": distinct,
                "spatial_bin_count": bins,
                "region_count": len(regions),
                "unit_status": unit_status,
                "reference_status": reference_status,
                "status": status,
                "confidence_score": float(model["confidence_score"]),
                "duplicate_count": int(audit["duplicate_count"]),
            }
        )

    for scene in all_scenes:
        truth_scale = float(scene["truth_unit_scale"])
        for anchor in scene["anchors"]:
            if anchor.get("display_value") is None:
                continue
            raw = float(anchor["raw_span"])
            display = float(anchor["display_value"])
            ratio_mismatch_count += not math.isclose(
                display / raw, truth_scale, rel_tol=1e-12, abs_tol=1e-12
            )

    population_ir_counts = Counter(str(scene["population_role"]) for scene in all_scenes)
    assignment_digest = canonical_sha256(
        [
            {
                "base_scene_index": row["base_scene_index"],
                "seed": row["seed"],
                "assignment_rank": row["assignment_rank"],
                "population": row["population"],
            }
            for row in assignment_rows
        ]
    )
    return {
        "population_assignment_rule": (
            "rank base indices by sha256(str(seed)+':loop_l1b_population'); "
            "first 10 single_span, remaining 40 anchor_rich"
        ),
        "population_assignment_digest": assignment_digest,
        "single_span_indices": sorted(SINGLE_SPAN_INDICES),
        "anchor_rich_indices": sorted(ANCHOR_RICH_INDICES),
        "base_population_counts": dict(sorted(population_counts.items())),
        "ir_population_counts": dict(sorted(population_ir_counts.items())),
        "assignment_rows": assignment_rows,
        "anchor_count_rule": (
            "single_span=2 repeated-span DIM anchors in one region; "
            "anchor_rich=5+(uint32_seed mod 4) independent-span DIM anchors"
        ),
        "base_anchor_count_distribution": dict(sorted(Counter(base_counts).items())),
        "ir_anchor_count_distribution": dict(sorted(Counter(all_counts).items())),
        "base_anchor_count_min": min(base_counts),
        "base_anchor_count_max": max(base_counts),
        "base_anchor_count_mean": statistics.fmean(base_counts),
        "base_anchor_count_median": statistics.median(base_counts),
        "base_spatial_bin_count_distribution": dict(sorted(Counter(spatial_bins).items())),
        "base_spatial_bin_count_min": min(spatial_bins),
        "base_anchor_rich_spatial_bin_below_3_scene_count": spatial_mismatches,
        "base_spatial_bin_below_3_scene_count": sum(value < 3 for value in spatial_bins),
        "base_distinct_span_count_distribution": dict(sorted(Counter(distinct_spans).items())),
        "base_distinct_span_rule_mismatch_scene_count": distinctness_mismatches,
        "base_distinct_span_mismatch_scene_count": distinctness_mismatches,
        "base_canonical_anchor_count_distribution": dict(sorted(Counter(canonical_counts).items())),
        "base_canonical_anchor_count_mismatch_scene_count": canonical_mismatches,
        "base_anchor_count_rule_mismatch_scene_count": rule_mismatches,
        "base_mutation_population_role_mismatch_scene_count": mutation_role_mismatches,
        "base_non_dim_text_ratio_anchor_count": non_ratio_anchor_count,
        "ir_display_raw_truth_ratio_mismatch_count": ratio_mismatch_count,
        "base_anchor_rich_minimum_adjacent_log_span_separation": min(rich_log_separations),
        "base_minimum_adjacent_log_span_separation": min(rich_log_separations),
        "c1_ransac_log_tolerance": C1.RANSAC_LOG_TOLERANCE,
        "base_unit_confidence_score": numeric_distribution(confidence_scores),
        "base_unit_status_counts": dict(sorted(unit_statuses.items())),
        "base_reference_status_counts": dict(sorted(reference_statuses.items())),
        "base_unit_status_counts_by_population": {
            key: dict(sorted(value.items())) for key, value in unit_status_by_population.items()
        },
        "base_reference_status_counts_by_population": {
            key: dict(sorted(value.items())) for key, value in reference_status_by_population.items()
        },
        "base_status_counts_by_population": {
            key: dict(sorted(value.items())) for key, value in status_by_population.items()
        },
        "single_span_unit_HIGH_scene_count": int(
            unit_status_by_population["single_span"].get("HIGH", 0)
        ),
        "single_span_unit_LOW_or_NONE_scene_count": int(
            unit_status_by_population["single_span"].get("LOW", 0)
            + unit_status_by_population["single_span"].get("NONE", 0)
        ),
        "anchor_rich_unit_HIGH_scene_count": int(
            unit_status_by_population["anchor_rich"].get("HIGH", 0)
        ),
    }


def nested_get(payload: Mapping[str, Any], dotted: str) -> Any:
    current: Any = payload
    for token in dotted.split("."):
        current = current[token]
    return current


def numeric_delta(before: Any, after: Any) -> Any:
    if (
        isinstance(before, (int, float))
        and not isinstance(before, bool)
        and isinstance(after, (int, float))
        and not isinstance(after, bool)
    ):
        return float(after) - float(before)
    return None


def comparison_row(metric: str, v1: Any, l1: Any, v3: Any) -> dict[str, Any]:
    return {
        "metric": metric,
        "v1": v1,
        "l1": l1,
        "v3": v3,
        "delta_v3_minus_v1": numeric_delta(v1, v3),
        "delta_v3_minus_l1": numeric_delta(l1, v3),
        # Compatibility columns consumed by L1's workbook augmentation.
        "v2": v3,
        "delta": numeric_delta(v1, v3),
    }


def c0_three_way_comparison_rows(
    v1: Mapping[str, Any], v3: Mapping[str, Any], population: Mapping[str, Any]
) -> list[dict[str, Any]]:
    del population
    l1 = json.loads(L1_C0_NUMBERS_PATH.read_text(encoding="utf-8"))
    metrics = [
        "scene_counts.base_scene_count",
        "scene_counts.ir_scene_count",
        "truth_pair_numbers.ir_truth_pair_count",
        "truth_validator.error_count",
        "fidelity_numbers_kappa_1.entity_mix_tv",
        "fidelity_numbers_kappa_1.thickness_histogram_ks",
        "determinism_and_scale_numbers.four_scale_topology_mismatch_base_scene_count",
        "determinism_and_scale_numbers.four_scale_normalized_geometry_mismatch_base_scene_count",
        "determinism_and_scale_numbers.four_scale_source_handle_mismatch_base_scene_count",
    ]
    metrics.extend(
        f"mutation_family_coverage.{family}.scene_count" for _, family in DOSSIER_FAMILY_ROWS
    )
    return [
        comparison_row(metric, nested_get(v1, metric), nested_get(l1, metric), nested_get(v3, metric))
        for metric in metrics
    ]


def c1_three_way_comparison_rows(
    baseline: Mapping[str, Any], aggregates: Mapping[str, Any], ticket: Mapping[str, Any]
) -> list[dict[str, Any]]:
    l1 = json.loads(L1_C1_RESULTS_PATH.read_text(encoding="utf-8"))
    l1_overall = l1["aggregates"]["overall"]
    v3_overall = aggregates["overall"]
    rows = [
        comparison_row("estimate_count", baseline["estimate_count"], l1_overall["estimate_count"], v3_overall["estimate_count"]),
        comparison_row("estimate_coverage", baseline["estimate_coverage"], l1_overall["estimate_coverage"], v3_overall["estimate_coverage"]),
        comparison_row("accuracy_within_5pct", baseline["accuracy_within_5pct"], l1_overall["accuracy_within_5pct"], v3_overall["accuracy_within_5pct"]),
        comparison_row("HIGH_scene_count", baseline["high_scene_count"], l1_overall["high_scene_count"], v3_overall["high_scene_count"]),
        comparison_row("HIGH_coverage", baseline["high_coverage"], l1_overall["high_coverage"], v3_overall["high_coverage"]),
        comparison_row("HIGH_accuracy_within_5pct", baseline["high_accuracy_within_5pct"], l1_overall["high_accuracy_within_5pct"], v3_overall["high_accuracy_within_5pct"]),
        comparison_row(
            "pair_label_mismatch_scene_count",
            baseline["pair_label_mismatching_anchor_artifact_scene_count"],
            l1["aggregates"]["pair_label_permutation"]["mismatching_anchor_artifact_scene_count"],
            aggregates["pair_label_permutation"]["mismatching_anchor_artifact_scene_count"],
        ),
        comparison_row(
            "single_outlier_LOW_to_HIGH_status_count",
            int(baseline["single_outlier"]["status_transitions"].get("LOW->HIGH", 0)),
            l1["ticket_single_outlier"]["v2_low_to_high_status_count"],
            ticket["v2_low_to_high_status_count"],
        ),
        comparison_row(
            "single_outlier_scale_estimate_unchanged_count",
            baseline["single_outlier"]["scale_estimate_unchanged_count"],
            l1["ticket_single_outlier"]["v2_scale_estimate_unchanged_count"],
            ticket["v2_scale_estimate_unchanged_count"],
        ),
        comparison_row(
            "single_outlier_confidence_or_status_increased_count",
            None,
            l1["ticket_single_outlier"]["v2_confidence_or_status_increased_count"],
            ticket["v2_confidence_or_status_increased_count"],
        ),
    ]
    return rows


def run_selftests(stream: io.TextIOBase) -> tuple[int, dict[str, Any]]:
    tests: list[dict[str, Any]] = []

    def check(name: str, condition: bool, detail: str) -> None:
        passed = bool(condition)
        print(f"SELFTEST {name}: {'PASS' if passed else 'FAIL'} | {detail}", file=stream)
        tests.append({"name": name, "passed": passed, "detail": detail})

    print("=== loop_l1b mixed-population selftest ===", file=stream)
    try:
        check(
            "dossier_family_mapping_one_to_one",
            tuple(row[1] for row in DOSSIER_FAMILY_ROWS) == tuple(C0.MUTATION_FAMILIES),
            f"dossier_rows={len(DOSSIER_FAMILY_ROWS)} c0_families={len(C0.MUTATION_FAMILIES)}",
        )
        check(
            "population_cardinality",
            len(ANCHOR_RICH_INDICES) == 40 and len(SINGLE_SPAN_INDICES) == 10,
            f"anchor_rich={len(ANCHOR_RICH_INDICES)} single_span={len(SINGLE_SPAN_INDICES)}",
        )
        check(
            "population_assignment_seed_deterministic",
            POPULATION_ORDER == tuple(sorted(range(C0.BASE_SCENE_COUNT), key=assignment_key)),
            f"single_indices={sorted(SINGLE_SPAN_INDICES)} digest={canonical_sha256(list(POPULATION_ORDER))}",
        )

        single_index = min(SINGLE_SPAN_INDICES)
        rich_index = min(ANCHOR_RICH_INDICES)
        schedule = C0.fidelity_schedule()
        fixtures: dict[str, dict[str, Any]] = {}
        for label, index in (("single_span", single_index), ("anchor_rich", rich_index)):
            first = build_mixed_structural_scene(index, schedule[index][:3]).finalize()
            enrich_base_scene(first)
            second = build_mixed_structural_scene(index, schedule[index][:3]).finalize()
            enrich_base_scene(second)
            check(
                f"{label}_same_j_same_sha",
                canonical_sha256(first) == canonical_sha256(second),
                f"j={index} sha256={canonical_sha256(first)}",
            )
            fixtures[label] = first

        single = fixtures["single_span"]
        single_prepared, single_audit = C1.prepare_anchors(single["anchors"])
        single_model = C1.fit_anchor_model(single["anchors"])
        check(
            "single_span_structure",
            len(single["anchors"]) == 2
            and len({anchor["region"] for anchor in single["anchors"]}) == 1
            and len({C1.round_key(float(anchor["raw_span"])) for anchor in single["anchors"]}) == 1,
            f"anchors={len(single['anchors'])} regions={len({anchor['region'] for anchor in single['anchors']})} distinct_spans={len({C1.round_key(float(anchor['raw_span'])) for anchor in single['anchors']})}",
        )
        check(
            "single_span_canonical_low_representation",
            len(single_prepared) == 2
            and single_audit["duplicate_count"] == 0
            and single_model["unit_status"] in ("LOW", "NONE"),
            f"canonical={len(single_prepared)} unit_status={single_model['unit_status']} confidence={single_model['confidence_score']:.12g}",
        )

        rich = fixtures["anchor_rich"]
        rich_prepared, rich_audit = C1.prepare_anchors(rich["anchors"])
        rich_spans = [float(anchor["raw_span"]) for anchor in rich["anchors"]]
        rich_model = C1.fit_anchor_model(rich["anchors"])
        check(
            "anchor_rich_l1_rule",
            len(rich["anchors"]) == L1.anchor_count_for_seed(int(rich["seed"]))
            and 5 <= len(rich["anchors"]) <= 8
            and len(rich_prepared) == len(rich["anchors"])
            and rich_audit["duplicate_count"] == 0,
            f"anchors={len(rich['anchors'])} canonical={len(rich_prepared)} duplicates={rich_audit['duplicate_count']}",
        )
        check(
            "anchor_rich_independent_spans_bins",
            len({C1.round_key(span) for span in rich_spans}) == len(rich_spans)
            and C1.spatial_bin_count(rich_prepared) >= 3,
            f"distinct_spans={len({C1.round_key(span) for span in rich_spans})} bins={C1.spatial_bin_count(rich_prepared)}",
        )
        check(
            "anchor_rich_scale_estimate",
            rich_model["display_per_raw"] is not None
            and math.isclose(float(rich_model["display_per_raw"]), 1.0, rel_tol=1e-12),
            f"estimate={rich_model['display_per_raw']} unit_status={rich_model['unit_status']} confidence={rich_model['confidence_score']:.12g}",
        )

        scaled = [C0.scale_scene(single, kappa, token) for kappa, token in C0.SCALES]
        check(
            "single_span_four_scale_topology",
            len({scene["topology_sha256"] for scene in scaled}) == 1
            and len({scene["normalized_geometry_sha256"] for scene in scaled}) == 1,
            f"topology_unique={len({scene['topology_sha256'] for scene in scaled})} normalized_geometry_unique={len({scene['normalized_geometry_sha256'] for scene in scaled})}",
        )
        reverse_transition_count = 0
        for scene in scaled:
            before = C1.fit_anchor_model(scene["anchors"])
            after = C1.fit_anchor_model(
                C1.apply_corruption(
                    scene["anchors"], "single_outlier", str(scene["base_scene_id"])
                )
            )
            reverse_transition_count += (
                L1.STATUS_RANK[str(after["status"])]
                > L1.STATUS_RANK[str(before["status"])]
                or L1.STATUS_RANK[str(after["unit_status"])]
                > L1.STATUS_RANK[str(before["unit_status"])]
                or L1.STATUS_RANK[str(after["reference_status"])]
                > L1.STATUS_RANK[str(before["reference_status"])]
                or float(after["confidence_score"])
                > float(before["confidence_score"]) + 1e-15
            )
        check(
            "single_span_single_outlier_no_reverse_transition",
            reverse_transition_count == 0,
            f"four_scale_reverse_transition_count={reverse_transition_count}",
        )
        positive_errors = [error for scene in scaled for error in C0.validate_scene(scene)]
        check(
            "truth_validator_positive_cases",
            not positive_errors,
            f"scene_count={len(scaled)} error_count={len(positive_errors)}",
        )
        corrupted = copy.deepcopy(scaled[0])
        if corrupted["truth_pairs"]:
            corrupted["truth_pairs"][0]["handles"][1] = "MISSING_SOURCE_HANDLE"
            corrupted["truth_pairs"][0]["pair_id"] = C0.pair_id(
                corrupted["truth_pairs"][0]["handles"]
            )
            negative_errors = C0.validate_scene(corrupted)
        else:
            positive_fixture = C0.scale_scene(rich, C0.SCALES[0][0], C0.SCALES[0][1])
            positive_fixture["truth_pairs"][0]["handles"][1] = "MISSING_SOURCE_HANDLE"
            positive_fixture["truth_pairs"][0]["pair_id"] = C0.pair_id(
                positive_fixture["truth_pairs"][0]["handles"]
            )
            negative_errors = C0.validate_scene(positive_fixture)
        check(
            "truth_validator_negative_case_honest_fail",
            bool(negative_errors),
            f"error_count={len(negative_errors)}",
        )

        distribution: Counter[int] = Counter()
        role_errors = 0
        rich_bin_errors = 0
        for index in range(C0.BASE_SCENE_COUNT):
            builder = MixedPopulationSceneBuilder(index)
            builder.add_reference_anchors()
            population = population_for_index(index)
            distribution[len(builder.anchors)] += 1
            prepared, _ = C1.prepare_anchors(builder.anchors)
            if population == "single_span":
                role_errors += not (
                    len(builder.anchors) == 2
                    and "single_reference_span_region" in builder.mutations
                    and "multiple_reference_span_regions" not in builder.mutations
                )
            else:
                role_errors += not (
                    len(builder.anchors) == L1.anchor_count_for_seed(builder.seed)
                    and "multiple_reference_span_regions" in builder.mutations
                    and "single_reference_span_region" not in builder.mutations
                )
                rich_bin_errors += C1.spatial_bin_count(prepared) < 3
        check(
            "all_base_population_rules",
            role_errors == 0 and rich_bin_errors == 0,
            f"distribution={dict(sorted(distribution.items()))} role_errors={role_errors} rich_bin_errors={rich_bin_errors}",
        )
        entity_types, wall_count = C0.gen2_library_probe()
        required = {"LINE", "ARC", "SPLINE", "HATCH", "LWPOLYLINE"}
        check(
            "permitted_gen2_synthetic_fixture",
            wall_count >= 2 and required.issubset(entity_types),
            f"wall_records={wall_count} entity_types={sorted(entity_types)}",
        )
    except Exception as exc:
        check("unexpected_exception", False, f"{type(exc).__name__}: {exc}")

    c1_stream = io.StringIO()
    c1_tests: dict[str, Any] = {"passed": 0, "total": 0, "tests": []}
    try:
        c1_tests = C1.run_selftests(c1_stream)
        print("=== imported sealed C1 selftests ===", file=stream)
        print(c1_stream.getvalue().rstrip("\n"), file=stream)
        check(
            "sealed_c1_selftests",
            c1_tests["passed"] == c1_tests["total"],
            f"passed={c1_tests['passed']} total={c1_tests['total']}",
        )
    except Exception as exc:
        print(c1_stream.getvalue().rstrip("\n"), file=stream)
        check("sealed_c1_selftests", False, f"{type(exc).__name__}: {exc}")

    passed = sum(test["passed"] for test in tests)
    print(f"SELFTEST SUMMARY: {passed}/{len(tests)} passed", file=stream)
    print(f"SELFTEST_RESULT: {'PASS' if passed == len(tests) else 'FAIL'}", file=stream)
    details = {
        "passed": passed,
        "total": len(tests),
        "tests": tests,
        "sealed_c1": c1_tests,
    }
    return (0 if passed == len(tests) else 1), details


_ORIGINAL_AUGMENT_EVIDENCE_WORKBOOK = L1.augment_evidence_workbook


def augment_evidence_workbook(
    c0_numbers: Mapping[str, Any],
    c0_comparison: Sequence[Mapping[str, Any]],
    c1_comparison: Sequence[Mapping[str, Any]],
    aggregates: Mapping[str, Any],
    ticket: Mapping[str, Any],
    source_manifest_before: Mapping[str, Any],
    source_manifest_after: Mapping[str, Any],
) -> None:
    _ORIGINAL_AUGMENT_EVIDENCE_WORKBOOK(
        c0_numbers,
        c0_comparison,
        c1_comparison,
        aggregates,
        ticket,
        source_manifest_before,
        source_manifest_after,
    )
    import openpyxl

    workbook = openpyxl.load_workbook(EVIDENCE_PATH)
    for name in (
        "population_assignment",
        "mutation_families",
        "c0_three_way",
        "c1_three_way",
    ):
        if name in workbook.sheetnames:
            del workbook[name]

    overview = workbook["overview"]
    overview["A1"] = "E2 Loop L1b — mixed-population C0v3 / sealed C1 reevaluation"
    for row in overview.iter_rows(min_row=3):
        if isinstance(row[3].value, str):
            row[3].value = row[3].value.replace("c0v2", "c0v3").replace("c1v2", "c1v3")

    if "anchor_distribution" in workbook.sheetnames:
        index = workbook.sheetnames.index("anchor_distribution")
        del workbook["anchor_distribution"]
    else:
        index = 1
    distribution = workbook.create_sheet("anchor_distribution", index)
    distribution.append(["anchor_count", "base_scene_count", "ir_scene_count"])
    population = c0_numbers["anchor_richness"]
    base_distribution = population["base_anchor_count_distribution"]
    ir_distribution = population["ir_anchor_count_distribution"]
    for count in sorted({int(key) for key in base_distribution} | {int(key) for key in ir_distribution}):
        distribution.append(
            [
                count,
                int(base_distribution.get(count, base_distribution.get(str(count), 0))),
                int(ir_distribution.get(count, ir_distribution.get(str(count), 0))),
            ]
        )
    L1.style_table_sheet(distribution, {"A": 18, "B": 22, "C": 20})

    assignment = workbook.create_sheet("population_assignment")
    assignment_headers = [
        "base_scene_index",
        "seed",
        "assignment_rank",
        "population",
        "anchor_count",
        "distinct_span_count",
        "spatial_bin_count",
        "region_count",
        "unit_status",
        "reference_status",
        "confidence_score",
        "assignment_sha256",
    ]
    assignment.append(assignment_headers)
    for row in population["assignment_rows"]:
        assignment.append([row.get(header) for header in assignment_headers])
    L1.style_table_sheet(
        assignment,
        {"A": 18, "B": 16, "C": 18, "D": 18, "E": 14, "F": 20, "G": 18, "H": 14, "I": 16, "J": 20, "K": 20, "L": 68},
    )

    v1 = json.loads(V1_C0_NUMBERS_PATH.read_text(encoding="utf-8"))
    l1 = json.loads(L1_C0_NUMBERS_PATH.read_text(encoding="utf-8"))
    families = workbook.create_sheet("mutation_families")
    families.append(
        [
            "dossier_family",
            "manifest_family",
            "v1_scene_count",
            "l1_scene_count",
            "v3_scene_count",
            "delta_v3_minus_v1",
            "delta_v3_minus_l1",
        ]
    )
    for dossier_label, family in DOSSIER_FAMILY_ROWS:
        v1_count = v1["mutation_family_coverage"][family]["scene_count"]
        l1_count = l1["mutation_family_coverage"][family]["scene_count"]
        v3_count = c0_numbers["mutation_family_coverage"][family]["scene_count"]
        families.append(
            [
                dossier_label,
                family,
                v1_count,
                l1_count,
                v3_count,
                v3_count - v1_count,
                v3_count - l1_count,
            ]
        )
    L1.style_table_sheet(
        families,
        {"A": 48, "B": 62, "C": 18, "D": 18, "E": 18, "F": 22, "G": 22},
    )

    for name, rows in (("c0_three_way", c0_comparison), ("c1_three_way", c1_comparison)):
        sheet = workbook.create_sheet(name)
        headers = [
            "metric",
            "v1",
            "l1",
            "v3",
            "delta_v3_minus_v1",
            "delta_v3_minus_l1",
        ]
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header) for header in headers])
        L1.style_table_sheet(
            sheet,
            {"A": 76, "B": 20, "C": 20, "D": 20, "E": 24, "F": 24},
        )

    workbook.properties.title = "E2 Loop L1b numeric evidence"
    workbook.properties.subject = "Mixed-population C0v3 and sealed C1 reevaluation"
    temporary = EVIDENCE_PATH.with_name(EVIDENCE_PATH.stem + ".l1b.tmp.xlsx")
    workbook.save(temporary)
    workbook.close()
    temporary.replace(EVIDENCE_PATH)


def format_cell(value: Any) -> Any:
    if isinstance(value, (Mapping, list, tuple)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return value


def three_way_table(rows: Sequence[Mapping[str, Any]]) -> str:
    return L1.markdown_table(
        ["metric", "v1", "L1", "L1b-v3", "v3-v1", "v3-L1"],
        [
            (
                row["metric"],
                format_cell(row["v1"]),
                format_cell(row["l1"]),
                format_cell(row["v3"]),
                format_cell(row["delta_v3_minus_v1"]),
                format_cell(row["delta_v3_minus_l1"]),
            )
            for row in rows
        ],
    )


def render_report(
    c0_numbers: Mapping[str, Any],
    c1_results: Mapping[str, Any],
    selftest_transcript: str,
    artifact_rows: Sequence[Mapping[str, Any]],
) -> str:
    v1_c0 = json.loads(V1_C0_NUMBERS_PATH.read_text(encoding="utf-8"))
    l1_c0 = json.loads(L1_C0_NUMBERS_PATH.read_text(encoding="utf-8"))
    aggregates = c1_results["aggregates"]
    overall = aggregates["overall"]
    population = c0_numbers["anchor_richness"]
    c0_compare = c0_numbers["v1_comparison_rows"]
    c1_compare = c1_results["v1_comparison_rows"]
    ticket = c1_results["ticket_single_outlier"]

    family_rows = []
    for dossier_label, family in DOSSIER_FAMILY_ROWS:
        v1_count = v1_c0["mutation_family_coverage"][family]["scene_count"]
        l1_count = l1_c0["mutation_family_coverage"][family]["scene_count"]
        v3_count = c0_numbers["mutation_family_coverage"][family]["scene_count"]
        family_rows.append(
            (
                dossier_label,
                family,
                v1_count,
                l1_count,
                v3_count,
                v3_count - v1_count,
                v3_count - l1_count,
            )
        )

    high_by_scale = [
        row for row in aggregates["scale_confidence_rows"] if row["confidence"] == "HIGH"
    ]
    lines: list[str] = [
        "# E2 Loop L1b — 혼합 모집단 C0/C1 재평가",
        "",
        "## 설계 및 실행 경계",
        "",
        f"- base 50은 seed-derived rank로 anchor-rich {ANCHOR_RICH_BASE_COUNT}, single-span {SINGLE_SPAN_BASE_COUNT}에 배정했다.",
        "- anchor-rich 장면은 L1의 5~8 independent-span DIM builder를 그대로 재사용했고, single-span 장면은 한 region의 동일 reference span DIM 2개를 사용했다.",
        "- C0 구조 생성, 4-scale 복제, fidelity, truth validator와 C1 estimator, 네 corruption, pair-label permutation은 read-only import로 실행했다.",
        f"- 쓰기는 `{CELL_DIR}` 아래에만 수행했고 source CAD/test split 접근 수는 각각 0이다.",
        "- 아래에는 측정값과 selftest 기록만 수록한다.",
        "",
        "## feyerabend_P2 지정 mutation family 1:1 대조",
        "",
        f"원문 위치: `{DOSSIER_PATH}` lines 550-561. 결합 표기된 단일·다중 reference span 항목은 두 manifest family로 분리했다.",
        "",
        L1.markdown_table(
            ["dossier family", "manifest family", "v1", "L1", "v3", "v3-v1", "v3-L1"],
            family_rows,
        ),
        "",
        "## 혼합 모집단 실측",
        "",
        L1.markdown_table(
            ["metric", "value"],
            [
                ("base_population_counts", json.dumps(population["base_population_counts"], sort_keys=True)),
                ("ir_population_counts", json.dumps(population["ir_population_counts"], sort_keys=True)),
                ("single_span_indices", json.dumps(population["single_span_indices"])),
                ("population_assignment_digest", population["population_assignment_digest"]),
                ("base_anchor_count_distribution", json.dumps(population["base_anchor_count_distribution"], sort_keys=True)),
                ("base_unit_status_counts", json.dumps(population["base_unit_status_counts"], sort_keys=True)),
                ("base_unit_status_counts_by_population", json.dumps(population["base_unit_status_counts_by_population"], sort_keys=True)),
                ("single_span_unit_HIGH_scene_count", population["single_span_unit_HIGH_scene_count"]),
                ("single_span_unit_LOW_or_NONE_scene_count", population["single_span_unit_LOW_or_NONE_scene_count"]),
                ("anchor_rich_unit_HIGH_scene_count", population["anchor_rich_unit_HIGH_scene_count"]),
                ("anchor_count_rule_mismatch_scene_count", population["base_anchor_count_rule_mismatch_scene_count"]),
                ("distinct_span_rule_mismatch_scene_count", population["base_distinct_span_rule_mismatch_scene_count"]),
                ("mutation_population_role_mismatch_scene_count", population["base_mutation_population_role_mismatch_scene_count"]),
            ],
        ),
        "",
        "## Selftest 전문",
        "",
        "```text",
        selftest_transcript.rstrip("\n"),
        "```",
        "",
        "## C0 v1 / L1 / L1b-v3 델타",
        "",
        three_way_table(c0_compare),
        "",
        "## C1 v1 / L1 / L1b-v3 델타",
        "",
        three_way_table(c1_compare),
        "",
        "## C0v3 fidelity·truth·4-scale topology 수치 전문",
        "",
        "```json",
        json.dumps(c0_numbers, ensure_ascii=False, indent=2, sort_keys=True, allow_nan=False),
        "```",
        "",
        "## C1v3 전체 scale 추정 수치",
        "",
        L1.markdown_table(
            ["metric", "value"],
            [
                ("scene_count", overall["scene_count"]),
                ("estimate_count", overall["estimate_count"]),
                ("estimate_coverage", overall["estimate_coverage"]),
                ("accuracy_within_5pct", overall["accuracy_within_5pct"]),
                ("HIGH_scene_count", overall["high_scene_count"]),
                ("HIGH_coverage", overall["high_coverage"]),
                ("HIGH_accuracy_within_5pct", overall["high_accuracy_within_5pct"]),
                ("e_s_min", overall["e_s"]["min"]),
                ("e_s_median", overall["e_s"]["median"]),
                ("e_s_p95", overall["e_s"]["p95"]),
                ("e_s_max", overall["e_s"]["max"]),
                ("relative_error_min", overall["relative_error"]["min"]),
                ("relative_error_median", overall["relative_error"]["median"]),
                ("relative_error_p95", overall["relative_error"]["p95"]),
                ("relative_error_max", overall["relative_error"]["max"]),
            ],
        ),
        "",
        "## Scale × confidence 전 행",
        "",
        L1.markdown_table(
            ["kappa", "unit_status", "n", "fraction", "n_est", "accuracy_5pct", "e_s_med", "e_s_p95", "relerr_med", "relerr_p95", "conf_med"],
            [
                (
                    row["scale_kappa"],
                    row["confidence"],
                    row["scene_count"],
                    row["fraction_of_scale"],
                    row["estimate_count"],
                    row["accuracy_within_5pct"],
                    row["e_s"]["median"],
                    row["e_s"]["p95"],
                    row["relative_error"]["median"],
                    row["relative_error"]["p95"],
                    row["confidence_score"]["median"],
                )
                for row in aggregates["scale_confidence_rows"]
            ],
        ),
        "",
        "## Scale별 HIGH coverage·accuracy",
        "",
        L1.markdown_table(
            ["kappa", "HIGH_n", "HIGH_coverage", "HIGH_accuracy_5pct", "HIGH_relerr_p95"],
            [
                (
                    row["scale_kappa"],
                    row["scene_count"],
                    row["fraction_of_scale"],
                    row["accuracy_within_5pct"],
                    row["relative_error"]["p95"],
                )
                for row in high_by_scale
            ],
        ),
        "",
        "## Numeric confidence-score bin별 accuracy",
        "",
        L1.markdown_table(
            ["scale", "score_bin", "n", "n_est", "accuracy_5pct", "relerr_med", "relerr_p95"],
            [
                (
                    row["scale_kappa"],
                    row["score_bin"],
                    row["scene_count"],
                    row["estimate_count"],
                    row["accuracy_within_5pct"],
                    row["relative_error"]["median"],
                    row["relative_error"]["p95"],
                )
                for row in aggregates["score_calibration_rows"]
            ],
        ),
        "",
        "## Corruption 4종 전이 수치",
        "",
        L1.markdown_table(
            ["corruption", "n", "unit_transition", "status_transition", "reference_transition", "scale_same", "scale_changed", "relerr_med", "relerr_p95", "conf_after_med"],
            [
                (
                    kind,
                    summary["scene_count"],
                    json.dumps(summary["unit_status_transitions"], sort_keys=True),
                    json.dumps(summary["status_transitions"], sort_keys=True),
                    json.dumps(summary["reference_status_transitions"], sort_keys=True),
                    summary["scale_estimate_unchanged_count"],
                    summary["scale_estimate_changed_count"],
                    summary["relative_error_after"]["median"],
                    summary["relative_error_after"]["p95"],
                    summary["confidence_score_after"]["median"],
                )
                for kind, summary in aggregates["corruption"]["all_four_applied"].items()
            ],
        ),
        "",
        "## Single-outlier 역방향 전이 수치",
        "",
        L1.markdown_table(
            ["metric", "value"],
            [
                (key.replace("v2_", "v3_"), format_cell(value))
                for key, value in ticket.items()
            ],
        ),
        "",
        "## Pair-label permutation digest 수치",
        "",
        L1.markdown_table(
            ["metric", "value"],
            [(key, value) for key, value in aggregates["pair_label_permutation"].items()],
        ),
        "",
        "## C1v3 aggregate 수치 전문",
        "",
        "```json",
        json.dumps(aggregates, ensure_ascii=False, indent=2, sort_keys=True, allow_nan=False),
        "```",
        "",
        "## 산출물 검증 수치",
        "",
        L1.markdown_table(
            ["artifact", "status", "bytes", "sha256_or_digest"],
            [
                (row["artifact"], row["status"], row["bytes"], row["sha256_or_digest"])
                for row in artifact_rows
            ],
        ),
        "",
        "## 미해결",
        "",
        f"- evidence workbook 구조 검증 required sheet 수는 {c1_results['evidence']['verification']['required_sheet_count']}, formula error 수는 {c1_results['evidence']['verification']['formula_error_count']}, raster render 수는 {c1_results['evidence']['verification']['visual_render_count']}이다.",
        f"- source read-only manifest mismatch 수는 {c1_results['source_readonly_manifest']['mismatch_count']}, v3 scene input manifest mismatch 수는 {c1_results['inputs']['manifest_mismatch_count']}이다.",
        f"- reference_status 분포는 `{json.dumps(overall['status_counts'], sort_keys=True)}`, unit_status 분포는 `{json.dumps(overall['unit_status_counts'], sort_keys=True)}`다.",
        "- 원본 CAD 접근 수와 test split 접근 수는 각각 0이다.",
        "",
        # L1.execute_full validates this temporary internal marker.  execute_full
        # atomically replaces it with the packet-required L1b marker afterward.
        "LOOP_COMPLETE: L1",
    ]
    report = "\n".join(lines) + "\n"
    if report.rstrip("\n").splitlines()[-1] != "LOOP_COMPLETE: L1":
        raise AssertionError("temporary report completion marker mismatch")
    return report


def artifact_row(path: Path, status: str = "GENERATED") -> dict[str, Any]:
    if path.parent == CELL_DIR and path.name == "loop_l1.py":
        path = SCRIPT_PATH
    return {
        "artifact": path.name,
        "status": status,
        "bytes": path.stat().st_size,
        "sha256_or_digest": L1.sha256_file(path),
    }


SOURCE_PATHS = (
    PACKET_PATH,
    DOSSIER_PATH,
    C0_PATH,
    C1_PATH,
    L1_PATH,
    V1_C0_NUMBERS_PATH,
    V1_C1_REPORT_PATH,
    L1_C0_NUMBERS_PATH,
    L1_C1_RESULTS_PATH,
    L1_REPORT_PATH,
    SCRIPT_PATH,
)


def configure_l1_extension() -> None:
    L1.CELL_DIR = CELL_DIR
    L1.SCENES_V2_DIR = SCENES_V3_DIR
    L1.C0V2_NUMBERS_PATH = C0V3_NUMBERS_PATH
    L1.C1V2_RESULTS_PATH = C1V3_RESULTS_PATH
    L1.EVIDENCE_PATH = EVIDENCE_PATH
    L1.EVIDENCE_FALLBACK_PATH = EVIDENCE_FALLBACK_PATH
    L1.REPORT_PATH = REPORT_PATH
    L1.PACKET_PATH = PACKET_PATH
    L1.SOURCE_PATHS = SOURCE_PATHS
    L1.LOOP_SCHEMA = LOOP_SCHEMA
    L1.C0V2_SCHEMA = C0V3_SCHEMA
    L1.C1V2_SCHEMA = C1V3_SCHEMA
    L1.enrich_base_scene = enrich_base_scene
    L1.build_v2_structural_scene = build_mixed_structural_scene
    L1.build_v2_base_corpus = build_mixed_base_corpus
    L1.anchor_richness_numbers = population_numbers
    L1.c0_v1_v2_comparison_rows = c0_three_way_comparison_rows
    L1.c1_v1_v2_comparison_rows = c1_three_way_comparison_rows
    L1.run_selftests = run_selftests
    L1.augment_evidence_workbook = augment_evidence_workbook
    L1.render_report = render_report
    L1.artifact_row = artifact_row


def finalize_report_marker() -> None:
    report = REPORT_PATH.read_text(encoding="utf-8")
    lines = report.rstrip("\n").splitlines()
    if not lines or lines[-1] != "LOOP_COMPLETE: L1":
        raise RuntimeError(f"unexpected temporary REPORT final line: {lines[-1] if lines else None!r}")
    lines[-1] = "LOOP_COMPLETE: L1b"
    L1.write_text_atomic(REPORT_PATH, "\n".join(lines) + "\n")


def verify_final_artifacts() -> dict[str, Any]:
    required = (
        SCRIPT_PATH,
        C0V3_NUMBERS_PATH,
        C1V3_RESULTS_PATH,
        EVIDENCE_PATH if EVIDENCE_PATH.is_file() else EVIDENCE_FALLBACK_PATH,
        REPORT_PATH,
    )
    missing = [str(path) for path in required if not path.is_file() or path.stat().st_size <= 0]
    scene_paths = sorted(SCENES_V3_DIR.glob("scene_*.json"))
    report_final_line = REPORT_PATH.read_text(encoding="utf-8").rstrip("\n").splitlines()[-1]
    c0_numbers = json.loads(C0V3_NUMBERS_PATH.read_text(encoding="utf-8"))
    c1_results = json.loads(C1V3_RESULTS_PATH.read_text(encoding="utf-8"))
    family_counts = {
        family: c0_numbers["mutation_family_coverage"][family]["scene_count"]
        for _, family in DOSSIER_FAMILY_ROWS
    }
    scale_rows = c1_results["aggregates"]["scale_confidence_rows"]
    corruption_kinds = sorted(
        c1_results["aggregates"]["corruption"]["all_four_applied"]
    )
    structural_errors = []
    if missing:
        structural_errors.append(f"missing_or_empty={missing}")
    if len(scene_paths) != C0.BASE_SCENE_COUNT * len(C0.SCALES):
        structural_errors.append(f"scene_count={len(scene_paths)}")
    if report_final_line != "LOOP_COMPLETE: L1b":
        structural_errors.append(f"report_final_line={report_final_line!r}")
    if len(scale_rows) != len(C0.SCALES) * len(C1.CONFIDENCE_LEVELS):
        structural_errors.append(f"scale_confidence_row_count={len(scale_rows)}")
    if corruption_kinds != sorted(C1.CORRUPTIONS):
        structural_errors.append(f"corruptions={corruption_kinds}")
    if any(count < 1 for count in family_counts.values()):
        structural_errors.append(f"zero_family_counts={family_counts}")
    if structural_errors:
        raise RuntimeError("; ".join(structural_errors))
    return {
        "scene_count": len(scene_paths),
        "family_scene_counts": family_counts,
        "base_population_counts": c0_numbers["anchor_richness"]["base_population_counts"],
        "scale_confidence_row_count": len(scale_rows),
        "corruption_kinds": corruption_kinds,
        "high_coverage": c1_results["aggregates"]["overall"]["high_coverage"],
        "high_accuracy_within_5pct": c1_results["aggregates"]["overall"]["high_accuracy_within_5pct"],
        "single_outlier_confidence_or_status_increased_count": c1_results["ticket_single_outlier"]["v2_confidence_or_status_increased_count"],
        "pair_digest_mismatch_count": c1_results["aggregates"]["pair_label_permutation"]["mismatching_anchor_artifact_scene_count"],
        "truth_validator_error_count": c0_numbers["truth_validator"]["error_count"],
        "source_manifest_mismatch_count": c1_results["source_readonly_manifest"]["mismatch_count"],
        "input_manifest_mismatch_count": c1_results["inputs"]["manifest_mismatch_count"],
        "report_final_line": report_final_line,
    }


def execute_full() -> int:
    configure_l1_extension()
    return_code = L1.execute_full()
    if return_code != 0:
        return return_code
    finalize_report_marker()
    verification = verify_final_artifacts()
    print(json.dumps(verification, ensure_ascii=False, sort_keys=True))
    return 0


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--selftest",
        action="store_true",
        help="run deterministic, truth-contact, population, and sealed C1 checks without writes",
    )
    arguments = parser.parse_args(argv)
    if arguments.selftest:
        stream = io.StringIO()
        return_code, _details = run_selftests(stream)
        print(stream.getvalue(), end="")
        return return_code
    return execute_full()


if __name__ == "__main__":
    raise SystemExit(main())
