#!/usr/bin/env python3
"""Feyerabend C1: label-free anchor scale estimation and calibration.

The estimator accepts anchor records only.  Truth scale and wall-pair labels are
used by the evaluation layer after estimation, never by the estimator.
"""

from __future__ import annotations

import argparse
import copy
import hashlib
import io
import json
import math
import platform
import statistics
import sys
import time
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

import numpy as np


CELL_DIR = Path(__file__).resolve().parent
INPUT_DIR = Path(r"D:\runs\e2_program\cells\feyerabend_c0\scenes")
RESULTS_PATH = CELL_DIR / "results.json"
EVIDENCE_PATH = CELL_DIR / "evidence.xlsx"
EVIDENCE_FALLBACK_PATH = CELL_DIR / "evidence_rows.json"
REPORT_PATH = CELL_DIR / "REPORT.md"

INPUT_SCHEMA = "ariadne.e2.feyerabend_c0.scene.v1"
RESULTS_SCHEMA = "ariadne.e2.feyerabend_c1.results.v1"
EXPECTED_SCENE_COUNT = 200
EXPECTED_BASE_COUNT = 50
SCALES = (0.001, 0.01, 1.0, 1000.0)
CONFIDENCE_LEVELS = ("HIGH", "LOW", "NONE")
CORRUPTIONS = ("duplicate", "stale_override", "suffix_removal", "single_outlier")

RANSAC_LOG_TOLERANCE = math.log(1.05)
HUBER_DELTA = 1.5
HIGH_CONFIDENCE_THRESHOLD = 0.75
CONSENSUS_THRESHOLD = 0.80
MIN_INDEPENDENT = 3
ACCURACY_RELATIVE_ERROR = 0.05
EPSILON = 1e-12

BASE_WEIGHTS = {"DIM": 1.0, "TEXT": 0.6, "GRID": 0.4}
UNIT_TO_MM = {"MM": 1.0, "CM": 10.0, "M": 1000.0, "IN": 25.4, "FT": 304.8}
SCORE_BINS = (
    ("[0.00,0.25)", 0.00, 0.25, False),
    ("[0.25,0.50)", 0.25, 0.50, False),
    ("[0.50,0.75)", 0.50, 0.75, False),
    ("[0.75,1.00]", 0.75, 1.00, True),
)


def canonical_bytes(value: Any) -> bytes:
    return json.dumps(
        value,
        sort_keys=True,
        separators=(",", ":"),
        ensure_ascii=False,
        allow_nan=False,
    ).encode("utf-8")


def canonical_sha256(value: Any) -> str:
    return hashlib.sha256(canonical_bytes(value)).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def finite_float(value: Any) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError, OverflowError):
        return None
    return number if math.isfinite(number) else None


def round_key(value: float) -> str:
    return format(float(value), ".12g")


def scale_token(value: float) -> str:
    return format(float(value), ".12g")


def weighted_median(values: Sequence[float], weights: Sequence[float]) -> float:
    if len(values) != len(weights) or not values:
        raise ValueError("weighted_median requires non-empty, aligned inputs")
    order = sorted(range(len(values)), key=lambda index: (values[index], index))
    total = math.fsum(weights)
    if total <= 0.0:
        raise ValueError("weighted_median requires positive total weight")
    threshold = 0.5 * total
    cumulative = 0.0
    for index in order:
        cumulative += weights[index]
        if cumulative >= threshold:
            return float(values[index])
    return float(values[order[-1]])


def weighted_log_mad(values: Sequence[float], weights: Sequence[float], center: float) -> float:
    deviations = [abs(float(value) - float(center)) for value in values]
    return weighted_median(deviations, weights)


def huber_location(
    values: Sequence[float], weights: Sequence[float], delta: float = HUBER_DELTA
) -> float:
    center = weighted_median(values, weights)
    scale = weighted_log_mad(values, weights, center)
    if scale <= EPSILON:
        return center
    for _ in range(64):
        effective: list[float] = []
        for value, weight in zip(values, weights):
            residual = abs((float(value) - center) / scale)
            factor = 1.0 if residual <= delta else delta / residual
            effective.append(float(weight) * factor)
        denominator = math.fsum(effective)
        if denominator <= 0.0:
            return center
        updated = math.fsum(
            float(value) * weight for value, weight in zip(values, effective)
        ) / denominator
        if abs(updated - center) <= 1e-14:
            return float(updated)
        center = float(updated)
    return center


def _point(value: Any) -> tuple[float, float] | None:
    if not isinstance(value, (list, tuple)) or len(value) < 2:
        return None
    x = finite_float(value[0])
    y = finite_float(value[1])
    if x is None or y is None:
        return None
    return (x, y)


def geometric_span(anchor: Mapping[str, Any]) -> float | None:
    p0 = _point(anchor.get("p0"))
    p1 = _point(anchor.get("p1"))
    if p0 is None or p1 is None:
        return None
    span = math.hypot(p1[0] - p0[0], p1[1] - p0[1])
    return span if math.isfinite(span) and span > EPSILON else None


def midpoint(anchor: Mapping[str, Any]) -> tuple[float, float]:
    p0 = _point(anchor["p0"])
    p1 = _point(anchor["p1"])
    assert p0 is not None and p1 is not None
    return (0.5 * (p0[0] + p1[0]), 0.5 * (p0[1] + p1[1]))


def canonical_anchor_key(anchor: Mapping[str, Any], span: float) -> tuple[Any, ...]:
    p0 = _point(anchor["p0"])
    p1 = _point(anchor["p1"])
    assert p0 is not None and p1 is not None
    endpoints = sorted(
        ((round_key(p0[0]), round_key(p0[1])), (round_key(p1[0]), round_key(p1[1])))
    )
    display = finite_float(anchor.get("display_value"))
    return (
        str(anchor.get("anchor_type", "")).upper(),
        tuple(endpoints),
        round_key(span),
        None if display is None else round_key(display),
        str(anchor.get("display_unit", "UNKNOWN")).upper(),
    )


def prepare_anchors(raw_anchors: Sequence[Mapping[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    prepared: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []
    duplicate_handles: list[str] = []
    seen: dict[tuple[Any, ...], str] = {}

    ordered = sorted(
        (copy.deepcopy(dict(anchor)) for anchor in raw_anchors if isinstance(anchor, Mapping)),
        key=lambda anchor: str(anchor.get("handle", "")),
    )
    for input_index, anchor in enumerate(ordered):
        handle = str(anchor.get("handle") or f"ANCHOR_{input_index:04d}")
        anchor_type = str(anchor.get("anchor_type", "")).upper()
        if anchor_type not in BASE_WEIGHTS:
            rejected.append({"handle": handle, "reason": "unsupported_anchor_type"})
            continue
        span = geometric_span(anchor)
        if span is None:
            rejected.append({"handle": handle, "reason": "invalid_geometry_span"})
            continue
        declared_span = finite_float(anchor.get("raw_span"))
        span_disagreement = None
        if declared_span is not None:
            span_disagreement = abs(declared_span - span) / max(span, EPSILON)
        weight = finite_float(anchor.get("weight"))
        if weight is None or weight <= 0.0:
            weight = BASE_WEIGHTS[anchor_type]
        display = finite_float(anchor.get("display_value"))
        if display is not None and display <= 0.0:
            display = None
        normalized = {
            "handle": handle,
            "anchor_type": anchor_type,
            "region": str(anchor.get("region", "UNKNOWN")),
            "p0": list(_point(anchor["p0"]) or ()),
            "p1": list(_point(anchor["p1"]) or ()),
            "raw_span": float(span),
            "declared_raw_span": declared_span,
            "span_relative_disagreement": span_disagreement,
            "display_value": display,
            "display_unit": str(anchor.get("display_unit", "UNKNOWN")).upper(),
            "text_height": finite_float(anchor.get("text_height")),
            "weight": float(weight),
        }
        key = canonical_anchor_key(normalized, span)
        if key in seen:
            duplicate_handles.append(handle)
            rejected.append(
                {
                    "handle": handle,
                    "reason": "duplicate_anchor",
                    "canonical_handle": seen[key],
                }
            )
            continue
        seen[key] = handle
        prepared.append(normalized)

    audit = {
        "input_anchor_count": len(raw_anchors),
        "canonical_anchor_count": len(prepared),
        "duplicate_count": len(duplicate_handles),
        "duplicate_handles": duplicate_handles,
        "rejections": rejected,
    }
    return prepared, audit


def spatial_bin_count(anchors: Sequence[Mapping[str, Any]]) -> int:
    if not anchors:
        return 0
    points = [midpoint(anchor) for anchor in anchors]
    xs = [point[0] for point in points]
    ys = [point[1] for point in points]
    xmin, xmax = min(xs), max(xs)
    ymin, ymax = min(ys), max(ys)

    def axis_bin(value: float, low: float, high: float) -> int:
        if abs(high - low) <= EPSILON:
            return 1
        return min(2, max(0, int(math.floor(3.0 * (value - low) / (high - low)))))

    bins = {
        (axis_bin(x, xmin, xmax), axis_bin(y, ymin, ymax)) for x, y in points
    }
    return len(bins)


def maximum_consensus(
    records: Sequence[Mapping[str, Any]], value_key: str, tolerance: float
) -> dict[str, Any] | None:
    if not records:
        return None
    total_weight = math.fsum(float(record["weight"]) for record in records)
    candidates: list[tuple[tuple[Any, ...], list[int]]] = []
    for candidate_index, candidate in enumerate(records):
        center = float(candidate[value_key])
        inliers = [
            index
            for index, record in enumerate(records)
            if abs(float(record[value_key]) - center) <= tolerance + 1e-15
        ]
        inlier_weight = math.fsum(float(records[index]["weight"]) for index in inliers)
        values = [float(records[index][value_key]) for index in inliers]
        weights = [float(records[index]["weight"]) for index in inliers]
        median = weighted_median(values, weights)
        mad = weighted_log_mad(values, weights, median)
        rank = (-inlier_weight, -len(inliers), mad, center, candidate_index)
        candidates.append((rank, inliers))
    candidates.sort(key=lambda item: item[0])
    selected = candidates[0][1]
    selected_values = [float(records[index][value_key]) for index in selected]
    selected_weights = [float(records[index]["weight"]) for index in selected]
    location = huber_location(selected_values, selected_weights)
    mad = weighted_log_mad(selected_values, selected_weights, location)
    inlier_weight = math.fsum(selected_weights)
    selected_set = set(selected)
    return {
        "inlier_indices": selected,
        "outlier_indices": [index for index in range(len(records)) if index not in selected_set],
        "location": float(location),
        "log_mad": float(mad),
        "inlier_weight": float(inlier_weight),
        "total_weight": float(total_weight),
        "consensus_weight": float(inlier_weight / total_weight),
    }


def confidence_from_consensus(
    consensus: Mapping[str, Any], inlier_anchors: Sequence[Mapping[str, Any]]
) -> dict[str, Any]:
    n_independent = len(inlier_anchors)
    n_regions = spatial_bin_count(inlier_anchors)
    score = (
        float(consensus["consensus_weight"])
        * math.exp(-float(consensus["log_mad"]) / RANSAC_LOG_TOLERANCE)
        * min(1.0, n_independent / 5.0)
        * min(1.0, n_regions / 3.0)
    )
    return {
        "score": float(score),
        "n_independent": n_independent,
        "n_spatial_bins": n_regions,
    }


def _status(
    exists: bool, consensus: Mapping[str, Any] | None, confidence: Mapping[str, Any] | None
) -> str:
    if not exists or consensus is None or confidence is None:
        return "NONE"
    high = (
        int(confidence["n_independent"]) >= MIN_INDEPENDENT
        and float(consensus["consensus_weight"]) >= CONSENSUS_THRESHOLD
        and float(consensus["log_mad"]) <= RANSAC_LOG_TOLERANCE
        and float(confidence["score"]) >= HIGH_CONFIDENCE_THRESHOLD
    )
    return "HIGH" if high else "LOW"


def fit_anchor_model(raw_anchors: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    """Fit from anchor-only input; no scene or truth object is accepted."""
    anchors, audit = prepare_anchors(raw_anchors)

    ratio_records: list[dict[str, Any]] = []
    for anchor in anchors:
        if anchor["anchor_type"] not in ("DIM", "TEXT"):
            continue
        display = anchor["display_value"]
        if display is None or display <= 0.0:
            continue
        ratio_records.append(
            {
                **anchor,
                "log_ratio": math.log(float(display) / float(anchor["raw_span"])),
            }
        )

    ratio_consensus = maximum_consensus(
        ratio_records, "log_ratio", RANSAC_LOG_TOLERANCE
    )
    if ratio_consensus is None:
        scale_estimate = None
        ratio_confidence = None
        unit_status = "NONE"
        ratio_inliers: list[dict[str, Any]] = []
        ratio_outliers: list[dict[str, Any]] = []
    else:
        ratio_inliers = [
            ratio_records[index] for index in ratio_consensus["inlier_indices"]
        ]
        ratio_outliers = [
            ratio_records[index] for index in ratio_consensus["outlier_indices"]
        ]
        ratio_confidence = confidence_from_consensus(ratio_consensus, ratio_inliers)
        unit_status = _status(True, ratio_consensus, ratio_confidence)
        scale_estimate = math.exp(float(ratio_consensus["location"]))

    known_units = sorted(
        {
            str(anchor["display_unit"])
            for anchor in ratio_inliers
            if str(anchor["display_unit"]) in UNIT_TO_MM
        }
    )
    if len(known_units) == 1:
        physical_unit = known_units[0]
        mm_per_raw = float(scale_estimate) * UNIT_TO_MM[physical_unit]
    elif len(known_units) > 1:
        physical_unit = "MIXED"
        mm_per_raw = None
    else:
        physical_unit = "UNKNOWN"
        mm_per_raw = None

    reference_records: list[dict[str, Any]] = []
    reference_rejections: list[dict[str, Any]] = []
    for anchor in anchors:
        span = float(anchor["raw_span"])
        text_height = anchor["text_height"]
        if text_height is not None and text_height > 0.0 and span / text_height < 10.0:
            reference_rejections.append(
                {"handle": anchor["handle"], "reason": "annotation_scale_candidate"}
            )
            continue
        reference_records.append({**anchor, "log_span": math.log(span)})

    reference_consensus = maximum_consensus(
        reference_records, "log_span", RANSAC_LOG_TOLERANCE
    )
    if reference_consensus is None:
        reference_span = None
        reference_confidence = None
        reference_status = "NONE"
        reference_inliers: list[dict[str, Any]] = []
        reference_outliers: list[dict[str, Any]] = []
    else:
        reference_inliers = [
            reference_records[index]
            for index in reference_consensus["inlier_indices"]
        ]
        reference_outliers = [
            reference_records[index]
            for index in reference_consensus["outlier_indices"]
        ]
        reference_confidence = confidence_from_consensus(
            reference_consensus, reference_inliers
        )
        reference_status = _status(
            True, reference_consensus, reference_confidence
        )
        reference_span = math.exp(float(reference_consensus["location"]))

    status = reference_status
    source_diversity = sorted({anchor["anchor_type"] for anchor in anchors})
    model = {
        "status": status,
        "unit_status": unit_status,
        "reference_status": reference_status,
        "display_per_raw": scale_estimate,
        "physical_unit": physical_unit,
        "mm_per_raw": mm_per_raw,
        "consensus_weight": None
        if ratio_consensus is None
        else float(ratio_consensus["consensus_weight"]),
        "log_mad": None
        if ratio_consensus is None
        else float(ratio_consensus["log_mad"]),
        "n_independent": 0
        if ratio_confidence is None
        else int(ratio_confidence["n_independent"]),
        "n_spatial_bins": 0
        if ratio_confidence is None
        else int(ratio_confidence["n_spatial_bins"]),
        "confidence_score": 0.0
        if ratio_confidence is None
        else float(ratio_confidence["score"]),
        "source_diversity": source_diversity,
        "reference_span": reference_span,
        "reference_consensus_weight": None
        if reference_consensus is None
        else float(reference_consensus["consensus_weight"]),
        "reference_log_mad": None
        if reference_consensus is None
        else float(reference_consensus["log_mad"]),
        "reference_n_independent": 0
        if reference_confidence is None
        else int(reference_confidence["n_independent"]),
        "reference_n_spatial_bins": 0
        if reference_confidence is None
        else int(reference_confidence["n_spatial_bins"]),
        "reference_confidence_score": 0.0
        if reference_confidence is None
        else float(reference_confidence["score"]),
        "provenance": {
            **audit,
            "ratio_inlier_handles": [anchor["handle"] for anchor in ratio_inliers],
            "ratio_outlier_handles": [anchor["handle"] for anchor in ratio_outliers],
            "reference_inlier_handles": [
                anchor["handle"] for anchor in reference_inliers
            ],
            "reference_outlier_handles": [
                anchor["handle"] for anchor in reference_outliers
            ],
            "reference_rejections": reference_rejections,
        },
    }
    return model


def anchor_artifact_from_scene(scene: Mapping[str, Any]) -> dict[str, Any]:
    # This is the only evaluation-to-estimator bridge and it reads one key only.
    model = fit_anchor_model(copy.deepcopy(scene["anchors"]))
    return {
        "artifact_schema": "ariadne.e2.feyerabend_c1.anchor_artifact.v1",
        "model": model,
    }


def assigned_corruption(base_scene_id: str) -> str:
    digest = hashlib.sha256(base_scene_id.encode("utf-8")).digest()
    return CORRUPTIONS[int.from_bytes(digest[:4], "big") % len(CORRUPTIONS)]


def deterministic_index(base_scene_id: str, label: str, size: int) -> int:
    if size <= 0:
        raise ValueError("size must be positive")
    digest = hashlib.sha256(f"{base_scene_id}:{label}".encode("utf-8")).digest()
    return int.from_bytes(digest[:8], "big") % size


def apply_corruption(
    raw_anchors: Sequence[Mapping[str, Any]], kind: str, base_scene_id: str
) -> list[dict[str, Any]]:
    if kind not in CORRUPTIONS:
        raise ValueError(f"unknown corruption: {kind}")
    anchors = [copy.deepcopy(dict(anchor)) for anchor in raw_anchors]
    ratio_indices = [
        index
        for index, anchor in enumerate(anchors)
        if str(anchor.get("anchor_type", "")).upper() in ("DIM", "TEXT")
        and finite_float(anchor.get("display_value")) not in (None, 0.0)
    ]
    if kind == "duplicate":
        if not anchors:
            return anchors
        index = deterministic_index(base_scene_id, kind, len(anchors))
        clone = copy.deepcopy(anchors[index])
        clone["handle"] = f"{clone.get('handle', 'ANCHOR')}__DUP_{canonical_sha256(clone)[:8]}"
        clone["diagnostic_mutation"] = kind
        anchors.append(clone)
    elif kind == "stale_override":
        if not ratio_indices:
            return anchors
        index = ratio_indices[deterministic_index(base_scene_id, kind, len(ratio_indices))]
        anchors[index]["display_value"] = float(anchors[index]["display_value"]) * 2.0
        anchors[index]["text_override"] = str(anchors[index]["display_value"])
        anchors[index]["diagnostic_mutation"] = kind
    elif kind == "suffix_removal":
        for index in ratio_indices:
            anchors[index]["display_unit"] = "UNKNOWN"
            anchors[index]["suffix_removed"] = True
            anchors[index]["diagnostic_mutation"] = kind
    elif kind == "single_outlier":
        if not ratio_indices:
            return anchors
        source_index = ratio_indices[
            deterministic_index(base_scene_id, kind, len(ratio_indices))
        ]
        outlier = copy.deepcopy(anchors[source_index])
        digest = hashlib.sha256(f"{base_scene_id}:{kind}".encode("utf-8")).hexdigest()
        shift = 10_000.0 + float(int(digest[:6], 16) % 10_000)
        for key in ("p0", "p1"):
            point = list(outlier[key])
            point[0] = float(point[0]) + shift
            point[1] = float(point[1]) + 0.5 * shift
            outlier[key] = point
        outlier["handle"] = f"{outlier.get('handle', 'ANCHOR')}__OUT_{digest[:8]}"
        outlier["display_value"] = float(outlier["display_value"]) * 10.0
        outlier["weight"] = 1.0
        outlier["diagnostic_mutation"] = kind
        anchors.append(outlier)
    return anchors


def permute_pair_labels(scene: Mapping[str, Any]) -> dict[str, Any]:
    mutated = copy.deepcopy(dict(scene))
    rows = [copy.deepcopy(dict(row)) for row in mutated.get("truth_pairs", [])]
    if not rows:
        mutated["truth_pairs"] = rows
        return mutated
    label_payloads = [
        {
            "pair_id": row.get("pair_id"),
            "handles": copy.deepcopy(row.get("handles")),
            "wall_id": row.get("wall_id"),
        }
        for row in rows
    ]
    shift = 1 if len(rows) > 1 else 0
    for index, row in enumerate(rows):
        payload = label_payloads[(index + shift) % len(rows)]
        row["pair_id"] = payload["pair_id"]
        row["handles"] = copy.deepcopy(payload["handles"])
        row["wall_id"] = f"PERMUTED::{payload['wall_id']}"
    mutated["truth_pairs"] = list(reversed(rows))
    return mutated


class GuardedScene(Mapping[str, Any]):
    def __init__(self, anchors: Sequence[Mapping[str, Any]]):
        self._anchors = copy.deepcopy(list(anchors))
        self.accessed: list[str] = []

    def __getitem__(self, key: str) -> Any:
        self.accessed.append(key)
        if key != "anchors":
            raise AssertionError(f"estimator bridge attempted forbidden key: {key}")
        return copy.deepcopy(self._anchors)

    def __iter__(self):
        return iter(("anchors",))

    def __len__(self) -> int:
        return 1


def exact_fixture(scale: float = 2.5) -> list[dict[str, Any]]:
    origins = ((0.0, 0.0), (200.0, 0.0), (0.0, 200.0), (200.0, 200.0), (100.0, 100.0))
    return [
        {
            "handle": f"FIXTURE_DIM_{index}",
            "anchor_type": "DIM",
            "region": f"R{index}",
            "p0": [x, y],
            "p1": [x + 100.0, y],
            "raw_span": 100.0,
            "display_value": 100.0 * scale,
            "display_unit": "MM",
            "text_height": 5.0,
            "weight": 1.0,
        }
        for index, (x, y) in enumerate(origins)
    ]


def run_selftests(stream: io.TextIOBase) -> dict[str, Any]:
    tests: list[dict[str, Any]] = []

    def check(name: str, condition: bool, detail: str) -> None:
        status = "PASS" if condition else "FAIL"
        print(f"SELFTEST {name}: {status} | {detail}", file=stream)
        tests.append({"name": name, "passed": bool(condition), "detail": detail})
        if not condition:
            raise AssertionError(f"selftest failed: {name}: {detail}")

    fixture = exact_fixture()
    exact_model = fit_anchor_model(fixture)
    estimate = exact_model["display_per_raw"]
    check(
        "exact_anchor_exact_scale",
        estimate is not None
        and math.isclose(float(estimate), 2.5, rel_tol=1e-12, abs_tol=1e-12),
        f"estimate={estimate!r} expected=2.5 unit_status={exact_model['unit_status']}",
    )
    check(
        "exact_anchor_high_confidence",
        exact_model["unit_status"] == "HIGH",
        f"unit_status={exact_model['unit_status']} confidence={exact_model['confidence_score']:.12g}",
    )

    empty_model = fit_anchor_model([])
    check(
        "no_anchor_honest_no_estimate",
        empty_model["display_per_raw"] is None
        and empty_model["unit_status"] == "NONE"
        and empty_model["status"] == "NONE",
        f"estimate={empty_model['display_per_raw']!r} unit_status={empty_model['unit_status']} status={empty_model['status']}",
    )

    reproducible = True
    mutation_digests: dict[str, str] = {}
    for kind in CORRUPTIONS:
        first = apply_corruption(fixture, kind, "fixture_base")
        second = apply_corruption(fixture, kind, "fixture_base")
        first_digest = canonical_sha256(first)
        second_digest = canonical_sha256(second)
        reproducible = reproducible and first_digest == second_digest
        mutation_digests[kind] = first_digest
    check(
        "corruption_reproducibility",
        reproducible,
        " ".join(f"{key}={value[:12]}" for key, value in mutation_digests.items()),
    )

    outlier_model = fit_anchor_model(
        apply_corruption(fixture, "single_outlier", "fixture_base")
    )
    check(
        "single_outlier_mode_or_downgrade",
        (
            outlier_model["display_per_raw"] is not None
            and math.isclose(
                float(outlier_model["display_per_raw"]), 2.5, rel_tol=1e-12, abs_tol=1e-12
            )
        )
        or outlier_model["unit_status"] == "LOW",
        f"estimate={outlier_model['display_per_raw']!r} unit_status={outlier_model['unit_status']}",
    )

    guarded = GuardedScene(fixture)
    guarded_artifact = anchor_artifact_from_scene(guarded)
    check(
        "truth_key_access_guard",
        guarded.accessed == ["anchors"] and guarded_artifact["model"]["display_per_raw"] is not None,
        f"accessed_keys={guarded.accessed!r}",
    )

    passed = sum(1 for test in tests if test["passed"])
    print(f"SELFTEST SUMMARY: {passed}/{len(tests)} passed", file=stream)
    return {"passed": passed, "total": len(tests), "tests": tests}


def input_manifest(paths: Sequence[Path]) -> dict[str, Any]:
    rows = [
        {"name": path.name, "size": path.stat().st_size, "sha256": sha256_file(path)}
        for path in paths
    ]
    return {"files": rows, "digest": canonical_sha256(rows)}


def load_and_validate_inputs() -> tuple[list[dict[str, Any]], list[Path], dict[str, Any]]:
    paths = sorted(INPUT_DIR.glob("scene_*.json"))
    errors: list[str] = []
    if len(paths) != EXPECTED_SCENE_COUNT:
        errors.append(f"expected {EXPECTED_SCENE_COUNT} scene files, found {len(paths)}")
    scenes: list[dict[str, Any]] = []
    for path in paths:
        with path.open("r", encoding="utf-8") as stream:
            scene = json.load(stream)
        if scene.get("schema") != INPUT_SCHEMA:
            errors.append(f"{path.name}: unexpected schema {scene.get('schema')!r}")
        if not isinstance(scene.get("anchors"), list):
            errors.append(f"{path.name}: anchors is not a list")
        kappa = finite_float(scene.get("scale_kappa"))
        truth_scale = finite_float(scene.get("truth_unit_scale"))
        if kappa is None or truth_scale is None or kappa <= 0.0:
            errors.append(f"{path.name}: invalid scale fields")
        elif not math.isclose(truth_scale, 1.0 / kappa, rel_tol=1e-12, abs_tol=0.0):
            errors.append(f"{path.name}: truth_unit_scale is not 1/kappa")
        scene["_input_file"] = path.name
        scenes.append(scene)

    base_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for scene in scenes:
        base_groups[str(scene.get("base_scene_id"))].append(scene)
    if len(base_groups) != EXPECTED_BASE_COUNT:
        errors.append(f"expected {EXPECTED_BASE_COUNT} base scenes, found {len(base_groups)}")
    expected_scale_tokens = {scale_token(scale) for scale in SCALES}
    for base_scene_id, group in sorted(base_groups.items()):
        observed = {scale_token(float(scene["scale_kappa"])) for scene in group}
        if observed != expected_scale_tokens:
            errors.append(f"{base_scene_id}: scale set {sorted(observed)}")
        seeds = {int(scene["seed"]) for scene in group}
        if len(seeds) != 1:
            errors.append(f"{base_scene_id}: paired copies do not share one seed")
        assigned = {assigned_corruption(str(scene["base_scene_id"])) for scene in group}
        if len(assigned) != 1:
            errors.append(f"{base_scene_id}: assigned corruption is not paired")
    if errors:
        raise RuntimeError("input validation failed:\n" + "\n".join(errors))
    validation = {
        "error_count": 0,
        "scene_count": len(scenes),
        "base_scene_count": len(base_groups),
        "seed_count": len({int(scene["seed"]) for scene in scenes}),
        "scale_counts": dict(
            sorted(
                Counter(scale_token(float(scene["scale_kappa"])) for scene in scenes).items()
            )
        ),
    }
    return scenes, paths, validation


def relative_error(estimate: float | None, truth: float) -> float | None:
    if estimate is None:
        return None
    return abs(float(estimate) / float(truth) - 1.0)


def log_error(estimate: float | None, truth: float) -> float | None:
    if estimate is None:
        return None
    return abs(math.log(float(estimate) / float(truth)))


def numeric_distribution(values: Iterable[float | None]) -> dict[str, Any]:
    array = np.asarray([float(value) for value in values if value is not None], dtype=float)
    if array.size == 0:
        return {
            "count": 0,
            "min": None,
            "p05": None,
            "p25": None,
            "median": None,
            "mean": None,
            "p75": None,
            "p95": None,
            "max": None,
        }
    return {
        "count": int(array.size),
        "min": float(np.min(array)),
        "p05": float(np.quantile(array, 0.05)),
        "p25": float(np.quantile(array, 0.25)),
        "median": float(np.median(array)),
        "mean": float(np.mean(array)),
        "p75": float(np.quantile(array, 0.75)),
        "p95": float(np.quantile(array, 0.95)),
        "max": float(np.max(array)),
    }


def summarize_rows(rows: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    estimate_rows = [row for row in rows if row["scale_estimate"] is not None]
    accurate = [
        row
        for row in estimate_rows
        if float(row["relative_error"]) <= ACCURACY_RELATIVE_ERROR
    ]
    return {
        "scene_count": len(rows),
        "estimate_count": len(estimate_rows),
        "estimate_coverage": len(estimate_rows) / len(rows) if rows else None,
        "accuracy_within_5pct": len(accurate) / len(estimate_rows)
        if estimate_rows
        else None,
        "e_s": numeric_distribution(row["e_s"] for row in estimate_rows),
        "relative_error": numeric_distribution(
            row["relative_error"] for row in estimate_rows
        ),
        "confidence_score": numeric_distribution(
            row["confidence_score"] for row in rows
        ),
    }


def score_bin_for(score: float) -> str:
    for label, lower, upper, inclusive_upper in SCORE_BINS:
        if score >= lower and (score <= upper if inclusive_upper else score < upper):
            return label
    raise AssertionError(f"confidence score outside [0,1]: {score}")


def evaluate_scene(scene: Mapping[str, Any]) -> dict[str, Any]:
    artifact_before = anchor_artifact_from_scene(scene)
    model = artifact_before["model"]
    truth_scale = float(scene["truth_unit_scale"])
    estimate = model["display_per_raw"]

    permuted = permute_pair_labels(scene)
    artifact_after = anchor_artifact_from_scene(permuted)
    before_digest = canonical_sha256(artifact_before)
    after_digest = canonical_sha256(artifact_after)
    truth_before_digest = canonical_sha256(scene.get("truth_pairs", []))
    truth_after_digest = canonical_sha256(permuted.get("truth_pairs", []))

    corruption_rows: dict[str, Any] = {}
    for kind in CORRUPTIONS:
        corrupted_anchors = apply_corruption(
            scene["anchors"], kind, str(scene["base_scene_id"])
        )
        corrupted_model = fit_anchor_model(corrupted_anchors)
        corrupted_estimate = corrupted_model["display_per_raw"]
        same_estimate = (
            estimate is None
            and corrupted_estimate is None
            or estimate is not None
            and corrupted_estimate is not None
            and math.isclose(
                float(estimate), float(corrupted_estimate), rel_tol=1e-12, abs_tol=1e-15
            )
        )
        corruption_rows[kind] = {
            "status_before": model["status"],
            "status_after": corrupted_model["status"],
            "unit_status_before": model["unit_status"],
            "unit_status_after": corrupted_model["unit_status"],
            "reference_status_before": model["reference_status"],
            "reference_status_after": corrupted_model["reference_status"],
            "physical_unit_before": model["physical_unit"],
            "physical_unit_after": corrupted_model["physical_unit"],
            "scale_estimate_before": estimate,
            "scale_estimate_after": corrupted_estimate,
            "scale_estimate_unchanged": same_estimate,
            "relative_error_after": relative_error(corrupted_estimate, truth_scale),
            "e_s_after": log_error(corrupted_estimate, truth_scale),
            "confidence_score_before": model["confidence_score"],
            "confidence_score_after": corrupted_model["confidence_score"],
            "artifact_digest": canonical_sha256(
                {
                    "artifact_schema": artifact_before["artifact_schema"],
                    "model": corrupted_model,
                }
            ),
            "input_anchor_count_after": len(corrupted_anchors),
        }

    score = float(model["confidence_score"])
    row = {
        "input_file": scene["_input_file"],
        "scene_id": scene["scene_id"],
        "base_scene_id": scene["base_scene_id"],
        "seed": int(scene["seed"]),
        "scale_kappa": float(scene["scale_kappa"]),
        "truth_unit_scale": truth_scale,
        "input_anchor_count": len(scene["anchors"]),
        "input_anchor_types": sorted(
            {str(anchor.get("anchor_type")) for anchor in scene["anchors"]}
        ),
        "scale_estimate": estimate,
        "e_s": log_error(estimate, truth_scale),
        "relative_error": relative_error(estimate, truth_scale),
        "status": model["status"],
        "unit_status": model["unit_status"],
        "reference_status": model["reference_status"],
        "confidence_score": score,
        "confidence_score_bin": score_bin_for(score),
        "physical_unit": model["physical_unit"],
        "anchor_model": model,
        "assigned_corruption": assigned_corruption(str(scene["base_scene_id"])),
        "corruption_diagnostics": corruption_rows,
        "pair_label_permutation": {
            "truth_label_digest_before": truth_before_digest,
            "truth_label_digest_after": truth_after_digest,
            "truth_label_input_changed": truth_before_digest != truth_after_digest,
            "anchor_artifact_digest_before": before_digest,
            "anchor_artifact_digest_after": after_digest,
            "anchor_artifact_digest_equal": before_digest == after_digest,
        },
    }
    return row


def transition_counter(
    rows: Sequence[Mapping[str, Any]], kind: str, prefix: str
) -> dict[str, int]:
    counter: Counter[str] = Counter()
    before_key = f"{prefix}_before"
    after_key = f"{prefix}_after"
    for row in rows:
        diagnostic = row["corruption_diagnostics"][kind]
        counter[f"{diagnostic[before_key]}->{diagnostic[after_key]}"] += 1
    return dict(sorted(counter.items()))


def corruption_summary(rows: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    output: dict[str, Any] = {}
    for kind in CORRUPTIONS:
        diagnostics = [row["corruption_diagnostics"][kind] for row in rows]
        output[kind] = {
            "scene_count": len(rows),
            "status_transitions": transition_counter(rows, kind, "status"),
            "unit_status_transitions": transition_counter(rows, kind, "unit_status"),
            "reference_status_transitions": transition_counter(
                rows, kind, "reference_status"
            ),
            "physical_unit_transitions": dict(
                sorted(
                    Counter(
                        f"{diagnostic['physical_unit_before']}->{diagnostic['physical_unit_after']}"
                        for diagnostic in diagnostics
                    ).items()
                )
            ),
            "scale_estimate_unchanged_count": sum(
                bool(diagnostic["scale_estimate_unchanged"])
                for diagnostic in diagnostics
            ),
            "scale_estimate_changed_count": sum(
                not bool(diagnostic["scale_estimate_unchanged"])
                for diagnostic in diagnostics
            ),
            "e_s_after": numeric_distribution(
                diagnostic["e_s_after"] for diagnostic in diagnostics
            ),
            "relative_error_after": numeric_distribution(
                diagnostic["relative_error_after"] for diagnostic in diagnostics
            ),
            "confidence_score_after": numeric_distribution(
                diagnostic["confidence_score_after"] for diagnostic in diagnostics
            ),
        }
    assigned_rows: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    for row in rows:
        assigned_rows[str(row["assigned_corruption"])].append(row)
    assigned: dict[str, Any] = {}
    for kind in CORRUPTIONS:
        subset = assigned_rows.get(kind, [])
        assigned[kind] = {
            "scene_count": len(subset),
            "base_scene_count": len({row["base_scene_id"] for row in subset}),
            "unit_status_transitions": transition_counter(subset, kind, "unit_status")
            if subset
            else {},
        }
    return {"all_four_applied": output, "hash_assigned_primary": assigned}


def aggregate_results(rows: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    overall = summarize_rows(rows)
    high_rows = [row for row in rows if row["unit_status"] == "HIGH"]
    overall["high_scene_count"] = len(high_rows)
    overall["high_coverage"] = len(high_rows) / len(rows) if rows else None
    overall["high_accuracy_within_5pct"] = summarize_rows(high_rows)[
        "accuracy_within_5pct"
    ]
    overall["high_relative_error"] = numeric_distribution(
        row["relative_error"] for row in high_rows
    )
    overall["high_e_s"] = numeric_distribution(row["e_s"] for row in high_rows)
    overall["unit_status_counts"] = dict(
        sorted(Counter(str(row["unit_status"]) for row in rows).items())
    )
    overall["status_counts"] = dict(
        sorted(Counter(str(row["status"]) for row in rows).items())
    )

    by_scale: dict[str, Any] = {}
    scale_confidence_rows: list[dict[str, Any]] = []
    for scale in SCALES:
        token = scale_token(scale)
        scale_rows = [
            row
            for row in rows
            if math.isclose(float(row["scale_kappa"]), scale, rel_tol=0.0, abs_tol=0.0)
        ]
        summary = summarize_rows(scale_rows)
        scale_high = [row for row in scale_rows if row["unit_status"] == "HIGH"]
        summary["high_scene_count"] = len(scale_high)
        summary["high_coverage"] = len(scale_high) / len(scale_rows)
        summary["high_accuracy_within_5pct"] = summarize_rows(scale_high)[
            "accuracy_within_5pct"
        ]
        summary["high_relative_error"] = numeric_distribution(
            row["relative_error"] for row in scale_high
        )
        summary["unit_status_counts"] = dict(
            sorted(Counter(str(row["unit_status"]) for row in scale_rows).items())
        )
        by_scale[token] = summary
        for confidence in CONFIDENCE_LEVELS:
            subset = [row for row in scale_rows if row["unit_status"] == confidence]
            subset_summary = summarize_rows(subset)
            scale_confidence_rows.append(
                {
                    "scale_kappa": scale,
                    "confidence": confidence,
                    "scene_count": len(subset),
                    "fraction_of_scale": len(subset) / len(scale_rows),
                    **{key: value for key, value in subset_summary.items() if key != "scene_count"},
                }
            )

    confidence_bins = {
        confidence: summarize_rows(
            [row for row in rows if row["unit_status"] == confidence]
        )
        for confidence in CONFIDENCE_LEVELS
    }
    score_calibration_rows: list[dict[str, Any]] = []
    for scale_scope in ("ALL", *(scale_token(scale) for scale in SCALES)):
        scoped = (
            list(rows)
            if scale_scope == "ALL"
            else [
                row
                for row in rows
                if scale_token(float(row["scale_kappa"])) == scale_scope
            ]
        )
        for label, lower, upper, inclusive_upper in SCORE_BINS:
            subset = [row for row in scoped if row["confidence_score_bin"] == label]
            summary = summarize_rows(subset)
            score_calibration_rows.append(
                {
                    "scale_kappa": scale_scope,
                    "score_bin": label,
                    "lower": lower,
                    "upper": upper,
                    "upper_inclusive": inclusive_upper,
                    **summary,
                }
            )

    before_digests = [
        row["pair_label_permutation"]["anchor_artifact_digest_before"] for row in rows
    ]
    after_digests = [
        row["pair_label_permutation"]["anchor_artifact_digest_after"] for row in rows
    ]
    pair_digest = {
        "scene_count": len(rows),
        "pair_label_changed_scene_count": sum(
            bool(row["pair_label_permutation"]["truth_label_input_changed"])
            for row in rows
        ),
        "matching_anchor_artifact_scene_count": sum(
            bool(row["pair_label_permutation"]["anchor_artifact_digest_equal"])
            for row in rows
        ),
        "mismatching_anchor_artifact_scene_count": sum(
            not bool(row["pair_label_permutation"]["anchor_artifact_digest_equal"])
            for row in rows
        ),
        "anchor_artifact_match_rate": sum(
            bool(row["pair_label_permutation"]["anchor_artifact_digest_equal"])
            for row in rows
        )
        / len(rows),
        "global_anchor_artifact_digest_before": canonical_sha256(before_digests),
        "global_anchor_artifact_digest_after": canonical_sha256(after_digests),
    }
    return {
        "overall": overall,
        "by_scale": by_scale,
        "confidence_grade_bins": confidence_bins,
        "scale_confidence_rows": scale_confidence_rows,
        "score_calibration_rows": score_calibration_rows,
        "corruption": corruption_summary(rows),
        "pair_label_permutation": pair_digest,
    }


def flatten_distribution(prefix: str, distribution: Mapping[str, Any]) -> dict[str, Any]:
    return {f"{prefix}_{key}": value for key, value in distribution.items()}


def workbook_rows(aggregates: Mapping[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in aggregates["scale_confidence_rows"]:
        rows.append(
            {
                "scale_kappa": row["scale_kappa"],
                "confidence": row["confidence"],
                "scene_count": row["scene_count"],
                "fraction_of_scale": row["fraction_of_scale"],
                "estimate_count": row["estimate_count"],
                "estimate_coverage": row["estimate_coverage"],
                "accuracy_within_5pct": row["accuracy_within_5pct"],
                **flatten_distribution("e_s", row["e_s"]),
                **flatten_distribution("relative_error", row["relative_error"]),
                **flatten_distribution("confidence_score", row["confidence_score"]),
            }
        )
    return rows


def workbook_score_rows(aggregates: Mapping[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in aggregates["score_calibration_rows"]:
        rows.append(
            {
                "scale_kappa": row["scale_kappa"],
                "score_bin": row["score_bin"],
                "lower": row["lower"],
                "upper": row["upper"],
                "upper_inclusive": row["upper_inclusive"],
                "scene_count": row["scene_count"],
                "estimate_count": row["estimate_count"],
                "estimate_coverage": row["estimate_coverage"],
                "accuracy_within_5pct": row["accuracy_within_5pct"],
                **flatten_distribution("e_s", row["e_s"]),
                **flatten_distribution("relative_error", row["relative_error"]),
                **flatten_distribution("confidence_score", row["confidence_score"]),
            }
        )
    return rows


def write_workbook(aggregates: Mapping[str, Any]) -> dict[str, Any]:
    rows = workbook_rows(aggregates)
    score_rows = workbook_score_rows(aggregates)
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        fallback = {
            "schema": "ariadne.e2.feyerabend_c1.evidence_rows.v1",
            "blocked_reason": "BLOCKED_XLSX: openpyxl unavailable",
            "scale_confidence_rows": rows,
            "score_calibration_rows": aggregates["score_calibration_rows"],
            "corruption": aggregates["corruption"],
            "pair_label_permutation": aggregates["pair_label_permutation"],
        }
        write_json_atomic(EVIDENCE_FALLBACK_PATH, fallback)
        return {
            "status": "BLOCKED_XLSX",
            "backend": None,
            "path": str(EVIDENCE_FALLBACK_PATH),
            "row_count": len(rows),
        }

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "scale_confidence"

    def append_table(target: Any, table_rows: Sequence[Mapping[str, Any]]) -> None:
        if not table_rows:
            target.append(["NO_ROWS"])
            return
        headers = list(table_rows[0].keys())
        target.append(headers)
        for cell in target[1]:
            cell.font = Font(bold=True)
        for item in table_rows:
            target.append([item.get(header) for header in headers])
        target.freeze_panes = "A2"
        target.auto_filter.ref = target.dimensions
        for column in target.columns:
            width = min(36, max(len(str(cell.value or "")) for cell in column) + 2)
            target.column_dimensions[column[0].column_letter].width = width

    append_table(sheet, rows)
    append_table(workbook.create_sheet("score_bins"), score_rows)

    transition_rows: list[dict[str, Any]] = []
    for kind, summary in aggregates["corruption"]["all_four_applied"].items():
        for dimension in (
            "status_transitions",
            "unit_status_transitions",
            "reference_status_transitions",
            "physical_unit_transitions",
        ):
            for transition, count in summary[dimension].items():
                transition_rows.append(
                    {
                        "corruption": kind,
                        "dimension": dimension,
                        "transition": transition,
                        "count": count,
                        "fraction": count / summary["scene_count"],
                    }
                )
    append_table(workbook.create_sheet("corruption_transitions"), transition_rows)

    digest_rows = [
        {"metric": key, "value": value}
        for key, value in aggregates["pair_label_permutation"].items()
    ]
    append_table(workbook.create_sheet("pair_label_digest"), digest_rows)
    readme_rows = [
        {"key": "schema", "value": "ariadne.e2.feyerabend_c1.evidence.v1"},
        {"key": "primary_confidence", "value": "unit_status"},
        {"key": "accuracy_definition", "value": "abs(scale_estimate / truth_unit_scale - 1) <= 0.05"},
        {"key": "high_threshold", "value": HIGH_CONFIDENCE_THRESHOLD},
        {"key": "ransac_log_tolerance", "value": RANSAC_LOG_TOLERANCE},
        {"key": "verdicts", "value": "not emitted by C1"},
    ]
    append_table(workbook.create_sheet("README"), readme_rows)

    temporary = EVIDENCE_PATH.with_suffix(".xlsx.tmp")
    workbook.save(temporary)
    temporary.replace(EVIDENCE_PATH)
    if EVIDENCE_FALLBACK_PATH.exists():
        EVIDENCE_FALLBACK_PATH.unlink()
    return {
        "status": "GENERATED",
        "backend": f"openpyxl {openpyxl.__version__}",
        "path": str(EVIDENCE_PATH),
        "row_count": len(rows),
        "sha256": sha256_file(EVIDENCE_PATH),
    }


def write_json_atomic(path: Path, payload: Any) -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    with temporary.open("w", encoding="utf-8", newline="\n") as stream:
        json.dump(payload, stream, indent=2, ensure_ascii=False, allow_nan=False)
        stream.write("\n")
    temporary.replace(path)


def format_number(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return format(value, ".12g")
    return str(value)


def markdown_table(headers: Sequence[str], rows: Sequence[Sequence[Any]]) -> str:
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(format_number(value) for value in row) + " |")
    return "\n".join(lines)


def render_report(results: Mapping[str, Any], selftest_transcript: str) -> str:
    aggregates = results["aggregates"]
    overall = aggregates["overall"]
    lines: list[str] = [
        "# Feyerabend C1 — Anchor scale estimation and confidence calibration",
        "",
        "## 실행 계약과 설계",
        "",
        f"- 입력: `{results['inputs']['directory']}`의 C0 IR {results['inputs']['validation']['scene_count']}개, base seed {results['inputs']['validation']['seed_count']}개.",
        "- estimator 경계: `anchor_artifact_from_scene`은 `anchors` 키 하나만 읽고, `fit_anchor_model`은 anchor sequence만 인자로 받는다. `truth_unit_scale`은 추정 뒤 오차 계산에만 사용했다.",
        "- scale 추정: DIM/TEXT의 `display_value / geometric_span` log-ratio에서 `log(1.05)` 최대 가중 합의군을 고르고 Huber location(delta 1.5)을 계산했다.",
        "- confidence: dossier의 consensus × exp(-logMAD/tau) × min(1,n/5) × min(1,spatial_bins/3), HIGH threshold 0.75를 그대로 사용했다. C1의 primary confidence bin은 `unit_status`다.",
        "- corruption: 모든 200 IR에 duplicate, stale override(한 표시값 ×2), suffix removal(명시 suffix 전부 제거), single outlier(독립 표시값 ×10) 네 종류를 각각 적용했다. 별도로 `sha256(base_scene_id) mod 4` 배정도 기록했다.",
        "- pair-label permutation은 `truth_pairs`의 label payload를 회전·재표기한 뒤 동일 anchor artifact를 다시 산출했다.",
        "- 합격선 또는 셀/theory 판정은 산출하지 않았다.",
        "",
        "## Selftest 전문",
        "",
        "```text",
        selftest_transcript.rstrip("\n"),
        "```",
        "",
        "## 입력 및 실행 수치",
        "",
        markdown_table(
            ["metric", "value"],
            [
                ("scene_count", results["inputs"]["validation"]["scene_count"]),
                ("base_scene_count", results["inputs"]["validation"]["base_scene_count"]),
                ("seed_count", results["inputs"]["validation"]["seed_count"]),
                ("input_manifest_digest_before", results["inputs"]["manifest_before"]["digest"]),
                ("input_manifest_digest_after", results["inputs"]["manifest_after"]["digest"]),
                ("input_manifest_mismatch_count", results["inputs"]["manifest_mismatch_count"]),
                ("elapsed_cpu_seconds", results["runtime"]["cpu_seconds"]),
                ("elapsed_wall_seconds", results["runtime"]["wall_seconds"]),
            ],
        ),
        "",
        "## 전체 scale 추정 수치",
        "",
        markdown_table(
            ["metric", "value"],
            [
                ("estimate_count", overall["estimate_count"]),
                ("estimate_coverage", overall["estimate_coverage"]),
                ("accuracy_within_5pct", overall["accuracy_within_5pct"]),
                ("HIGH_scene_count", overall["high_scene_count"]),
                ("HIGH_coverage", overall["high_coverage"]),
                ("HIGH_accuracy_within_5pct", overall["high_accuracy_within_5pct"]),
                ("e_s_min", overall["e_s"]["min"]),
                ("e_s_median", overall["e_s"]["median"]),
                ("e_s_p95", overall["e_s"]["p95"]),
                ("e_s_max", overall["e_s"]["max"]),
                ("relative_error_min", overall["relative_error"]["min"]),
                ("relative_error_median", overall["relative_error"]["median"]),
                ("relative_error_p95", overall["relative_error"]["p95"]),
                ("relative_error_max", overall["relative_error"]["max"]),
            ],
        ),
        "",
        "## Scale × confidence 전 행",
        "",
    ]
    scale_rows = []
    for row in aggregates["scale_confidence_rows"]:
        scale_rows.append(
            (
                row["scale_kappa"],
                row["confidence"],
                row["scene_count"],
                row["fraction_of_scale"],
                row["estimate_count"],
                row["accuracy_within_5pct"],
                row["e_s"]["median"],
                row["e_s"]["p95"],
                row["relative_error"]["median"],
                row["relative_error"]["p95"],
                row["confidence_score"]["median"],
            )
        )
    lines.extend(
        [
            markdown_table(
                [
                    "kappa",
                    "unit_status",
                    "n",
                    "fraction",
                    "n_est",
                    "accuracy_5pct",
                    "e_s_med",
                    "e_s_p95",
                    "relerr_med",
                    "relerr_p95",
                    "conf_med",
                ],
                scale_rows,
            ),
            "",
            "## Numeric confidence-score bin별 accuracy",
            "",
        ]
    )
    score_rows = []
    for row in aggregates["score_calibration_rows"]:
        score_rows.append(
            (
                row["scale_kappa"],
                row["score_bin"],
                row["scene_count"],
                row["estimate_count"],
                row["accuracy_within_5pct"],
                row["relative_error"]["median"],
                row["relative_error"]["p95"],
            )
        )
    lines.extend(
        [
            markdown_table(
                ["scale", "score_bin", "n", "n_est", "accuracy_5pct", "relerr_med", "relerr_p95"],
                score_rows,
            ),
            "",
            "## Corruption 전후 수치",
            "",
        ]
    )
    corruption_rows = []
    for kind, summary in aggregates["corruption"]["all_four_applied"].items():
        corruption_rows.append(
            (
                kind,
                summary["scene_count"],
                json.dumps(summary["unit_status_transitions"], sort_keys=True),
                json.dumps(summary["status_transitions"], sort_keys=True),
                summary["scale_estimate_unchanged_count"],
                summary["scale_estimate_changed_count"],
                summary["relative_error_after"]["median"],
                summary["relative_error_after"]["p95"],
                summary["confidence_score_after"]["median"],
            )
        )
    lines.extend(
        [
            markdown_table(
                [
                    "corruption",
                    "n",
                    "unit_transition",
                    "overall_transition",
                    "scale_same",
                    "scale_changed",
                    "relerr_med",
                    "relerr_p95",
                    "conf_after_med",
                ],
                corruption_rows,
            ),
            "",
            "### Hash-assigned corruption 분포",
            "",
            markdown_table(
                ["corruption", "scene_count", "base_scene_count", "unit_transitions"],
                [
                    (
                        kind,
                        summary["scene_count"],
                        summary["base_scene_count"],
                        json.dumps(summary["unit_status_transitions"], sort_keys=True),
                    )
                    for kind, summary in aggregates["corruption"]["hash_assigned_primary"].items()
                ],
            ),
            "",
            "## Pair-label permutation digest 수치",
            "",
            markdown_table(
                ["metric", "value"],
                list(aggregates["pair_label_permutation"].items()),
            ),
            "",
            "## 산출물 검증 수치",
            "",
            markdown_table(
                ["artifact", "status", "bytes", "sha256"],
                [
                    (
                        item["name"],
                        item["status"],
                        item.get("bytes"),
                        item.get("sha256"),
                    )
                    for item in results["artifacts"]
                ],
            ),
            "",
            "## 미해결",
            "",
            "- dossier confidence 식의 `min(1,n/5)` 항 때문에 C0의 독립 DIM ratio anchor 3개는 완전 합의·zero MAD에서도 confidence 0.60이며 HIGH threshold 0.75에 닿지 않는다. HIGH subset 수치가 비어 있으면 null로 보존했다.",
            "- C0 IR은 parser 이전의 정규화된 anchor schema다. stale override와 suffix 제거 diagnostic은 각각 `display_value` 및 `display_unit` 정규화 필드에서 deterministic mutation으로 구현했다.",
            "- single-outlier diagnostic에서 scale estimate는 200/200 동일했고 unit_status도 200/200 LOW 유지였지만, 추가 anchor의 geometry span이 reference 독립성 수를 늘려 reference/overall status가 26/200에서 LOW→HIGH로 변했다. 이 transition은 수치 그대로 보존했다.",
            "- 첫 full-process 시도는 200-scene 계산 뒤 nested distribution을 Excel cell로 직렬화하는 단계에서 중단됐다. estimator/config/metric은 바꾸지 않고 workbook 행만 평탄화한 뒤 동일 deterministic 평가를 재실행했다.",
            "- 첫 successful export의 독립 검증에서 `results.json` 안의 self-hash가 final rewrite 전 값을 가리키는 순환 manifest 문제가 발견됐다. 수치 계산은 바꾸지 않고 finalized results를 REPORT가 단방향으로 hash하도록 export finalization만 수정했다.",
            "- packet을 처음 읽기 위한 병렬 진단에 read-only `git status`가 잘못 포함되었고, target cell이 아닌 기본 `D:\\dev`에서 `not a git repository`로 종료했다. Git 상태와 파일은 바뀌지 않았으며 packet 확인 뒤 Git 명령을 다시 실행하지 않았다.",
            "- 이 셀은 수치와 진단만 산출하며 제안 합격선, C2 개방 여부, reigning/counter theory 판정을 출력하지 않는다.",
            "",
            "CELL_COMPLETE: feyerabend_c1",
        ]
    )
    return "\n".join(lines) + "\n"


def verify_workbook(expected_rows: int) -> dict[str, Any]:
    if EVIDENCE_PATH.exists():
        import openpyxl

        workbook = openpyxl.load_workbook(EVIDENCE_PATH, read_only=True, data_only=True)
        required = {
            "scale_confidence",
            "score_bins",
            "corruption_transitions",
            "pair_label_digest",
            "README",
        }
        missing = sorted(required - set(workbook.sheetnames))
        sheet = workbook["scale_confidence"]
        observed_rows = max(0, sheet.max_row - 1)
        workbook.close()
        if missing or observed_rows != expected_rows:
            raise RuntimeError(
                f"workbook verification failed: missing={missing}, rows={observed_rows}, expected={expected_rows}"
            )
        return {
            "sheet_count": len(required),
            "scale_confidence_data_rows": observed_rows,
            "missing_required_sheets": missing,
        }
    if EVIDENCE_FALLBACK_PATH.exists():
        with EVIDENCE_FALLBACK_PATH.open("r", encoding="utf-8") as stream:
            payload = json.load(stream)
        observed_rows = len(payload.get("scale_confidence_rows", []))
        if observed_rows != expected_rows:
            raise RuntimeError("fallback evidence row count mismatch")
        return {
            "sheet_count": 0,
            "scale_confidence_data_rows": observed_rows,
            "missing_required_sheets": ["BLOCKED_XLSX"],
        }
    raise RuntimeError("no evidence workbook or fallback exists")


def execute_full() -> int:
    wall_start = time.perf_counter()
    cpu_start = time.process_time()

    transcript_stream = io.StringIO()
    selftests = run_selftests(transcript_stream)
    transcript = transcript_stream.getvalue()
    print(transcript, end="")

    scenes, paths, validation = load_and_validate_inputs()
    manifest_before = input_manifest(paths)
    rows = [evaluate_scene(scene) for scene in scenes]
    aggregates = aggregate_results(rows)
    workbook_status = write_workbook(aggregates)
    workbook_verification = verify_workbook(
        len(aggregates["scale_confidence_rows"])
    )
    manifest_after = input_manifest(paths)
    mismatch_count = sum(
        before != after
        for before, after in zip(
            manifest_before["files"], manifest_after["files"]
        )
    )
    if manifest_before["digest"] != manifest_after["digest"]:
        mismatch_count = max(1, mismatch_count)

    wall_seconds = time.perf_counter() - wall_start
    cpu_seconds = time.process_time() - cpu_start
    results: dict[str, Any] = {
        "schema": RESULTS_SCHEMA,
        "contract": {
            "source_packet": r"D:\runs\e2_program\build\PACKET_feyerabend_c1.md",
            "source_dossier": r"D:\dev\99_tools\autocad-sdk-router\reports\e2\dossiers\feyerabend_P2.md",
            "estimator_input": ["anchors"],
            "truth_evaluation_only": ["truth_unit_scale"],
            "pair_labels_excluded_from_estimator": ["truth_pairs", "candidate_pairs"],
            "random_search_count": 0,
            "additional_seed_count": 0,
            "serialization_only_retry_count": 1,
            "artifact_finalization_retry_count": 1,
            "cell_verdict_emitted": False,
        },
        "sealed_configuration": {
            "ransac_log_tolerance": RANSAC_LOG_TOLERANCE,
            "huber_delta": HUBER_DELTA,
            "consensus_threshold": CONSENSUS_THRESHOLD,
            "high_confidence_threshold": HIGH_CONFIDENCE_THRESHOLD,
            "minimum_independent": MIN_INDEPENDENT,
            "accuracy_relative_error": ACCURACY_RELATIVE_ERROR,
            "confidence_formula": "consensus*exp(-log_mad/tau)*min(1,n_independent/5)*min(1,n_spatial_bins/3)",
            "primary_confidence_bin": "unit_status",
            "corruptions": list(CORRUPTIONS),
            "corruption_assignment": "uint32_be(sha256(base_scene_id)[0:4]) mod 4",
        },
        "environment": {
            "python": platform.python_version(),
            "numpy": np.__version__,
            "platform": platform.platform(),
        },
        "selftest": {**selftests, "transcript": transcript},
        "inputs": {
            "directory": str(INPUT_DIR),
            "validation": validation,
            "manifest_before": manifest_before,
            "manifest_after": manifest_after,
            "manifest_mismatch_count": mismatch_count,
        },
        "aggregates": aggregates,
        "scenes": rows,
        "evidence": {
            **workbook_status,
            "verification": workbook_verification,
        },
        "runtime": {
            "cpu_seconds": cpu_seconds,
            "wall_seconds": wall_seconds,
            "cpu_budget_seconds": 3600,
            "cpu_budget_exceeded": cpu_seconds > 3600.0,
        },
        "artifact_manifest_scope": "stable predecessor artifacts only; results self-hash and later REPORT are excluded",
        "artifacts": [
            {
                "name": Path(workbook_status["path"]).name,
                "status": workbook_status["status"],
                "bytes": Path(workbook_status["path"]).stat().st_size,
                "sha256": sha256_file(Path(workbook_status["path"])),
            }
        ],
        "report_path": str(REPORT_PATH),
    }

    # Finalize results once.  The later REPORT may hash results, but results never hashes itself.
    write_json_atomic(RESULTS_PATH, results)
    report_artifacts = [
        {
            "name": RESULTS_PATH.name,
            "status": "GENERATED",
            "bytes": RESULTS_PATH.stat().st_size,
            "sha256": sha256_file(RESULTS_PATH),
        },
        *results["artifacts"],
    ]
    report_context = dict(results)
    report_context["artifacts"] = report_artifacts
    report = render_report(report_context, transcript)
    temporary_report = REPORT_PATH.with_suffix(".md.tmp")
    with temporary_report.open("w", encoding="utf-8", newline="\n") as stream:
        stream.write(report)
    temporary_report.replace(REPORT_PATH)
    print(
        json.dumps(
            {
                "scene_count": validation["scene_count"],
                "estimate_count": aggregates["overall"]["estimate_count"],
                "high_scene_count": aggregates["overall"]["high_scene_count"],
                "pair_digest_mismatch_count": aggregates["pair_label_permutation"]["mismatching_anchor_artifact_scene_count"],
                "input_manifest_mismatch_count": mismatch_count,
                "cpu_seconds": cpu_seconds,
                "evidence_status": workbook_status["status"],
            },
            sort_keys=True,
        )
    )
    return 0


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--selftest",
        action="store_true",
        help="run deterministic selftests without writing evaluation artifacts",
    )
    arguments = parser.parse_args(argv)
    if arguments.selftest:
        stream = io.StringIO()
        run_selftests(stream)
        print(stream.getvalue(), end="")
        return 0
    return execute_full()


if __name__ == "__main__":
    raise SystemExit(main())
