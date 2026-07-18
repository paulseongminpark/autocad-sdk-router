#!/usr/bin/env python3
"""Execute E2 loop L1f Phase B inside the sealed run-local cell.

Repository inputs, prior estimators, prior reports, CAD data, and tests are
read-only.  Every emitted artifact is an exclusive-create file in this script's
directory.  The program reports numeric observations and emits no gate verdict.
"""

from __future__ import annotations

import argparse
import copy
import csv
import hashlib
import importlib.util
import io
import json
import math
import platform
import random
import sys
from collections import Counter
from pathlib import Path
from typing import Any, Mapping, Sequence

import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


sys.dont_write_bytecode = True

CELL_DIR = Path(__file__).resolve().parent
SCRIPT_PATH = CELL_DIR / "loop_l1f.py"
ESTIMATOR_PATH = CELL_DIR / "feyerabend_c1_v5.py"
PREREG_PATH = CELL_DIR / "prereg.json"
SEALED_CSV_PATH = CELL_DIR / "PREREG_SEALED.csv"
SEAL_MANIFEST_PATH = CELL_DIR / "SEAL_MANIFEST.txt"
RESULTS_PATH = CELL_DIR / "c1v7_results.json"
REPLAY_PATH = CELL_DIR / "replay_delta.json"
FLEET_PATH = CELL_DIR / "fleet_probe_results.json"
WITNESS_PATH = CELL_DIR / "witness_classifications.json"
PREDICATE_PATH = CELL_DIR / "predicate_registry.json"
ENVELOPE_PATH = CELL_DIR / "honest_envelope.json"
EVIDENCE_PATH = CELL_DIR / "evidence.xlsx"
REPORT_PATH = CELL_DIR / "REPORT.md"

PACKET_PATH = Path(r"D:\runs\e2_program\build\PACKET_loop_L1f_phaseB.md")
PHASE_A_PACKET_PATH = Path(r"D:\runs\e2_program\build\PACKET_loop_L1f_sealA.md")
CHAIN_DIR = Path(r"D:\runs\e2_program\chainverify_L1e")
SYNTHESIS_PATH = CHAIN_DIR / "SYNTHESIS.md"
LENS1_PATH = CHAIN_DIR / "lens1_witness.md"
LENS2_PATH = CHAIN_DIR / "lens2_formula.md"
SEAT4_PATH = CHAIN_DIR / "seat4_sol.md"
REPO_ROOT = Path(r"D:\dev\99_tools\autocad-sdk-router")
REPO_CELL_DIR = REPO_ROOT / "tools" / "e2" / "cells"
REPO_SEAL_DIR = REPO_ROOT / "reports" / "e2" / "cells" / "loop_l1f"
L1E_RUNNER_PATH = Path(r"D:\runs\e2_program\cells\loop_l1e\loop_l1e.py")
V1_PATH = REPO_CELL_DIR / "feyerabend_c1.py"
V2_PATH = REPO_CELL_DIR / "feyerabend_c1_v2.py"
V3_PATH = REPO_CELL_DIR / "feyerabend_c1_v3.py"
V4_PATH = REPO_CELL_DIR / "feyerabend_c1_v4.py"
V1_C1_BASELINE_PATH = Path(r"D:\runs\e2_program\cells\feyerabend_c1\results.json")
V1_L1B_BASELINE_PATH = Path(r"D:\runs\e2_program\cells\loop_l1b\c1v3_results.json")

WITNESS_COMMIT_SHA = "c896068"
EXPECTED_PREREG_SHA256 = "76AC2A58D74C644A3BF7897325818F1E12151596DC3316BA7CA488BDEB207861"
EXPECTED_SEALED_CSV_SHA256 = "94356AF8F4D219AF65A96825E3A08B29245454EDA1B29B2C3AE83F4B19A8F266"
EXPECTED_MANIFEST_SHA256 = "2C16BA1E66A2CA491364DD533A914F94592E3EC1558ED35A8A1D02E077F92501"
EXPECTED_SOURCE_HASHES = {
    str(SYNTHESIS_PATH): "5FBCBB61D5380469E0DBC6C99C303426B86AD83161A411E66FB261950B0BA6C7",
    str(LENS1_PATH): "836F76A62C6A7B0641CC1CC18273965603C60D4C2F7F5CB71EEE37334CF43C9B",
    str(LENS2_PATH): "FB73840EE6B8555664F3D85A9DC3D5FACE70A53D8A412DEE8A835659B187ED23",
    str(SEAT4_PATH): "858D41A38CCF51579FCD3A15BE431BBD4BA548E7685942A909F77F901585E767",
}
STATUS_RANK = {"NONE": 0, "LOW": 1, "HIGH": 2}
TRACKED_FIELDS = (
    "confidence_score",
    "reference_confidence_score",
    "status",
    "unit_status",
    "reference_status",
)
PROPERTY_RANDOM_CASES = 630
NON_DILUTION_N = (0, 3, 10, 20, 40)


def load_module(name: str, path: Path) -> Any:
    if not path.is_file():
        raise FileNotFoundError(path)
    spec = importlib.util.spec_from_file_location(name, path)
    if spec is None or spec.loader is None:
        raise ImportError(path)
    module = importlib.util.module_from_spec(spec)
    sys.modules[name] = module
    spec.loader.exec_module(module)
    return module


V5 = load_module("loop_l1f_estimator", ESTIMATOR_PATH)
L1E = load_module("loop_l1f_l1e_readonly", L1E_RUNNER_PATH)
V4 = V5.V4
V1 = V5.ORIGINAL


def canonical_bytes(value: Any) -> bytes:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8")


def canonical_sha256(value: Any) -> str:
    return hashlib.sha256(canonical_bytes(value)).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def json_text(payload: Any) -> str:
    return json.dumps(
        payload, ensure_ascii=False, sort_keys=True, indent=2, allow_nan=False
    ) + "\n"


def file_record(path: Path) -> dict[str, Any]:
    return {"path": str(path), "bytes": path.stat().st_size, "sha256": sha256_file(path)}


def directory_record(path: Path) -> dict[str, Any]:
    rows = [file_record(item) for item in sorted(path.glob("scene_*.json"))]
    return {
        "path": str(path),
        "file_count": len(rows),
        "bytes": sum(int(row["bytes"]) for row in rows),
        "digest": canonical_sha256(rows),
    }


def readonly_files() -> tuple[Path, ...]:
    harnesses = []
    for subdir in ("lens1_work", "lens2_work", "seat4_work"):
        harnesses.extend(sorted((CHAIN_DIR / subdir).glob("*.py")))
    fixed = [
        PACKET_PATH,
        PHASE_A_PACKET_PATH,
        SYNTHESIS_PATH,
        LENS1_PATH,
        LENS2_PATH,
        SEAT4_PATH,
        V1_PATH,
        V2_PATH,
        V3_PATH,
        V4_PATH,
        L1E_RUNNER_PATH,
        V1_C1_BASELINE_PATH,
        V1_L1B_BASELINE_PATH,
    ]
    fixed.extend(REPO_SEAL_DIR / name for name in ("prereg.json", "PREREG_SEALED.csv", "SEAL_MANIFEST.txt"))
    fixed.extend((PREREG_PATH, SEALED_CSV_PATH, SEAL_MANIFEST_PATH))
    return tuple(fixed + harnesses)


def readonly_manifest() -> dict[str, Any]:
    files = [file_record(path) for path in readonly_files()]
    directories = [directory_record(V5.L1B_SCENE_DIR), directory_record(V5.V1_SCENE_DIR)]
    return {
        "files": files,
        "directories": directories,
        "digest": canonical_sha256({"files": files, "directories": directories}),
    }


def compare_manifests(before: Mapping[str, Any], after: Mapping[str, Any]) -> dict[str, Any]:
    mismatches = []
    for kind in ("files", "directories"):
        left = {row["path"]: row for row in before[kind]}
        right = {row["path"]: row for row in after[kind]}
        for path in sorted(set(left) | set(right)):
            if left.get(path) != right.get(path):
                mismatches.append({"kind": kind[:-1], "path": path, "before": left.get(path), "after": right.get(path)})
    return {"before": dict(before), "after": dict(after), "mismatch_count": len(mismatches), "mismatches": mismatches}


def verify_seals() -> dict[str, Any]:
    prereg = json.loads(PREREG_PATH.read_text(encoding="utf-8"))
    with SEALED_CSV_PATH.open("r", encoding="utf-8", newline="") as stream:
        csv_rows = list(csv.reader(stream))
    csv_map = {row[0]: row[1] for row in csv_rows if len(row) == 2}
    csv_payload = json.loads(csv_map["canonical_prereg_json"])
    target = {
        "prereg.json": sha256_file(PREREG_PATH),
        "PREREG_SEALED.csv": sha256_file(SEALED_CSV_PATH),
        "SEAL_MANIFEST.txt": sha256_file(SEAL_MANIFEST_PATH),
    }
    repo = {name: sha256_file(REPO_SEAL_DIR / name) for name in target}
    expected = {
        "prereg.json": EXPECTED_PREREG_SHA256,
        "PREREG_SEALED.csv": EXPECTED_SEALED_CSV_SHA256,
        "SEAL_MANIFEST.txt": EXPECTED_MANIFEST_SHA256,
    }
    source_hashes = {path: sha256_file(Path(path)) for path in EXPECTED_SOURCE_HASHES}
    manifest_lines = [line.strip() for line in SEAL_MANIFEST_PATH.read_text(encoding="utf-8").splitlines() if line.strip()]
    observations = {
        "target_hash_match_count": sum(target[name] == expected[name] for name in expected),
        "repo_hash_match_count": sum(repo[name] == expected[name] for name in expected),
        "target_repo_equal_count": sum(target[name] == repo[name] for name in expected),
        "csv_row_count": len(csv_rows),
        "csv_payload_equal": csv_payload == prereg,
        "manifest_line_count": len(manifest_lines),
        "manifest_expected_hash_mentions": sum(any(value in line for line in manifest_lines) for value in (EXPECTED_PREREG_SHA256, EXPECTED_SEALED_CSV_SHA256)),
        "source_hash_match_count": sum(source_hashes[path] == expected_hash for path, expected_hash in EXPECTED_SOURCE_HASHES.items()),
    }
    if observations != {
        "target_hash_match_count": 3,
        "repo_hash_match_count": 3,
        "target_repo_equal_count": 3,
        "csv_row_count": 2,
        "csv_payload_equal": True,
        "manifest_line_count": 2,
        "manifest_expected_hash_mentions": 2,
        "source_hash_match_count": 4,
    }:
        raise RuntimeError("seal/source mismatch: " + json.dumps(observations, sort_keys=True))
    return {
        "witness_commit_citation": WITNESS_COMMIT_SHA,
        "witness_commit_verification_performed": False,
        "witness_commit_verification_reason": "packet prohibits git",
        "target_hashes": target,
        "repo_hashes": repo,
        "expected_hashes": expected,
        "source_hashes": source_hashes,
        "observations": observations,
        "payload": prereg,
    }


def load_scenes(path: Path) -> list[dict[str, Any]]:
    files = sorted(path.glob("scene_*.json"))
    if len(files) != 200:
        raise RuntimeError(f"{path}: expected 200 scenes, found {len(files)}")
    rows = []
    for item in files:
        scene = json.loads(item.read_text(encoding="utf-8"))
        scene["_input_file"] = item.name
        rows.append(scene)
    return rows


def snapshot(module: Any, anchors: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    model = module.fit_anchor_model(copy.deepcopy(list(anchors)))
    return {
        "confidence_score": float(model.get("confidence_score", 0.0)),
        "reference_confidence_score": float(model.get("reference_confidence_score", 0.0)),
        "status": str(model["status"]),
        "unit_status": str(model["unit_status"]),
        "reference_status": str(model["reference_status"]),
        "display_per_raw": model.get("display_per_raw"),
        "mm_per_raw": model.get("mm_per_raw"),
        "physical_unit": model.get("physical_unit"),
        "reference_span": model.get("reference_span"),
        "consensus_weight": model.get("consensus_weight"),
        "reference_consensus_weight": model.get("reference_consensus_weight"),
        "n_independent": int(model.get("n_independent", 0)),
        "n_candidate_handles": int(model.get("n_candidate_handles", model.get("n_independent", 0))),
        "reference_n_independent": int(model.get("reference_n_independent", 0)),
        "reference_n_candidate_handles": int(model.get("reference_n_candidate_handles", model.get("reference_n_independent", 0))),
        "provenance": copy.deepcopy(model.get("provenance", {})),
    }


def increases(before: Mapping[str, Any], after: Mapping[str, Any]) -> dict[str, bool]:
    return {
        "confidence_score": float(after["confidence_score"]) > float(before["confidence_score"]) + 1e-15,
        "reference_confidence_score": float(after["reference_confidence_score"]) > float(before["reference_confidence_score"]) + 1e-15,
        "status": STATUS_RANK[str(after["status"])] > STATUS_RANK[str(before["status"])],
        "unit_status": STATUS_RANK[str(after["unit_status"])] > STATUS_RANK[str(before["unit_status"])],
        "reference_status": STATUS_RANK[str(after["reference_status"])] > STATUS_RANK[str(before["reference_status"])],
    }


def classify_upward(
    classifications: list[dict[str, Any]],
    *,
    scope: str,
    case_id: str,
    family: str,
    before_anchors: Sequence[Mapping[str, Any]],
    after_anchors: Sequence[Mapping[str, Any]],
    before: Mapping[str, Any],
    after: Mapping[str, Any],
    rise_map: Mapping[str, bool],
) -> str | None:
    fields = [field for field in TRACKED_FIELDS if rise_map.get(field)]
    if not fields:
        return None
    surface = V5.estimator_input_surface(after_anchors)
    serialized = canonical_bytes(surface).decode("utf-8")
    surface_digest = hashlib.sha256(serialized.encode("utf-8")).hexdigest()
    analysis = V5.suspicion_analysis(after_anchors)
    witness = V5.build_honest_witness(after_anchors)
    exact = bool(witness is not None and witness["witness_surface"] == surface and witness["witness_serialization"] == serialized)
    legitimate = bool(exact and analysis["tier_B_signal_count"] == 0)
    label = "information_limit_record" if legitimate else "violation"
    identifier = f"W{len(classifications):06d}"
    nonzero_rules = [
        {"rule_id": row["rule_id"], "observed_count": row["observed_count"], "measurements": row["measurements"]}
        for row in analysis["executed_rules"] if int(row["observed_count"]) > 0
    ]
    narrative = (
        f"Case {case_id} freezes {analysis['record_count']} ordered records at surface {surface_digest}; "
        f"the semantic generator rebuilds raw_span from each p0/p1 relation and reproduces all "
        f"{len(analysis['declared_field_inventory'])} declared field names exactly. "
        f"Observed nonzero rule counts are {json.dumps({r['rule_id']: r['observed_count'] for r in nonzero_rules}, sort_keys=True)}; "
        f"Tier-A admissions are {json.dumps(analysis['tier_A_signals'], sort_keys=True)}."
        if legitimate
        else f"Case {case_id} freezes {analysis['record_count']} ordered records at surface {surface_digest}; "
        f"the honest generator rejects the surface because Tier-B findings are {json.dumps(analysis['tier_B_signals'], sort_keys=True)}."
    )
    classifications.append(
        {
            "classification_id": identifier,
            "scope": scope,
            "case_id": case_id,
            "family": family,
            "increased_fields": fields,
            "field_events": [{"field": field, "before": before[field], "after": after[field]} for field in fields],
            "witness_scene_identifier": witness["witness_scene_identifier"] if witness else f"REJECTED::{surface_digest}",
            "honest_generation_spec": None if witness is None else witness["honest_generation_spec"],
            "honest_generation_spec_digest": None if witness is None else witness["honest_generation_spec_digest"],
            "witness_scene_complete_estimator_input_surface_serialization": None if witness is None else witness["witness_serialization"],
            "post_perturbation_complete_estimator_input_surface_serialization": serialized,
            "exact_all_field_equality": exact,
            "witness_serialization_sha256": None if witness is None else witness["witness_serialization_sha256"],
            "post_serialization_sha256": surface_digest,
            "city_semantic_legitimacy": legitimate,
            "scene_specific_honest_generation_narrative": narrative,
            "complete_ordered_executed_rule_list": analysis["executed_rules"],
            "per_rule_observed_counts_including_zeros": analysis["rule_counts"],
            "nonzero_rule_measurements": nonzero_rules,
            "sealed_envelope_admissions": analysis["tier_A_signals"],
            "frozen_declared_field_inventory": analysis["declared_field_inventory"],
            "uncovered_declared_field_count": analysis["uncovered_declared_field_count"],
            "automated_classification_result": label,
            "manual_suppression_used": False,
            "unclassified_field_count": 0,
        }
    )
    return identifier


def case_record(
    classifications: list[dict[str, Any]],
    *,
    scope: str,
    case_id: str,
    family: str,
    before_anchors: Sequence[Mapping[str, Any]],
    after_anchors: Sequence[Mapping[str, Any]],
    include_v1: bool = False,
) -> dict[str, Any]:
    versions = {}
    modules = (("v1", V1), ("v4", V4), ("v5", V5)) if include_v1 else (("v4", V4), ("v5", V5))
    for name, module in modules:
        before = snapshot(module, before_anchors)
        after = snapshot(module, after_anchors)
        versions[name] = {"before": before, "after": after, "increases": increases(before, after)}
    v5 = versions["v5"]
    classification_id = classify_upward(
        classifications,
        scope=scope,
        case_id=case_id,
        family=family,
        before_anchors=before_anchors,
        after_anchors=after_anchors,
        before=v5["before"],
        after=v5["after"],
        rise_map=v5["increases"],
    )
    record = {
        "case_id": case_id,
        "family": family,
        "before_input_surface": V5.estimator_input_surface(before_anchors),
        "after_input_surface": V5.estimator_input_surface(after_anchors),
        "before_surface_digest": canonical_sha256(V5.estimator_input_surface(before_anchors)),
        "after_surface_digest": canonical_sha256(V5.estimator_input_surface(after_anchors)),
        "before_suspicion": V5.suspicion_analysis(before_anchors),
        "after_suspicion": V5.suspicion_analysis(after_anchors),
        "versions": versions,
        "classification_id": classification_id,
    }
    record["case_digest"] = canonical_sha256(record)
    return record


def mk(handle: str, span: float, display: Any, x: float, y: float, *, unit: str = "MM", anchor_type: str = "DIM") -> dict[str, Any]:
    return V4._make_anchor(handle, span, display, x, y, unit=unit, anchor_type=anchor_type)


BASE_POS = {"A": (0.0, 0.0), "B": (0.0, 20000.0), "C": (20000.0, 0.0)}
CENTER = (10000.0, 10000.0)
DIL_POS = [(x, y) for x in (1000.0, 5500.0, 10000.0, 14500.0, 19000.0) for y in (500.0, 6500.0, 12500.0, 18500.0)]


def clean_dims() -> list[dict[str, Any]]:
    return [mk(handle, 1000.0, 2500.0, *BASE_POS[handle]) for handle in ("A", "B", "C")]


def dil_dims(count: int) -> list[dict[str, Any]]:
    rows = []
    for index in range(count):
        x, y = DIL_POS[index % len(DIL_POS)]
        rows.append(mk(f"D{index:02d}", 1000.0, 2500.0, x + 7.0 * (index // len(DIL_POS)), y))
    return rows


def ratio_outlier(handle: str, distance_tau: float, position: tuple[float, float] = CENTER) -> dict[str, Any]:
    return mk(handle, 1000.0, 2500.0 * math.exp(distance_tau * V5.RANSAC_LOG_TOLERANCE), *position)


def run_honest_envelope(scenes_by_cohort: Mapping[str, Sequence[Mapping[str, Any]]]) -> dict[str, Any]:
    rows = []
    cohort_summary = {}
    field_counts: Counter[str] = Counter()
    total_anchors = 0
    max_ratio = 0.0
    raw_mismatch = 0
    for cohort, scenes in scenes_by_cohort.items():
        cohort_anchors = 0
        cohort_max = 0.0
        cohort_mismatch = 0
        for scene in scenes:
            anchors = scene["anchors"]
            analysis = V5.suspicion_analysis(anchors)
            distance = max(analysis["ratio_distances_tau"].values(), default=0.0)
            mismatch = int(analysis["rule_counts"]["raw_span_equals_euclidean_geometry"])
            inventory = Counter(analysis["declared_field_inventory"])
            field_counts.update(inventory)
            cohort_anchors += len(anchors)
            cohort_max = max(cohort_max, distance)
            cohort_mismatch += mismatch
            rows.append({
                "cohort": cohort,
                "scene_id": scene["scene_id"],
                "input_file": scene["_input_file"],
                "anchor_count": len(anchors),
                "maximum_ratio_deviation_tau": distance,
                "raw_span_geometry_mismatch_count": mismatch,
                "declared_field_inventory": analysis["declared_field_inventory"],
                "frozen_surface_digest": analysis["surface_digest"],
            })
        total_anchors += cohort_anchors
        max_ratio = max(max_ratio, cohort_max)
        raw_mismatch += cohort_mismatch
        cohort_summary[cohort] = {"scene_count": len(scenes), "anchor_count": cohort_anchors, "maximum_ratio_deviation_tau": cohort_max, "raw_span_geometry_mismatch_count": cohort_mismatch}

    o1 = [
        mk("O1_MM0", 100.0, 250.0, 0.0, 0.0, unit="MM"),
        mk("O1_MM1", 100.0, 250.0, 1000.0, 0.0, unit="MM"),
        mk("O1_MM2", 100.0, 250.0, 0.0, 1000.0, unit="MM"),
        mk("O1_M0", 100.0, 0.25, 1000.0, 1000.0, unit="M"),
        mk("O1_M1", 100.0, 0.25, 500.0, 500.0, unit="M"),
    ]
    o2 = [
        mk("O2_0", 100.0, 250.0, 0.0, 0.0),
        mk("O2_1", 100.0, 250.0, 1000.0, 0.0),
        mk("O2_2", 100.0, 250.0, 0.0, 1000.0),
        mk("O2_STALE", 100.0, 225.0, 1000.0, 1000.0),
    ]
    fixtures = {}
    status_downgrades = 0
    coverage_losses = 0
    for fixture_id, anchors in (("O1", o1), ("O2", o2)):
        v4 = snapshot(V4, anchors)
        v5 = snapshot(V5, anchors)
        downgraded = sum(STATUS_RANK[v5[field]] < STATUS_RANK[v4[field]] for field in ("status", "unit_status", "reference_status"))
        loss = int(v4["unit_status"] != "NONE" and v5["unit_status"] == "NONE")
        status_downgrades += downgraded
        coverage_losses += loss
        fixtures[fixture_id] = {
            "input_surface": V5.estimator_input_surface(anchors),
            "input_surface_digest": canonical_sha256(V5.estimator_input_surface(anchors)),
            "v4": v4,
            "v5": v5,
            "v5_suspicion": V5.suspicion_analysis(anchors),
            "status_downgrade_count": downgraded,
            "coverage_loss_count": loss,
        }
    sealed = {
        "scene_count": 400,
        "anchor_count": 2008,
        "maximum_within_scene_ratio_deviation_tau": V5.HONEST_RATIO_ROUNDOFF_MAX_TAU,
        "declared_raw_span_geometry_mismatch_count": 0,
        "O2_stale_maximum_distance_tau": V5.O2_STALE_MAX_TAU,
        "O2_type_to_grid_maximum_distance_tau": V5.O2_TYPE_TO_GRID_MAX_TAU,
    }
    measured = {
        "scene_count": sum(len(value) for value in scenes_by_cohort.values()),
        "anchor_count": total_anchors,
        "maximum_within_scene_ratio_deviation_tau": max_ratio,
        "declared_raw_span_geometry_mismatch_count": raw_mismatch,
    }
    if measured != {"scene_count": 400, "anchor_count": 2008, "maximum_within_scene_ratio_deviation_tau": V5.HONEST_RATIO_ROUNDOFF_MAX_TAU, "declared_raw_span_geometry_mismatch_count": 0}:
        raise RuntimeError("honest envelope measurement mismatch: " + json.dumps(measured, sort_keys=True))
    if status_downgrades or coverage_losses:
        raise RuntimeError("O1/O2 numeric loss")
    return {
        "schema": "ariadne.e2.loop_l1f.honest_envelope.v1",
        "tau": V5.RANSAC_LOG_TOLERANCE,
        "sealed_limits": sealed,
        "measured": measured,
        "cohorts": cohort_summary,
        "declared_field_presence_counts": dict(sorted(field_counts.items())),
        "O1_O2": fixtures,
        "O1_O2_status_downgrade_count": status_downgrades,
        "O1_O2_coverage_loss_count": coverage_losses,
        "per_scene_measurements": rows,
        "fixture_generator_invariants": {
            "raw_span_dependency_expression": "raw_span=EuclideanDistance(p0,p1)",
            "honest_corpus_mismatch_count": raw_mismatch,
            "frozen_surface_before_canonicalization": True,
        },
    }


def third_fleet_definitions(fx: Mapping[str, Sequence[Mapping[str, Any]]]) -> list[tuple[str, str, Sequence[Mapping[str, Any]], Sequence[Mapping[str, Any]]]]:
    mutate = lambda anchors, family, seed: V4.randomized_corruption(anchors, family, random.Random(seed), 0)
    rows = [
        ("CE-A", "suffix_removal", fx["CE-A"], mutate(fx["CE-A"], "suffix_removal", 0)),
        ("CE-B", "type_to_grid", fx["CE-B"], mutate(fx["CE-B"], "type_to_grid", 0)),
        ("CE-B2", "type_to_grid", fx["CE-B2"], mutate(fx["CE-B2"], "type_to_grid", 5)),
        ("CE-C", "geometry_ratio_break", fx["CE-C"], mutate(fx["CE-C"], "geometry_ratio_break", 9)),
        ("CE-D", "outlier_clone", fx["CE-D"], mutate(fx["CE-D"], "outlier_clone", 0)),
        ("CE-E", "stale_override", fx["CE-E"], mutate(fx["CE-E"], "stale_override", 384)),
        ("CE-F", "type_to_grid", fx["CE-F"], mutate(fx["CE-F"], "type_to_grid", 5)),
    ]
    d1_after = copy.deepcopy(list(fx["D1"]))
    repeat = copy.deepcopy(d1_after[0])
    x, y = V4.POSITIONS[3]
    repeat["p0"], repeat["p1"] = [x, y], [x + 100.0, y]
    d1_after.append(repeat)
    rows.extend([
        ("D1", "consistent_same_handle_repeat", fx["D1"], d1_after),
        ("D2", "space_tie_observation", fx["D2"], copy.deepcopy(list(fx["D2"]))),
        ("D3", "tau_boundary", fx["D3-below"], fx["D3-above"]),
        ("D4", "eligible_empty_observation", fx["D4"], copy.deepcopy(list(fx["D4"]))),
    ])
    return rows


def run_property(classifications: list[dict[str, Any]]) -> tuple[dict[str, Any], dict[str, Any]]:
    fx = L1E.fixtures()
    third_cases = [
        case_record(classifications, scope="third_fleet", case_id=case_id, family=family, before_anchors=before, after_anchors=after, include_v1=True)
        for case_id, family, before, after in third_fleet_definitions(fx)
    ]
    strata = [
        ("zero_cliff_start", "CE-A", fx["CE-A"]),
        ("ratio_outlier", "CE-B2", fx["CE-B2"]),
        ("mixed_space", "D2", fx["D2"]),
        ("handle_collision", "CE-D", fx["CE-D"]),
        ("near_tau_spread", "D3-below", fx["D3-below"]),
        ("near_tau_spread", "D3-above", fx["D3-above"]),
    ]
    rng = random.Random(20260719)
    cases = []
    family_counts: Counter[str] = Counter()
    stratum_counts: Counter[str] = Counter()
    upward_counts: Counter[str] = Counter({field: 0 for field in TRACKED_FIELDS})
    for index in range(PROPERTY_RANDOM_CASES):
        family = V4.PROPERTY_FAMILIES[index % len(V4.PROPERTY_FAMILIES)]
        stratum, base_id, before = strata[index % len(strata)]
        after = V4.randomized_corruption(before, family, rng, index)
        record = case_record(classifications, scope="property_20260719", case_id=f"P{index:04d}", family=family, before_anchors=before, after_anchors=after)
        record.update({"case_index": index, "seed": 20260719, "stratum": stratum, "base_scene_id": base_id})
        record["case_digest"] = canonical_sha256(record)
        cases.append(record)
        family_counts[family] += 1
        stratum_counts[stratum] += 1
        for field, rose in record["versions"]["v5"]["increases"].items():
            upward_counts[field] += int(rose)
    property_result = {
        "schema": "ariadne.e2.loop_l1f.property.v1",
        "seed": 20260719,
        "randomized_case_count": len(cases),
        "third_fleet_regression_case_count": len(third_cases),
        "case_count": len(cases) + len(third_cases),
        "required_minimum_case_count": 600,
        "family_counts": dict(sorted(family_counts.items())),
        "stratum_counts": dict(sorted(stratum_counts.items())),
        "v5_upward_field_counts": dict(upward_counts),
        "third_fleet_classes": [row["case_id"] for row in third_cases],
        "case_manifest": cases + [{"case_id": "REG::" + row["case_id"], **{k: v for k, v in row.items() if k != "case_id"}} for row in third_cases],
    }
    property_result["case_manifest_digest"] = canonical_sha256([row["case_digest"] for row in property_result["case_manifest"]])
    third = {
        "schema": "ariadne.e2.loop_l1f.third_fleet.v1",
        "class_count": len(third_cases),
        "classes": [row["case_id"] for row in third_cases],
        "cases": third_cases,
        "cases_digest": canonical_sha256(third_cases),
    }
    return property_result, third


def run_non_dilution() -> dict[str, Any]:
    tau = V5.RANSAC_LOG_TOLERANCE
    geometry_bad = mk("B_GEOM", 300.0, 750.0, *CENTER)
    geometry_bad["raw_span"] = 100.0
    signal_sets = [
        ("TierA_O2_moderate_display_stale", "A", ("ratio", "reference"), [ratio_outlier("A_STALE", 2.0)]),
        ("TierA_O2_type_to_grid_display_residual", "A", ("ratio", "reference"), [ratio_outlier("A_T2G", 0.5), mk("A_GRID", 1000.0, None, 12000.0, 12000.0, anchor_type="GRID")]),
        ("TierB_declared_field_geometry_contradiction", "B", ("ratio", "reference"), [geometry_bad]),
        ("TierB_source_handle_ratio_conflict", "B", ("ratio", "reference"), [mk("B_RC", 1000.0, 2500.0, *CENTER), mk("B_RC", 1000.0, 2500.0 * math.exp(2.0 * tau), 11000.0, 11000.0)]),
        ("TierB_source_handle_span_conflict", "B", ("reference",), [mk("B_SC", 1000.0, None, *CENTER, anchor_type="GRID"), mk("B_SC", 3000.0, None, 11000.0, 11000.0, anchor_type="GRID")]),
        ("TierB_mixed_estimator_ratio_space_trace", "B", ("ratio", "reference"), [mk("B_MIX", 1000.0, 2500.0, *CENTER, unit="PX")]),
        ("TierB_severe_distance_ratio", "B", ("ratio", "reference"), [ratio_outlier("B_SR", 5.0)]),
        ("TierB_severe_distance_span", "B", ("reference",), [mk("B_SS", 1000.0 * math.exp(10.0 * tau), 2500.0 * math.exp(10.0 * tau), *CENTER)]),
        ("TierB_unregistered_ratio_residual", "B", ("ratio", "reference"), [ratio_outlier("B_UR", 2.4)]),
        ("TierB_unregistered_span_residual", "B", ("reference",), [mk("B_US", 1000.0 * math.exp(2.4 * tau), 2500.0 * math.exp(2.4 * tau), *CENTER)]),
        ("TierB_missing_or_ambiguous_ratio", "B", ("ratio", "reference"), [mk("B_MISS", 1000.0, None, *CENTER)]),
    ]
    cases = []
    for signal_id, tier, paths, suspicious in signal_sets:
        suspicious_surface = V5.estimator_input_surface(suspicious)
        rows = []
        kind_sets = []
        for n in NON_DILUTION_N:
            clean = clean_dims() + dil_dims(n)
            with_s = clean + copy.deepcopy(suspicious)
            matched = snapshot(V5, clean)
            observed = snapshot(V5, with_s)
            analysis = V5.suspicion_analysis(with_s)
            kinds = sorted((row["tier"], row["signal_kind"]) for row in analysis["tier_A_signals"] + analysis["tier_B_signals"])
            kind_sets.append(kinds)
            rows.append({
                "clean_N": n,
                "fixed_S_surface": suspicious_surface,
                "fixed_S_digest": canonical_sha256(suspicious_surface),
                "matched_clean_input_surface": V5.estimator_input_surface(clean),
                "with_S_input_surface": V5.estimator_input_surface(with_s),
                "matched_clean": matched,
                "with_S": observed,
                "penalty": {
                    "confidence_score": float(matched["confidence_score"]) - float(observed["confidence_score"]),
                    "reference_confidence_score": float(matched["reference_confidence_score"]) - float(observed["reference_confidence_score"]),
                },
                "suspicion": analysis,
                "signal_kinds": kinds,
            })
        reclassified = any(kinds != kind_sets[0] for kinds in kind_sets[1:])
        tier_a_floors = {
            path: min(float(row["penalty"]["confidence_score" if path == "ratio" else "reference_confidence_score"]) for row in rows)
            for path in paths
        } if tier == "A" else {}
        cases.append({
            "signal_id": signal_id,
            "tier": tier,
            "affected_paths": list(paths),
            "fixed_S_surface": suspicious_surface,
            "fixed_S_digest": canonical_sha256(suspicious_surface),
            "reclassification_count": int(reclassified),
            "observed_candidate_count_independent_penalty_floors": tier_a_floors,
            "conceptual_N_to_infinity": {
                "formula": V5.NON_DILUTION_FORMULA,
                "candidate_count_in_suspicion_penalty": False,
                "tier_B_affected_paths_remain_exact_zero_NONE": tier == "B",
            },
            "sweep": rows,
        })
    return {
        "schema": "ariadne.e2.loop_l1f.non_dilution.v1",
        "clean_evidence_counts_N": list(NON_DILUTION_N),
        "registered_Tier_A_kind_count": 2,
        "enumerated_Tier_B_path_case_count": 9,
        "case_count": len(cases),
        "cases": cases,
        "cases_digest": canonical_sha256(cases),
    }


SEARCH_POSITIONS = (
    (0.0, 0.0),
    (0.0, 1000.0),
    (1000.0, 0.0),
    (1000.0, 1000.0),
    (500.0, 500.0),
    (2000.0, 0.0),
)


def simple_anchor(handle: str, ratio: float, position: tuple[float, float], *, unit: str = "MM") -> dict[str, Any]:
    x, y = position
    return {
        "handle": handle,
        "anchor_type": "DIM",
        "region": "P",
        "p0": [float(x), float(y)],
        "p1": [float(x) + 100.0, float(y)],
        "raw_span": 100.0,
        "display_value": float(ratio) * 100.0,
        "display_unit": unit,
        "text_height": None,
        "weight": 1.0,
    }


def run_seat4_regressions(classifications: list[dict[str, Any]]) -> dict[str, Any]:
    tau = V5.RANSAC_LOG_TOLERANCE
    good = 2.5
    p = SEARCH_POSITIONS
    base = [simple_anchor(f"G{i}", good, p[i]) for i in range(3)]

    dilution_before = base + [simple_anchor("OUT", good * math.exp(2.0 * tau), p[3])]
    dilution_after = copy.deepcopy(dilution_before) + [simple_anchor("G3", good, p[3])]
    suffix_before = [
        simple_anchor("SG0", good, p[0], unit="MM"),
        simple_anchor("SG1", good, p[1], unit="MM"),
        simple_anchor("SG2", good, p[2], unit="MM"),
        simple_anchor("SOUT", good * math.exp(2.0 * tau), p[3], unit="UNKNOWN"),
    ]
    suffix_after = copy.deepcopy(suffix_before)
    for anchor in suffix_after:
        anchor["display_unit"] = "UNKNOWN"
    grid_before = base + [
        simple_anchor("SEVERE", good * math.exp(5.0 * tau), p[3]),
        simple_anchor("MODERATE", good * math.exp(2.0 * tau), p[4]),
    ]
    grid_after = copy.deepcopy(grid_before)
    for anchor in grid_after:
        if anchor["handle"] == "SEVERE":
            anchor["anchor_type"] = "GRID"
            anchor.pop("display_value", None)
            anchor.pop("display_unit", None)
    stale_before = base + [simple_anchor("STALE", good * math.exp(-3.4 * tau), p[3])]
    stale_after = copy.deepcopy(stale_before)
    stale_after[-1]["display_value"] *= math.exp(tau)
    geometry_before = base + [simple_anchor("GEOM", good * math.exp(3.4 * tau), p[3])]
    geometry_after = copy.deepcopy(geometry_before)
    p0 = geometry_after[-1]["p0"]
    geometry_after[-1]["p1"] = [p0[0] + 100.0 * math.exp(tau), p0[1]]

    definitions = [
        ("clean_handle_dilution_spatial_mask", "free_clean_handle_addition", dilution_before, dilution_after),
        ("suffix_merge_leaves_moderate_outlier", "suffix_removal", suffix_before, suffix_after),
        ("mixed_severity_retype_leaves_moderate_outlier", "type_to_grid", grid_before, grid_after),
        ("partial_stale_correction_still_outlier", "stale_override", stale_before, stale_after),
        ("partial_geometry_correction_still_outlier", "geometry_ratio_break", geometry_before, geometry_after),
    ]
    targeted = [
        case_record(classifications, scope="regression_seat4_targeted", case_id=case_id, family=family, before_anchors=before, after_anchors=after)
        for case_id, family, before, after in definitions
    ]

    parent = []
    known_positive = []
    for hundredths in range(101, 350):
        distance = hundredths / 100.0
        before = base + [simple_anchor("OUT", good * math.exp(distance * tau), p[3])]
        for position_index, position in enumerate(p):
            after = copy.deepcopy(before) + [simple_anchor(f"ADD_{position_index}", good, position)]
            v4_before, v4_after = snapshot(V4, before), snapshot(V4, after)
            v5_before, v5_after = snapshot(V5, before), snapshot(V5, after)
            v4_rise, v5_rise = increases(v4_before, v4_after), increases(v5_before, v5_after)
            case_id = f"broad_dilution_k{distance:.2f}_p{position_index}"
            classification_id = classify_upward(
                classifications,
                scope="regression_seat4_747_window",
                case_id=case_id,
                family="free_clean_handle_addition",
                before_anchors=before,
                after_anchors=after,
                before=v5_before,
                after=v5_after,
                rise_map=v5_rise,
            )
            record = {
                "case_id": case_id,
                "distance_tau": distance,
                "clean_handle_position_index": position_index,
                "before_input_surface": V5.estimator_input_surface(before),
                "after_input_surface": V5.estimator_input_surface(after),
                "before_surface_digest": canonical_sha256(V5.estimator_input_surface(before)),
                "after_surface_digest": canonical_sha256(V5.estimator_input_surface(after)),
                "v4": {"before": v4_before, "after": v4_after, "increases": v4_rise},
                "v5": {"before": v5_before, "after": v5_after, "increases": v5_rise},
                "v5_after_suspicion": V5.suspicion_analysis(after),
                "classification_id": classification_id,
            }
            record["case_digest"] = canonical_sha256(record)
            parent.append(record)
            if any(v4_rise.values()):
                known_positive.append(record)
    return {
        "targeted_five": {
            "case_count": len(targeted),
            "v4_rise_case_count": sum(any(row["versions"]["v4"]["increases"].values()) for row in targeted),
            "v5_rise_case_count": sum(any(row["versions"]["v5"]["increases"].values()) for row in targeted),
            "cases": targeted,
        },
        "window_747": {
            "parent_case_count": len(parent),
            "known_positive_case_count": len(known_positive),
            "v5_rise_case_count_within_known_positive": sum(any(row["v5"]["increases"].values()) for row in known_positive),
            "parent_cases": parent,
            "known_positive_case_ids": [row["case_id"] for row in known_positive],
            "parent_digest": canonical_sha256([row["case_digest"] for row in parent]),
        },
    }


def run_lens1_regressions(classifications: list[dict[str, Any]]) -> dict[str, Any]:
    fx = L1E.fixtures()
    w_before = fx["CE-C"]
    w_after = V4.randomized_corruption(w_before, "geometry_ratio_break", random.Random(9), 0)
    b_before = [mk("C0", 100.0, 250.0, *SEARCH_POSITIONS[0])]
    b_after = [mk(f"C{i}", 100.0, 250.0, *SEARCH_POSITIONS[i]) for i in range(2)]
    forged = mk("FORGE", 300.0, 750.0, *SEARCH_POSITIONS[2])
    forged["raw_span"] = 100.0
    b_after.append(forged)
    return {
        "W000002": case_record(classifications, scope="regression_lens1", case_id="W000002_CE-C", family="geometry_ratio_break", before_anchors=w_before, after_anchors=w_after, include_v1=True),
        "B5": case_record(classifications, scope="regression_lens1", case_id="B5_raw_span_forgery", family="declared_field_geometry_forgery", before_anchors=b_before, after_anchors=b_after, include_v1=True),
    }


def p5_base_scene(kind: str, rng: random.Random) -> list[dict[str, Any]]:
    tau = V5.RANSAC_LOG_TOLERANCE
    anchors = clean_dims()
    if kind == "midband":
        anchors += [ratio_outlier("X", 1.2 + 1.2 * rng.random())]
    elif kind == "severe":
        anchors += [ratio_outlier("X", 3.6 + 3.0 * rng.random())]
    elif kind == "span_collision":
        anchors += [
            mk("HB", 1000.0, None, *CENTER, anchor_type="GRID"),
            mk("HB", 1000.0 * math.exp((1.5 + 5.0 * rng.random()) * tau), None, 11000.0, 11000.0, anchor_type="GRID"),
        ]
    elif kind == "ratio_collision":
        anchors += [
            mk("X", 1000.0, 2500.0, *CENTER),
            mk("X", 1000.0, 2500.0 * math.exp((1.5 + 4.0 * rng.random()) * tau), 11000.0, 11000.0),
        ]
    elif kind == "missing":
        anchors += [mk("X", 1000.0, None, *CENTER)]
    elif kind == "mixed":
        anchors += [mk("X", 1000.0, 2500.0, *CENTER, unit="PX")]
    elif kind == "span_outlier":
        anchors += [mk("S", 1000.0 * math.exp(8.0 * tau), 2500.0 * math.exp(8.0 * tau), *CENTER)]
    elif kind == "grid_rich":
        anchors += [mk(f"G{i}", 1000.0, None, 2000.0 + 2500.0 * i, 15000.0, anchor_type="GRID") for i in range(4)]
    return anchors


def p5_perturb(anchors: Sequence[Mapping[str, Any]], rng: random.Random) -> tuple[str, list[dict[str, Any]]]:
    op = rng.choice(("add_clean", "add_clean_many", "add_grid", "toward_consensus", "away_consensus", "retype_grid", "remove_one", "collide", "dup_exact", "jitter", "unit_flip"))
    after = copy.deepcopy(list(anchors))
    if op == "add_clean":
        k = rng.randrange(1, 6)
        offset = rng.randrange(0, 1000)
        after += [mk(f"N{offset}_{i}", 1000.0, 2500.0, 300.0 + 173.0 * ((offset + i) % 113), 300.0 + 197.0 * ((offset + 3 * i) % 101)) for i in range(k)]
    elif op == "add_clean_many":
        after += dil_dims(rng.randrange(10, 41))
    elif op == "add_grid":
        offset = rng.randrange(0, 1000)
        after += [mk(f"NG{offset}", 1000.0, None, 500.0 + 19.0 * offset, 18000.0, anchor_type="GRID")]
    elif op in ("toward_consensus", "away_consensus"):
        targets = [a for a in after if a["anchor_type"] == "DIM" and a.get("display_value")]
        if targets:
            target = rng.choice(targets)
            z = math.log(float(target["display_value"]) / float(target["raw_span"]))
            zc = math.log(2.5)
            factor = rng.uniform(0.3, 0.9) if op == "toward_consensus" else rng.uniform(1.1, 2.5)
            target["display_value"] = float(target["raw_span"]) * math.exp(zc + (z - zc) * factor)
    elif op == "retype_grid":
        targets = [a for a in after if a["anchor_type"] == "DIM"]
        if targets:
            rng.choice(targets)["anchor_type"] = "GRID"
    elif op == "remove_one" and len(after) > 1:
        after.pop(rng.randrange(len(after)))
    elif op == "collide":
        targets = [a for a in after if a["anchor_type"] == "DIM"]
        if len(targets) >= 2:
            a0, a1 = rng.sample(targets, 2)
            a1["handle"] = a0["handle"]
    elif op == "dup_exact":
        after.append(copy.deepcopy(rng.choice(after)))
    elif op == "jitter":
        target = rng.choice(after)
        dx, dy = rng.uniform(-2000, 2000), rng.uniform(-2000, 2000)
        target["p0"] = [target["p0"][0] + dx, target["p0"][1] + dy]
        target["p1"] = [target["p1"][0] + dx, target["p1"][1] + dy]
    elif op == "unit_flip":
        targets = [a for a in after if a["anchor_type"] == "DIM" and a.get("display_value")]
        if targets:
            target = rng.choice(targets)
            target["display_unit"] = "CM"
            target["display_value"] = float(target["display_value"]) / 10.0
    return op, after


def run_lens2_regressions(classifications: list[dict[str, Any]]) -> dict[str, Any]:
    tau = V5.RANSAC_LOG_TOLERANCE
    hb = [mk("HB", 1000.0, None, *CENTER, anchor_type="GRID"), mk("HB", 3000.0, None, 11000.0, 11000.0, anchor_type="GRID")]
    base_a = hb + clean_dims()
    base_c = clean_dims() + [ratio_outlier("X", 2.4)]
    base_s = clean_dims() + [mk("S", 1000.0 * math.exp(10.0 * tau), 2500.0 * math.exp(10.0 * tau), *CENTER)]

    traces = {}
    for case_id, base, paths in (("T_A", base_a, ("reference",)), ("T_C", base_c, ("ratio", "reference")), ("T_S", base_s, ("reference",))):
        sweep = []
        for n in NON_DILUTION_N:
            anchors = base + dil_dims(n)
            sweep.append({"clean_N": n, "input_surface": V5.estimator_input_surface(anchors), "v4": snapshot(V4, anchors), "v5": snapshot(V5, anchors), "v5_suspicion": V5.suspicion_analysis(anchors)})
        traces[case_id] = {
            "affected_paths": list(paths),
            "sweep": sweep,
            "transition": case_record(classifications, scope="regression_lens2", case_id=case_id, family="fixed_suspicion_clean_addition", before_anchors=base, after_anchors=base + dil_dims(40)),
        }

    t_b_before = clean_dims() + [ratio_outlier("X", 3.6)]
    t_b_after = clean_dims() + [ratio_outlier("X", 3.4)]
    t_b = case_record(classifications, scope="regression_lens2", case_id="T_B_3.6_to_3.4_tau", family="display_move_toward_consensus", before_anchors=t_b_before, after_anchors=t_b_after)

    grids = [mk(f"G{i}", 1000.0, None, 2000.0 + 3000.0 * i, 15000.0, anchor_type="GRID") for i in range(5)]
    p4 = case_record(classifications, scope="regression_lens2", case_id="P4_ratio_absent_GRID_neutrality", family="honest_GRID_addition", before_anchors=clean_dims(), after_anchors=clean_dims() + grids)
    p4_before, p4_after = p4["versions"]["v5"]["before"], p4["versions"]["v5"]["after"]
    p4["reference_confidence_loss"] = float(p4_before["reference_confidence_score"]) - float(p4_after["reference_confidence_score"])
    p4["reference_status_downgrade_count"] = int(STATUS_RANK[p4_after["reference_status"]] < STATUS_RANK[p4_before["reference_status"]])
    p4["status_downgrade_count"] = int(STATUS_RANK[p4_after["status"]] < STATUS_RANK[p4_before["status"]])

    rng = random.Random(20260719)
    kinds = ("clean", "midband", "severe", "span_collision", "ratio_collision", "missing", "mixed", "span_outlier", "grid_rich")
    p5_cases = []
    v4_rises = 0
    v4_violation = 0
    v5_rises = 0
    v5_adverse_info = 0
    for trial in range(2000):
        kind = kinds[trial % len(kinds)]
        before_anchors = p5_base_scene(kind, rng)
        op, after_anchors = p5_perturb(before_anchors, rng)
        v4_before, v4_after = snapshot(V4, before_anchors), snapshot(V4, after_anchors)
        v5_before, v5_after = snapshot(V5, before_anchors), snapshot(V5, after_anchors)
        v4_rise, v5_rise = increases(v4_before, v4_after), increases(v5_before, v5_after)
        v4_analysis = V4.suspicion_analysis(after_anchors)
        v5_analysis = V5.suspicion_analysis(after_anchors)
        v4_label = None
        if any(v4_rise.values()):
            v4_rises += 1
            v4_label = "information_limit_record" if int(v4_analysis["residual_suspicion_count"]) == 0 else "violation"
            v4_violation += int(v4_label == "violation")
        case_id = f"P5_{trial:04d}_{kind}_{op}"
        classification_id = classify_upward(
            classifications,
            scope="regression_lens2_P5",
            case_id=case_id,
            family=op,
            before_anchors=before_anchors,
            after_anchors=after_anchors,
            before=v5_before,
            after=v5_after,
            rise_map=v5_rise,
        )
        v5_label = None
        if classification_id is not None:
            v5_rises += 1
            v5_label = classifications[int(classification_id[1:])]["automated_classification_result"]
            v5_adverse_info += int(v5_label == "information_limit_record" and v4_label == "violation")
        record = {
            "trial": trial,
            "case_id": case_id,
            "seed": 20260719,
            "base_kind": kind,
            "operation": op,
            "before_input_surface": V5.estimator_input_surface(before_anchors),
            "after_input_surface": V5.estimator_input_surface(after_anchors),
            "before_surface_digest": canonical_sha256(V5.estimator_input_surface(before_anchors)),
            "after_surface_digest": canonical_sha256(V5.estimator_input_surface(after_anchors)),
            "v4": {"before": v4_before, "after": v4_after, "increases": v4_rise, "post_suspicion": v4_analysis, "classification": v4_label},
            "v5": {"before": v5_before, "after": v5_after, "increases": v5_rise, "post_suspicion": v5_analysis, "classification": v5_label, "classification_id": classification_id},
        }
        record["case_digest"] = canonical_sha256(record)
        p5_cases.append(record)
    return {
        **traces,
        "T_B": t_b,
        "P4": p4,
        "P5": {
            "parent_hunt_case_count": len(p5_cases),
            "v4_upward_case_count": v4_rises,
            "v4_known_violation_case_count": v4_violation,
            "v5_upward_case_count": v5_rises,
            "v5_information_limit_on_v4_known_violation_count": v5_adverse_info,
            "v5_prevented_v4_known_violation_count": sum(row["v4"]["classification"] == "violation" and not any(row["v5"]["increases"].values()) for row in p5_cases),
            "cases": p5_cases,
            "cases_digest": canonical_sha256([row["case_digest"] for row in p5_cases]),
        },
    }


def run_regressions(classifications: list[dict[str, Any]]) -> dict[str, Any]:
    seat4 = run_seat4_regressions(classifications)
    lens1 = run_lens1_regressions(classifications)
    lens2 = run_lens2_regressions(classifications)
    adverse = []
    adverse.extend(seat4["targeted_five"]["cases"])
    adverse.extend(row for row in seat4["window_747"]["parent_cases"] if row["case_id"] in set(seat4["window_747"]["known_positive_case_ids"]))
    adverse.extend(lens1.values())
    adverse.extend(lens2[name]["transition"] for name in ("T_A", "T_C", "T_S"))
    adverse.append(lens2["T_B"])
    adverse_rises = 0
    adverse_info = 0
    for row in adverse:
        transition = row.get("versions", {}).get("v5", row.get("v5", {}))
        rise = transition.get("increases", {})
        if any(rise.values()):
            adverse_rises += 1
            cid = row.get("classification_id")
            if cid is not None:
                adverse_info += int(classifications[int(cid[1:])]["automated_classification_result"] == "information_limit_record")
    return {
        "schema": "ariadne.e2.loop_l1f.regressions.v1",
        "seat4": seat4,
        "lens1": lens1,
        "lens2": lens2,
        "numeric_summary": {
            "seat4_targeted_case_count": seat4["targeted_five"]["case_count"],
            "seat4_window_parent_case_count": seat4["window_747"]["parent_case_count"],
            "seat4_window_v4_known_positive_case_count": seat4["window_747"]["known_positive_case_count"],
            "seat4_window_v5_rise_case_count": seat4["window_747"]["v5_rise_case_count_within_known_positive"],
            "lens1_case_count": len(lens1),
            "lens2_P5_parent_case_count": lens2["P5"]["parent_hunt_case_count"],
            "lens2_P5_v4_known_violation_case_count": lens2["P5"]["v4_known_violation_case_count"],
            "lens2_P5_v5_rise_case_count": lens2["P5"]["v5_upward_case_count"],
            "P4_reference_confidence_loss": lens2["P4"]["reference_confidence_loss"],
            "P4_status_downgrade_count": lens2["P4"]["status_downgrade_count"] + lens2["P4"]["reference_status_downgrade_count"],
            "known_adverse_v5_rise_case_count_excluding_P5": adverse_rises,
            "known_adverse_information_limit_count": adverse_info + lens2["P5"]["v5_information_limit_on_v4_known_violation_count"],
        },
    }


def flatten_json(value: Any, path: str = "$") -> dict[str, dict[str, Any]]:
    output: dict[str, dict[str, Any]] = {}
    if isinstance(value, dict):
        output[path] = {"present": True, "kind": "object", "size": len(value)}
        for key in sorted(value):
            token = str(key).replace("~", "~0").replace("/", "~1")
            output.update(flatten_json(value[key], path + "/" + token))
    elif isinstance(value, list):
        output[path] = {"present": True, "kind": "array", "size": len(value)}
        for index, item in enumerate(value):
            output.update(flatten_json(item, path + f"/{index}"))
    else:
        kind = "boolean" if isinstance(value, bool) else "number" if isinstance(value, (int, float)) and not isinstance(value, bool) else "null" if value is None else "string"
        output[path] = {"present": True, "kind": kind, "value": value}
    return output


MISSING = {"present": False, "kind": "missing"}


def numeric_delta(left: Mapping[str, Any], right: Mapping[str, Any]) -> float | None:
    if left.get("kind") == right.get("kind") == "number":
        return float(right["value"]) - float(left["value"])
    return None


def aggregate_rows(rows: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    by_scale = {}
    for scale in V5.SCALES:
        selected = [row for row in rows if math.isclose(float(row["scale_kappa"]), float(scale), rel_tol=0.0, abs_tol=1e-15)]
        high = [row for row in selected if row["unit_status"] == "HIGH"]
        accurate = [row for row in high if row["relative_error"] is not None and float(row["relative_error"]) <= 0.05]
        by_scale[str(scale)] = {
            "scene_count": len(selected),
            "high_scene_count": len(high),
            "high_coverage": 0.0 if not selected else len(high) / len(selected),
            "high_accuracy_count": len(accurate),
            "high_accuracy_within_5pct": 0.0 if not high else len(accurate) / len(high),
            "unit_status_counts": dict(Counter(str(row["unit_status"]) for row in selected)),
            "maximum_high_relative_error": max((float(row["relative_error"]) for row in high if row["relative_error"] is not None), default=None),
        }
    high_all = [row for row in rows if row["unit_status"] == "HIGH"]
    accurate_all = [row for row in high_all if row["relative_error"] is not None and float(row["relative_error"]) <= 0.05]
    return {
        "scene_count": len(rows),
        "high_scene_count": len(high_all),
        "high_coverage": 0.0 if not rows else len(high_all) / len(rows),
        "high_accuracy_count": len(accurate_all),
        "high_accuracy_within_5pct": 0.0 if not high_all else len(accurate_all) / len(high_all),
        "unit_status_counts": dict(Counter(str(row["unit_status"]) for row in rows)),
        "status_counts": dict(Counter(str(row["status"]) for row in rows)),
        "by_scale": by_scale,
    }


def run_replay(scenes_by_cohort: Mapping[str, Sequence[Mapping[str, Any]]]) -> dict[str, Any]:
    cohorts = {}
    total_fields = 0
    zero_rows = 0
    for cohort_name, scenes in scenes_by_cohort.items():
        version_rows = {
            "v1": [V1.evaluate_scene(copy.deepcopy(scene)) for scene in scenes],
            "v4": [V4.evaluate_scene(copy.deepcopy(scene)) for scene in scenes],
            "v5": [V5.evaluate_scene(copy.deepcopy(scene)) for scene in scenes],
        }
        transcripts = []
        for index, scene in enumerate(scenes):
            complete = {version: version_rows[version][index] for version in ("v1", "v4", "v5")}
            flattened = {version: flatten_json(row) for version, row in complete.items()}
            paths = sorted(set().union(*(set(rows) for rows in flattened.values())))
            field_rows = []
            for path in paths:
                descriptors = {version: flattened[version].get(path, MISSING) for version in ("v1", "v4", "v5")}
                equal14 = descriptors["v1"] == descriptors["v4"]
                equal45 = descriptors["v4"] == descriptors["v5"]
                equal15 = descriptors["v1"] == descriptors["v5"]
                field_rows.append({
                    "path": path,
                    "v1": descriptors["v1"],
                    "v4": descriptors["v4"],
                    "v5": descriptors["v5"],
                    "v1_to_v4_equal": equal14,
                    "v4_to_v5_equal": equal45,
                    "v1_to_v5_equal": equal15,
                    "v1_to_v4_numeric_delta": numeric_delta(descriptors["v1"], descriptors["v4"]),
                    "v4_to_v5_numeric_delta": numeric_delta(descriptors["v4"], descriptors["v5"]),
                    "v1_to_v5_numeric_delta": numeric_delta(descriptors["v1"], descriptors["v5"]),
                })
                zero_rows += int(equal14 and equal45)
            total_fields += len(field_rows)
            transcripts.append({
                "scene_id": str(scene["scene_id"]),
                "input_file": str(scene["_input_file"]),
                "complete_input_scene": scene,
                "complete_input_scene_digest": canonical_sha256(scene),
                "complete_version_rows": complete,
                "complete_version_row_digests": {version: canonical_sha256(row) for version, row in complete.items()},
                "per_field_delta_transcript": field_rows,
                "field_row_count": len(field_rows),
            })
        cohorts[cohort_name] = {
            "scene_count": len(scenes),
            "versions": {version: {"aggregate": aggregate_rows(version_rows[version]), "rows_digest": canonical_sha256(version_rows[version])} for version in ("v1", "v4", "v5")},
            "scene_transcripts": transcripts,
            "scene_transcript_count": len(transcripts),
            "field_row_count": sum(row["field_row_count"] for row in transcripts),
        }
    return {
        "schema": "ariadne.e2.loop_l1f.replay_delta.v1",
        "versions": ["v1", "v4", "v5"],
        "cohorts": cohorts,
        "publication": {
            "scene_count": sum(row["scene_count"] for row in cohorts.values()),
            "scene_transcript_count": sum(row["scene_transcript_count"] for row in cohorts.values()),
            "per_scene_per_field_row_count": total_fields,
            "zero_delta_field_row_count": zero_rows,
            "zero_delta_rows_included": True,
            "complete_version_rows_included": True,
            "complete_input_scenes_included": True,
        },
    }


def build_predicates(evidence: Mapping[str, Any]) -> dict[str, Any]:
    inherited = L1E.build_predicate_registry()["registry"]
    inherited_rows = [
        {
            **copy.deepcopy(row),
            "inheritance": "loop_l1e_verbatim_logic",
        }
        for row in inherited
    ]

    def tier_b_hard_block(score: float, status: str) -> bool:
        return float(score) == 0.0 and str(status) == "NONE"

    def fixed_s_non_dilution(n0_floor: float, n20_penalty: float) -> bool:
        return float(n20_penalty) + 1e-15 >= float(n0_floor)

    def precanonical_declared_field_detection(original_mismatch: bool, recorded_count: int) -> bool:
        return not bool(original_mismatch) or int(recorded_count) > 0

    def grid_neutrality(before_score: float, after_score: float, before_status: str, after_status: str) -> bool:
        return float(after_score) + 1e-15 >= float(before_score) and STATUS_RANK[str(after_status)] >= STATUS_RANK[str(before_status)]

    def witness_substantiation(expected_rules: int, observed_rules: int, zero_counts_explicit: bool, scene_specific_narrative: bool) -> bool:
        return int(observed_rules) == int(expected_rules) and bool(zero_counts_explicit) and bool(scene_specific_narrative)

    def regression_obligations(adverse_information_limit_count: int, p4_downgrade_count: int) -> bool:
        return int(adverse_information_limit_count) == 0 and int(p4_downgrade_count) == 0

    extension_defs = [
        ("tier_B_hard_block", "score == 0 and status == NONE", ["affected_score", "affected_status"], "A frozen raw_span/geometry contradiction with a positive affected score or non-NONE affected status must return false.", {"affected_score": 0.1, "affected_status": "LOW"}, tier_b_hard_block(0.1, "LOW")),
        ("fixed_S_non_dilution", "penalty(S,N) >= sealed_N0_floor", ["sealed_N0_floor", "penalty_S_N20"], "If the matched-clean penalty for the same byte-identical S is lower at N=20 than its sealed N=0 floor, the predicate must return false.", {"sealed_N0_floor": 0.2, "penalty_S_N20": 0.1}, fixed_s_non_dilution(0.2, 0.1)),
        ("precanonical_declared_field_detection", "not original_mismatch or recorded_count > 0", ["original_mismatch", "recorded_count"], "If canonicalization recomputes raw_span and the original mismatch is absent from the suspicion record, the predicate must return false.", {"original_mismatch": True, "recorded_count": 0}, precanonical_declared_field_detection(True, 0)),
        ("ratio_absent_GRID_neutrality", "after_reference >= before_reference and after_status >= before_status", ["before_reference", "after_reference", "before_status", "after_status"], "Adding five honest ratio-absent GRID handles that lowers reference confidence or status must return false.", {"before_reference": 1.0, "after_reference": 0.375, "before_status": "HIGH", "after_status": "LOW"}, grid_neutrality(1.0, 0.375, "HIGH", "LOW")),
        ("scene_specific_witness_substantiation", "all rules listed and zero counts explicit and narrative scene specific", ["expected_rule_count", "observed_rule_count", "zero_counts_explicit", "scene_specific_narrative"], "A repeated boilerplate rationale lacking the full executed-rule list, per-rule zero counts, or scene-specific honest generation narrative must return false.", {"expected_rule_count": 18, "observed_rule_count": 5, "zero_counts_explicit": False, "scene_specific_narrative": False}, witness_substantiation(18, 5, False, False)),
        ("l1f_regression_obligations", "adverse_information_limit_count == 0 and P4_downgrade_count == 0", ["adverse_information_limit_count", "P4_downgrade_count"], "Any named adverse regression that both rises and is certified as information_limit_record, or the P4 no-loss regression that downgrades, must return false.", {"adverse_information_limit_count": 1, "P4_downgrade_count": 0}, regression_obligations(1, 0)),
    ]
    extension = []
    for identifier, expression, fields, annotation, inputs, observed in extension_defs:
        extension.append({
            "predicate_id": identifier,
            "predicate_expression": expression,
            "observed_input_fields": fields,
            "counterexample_input_annotation": annotation,
            "counterexample_input": inputs,
            "expected_counterexample_result_false": True,
            "counterexample_return_value": bool(observed),
            "observed_counterexample_result_false": not bool(observed),
            "inheritance": "loop_l1f_extension",
        })
    registry = inherited_rows + extension
    actual = {
        "tier_B_hard_block": all(
            (row["tier"] != "B") or all(
                (sweep["with_S"]["confidence_score"] == 0.0 and sweep["with_S"]["unit_status"] == "NONE") if path == "ratio" else (sweep["with_S"]["reference_confidence_score"] == 0.0 and sweep["with_S"]["reference_status"] == "NONE")
                for sweep in row["sweep"] for path in row["affected_paths"]
            )
            for row in evidence["non_dilution"]["cases"]
        ),
        "fixed_S_non_dilution": all(row["reclassification_count"] == 0 for row in evidence["non_dilution"]["cases"]),
        "precanonical_declared_field_detection": evidence["regressions"]["lens1"]["B5"]["after_suspicion"]["rule_counts"]["raw_span_equals_euclidean_geometry"] == 1,
        "ratio_absent_GRID_neutrality": evidence["regressions"]["lens2"]["P4"]["reference_confidence_loss"] <= 1e-15 and evidence["regressions"]["lens2"]["P4"]["reference_status_downgrade_count"] == 0,
        "scene_specific_witness_substantiation": all(len(row["complete_ordered_executed_rule_list"]) == len(V5.DETECTION_RULE_ORDER) and len(row["per_rule_observed_counts_including_zeros"]) == len(V5.DETECTION_RULE_ORDER) and bool(row["scene_specific_honest_generation_narrative"]) for row in evidence["classifications"]),
        "l1f_regression_obligations": evidence["regressions"]["numeric_summary"]["known_adverse_information_limit_count"] == 0 and evidence["regressions"]["numeric_summary"]["P4_status_downgrade_count"] == 0,
    }
    if any(row["counterexample_return_value"] for row in registry) or not all(actual.values()):
        raise RuntimeError("predicate execution mismatch: " + json.dumps(actual, sort_keys=True))
    return {
        "schema": "ariadne.e2.loop_l1f.predicate_registry.v1",
        "predicate_count": len(registry),
        "inherited_predicate_count": len(inherited_rows),
        "extension_predicate_count": len(extension),
        "registry": registry,
        "counterexample_execution_transcript": [json.dumps(row, ensure_ascii=False, sort_keys=True) for row in registry],
        "observed_evidence_evaluations": actual,
    }


def witness_payload(classifications: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    labels = Counter(str(row["automated_classification_result"]) for row in classifications)
    return {
        "schema": "ariadne.e2.loop_l1f.witness_classifications.v1",
        "classification_count": len(classifications),
        "upward_field_event_count": sum(len(row["increased_fields"]) for row in classifications),
        "label_counts": dict(labels),
        "manual_suppression_count": sum(bool(row["manual_suppression_used"]) for row in classifications),
        "unclassified_field_count": sum(int(row["unclassified_field_count"]) for row in classifications),
        "uncovered_declared_field_count": sum(int(row["uncovered_declared_field_count"]) for row in classifications),
        "complete_rule_count_per_classification": len(V5.DETECTION_RULE_ORDER),
        "classifications": list(classifications),
        "classifications_digest": canonical_sha256(classifications),
    }


def compute_all() -> dict[str, Any]:
    seals = verify_seals()
    source_before = readonly_manifest()
    scenes = {
        "l1b": load_scenes(V5.L1B_SCENE_DIR),
        "c1_original": load_scenes(V5.V1_SCENE_DIR),
    }
    envelope = run_honest_envelope(scenes)
    classifications: list[dict[str, Any]] = []
    property_result, third = run_property(classifications)
    non_dilution = run_non_dilution()
    regressions = run_regressions(classifications)
    replay = run_replay(scenes)
    source_after = readonly_manifest()
    source_integrity = compare_manifests(source_before, source_after)
    witnesses = witness_payload(classifications)
    predicate_evidence = {
        "non_dilution": non_dilution,
        "regressions": regressions,
        "classifications": classifications,
    }
    predicates = build_predicates(predicate_evidence)
    results = {
        "schema": "ariadne.e2.loop_l1f.c1v7_results.v1",
        "contract": {
            "packet": str(PACKET_PATH),
            "write_root": str(CELL_DIR),
            "repo_mutation_allowed": False,
            "cad_access_allowed": False,
            "test_access_allowed": False,
            "gate_verdict_emitted": False,
            "subagents_used": False,
            "git_used_after_packet_read": False,
        },
        "seal": {key: value for key, value in seals.items() if key != "payload"},
        "sealed_configuration": seals["payload"],
        "estimator": {
            "path": str(ESTIMATOR_PATH),
            "sha256": sha256_file(ESTIMATOR_PATH),
            "revision": V5.REVISION,
            "frozen_surface_capture": "before canonicalization/preparation",
            "complete_detection_rule_order": list(V5.DETECTION_RULE_ORDER),
            "non_dilution_formula": V5.NON_DILUTION_FORMULA,
            "boundary_behavior": V5.BOUNDARY_BEHAVIOR,
            "Tier_A_limits_tau": {
                "honest_roundoff": V5.HONEST_RATIO_ROUNDOFF_MAX_TAU,
                "O2_stale": V5.O2_STALE_MAX_TAU,
                "O2_type_to_grid": V5.O2_TYPE_TO_GRID_MAX_TAU,
            },
            "Tier_B_severe_boundary_tau_inclusive": V5.SEVERE_DISTANCE_TAU,
        },
        "honest_envelope_numeric": {
            "measured": envelope["measured"],
            "O1_O2_status_downgrade_count": envelope["O1_O2_status_downgrade_count"],
            "O1_O2_coverage_loss_count": envelope["O1_O2_coverage_loss_count"],
        },
        "property_test": property_result,
        "non_dilution": non_dilution,
        "regression_numeric_summary": regressions["numeric_summary"],
        "witness_classification_summary": {key: value for key, value in witnesses.items() if key != "classifications"},
        "replay_numeric_summary": {cohort: {version: row["versions"][version]["aggregate"] for version in ("v1", "v4", "v5")} for cohort, row in replay["cohorts"].items()},
        "predicate_counterexamples": predicates,
        "source_readonly_manifest": source_integrity,
        "environment": {"python": platform.python_version(), "numpy": np.__version__},
    }
    fleet = {
        "schema": "ariadne.e2.loop_l1f.fleet_probe_results.v1",
        "third_fleet": third,
        "non_dilution": non_dilution,
        "regressions": regressions,
    }
    return {
        "seals": seals,
        "results": results,
        "replay": replay,
        "fleet": fleet,
        "witnesses": witnesses,
        "predicates": predicates,
        "envelope": envelope,
    }


def numerical_validation(bundle: Mapping[str, Any]) -> dict[str, Any]:
    results = bundle["results"]
    prop = results["property_test"]
    non_dilution = results["non_dilution"]
    regression = bundle["fleet"]["regressions"]
    replay = bundle["replay"]
    witnesses = bundle["witnesses"]
    predicates = bundle["predicates"]
    envelope = bundle["envelope"]
    v5_aggregates = {cohort: replay["cohorts"][cohort]["versions"]["v5"]["aggregate"] for cohort in replay["cohorts"]}
    per_scale_coverages = [row["high_coverage"] for aggregate in v5_aggregates.values() for row in aggregate["by_scale"].values()]
    high_count = sum(aggregate["high_scene_count"] for aggregate in v5_aggregates.values())
    high_accurate = sum(aggregate["high_accuracy_count"] for aggregate in v5_aggregates.values())
    tier_a_cases = [row for row in non_dilution["cases"] if row["tier"] == "A"]
    tier_b_cases = [row for row in non_dilution["cases"] if row["tier"] == "B"]
    tier_a_floor_positive = all(all(value > 0.0 for value in row["observed_candidate_count_independent_penalty_floors"].values()) for row in tier_a_cases)
    tier_b_exact = all(
        all(
            (sweep["with_S"]["confidence_score"] == 0.0 and sweep["with_S"]["unit_status"] == "NONE") if path == "ratio" else (sweep["with_S"]["reference_confidence_score"] == 0.0 and sweep["with_S"]["reference_status"] == "NONE")
            for sweep in row["sweep"] for path in row["affected_paths"]
        )
        for row in tier_b_cases
    )
    numeric = {
        "property_case_count": prop["case_count"],
        "property_randomized_case_count": prop["randomized_case_count"],
        "property_family_count": len(prop["family_counts"]),
        "property_stratum_count": len(prop["stratum_counts"]),
        "third_fleet_class_count": len(prop["third_fleet_classes"]),
        "non_dilution_case_count": non_dilution["case_count"],
        "non_dilution_sweep_row_count": sum(len(row["sweep"]) for row in non_dilution["cases"]),
        "replay_scene_count": replay["publication"]["scene_count"],
        "replay_field_row_count": replay["publication"]["per_scene_per_field_row_count"],
        "replay_zero_delta_row_count": replay["publication"]["zero_delta_field_row_count"],
        "v5_high_scene_count": high_count,
        "v5_high_accuracy_count": high_accurate,
        "v5_high_accuracy_fraction": 0.0 if not high_count else high_accurate / high_count,
        "v5_minimum_per_scale_high_coverage": min(per_scale_coverages),
        "seat4_targeted_case_count": regression["numeric_summary"]["seat4_targeted_case_count"],
        "seat4_window_parent_case_count": regression["numeric_summary"]["seat4_window_parent_case_count"],
        "seat4_window_known_positive_case_count": regression["numeric_summary"]["seat4_window_v4_known_positive_case_count"],
        "seat4_window_v5_rise_case_count": regression["numeric_summary"]["seat4_window_v5_rise_case_count"],
        "lens2_P5_parent_case_count": regression["numeric_summary"]["lens2_P5_parent_case_count"],
        "lens2_P5_v4_known_violation_case_count": regression["numeric_summary"]["lens2_P5_v4_known_violation_case_count"],
        "lens2_P5_v5_rise_case_count": regression["numeric_summary"]["lens2_P5_v5_rise_case_count"],
        "known_adverse_information_limit_count": regression["numeric_summary"]["known_adverse_information_limit_count"],
        "P4_reference_confidence_loss": regression["numeric_summary"]["P4_reference_confidence_loss"],
        "P4_status_downgrade_count": regression["numeric_summary"]["P4_status_downgrade_count"],
        "witness_classification_count": witnesses["classification_count"],
        "witness_violation_count": witnesses["label_counts"].get("violation", 0),
        "witness_manual_suppression_count": witnesses["manual_suppression_count"],
        "witness_unclassified_field_count": witnesses["unclassified_field_count"],
        "witness_uncovered_declared_field_count": witnesses["uncovered_declared_field_count"],
        "predicate_count": predicates["predicate_count"],
        "predicate_counterexample_false_count": sum(int(row["observed_counterexample_result_false"]) for row in predicates["registry"]),
        "source_manifest_mismatch_count": results["source_readonly_manifest"]["mismatch_count"],
        "honest_scene_count": envelope["measured"]["scene_count"],
        "honest_anchor_count": envelope["measured"]["anchor_count"],
        "honest_raw_span_geometry_mismatch_count": envelope["measured"]["declared_raw_span_geometry_mismatch_count"],
        "O1_O2_status_downgrade_count": envelope["O1_O2_status_downgrade_count"],
        "O1_O2_coverage_loss_count": envelope["O1_O2_coverage_loss_count"],
    }
    checks = {
        "property_minimum": prop["case_count"] >= 600 and prop["randomized_case_count"] >= 600,
        "family_coverage": set(prop["family_counts"]) == set(V4.PROPERTY_FAMILIES),
        "strata_coverage": set(prop["stratum_counts"]) == {"zero_cliff_start", "ratio_outlier", "mixed_space", "handle_collision", "near_tau_spread"},
        "third_fleet_coverage": set(prop["third_fleet_classes"]) == {"CE-A", "CE-B", "CE-B2", "CE-C", "CE-D", "CE-E", "CE-F", "D1", "D2", "D3", "D4"},
        "non_dilution_counts": non_dilution["case_count"] == 11 and all(len(row["sweep"]) == 5 for row in non_dilution["cases"]),
        "non_dilution_fixed_class": all(row["reclassification_count"] == 0 for row in non_dilution["cases"]),
        "tier_A_positive_floor": tier_a_floor_positive,
        "tier_B_exact_zero_NONE": tier_b_exact,
        "replay_complete": replay["publication"]["scene_count"] == 400 and replay["publication"]["scene_transcript_count"] == 400 and replay["publication"]["zero_delta_rows_included"],
        "coverage_numeric": min(per_scale_coverages) >= 0.6,
        "accuracy_numeric": (high_accurate / high_count) >= 0.95,
        "O1_O2_no_loss": envelope["O1_O2_status_downgrade_count"] == 0 and envelope["O1_O2_coverage_loss_count"] == 0,
        "seat4_counts": regression["numeric_summary"]["seat4_targeted_case_count"] == 5 and regression["numeric_summary"]["seat4_window_parent_case_count"] == 1494 and regression["numeric_summary"]["seat4_window_v4_known_positive_case_count"] == 747,
        "seat4_known_adverse_blocked": regression["seat4"]["targeted_five"]["v5_rise_case_count"] == 0 and regression["numeric_summary"]["seat4_window_v5_rise_case_count"] == 0,
        "lens1_blocked": all(not any(row["versions"]["v5"]["increases"].values()) for row in regression["lens1"].values()),
        "lens2_named_blocked": all(not any(regression["lens2"][name]["transition"]["versions"]["v5"]["increases"].values()) for name in ("T_A", "T_C", "T_S")) and not any(regression["lens2"]["T_B"]["versions"]["v5"]["increases"].values()),
        "P4_no_loss": regression["numeric_summary"]["P4_reference_confidence_loss"] <= 1e-15 and regression["numeric_summary"]["P4_status_downgrade_count"] == 0,
        "P5_counts": regression["numeric_summary"]["lens2_P5_parent_case_count"] == 2000 and regression["numeric_summary"]["lens2_P5_v4_known_violation_case_count"] == 50 and regression["lens2"]["P5"]["v5_prevented_v4_known_violation_count"] == 50,
        "adverse_not_certified": regression["numeric_summary"]["known_adverse_information_limit_count"] == 0,
        "classification_complete": witnesses["manual_suppression_count"] == 0 and witnesses["unclassified_field_count"] == 0 and witnesses["uncovered_declared_field_count"] == 0,
        "global_violation_zero": witnesses["label_counts"].get("violation", 0) == 0,
        "predicate_counterexamples": predicates["predicate_count"] == numeric["predicate_counterexample_false_count"],
        "source_unchanged": results["source_readonly_manifest"]["mismatch_count"] == 0,
    }
    if not all(checks.values()):
        raise RuntimeError("numeric invariant mismatch: " + json.dumps({"numeric": numeric, "checks": checks}, sort_keys=True))
    return {"numeric": numeric, "checks": checks}


def style_sheet(sheet: Any) -> None:
    if sheet.max_row:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.freeze_panes = "A2"
    for column in sheet.columns:
        letter = column[0].column_letter
        width = min(60, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
        sheet.column_dimensions[letter].width = width


def build_workbook_bytes(bundle: Mapping[str, Any], validation: Mapping[str, Any], predicted_hashes: Mapping[str, str]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    ws.append(["field", "value"])
    ws.append(["cell", "loop_l1f"])
    ws.append(["scope", "numeric evidence; no gate verdict"])
    ws.append(["packet", str(PACKET_PATH)])
    ws.append(["witness_commit_citation", WITNESS_COMMIT_SHA])
    ws.append(["prereg_sha256", EXPECTED_PREREG_SHA256])
    ws.append(["sealed_csv_sha256", EXPECTED_SEALED_CSV_SHA256])

    numbers = wb.create_sheet("NUMBERS")
    numbers.append(["metric", "value"])
    for key, value in validation["numeric"].items():
        numbers.append([key, value])

    envelope = wb.create_sheet("ENVELOPE")
    envelope.append(["cohort", "scene_count", "anchor_count", "max_ratio_deviation_tau", "raw_geometry_mismatch_count"])
    for cohort, row in bundle["envelope"]["cohorts"].items():
        envelope.append([cohort, row["scene_count"], row["anchor_count"], row["maximum_ratio_deviation_tau"], row["raw_span_geometry_mismatch_count"]])
    for fixture_id, row in bundle["envelope"]["O1_O2"].items():
        envelope.append([fixture_id, 1, len(row["input_surface"]), None, row["v5_suspicion"]["rule_counts"]["raw_span_equals_euclidean_geometry"]])

    prop = wb.create_sheet("PROPERTY")
    prop.append(["dimension", "name", "count"])
    for name, value in bundle["results"]["property_test"]["family_counts"].items():
        prop.append(["family", name, value])
    for name, value in bundle["results"]["property_test"]["stratum_counts"].items():
        prop.append(["stratum", name, value])
    prop.append(["total", "randomized", bundle["results"]["property_test"]["randomized_case_count"]])
    prop.append(["total", "with_third_fleet", bundle["results"]["property_test"]["case_count"]])

    nd = wb.create_sheet("NON_DILUTION")
    nd.append(["signal_id", "tier", "affected_paths", "reclassification_count", "minimum_penalty_floors", "sweep_rows"])
    for row in bundle["results"]["non_dilution"]["cases"]:
        nd.append([row["signal_id"], row["tier"], ",".join(row["affected_paths"]), row["reclassification_count"], json.dumps(row["observed_candidate_count_independent_penalty_floors"], sort_keys=True), len(row["sweep"])])

    reg = wb.create_sheet("REGRESSIONS")
    reg.append(["metric", "value"])
    for key, value in bundle["fleet"]["regressions"]["numeric_summary"].items():
        reg.append([key, value])

    witness = wb.create_sheet("WITNESS")
    witness.append(["classification_id", "scope", "case_id", "label", "field_event_count", "tier_A_count", "tier_B_count", "rule_count", "uncovered_fields"])
    for row in bundle["witnesses"]["classifications"]:
        witness.append([row["classification_id"], row["scope"], row["case_id"], row["automated_classification_result"], len(row["increased_fields"]), len(row["sealed_envelope_admissions"]), sum(1 for item in row["nonzero_rule_measurements"] if item["rule_id"] not in ("declared_field_inventory", "O2_class_moderate_display_stale", "O2_class_type_to_grid_display_residual", "ratio_absent_GRID_neutrality")), len(row["complete_ordered_executed_rule_list"]), row["uncovered_declared_field_count"]])

    replay = wb.create_sheet("REPLAY")
    replay.append(["cohort", "version", "scene_count", "high_count", "high_coverage", "high_accuracy", "field_rows"])
    for cohort, row in bundle["replay"]["cohorts"].items():
        for version in ("v1", "v4", "v5"):
            aggregate = row["versions"][version]["aggregate"]
            replay.append([cohort, version, aggregate["scene_count"], aggregate["high_scene_count"], aggregate["high_coverage"], aggregate["high_accuracy_within_5pct"], row["field_row_count"]])

    selftest = wb.create_sheet("SELFTEST")
    selftest.append(["predicate_id", "counterexample_return_value", "observed_false", "annotation"])
    for row in bundle["predicates"]["registry"]:
        selftest.append([row["predicate_id"], row["counterexample_return_value"], row["observed_counterexample_result_false"], row["counterexample_input_annotation"]])

    files = wb.create_sheet("FILES")
    files.append(["name", "sha256"])
    for name, digest in predicted_hashes.items():
        files.append([name, digest])

    for sheet in wb.worksheets:
        style_sheet(sheet)
    stream = io.BytesIO()
    wb.save(stream)
    wb.close()
    payload = stream.getvalue()
    check = load_workbook(io.BytesIO(payload), read_only=True, data_only=False)
    names = list(check.sheetnames)
    check.close()
    expected = ["README", "NUMBERS", "ENVELOPE", "PROPERTY", "NON_DILUTION", "REGRESSIONS", "WITNESS", "REPLAY", "SELFTEST", "FILES"]
    if names != expected:
        raise RuntimeError("workbook sheet mismatch")
    return payload


def render_report(bundle: Mapping[str, Any], validation: Mapping[str, Any], artifact_hashes: Mapping[str, str]) -> str:
    n = validation["numeric"]
    reg = bundle["fleet"]["regressions"]["numeric_summary"]
    lines = [
        "# E2 loop L1f Phase B numeric evidence",
        "",
        f"Witness commit citation: `{WITNESS_COMMIT_SHA}` (git verification not performed because the packet prohibits git).",
        f"Sealed `prereg.json` SHA-256: `{EXPECTED_PREREG_SHA256}`.",
        f"Sealed `PREREG_SEALED.csv` SHA-256: `{EXPECTED_SEALED_CSV_SHA256}`.",
        f"Sealed `SEAL_MANIFEST.txt` SHA-256: `{EXPECTED_MANIFEST_SHA256}`.",
        "",
        "This report records executed numeric evidence only. It does not emit a gate verdict.",
        "",
        "## Estimator and honest envelope",
        "",
        f"Revision: `{V5.REVISION}`. Detection freezes the complete declared surface before canonicalization. Tier-A attenuation uses `{V5.NON_DILUTION_FORMULA}`.",
        f"Measured honest corpus: {n['honest_scene_count']} scenes, {n['honest_anchor_count']} anchors, maximum ratio deviation {bundle['envelope']['measured']['maximum_within_scene_ratio_deviation_tau']:.16g} tau, raw-span/geometry mismatches {n['honest_raw_span_geometry_mismatch_count']}.",
        f"O1/O2 status downgrades: {n['O1_O2_status_downgrade_count']}; coverage losses: {n['O1_O2_coverage_loss_count']}.",
        "",
        "## Property and non-dilution execution",
        "",
        f"Randomized cases: {n['property_randomized_case_count']}; total with third-fleet regressions: {n['property_case_count']}; families: {n['property_family_count']}; strata: {n['property_stratum_count']}; third-fleet classes: {n['third_fleet_class_count']}.",
        f"Fixed-suspicion cases: {n['non_dilution_case_count']}; executed N-sweep rows: {n['non_dilution_sweep_row_count']} at N={{0,3,10,20,40}}, with conceptual N-to-infinity formula recorded per case.",
        "",
        "## Named regression execution",
        "",
        f"Seat4 targeted cases: {reg['seat4_targeted_case_count']}; v5 rises: {bundle['fleet']['regressions']['seat4']['targeted_five']['v5_rise_case_count']}.",
        f"Seat4 window: {reg['seat4_window_parent_case_count']} parent cases, {reg['seat4_window_v4_known_positive_case_count']} v4 known-positive cases, {reg['seat4_window_v5_rise_case_count']} v5 rises within those cases.",
        f"Lens1 W000002/B5 cases: {reg['lens1_case_count']}; both complete frozen surfaces and rule transcripts are in `fleet_probe_results.json`.",
        f"Lens2 P5: {reg['lens2_P5_parent_case_count']} parent cases, {reg['lens2_P5_v4_known_violation_case_count']} reproduced v4 known-violation cases, {bundle['fleet']['regressions']['lens2']['P5']['v5_prevented_v4_known_violation_count']} prevented by v5, {reg['lens2_P5_v5_rise_case_count']} total v5 rises across the full parent hunt.",
        f"P4 reference-confidence loss: {reg['P4_reference_confidence_loss']:.16g}; status downgrades: {reg['P4_status_downgrade_count']}.",
        f"Known-adverse information-limit classifications: {reg['known_adverse_information_limit_count']}.",
        "",
        "## Replay disclosure",
        "",
        f"Replay scenes: {n['replay_scene_count']}; full per-scene/per-field rows: {n['replay_field_row_count']}; all-version zero-delta rows: {n['replay_zero_delta_row_count']}.",
        f"v5 HIGH scenes: {n['v5_high_scene_count']}; HIGH estimates within 5%: {n['v5_high_accuracy_count']} ({n['v5_high_accuracy_fraction']:.6f}); minimum cohort-scale HIGH coverage: {n['v5_minimum_per_scale_high_coverage']:.6f}.",
        "",
        "## Witness and predicate records",
        "",
        f"Upward classifications: {n['witness_classification_count']}; violations: {n['witness_violation_count']}; manual suppressions: {n['witness_manual_suppression_count']}; unclassified fields: {n['witness_unclassified_field_count']}; uncovered declared fields: {n['witness_uncovered_declared_field_count']}.",
        f"Executed predicates: {n['predicate_count']}; counterexamples observed false: {n['predicate_counterexample_false_count']}.",
        "Each witness row contains the complete ordered 18-rule transcript, explicit zero counts, measurements, frozen surface, and a surface-specific generation narrative.",
        "",
        "## Read-only and artifact records",
        "",
        f"Source-manifest mismatches before/after execution: {n['source_manifest_mismatch_count']}.",
        "Original CAD and test surfaces were not accessed. No repository file or Phase-A seal artifact was written.",
        "",
        "| Artifact | SHA-256 |",
        "|---|---|",
    ]
    lines.extend(f"| `{name}` | `{digest}` |" for name, digest in sorted(artifact_hashes.items()))
    lines.extend(["", "Complete unabridged case and replay records are in the JSON artifacts; the workbook is a numeric index to those records.", "", "LOOP_COMPLETE: L1f"])
    return "\n".join(lines) + "\n"


def emit_outputs(bundle: Mapping[str, Any], validation: Mapping[str, Any]) -> dict[str, str]:
    generated = (RESULTS_PATH, REPLAY_PATH, FLEET_PATH, WITNESS_PATH, PREDICATE_PATH, ENVELOPE_PATH, EVIDENCE_PATH, REPORT_PATH)
    collisions = [str(path) for path in generated if path.exists()]
    if collisions:
        raise FileExistsError("existing Phase B artifact(s): " + json.dumps(collisions))
    payloads = {
        RESULTS_PATH.name: bundle["results"],
        REPLAY_PATH.name: bundle["replay"],
        FLEET_PATH.name: bundle["fleet"],
        WITNESS_PATH.name: bundle["witnesses"],
        PREDICATE_PATH.name: bundle["predicates"],
        ENVELOPE_PATH.name: bundle["envelope"],
    }
    texts = {name: json_text(payload) for name, payload in payloads.items()}
    predicted = {name: hashlib.sha256(text.encode("utf-8")).hexdigest().upper() for name, text in texts.items()}
    for path in (ESTIMATOR_PATH, SCRIPT_PATH, PREREG_PATH, SEALED_CSV_PATH, SEAL_MANIFEST_PATH):
        predicted[path.name] = sha256_file(path)
    workbook = build_workbook_bytes(bundle, validation, predicted)
    predicted[EVIDENCE_PATH.name] = hashlib.sha256(workbook).hexdigest().upper()
    report = render_report(bundle, validation, predicted)
    predicted[REPORT_PATH.name] = hashlib.sha256(report.encode("utf-8")).hexdigest().upper()

    for path in (RESULTS_PATH, REPLAY_PATH, FLEET_PATH, WITNESS_PATH, PREDICATE_PATH, ENVELOPE_PATH):
        with path.open("x", encoding="utf-8", newline="\n") as stream:
            stream.write(texts[path.name])
    with EVIDENCE_PATH.open("xb") as stream:
        stream.write(workbook)
    with REPORT_PATH.open("x", encoding="utf-8", newline="\n") as stream:
        stream.write(report)
    return predicted


def verify_outputs() -> dict[str, Any]:
    expected_names = {
        "prereg.json", "PREREG_SEALED.csv", "SEAL_MANIFEST.txt",
        "feyerabend_c1_v5.py", "loop_l1f.py", "c1v7_results.json",
        "replay_delta.json", "fleet_probe_results.json", "witness_classifications.json",
        "predicate_registry.json", "honest_envelope.json", "evidence.xlsx", "REPORT.md",
    }
    observed_names = {item.name for item in CELL_DIR.iterdir()}
    if observed_names != expected_names:
        raise RuntimeError("output inventory mismatch: " + json.dumps({"missing": sorted(expected_names - observed_names), "extra": sorted(observed_names - expected_names)}))
    verify_seals()
    required = [CELL_DIR / name for name in sorted(expected_names)]
    files = [file_record(path) for path in required]
    report = REPORT_PATH.read_text(encoding="utf-8")
    replay = json.loads(REPLAY_PATH.read_text(encoding="utf-8"))
    fleet = json.loads(FLEET_PATH.read_text(encoding="utf-8"))
    witness = json.loads(WITNESS_PATH.read_text(encoding="utf-8"))
    predicates = json.loads(PREDICATE_PATH.read_text(encoding="utf-8"))
    envelope = json.loads(ENVELOPE_PATH.read_text(encoding="utf-8"))
    workbook = load_workbook(EVIDENCE_PATH, read_only=True, data_only=False)
    sheet_names = list(workbook.sheetnames)
    workbook.close()
    numeric = {
        "file_count": len(files),
        "report_terminal_line_count": sum(line == "LOOP_COMPLETE: L1f" for line in report.splitlines()),
        "report_gate_verdict_marker_count": sum(token in report for token in ("VERDICT:", "OVERALL PASS", "OVERALL FAIL")),
        "replay_scene_count": replay["publication"]["scene_count"],
        "replay_version_count": len(replay["versions"]),
        "seat4_window_known_positive_count": fleet["regressions"]["seat4"]["window_747"]["known_positive_case_count"],
        "P5_v4_known_violation_count": fleet["regressions"]["lens2"]["P5"]["v4_known_violation_case_count"],
        "witness_violation_count": witness["label_counts"].get("violation", 0),
        "witness_unclassified_field_count": witness["unclassified_field_count"],
        "predicate_count": predicates["predicate_count"],
        "predicate_observed_false_count": sum(int(row["observed_counterexample_result_false"]) for row in predicates["registry"]),
        "honest_scene_count": envelope["measured"]["scene_count"],
        "workbook_sheet_count": len(sheet_names),
    }
    if not (
        report.rstrip().endswith("LOOP_COMPLETE: L1f")
        and numeric["report_terminal_line_count"] == 1
        and numeric["report_gate_verdict_marker_count"] == 0
        and numeric["replay_scene_count"] == 400
        and numeric["replay_version_count"] == 3
        and numeric["seat4_window_known_positive_count"] == 747
        and numeric["P5_v4_known_violation_count"] == 50
        and numeric["witness_violation_count"] == 0
        and numeric["witness_unclassified_field_count"] == 0
        and numeric["predicate_count"] == numeric["predicate_observed_false_count"]
        and numeric["honest_scene_count"] == 400
        and numeric["workbook_sheet_count"] == 10
    ):
        raise RuntimeError("output verification mismatch: " + json.dumps(numeric, sort_keys=True))
    return {"numeric": numeric, "files": files, "workbook_sheet_names": sheet_names}


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--compute-only", action="store_true")
    parser.add_argument("--verify-only", action="store_true")
    args = parser.parse_args(argv)
    if args.compute_only and args.verify_only:
        parser.error("choose at most one mode")
    if args.verify_only:
        verification = verify_outputs()
        print(json.dumps(verification["numeric"], ensure_ascii=False, sort_keys=True))
        return 0
    bundle = compute_all()
    validation = numerical_validation(bundle)
    if args.compute_only:
        print(json.dumps(validation, ensure_ascii=False, sort_keys=True))
        return 0
    emit_outputs(bundle, validation)
    verification = verify_outputs()
    print(json.dumps(verification["numeric"], ensure_ascii=False, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
