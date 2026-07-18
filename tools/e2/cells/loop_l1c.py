#!/usr/bin/env python3
"""E2 loop L1c: sealed estimator repair and two-cohort replay.

All writes are confined to this file's directory.  Existing C1/L1/L1b outputs
are read-only inputs; no CAD drawing or held-out/test surface is opened.
"""

from __future__ import annotations

import argparse
import hashlib
import importlib.util
import io
import json
import math
import platform
import sys
import time
from collections import Counter
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import load_workbook


sys.dont_write_bytecode = True

CELL_DIR = Path(__file__).resolve().parent
SCRIPT_PATH = CELL_DIR / "loop_l1c.py"
ESTIMATOR_PATH = CELL_DIR / "feyerabend_c1_v2.py"
PREREG_PATH = CELL_DIR / "prereg.json"
EVIDENCE_PATH = CELL_DIR / "evidence.xlsx"
C1V4_RESULTS_PATH = CELL_DIR / "c1v4_results.json"
REPLAY_DELTA_PATH = CELL_DIR / "replay_delta.json"
REPORT_PATH = CELL_DIR / "REPORT.md"

PACKET_PATH = Path(r"D:\runs\e2_program\build\PACKET_loop_L1c_estimator_repair.md")
LENS2_PATH = Path(r"D:\runs\e2_program\chainverify_L1b\lens2_stats.md")
DOSSIER_PATH = Path(
    r"D:\dev\99_tools\autocad-sdk-router\reports\e2\dossiers\feyerabend_P2.md"
)
REPO_C1_PATH = Path(
    r"D:\dev\99_tools\autocad-sdk-router\tools\e2\cells\feyerabend_c1.py"
)
ARTIFACT_C1_PATH = Path(
    r"D:\runs\e2_program\cells\feyerabend_c1\feyerabend_c1.py"
)
V1_SCENE_DIR = Path(r"D:\runs\e2_program\cells\feyerabend_c0\scenes")
V1_BASELINE_PATH = Path(r"D:\runs\e2_program\cells\feyerabend_c1\results.json")
L1B_DIR = Path(r"D:\runs\e2_program\cells\loop_l1b")
L1B_SCRIPT_PATH = L1B_DIR / "loop_l1b.py"
L1B_SCENE_DIR = L1B_DIR / "scenes_v3"
L1B_BASELINE_PATH = L1B_DIR / "c1v3_results.json"

EXPECTED_PREREG_SHA256 = "30f6d0f7db9c5a9531183ec317936d4c5d3dda98139299d8ff43aeee68183fa8"
EXPECTED_EVIDENCE_PREREPAIR_SHA256 = "3013a276aa1c1dd0a4f8869cb4a83eff8ce31e980f49a7a7690cdda2b2f87041"
EXPECTED_SCENE_COUNT = 200
STATUS_RANK = {"NONE": 0, "LOW": 1, "HIGH": 2}


def load_module(name: str, path: Path) -> Any:
    if not path.is_file():
        raise FileNotFoundError(path)
    spec = importlib.util.spec_from_file_location(name, path)
    if spec is None or spec.loader is None:
        raise ImportError(f"cannot load module from {path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[name] = module
    spec.loader.exec_module(module)
    return module


C1 = load_module("loop_l1c_estimator", ESTIMATOR_PATH)
L1B = load_module("loop_l1c_l1b_readonly", L1B_SCRIPT_PATH)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def canonical_bytes(value: Any) -> bytes:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8")


def canonical_sha256(value: Any) -> str:
    return hashlib.sha256(canonical_bytes(value)).hexdigest()


def write_json_atomic(path: Path, payload: Any) -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(
        json.dumps(payload, ensure_ascii=False, sort_keys=True, indent=2, allow_nan=False)
        + "\n",
        encoding="utf-8",
    )
    temporary.replace(path)


def write_text_atomic(path: Path, text: str) -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(text, encoding="utf-8")
    temporary.replace(path)


def verify_dual_seal() -> dict[str, Any]:
    prereg_sha = sha256_file(PREREG_PATH)
    evidence_sha = sha256_file(EVIDENCE_PATH)
    if prereg_sha != EXPECTED_PREREG_SHA256:
        raise RuntimeError(f"prereg seal mismatch: {prereg_sha}")
    prereg = json.loads(PREREG_PATH.read_text(encoding="utf-8"))
    canonical = json.dumps(
        prereg, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    )
    workbook = load_workbook(EVIDENCE_PATH, data_only=False, read_only=False)
    if workbook.sheetnames != ["PREREG"]:
        raise RuntimeError(f"unexpected evidence sheets: {workbook.sheetnames}")
    sheet = workbook["PREREG"]
    if sheet["B5"].value != canonical:
        raise RuntimeError("PREREG worksheet canonical JSON differs from prereg.json")
    if PREREG_PATH.stat().st_mtime_ns >= ESTIMATOR_PATH.stat().st_mtime_ns:
        raise RuntimeError("prereg.json was not created before estimator repair")
    if (
        evidence_sha == EXPECTED_EVIDENCE_PREREPAIR_SHA256
        and EVIDENCE_PATH.stat().st_mtime_ns >= ESTIMATOR_PATH.stat().st_mtime_ns
    ):
        raise RuntimeError("evidence.xlsx was not created before estimator repair")
    return {
        "prereg_json_sha256": prereg_sha,
        "evidence_xlsx_pre_repair_sha256": EXPECTED_EVIDENCE_PREREPAIR_SHA256,
        "evidence_xlsx_sha256": evidence_sha,
        "prereg_canonical_sha256": hashlib.sha256(canonical.encode("utf-8")).hexdigest(),
        "sheet_names": workbook.sheetnames,
        "prereg_rows": sheet.max_row,
        "prereg_columns": sheet.max_column,
        "created_before_estimator": True,
        "content_equal": True,
        "payload": prereg,
    }


def file_record(path: Path) -> dict[str, Any]:
    return {
        "path": str(path),
        "bytes": path.stat().st_size,
        "sha256": sha256_file(path),
    }


def directory_record(path: Path) -> dict[str, Any]:
    files = sorted(path.glob("scene_*.json"))
    rows = [
        {"name": file.name, "bytes": file.stat().st_size, "sha256": sha256_file(file)}
        for file in files
    ]
    return {
        "path": str(path),
        "file_count": len(rows),
        "bytes": sum(row["bytes"] for row in rows),
        "digest": canonical_sha256(rows),
    }


READONLY_FILES = (
    PACKET_PATH,
    LENS2_PATH,
    DOSSIER_PATH,
    REPO_C1_PATH,
    ARTIFACT_C1_PATH,
    L1B_SCRIPT_PATH,
    L1B_BASELINE_PATH,
    V1_BASELINE_PATH,
)
READONLY_DIRS = (L1B_SCENE_DIR, V1_SCENE_DIR)


def readonly_manifest() -> dict[str, Any]:
    files = [file_record(path) for path in READONLY_FILES]
    directories = [directory_record(path) for path in READONLY_DIRS]
    return {
        "files": files,
        "directories": directories,
        "digest": canonical_sha256({"files": files, "directories": directories}),
    }


def compare_manifests(before: Mapping[str, Any], after: Mapping[str, Any]) -> dict[str, Any]:
    before_files = {row["path"]: row for row in before["files"]}
    after_files = {row["path"]: row for row in after["files"]}
    before_dirs = {row["path"]: row for row in before["directories"]}
    after_dirs = {row["path"]: row for row in after["directories"]}
    mismatches: list[dict[str, Any]] = []
    for key in sorted(set(before_files) | set(after_files)):
        if before_files.get(key) != after_files.get(key):
            mismatches.append(
                {"kind": "file", "path": key, "before": before_files.get(key), "after": after_files.get(key)}
            )
    for key in sorted(set(before_dirs) | set(after_dirs)):
        if before_dirs.get(key) != after_dirs.get(key):
            mismatches.append(
                {"kind": "directory", "path": key, "before": before_dirs.get(key), "after": after_dirs.get(key)}
            )
    return {
        "before": before,
        "after": after,
        "mismatch_count": len(mismatches),
        "mismatches": mismatches,
    }


def run_packet_selftests(stream: io.TextIOBase) -> dict[str, Any]:
    original_l1b_c1 = L1B.C1
    original_c1_selftests = C1.run_selftests
    try:
        L1B.C1 = C1
        C1.run_selftests = C1.run_legacy_selftests
        inherited_stream = io.StringIO()
        return_code, inherited = L1B.run_selftests(inherited_stream)
    finally:
        L1B.C1 = original_l1b_c1
        C1.run_selftests = original_c1_selftests
    print(inherited_stream.getvalue().rstrip("\n"), file=stream)
    if return_code != 0 or inherited["passed"] != 17 or inherited["total"] != 17:
        raise AssertionError(
            f"inherited L1b selftests: rc={return_code} passed={inherited['passed']} total={inherited['total']}"
        )

    counterexample = C1.counterexample_regression()
    property_result = C1.monotonicity_property_test()
    print(
        "SELFTEST live_counterexample_regression: "
        f"{'PASS' if counterexample['passed'] else 'FAIL'} | "
        f"old_status_rise={counterexample['old_status_rise_count']} "
        f"new_upward={sum(counterexample['new_upward_counts'].values())}",
        file=stream,
    )
    print(
        "SELFTEST seeded_monotonicity_300: "
        f"{'PASS' if property_result['passed'] else 'FAIL'} | "
        f"seed={property_result['seed']} cases={property_result['case_count']} "
        f"upward={sum(property_result['upward_counts'].values())}",
        file=stream,
    )
    tests = [
        *inherited["tests"],
        {
            "name": "live_counterexample_regression",
            "passed": bool(counterexample["passed"]),
            "detail": counterexample,
        },
        {
            "name": "seeded_monotonicity_300",
            "passed": bool(property_result["passed"]),
            "detail": property_result,
        },
    ]
    passed = sum(bool(test["passed"]) for test in tests)
    print(f"SELFTEST SUMMARY L1C: {passed}/{len(tests)} passed", file=stream)
    if passed != len(tests):
        raise AssertionError(f"L1c selftests: {passed}/{len(tests)}")
    return {
        "passed": passed,
        "total": len(tests),
        "inherited_l1b": inherited,
        "counterexample": counterexample,
        "monotonicity_property": property_result,
        "tests": tests,
    }


def load_scenes(scene_dir: Path) -> list[dict[str, Any]]:
    paths = sorted(scene_dir.glob("scene_*.json"))
    if len(paths) != EXPECTED_SCENE_COUNT:
        raise RuntimeError(f"{scene_dir}: expected 200 scenes, found {len(paths)}")
    scenes: list[dict[str, Any]] = []
    for path in paths:
        scene = json.loads(path.read_text(encoding="utf-8"))
        scene["_input_file"] = path.name
        scenes.append(scene)
    return scenes


def replay_cohort(scene_dir: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    rows = [C1.evaluate_scene(scene) for scene in load_scenes(scene_dir)]
    if len({row["scene_id"] for row in rows}) != EXPECTED_SCENE_COUNT:
        raise RuntimeError(f"duplicate scene ids in {scene_dir}")
    return rows, C1.aggregate_results(rows)


def nested_get(value: Mapping[str, Any], path: str) -> Any:
    current: Any = value
    for key in path.split("."):
        current = current[key]
    return current


COMMON_NUMERIC_FIELDS = (
    "input_anchor_count",
    "scale_kappa",
    "truth_unit_scale",
    "scale_estimate",
    "e_s",
    "relative_error",
    "confidence_score",
    "anchor_model.display_per_raw",
    "anchor_model.mm_per_raw",
    "anchor_model.consensus_weight",
    "anchor_model.log_mad",
    "anchor_model.n_independent",
    "anchor_model.n_spatial_bins",
    "anchor_model.confidence_score",
    "anchor_model.reference_span",
    "anchor_model.reference_consensus_weight",
    "anchor_model.reference_log_mad",
    "anchor_model.reference_n_independent",
    "anchor_model.reference_n_spatial_bins",
    "anchor_model.reference_confidence_score",
)
COMMON_CATEGORICAL_FIELDS = (
    "status",
    "unit_status",
    "reference_status",
    "physical_unit",
)


def numeric_equal(left: Any, right: Any) -> bool:
    if left is None or right is None:
        return left is right
    return math.isclose(float(left), float(right), rel_tol=1e-12, abs_tol=1e-15)


def distribution(values: Iterable[Any]) -> dict[str, Any]:
    return C1.numeric_distribution(
        float(value) for value in values if value is not None
    )


def field_delta(
    baseline_rows: Sequence[Mapping[str, Any]], repaired_rows: Sequence[Mapping[str, Any]]
) -> dict[str, Any]:
    baseline_by_id = {str(row["scene_id"]): row for row in baseline_rows}
    repaired_by_id = {str(row["scene_id"]): row for row in repaired_rows}
    if set(baseline_by_id) != set(repaired_by_id):
        raise RuntimeError("baseline and replay scene ids differ")
    numeric: dict[str, Any] = {}
    categorical: dict[str, Any] = {}
    changed_scenes: set[str] = set()
    digest_before_rows: list[dict[str, Any]] = []
    digest_after_rows: list[dict[str, Any]] = []
    for path in COMMON_NUMERIC_FIELDS:
        before_values = [nested_get(baseline_by_id[key], path) for key in sorted(baseline_by_id)]
        after_values = [nested_get(repaired_by_id[key], path) for key in sorted(repaired_by_id)]
        changed_ids = [
            key
            for key in sorted(baseline_by_id)
            if not numeric_equal(
                nested_get(baseline_by_id[key], path),
                nested_get(repaired_by_id[key], path),
            )
        ]
        changed_scenes.update(changed_ids)
        numeric[path] = {
            "changed_scene_count": len(changed_ids),
            "before": distribution(before_values),
            "after": distribution(after_values),
        }
    for path in COMMON_CATEGORICAL_FIELDS:
        before_values = [str(nested_get(baseline_by_id[key], path)) for key in sorted(baseline_by_id)]
        after_values = [str(nested_get(repaired_by_id[key], path)) for key in sorted(repaired_by_id)]
        changed_ids = [
            key
            for key in sorted(baseline_by_id)
            if nested_get(baseline_by_id[key], path)
            != nested_get(repaired_by_id[key], path)
        ]
        changed_scenes.update(changed_ids)
        categorical[path] = {
            "changed_scene_count": len(changed_ids),
            "before_counts": dict(sorted(Counter(before_values).items())),
            "after_counts": dict(sorted(Counter(after_values).items())),
        }
    for key in sorted(baseline_by_id):
        digest_before_rows.append(
            {
                "scene_id": key,
                "numeric": {
                    path: nested_get(baseline_by_id[key], path)
                    for path in COMMON_NUMERIC_FIELDS
                },
            }
        )
        digest_after_rows.append(
            {
                "scene_id": key,
                "numeric": {
                    path: nested_get(repaired_by_id[key], path)
                    for path in COMMON_NUMERIC_FIELDS
                },
            }
        )
    before_digest = canonical_sha256(digest_before_rows)
    after_digest = canonical_sha256(digest_after_rows)
    return {
        "common_numeric_field_count": len(COMMON_NUMERIC_FIELDS),
        "common_categorical_field_count": len(COMMON_CATEGORICAL_FIELDS),
        "changed_unperturbed_scene_count": len(changed_scenes),
        "numeric_fields": numeric,
        "categorical_fields": categorical,
        "common_numeric_digest_before": before_digest,
        "common_numeric_digest_after": after_digest,
        "common_numeric_digest_equal": before_digest == after_digest,
    }


def monotonic_counts(rows: Sequence[Mapping[str, Any]], kind: str) -> dict[str, int]:
    counts = {
        "confidence_score": 0,
        "status": 0,
        "unit_status": 0,
        "reference_status": 0,
    }
    for row in rows:
        diagnostic = row["corruption_diagnostics"][kind]
        counts["confidence_score"] += int(
            float(diagnostic["confidence_score_after"])
            > float(diagnostic["confidence_score_before"]) + 1e-15
        )
        for field in ("status", "unit_status", "reference_status"):
            counts[field] += int(
                STATUS_RANK[str(diagnostic[f"{field}_after"])]
                > STATUS_RANK[str(diagnostic[f"{field}_before"])]
            )
    return counts


def corruption_delta(
    baseline_rows: Sequence[Mapping[str, Any]], repaired_rows: Sequence[Mapping[str, Any]]
) -> dict[str, Any]:
    baseline_by_id = {str(row["scene_id"]): row for row in baseline_rows}
    repaired_by_id = {str(row["scene_id"]): row for row in repaired_rows}
    output: dict[str, Any] = {}
    for kind in C1.CORRUPTIONS:
        before_counts = monotonic_counts(baseline_rows, kind)
        after_counts = monotonic_counts(repaired_rows, kind)
        changed_after_fields: dict[str, int] = {}
        for field in (
            "confidence_score_after",
            "status_after",
            "unit_status_after",
            "reference_status_after",
        ):
            changed_after_fields[field] = sum(
                (
                    not numeric_equal(
                        baseline_by_id[key]["corruption_diagnostics"][kind][field],
                        repaired_by_id[key]["corruption_diagnostics"][kind][field],
                    )
                    if field == "confidence_score_after"
                    else baseline_by_id[key]["corruption_diagnostics"][kind][field]
                    != repaired_by_id[key]["corruption_diagnostics"][kind][field]
                )
                for key in sorted(baseline_by_id)
            )
        output[kind] = {
            "before_upward_counts": before_counts,
            "after_upward_counts": after_counts,
            "upward_count_delta": {
                field: after_counts[field] - before_counts[field]
                for field in before_counts
            },
            "changed_after_field_scene_counts": changed_after_fields,
        }
    return output


def scale_summary(aggregates: Mapping[str, Any]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for scale, summary in aggregates["by_scale"].items():
        result[scale] = {
            "scene_count": summary["scene_count"],
            "high_scene_count": summary["high_scene_count"],
            "high_coverage": summary["high_coverage"],
            "high_accuracy_within_5pct": summary["high_accuracy_within_5pct"],
            "relative_error": summary["relative_error"],
            "high_relative_error": summary["high_relative_error"],
            "confidence_score": summary["confidence_score"],
        }
    return result


def scalar_delta(before: Any, after: Any) -> Any:
    if before is None or after is None:
        return None if before is after else {"before": before, "after": after}
    return float(after) - float(before)


def cohort_delta(
    name: str,
    baseline_payload: Mapping[str, Any],
    repaired_rows: Sequence[Mapping[str, Any]],
    repaired_aggregates: Mapping[str, Any],
) -> dict[str, Any]:
    baseline_rows = baseline_payload["scenes"]
    baseline_aggregates = baseline_payload["aggregates"]
    normal = field_delta(baseline_rows, repaired_rows)
    corruptions = corruption_delta(baseline_rows, repaired_rows)
    before_overall = baseline_aggregates["overall"]
    after_overall = repaired_aggregates["overall"]
    scales: dict[str, Any] = {}
    for scale in sorted(repaired_aggregates["by_scale"], key=float):
        before = baseline_aggregates["by_scale"][scale]
        after = repaired_aggregates["by_scale"][scale]
        scales[scale] = {
            "high_coverage": {
                "before": before["high_coverage"],
                "after": after["high_coverage"],
                "delta": scalar_delta(before["high_coverage"], after["high_coverage"]),
            },
            "high_accuracy_within_5pct": {
                "before": before["high_accuracy_within_5pct"],
                "after": after["high_accuracy_within_5pct"],
                "delta": scalar_delta(
                    before["high_accuracy_within_5pct"],
                    after["high_accuracy_within_5pct"],
                ),
            },
            "relative_error_median": {
                "before": before["relative_error"]["median"],
                "after": after["relative_error"]["median"],
                "delta": scalar_delta(
                    before["relative_error"]["median"],
                    after["relative_error"]["median"],
                ),
            },
            "relative_error_max": {
                "before": before["relative_error"]["max"],
                "after": after["relative_error"]["max"],
                "delta": scalar_delta(
                    before["relative_error"]["max"],
                    after["relative_error"]["max"],
                ),
            },
        }
    return {
        "cohort": name,
        "scene_count": len(repaired_rows),
        "baseline_schema": baseline_payload.get("schema"),
        "normal_unperturbed": normal,
        "overall": {
            "high_coverage": {
                "before": before_overall["high_coverage"],
                "after": after_overall["high_coverage"],
                "delta": scalar_delta(
                    before_overall["high_coverage"], after_overall["high_coverage"]
                ),
            },
            "high_accuracy_within_5pct": {
                "before": before_overall["high_accuracy_within_5pct"],
                "after": after_overall["high_accuracy_within_5pct"],
                "delta": scalar_delta(
                    before_overall["high_accuracy_within_5pct"],
                    after_overall["high_accuracy_within_5pct"],
                ),
            },
            "relative_error": {
                "before": before_overall["relative_error"],
                "after": after_overall["relative_error"],
            },
            "high_relative_error": {
                "before": before_overall["high_relative_error"],
                "after": after_overall["high_relative_error"],
            },
            "unit_status_counts": {
                "before": before_overall["unit_status_counts"],
                "after": after_overall["unit_status_counts"],
            },
            "status_counts": {
                "before": before_overall["status_counts"],
                "after": after_overall["status_counts"],
            },
        },
        "by_scale": scales,
        "corruptions": corruptions,
        "repaired_aggregate_snapshot": {
            "overall": repaired_aggregates["overall"],
            "by_scale": scale_summary(repaired_aggregates),
        },
    }


def format_value(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, float):
        return format(value, ".12g")
    return str(value)


def markdown_table(headers: Sequence[str], rows: Sequence[Sequence[Any]]) -> str:
    def escape(value: Any) -> str:
        return format_value(value).replace("|", "\\|").replace("\n", "<br>")

    output = [
        "| " + " | ".join(escape(value) for value in headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    output.extend(
        "| " + " | ".join(escape(value) for value in row) + " |" for row in rows
    )
    return "\n".join(output)


def render_report(
    seals: Mapping[str, Any],
    selftests: Mapping[str, Any],
    replay_delta: Mapping[str, Any],
    source_manifest: Mapping[str, Any],
    artifact_hashes: Mapping[str, str],
) -> str:
    counterexample = selftests["counterexample"]
    property_result = selftests["monotonicity_property"]
    lines = [
        "# E2 Loop L1c — estimator repair and cohort replay",
        "",
        f"PREREG_JSON_SHA256: `{seals['prereg_json_sha256']}`",
        f"PREREG_CANONICAL_SHA256: `{seals['prereg_canonical_sha256']}`",
        f"EVIDENCE_XLSX_PRE_REPAIR_SHA256: `{seals['evidence_xlsx_pre_repair_sha256']}`",
        f"EVIDENCE_XLSX_FINAL_SHA256: `{seals['evidence_xlsx_sha256']}`",
        "",
        "## 이중 봉인 기록",
        "",
        f"- prereg.json과 evidence.xlsx PREREG 시트 내용 일치: {int(seals['content_equal'])}",
        f"- 수리 코드 생성 전 두 파일 생성: {int(seals['created_before_estimator'])}",
        f"- PREREG 시트 크기: {seals['prereg_rows']} rows × {seals['prereg_columns']} columns",
        "- 봉인 수치: scale별 HIGH coverage minimum 0.60; HIGH accuracy minimum 0.95; perturbation confidence/status/unit_status upward count 0.",
        "- 개선 크레딧: 0.",
        "",
        "## 수리 설계",
        "",
        "- 도시에 2.3(194–229행)의 ratio log-space 최대 합의와 reference confidence 공식을 유지했다.",
        "- 도시에 2.4(231–247행)의 reference span 및 annotation-scale 제외 규칙을 유지했다.",
        "- 신규 구조 가드: 유효 ratio를 가진 DIM/TEXT anchor는 선택된 ratio mode의 inlier일 때만 reference 합의의 n_independent 및 spatial-bin support에 참여한다.",
        "- ratio-space outlier는 span-space inlier여도 `ratio_space_outlier_guard`로 격리한다. GRID 및 ratio가 없는 reference-only anchor 경로는 유지한다.",
        "- 원본 feyerabend_c1.py 수정 수: 0 files.",
        "",
        "## selftest 수치",
        "",
        markdown_table(
            ["scope", "count", "passed", "failed"],
            [
                ["기존 L1b 17종", 17, selftests["inherited_l1b"]["passed"], 17 - selftests["inherited_l1b"]["passed"]],
                ["live counterexample regression", 1, int(counterexample["passed"]), 1 - int(counterexample["passed"])],
                ["seeded random perturbations", property_result["case_count"], property_result["case_count"] - sum(property_result["upward_counts"].values()), sum(property_result["upward_counts"].values())],
            ],
        ),
        "",
        "### live 반례 회귀 전문",
        "",
        markdown_table(
            ["estimator", "phase", "confidence", "status", "unit_status", "reference_status", "ref_conf", "ref_n", "ref_bins", "guarded"],
            [
                ["original", "before", *[counterexample["old_before"][key] for key in ("confidence_score", "status", "unit_status", "reference_status", "reference_confidence_score", "reference_n_independent", "reference_n_spatial_bins", "guarded_rejection_count")]],
                ["original", "after", *[counterexample["old_after"][key] for key in ("confidence_score", "status", "unit_status", "reference_status", "reference_confidence_score", "reference_n_independent", "reference_n_spatial_bins", "guarded_rejection_count")]],
                ["C1v4", "before", *[counterexample["new_before"][key] for key in ("confidence_score", "status", "unit_status", "reference_status", "reference_confidence_score", "reference_n_independent", "reference_n_spatial_bins", "guarded_rejection_count")]],
                ["C1v4", "after", *[counterexample["new_after"][key] for key in ("confidence_score", "status", "unit_status", "reference_status", "reference_confidence_score", "reference_n_independent", "reference_n_spatial_bins", "guarded_rejection_count")]],
            ],
        ),
        "",
        f"- original status upward count: {counterexample['old_status_rise_count']}",
        f"- C1v4 upward counts: {json.dumps(counterexample['new_upward_counts'], ensure_ascii=False, sort_keys=True)}",
        "",
        "### 300종 고정 시드 단조성",
        "",
        f"- seed: {property_result['seed']}",
        f"- family counts: `{json.dumps(property_result['family_counts'], ensure_ascii=False, sort_keys=True)}`",
        f"- upward counts: `{json.dumps(property_result['upward_counts'], ensure_ascii=False, sort_keys=True)}`",
        f"- cases digest: `{property_result['cases_digest']}`",
        "",
        "## 코호트 replay 델타 전문",
        "",
    ]
    for cohort_name in ("l1b_200", "c1_v1_200"):
        cohort = replay_delta["cohorts"][cohort_name]
        overall = cohort["overall"]
        normal = cohort["normal_unperturbed"]
        lines.extend(
            [
                f"### {cohort_name}",
                "",
                markdown_table(
                    ["metric", "before", "after", "delta"],
                    [
                        ["scene_count", cohort["scene_count"], cohort["scene_count"], 0],
                        ["HIGH coverage", overall["high_coverage"]["before"], overall["high_coverage"]["after"], overall["high_coverage"]["delta"]],
                        ["HIGH accuracy within 5%", overall["high_accuracy_within_5pct"]["before"], overall["high_accuracy_within_5pct"]["after"], overall["high_accuracy_within_5pct"]["delta"]],
                        ["relative error median", overall["relative_error"]["before"]["median"], overall["relative_error"]["after"]["median"], scalar_delta(overall["relative_error"]["before"]["median"], overall["relative_error"]["after"]["median"])],
                        ["relative error max", overall["relative_error"]["before"]["max"], overall["relative_error"]["after"]["max"], scalar_delta(overall["relative_error"]["before"]["max"], overall["relative_error"]["after"]["max"])],
                        ["unperturbed changed scenes", 0, normal["changed_unperturbed_scene_count"], normal["changed_unperturbed_scene_count"]],
                    ],
                ),
                "",
                "Per-scale:",
                "",
                markdown_table(
                    ["scale", "coverage before", "coverage after", "accuracy before", "accuracy after", "relerr med before", "relerr med after", "relerr max before", "relerr max after"],
                    [
                        [
                            scale,
                            row["high_coverage"]["before"],
                            row["high_coverage"]["after"],
                            row["high_accuracy_within_5pct"]["before"],
                            row["high_accuracy_within_5pct"]["after"],
                            row["relative_error_median"]["before"],
                            row["relative_error_median"]["after"],
                            row["relative_error_max"]["before"],
                            row["relative_error_max"]["after"],
                        ]
                        for scale, row in cohort["by_scale"].items()
                    ],
                ),
                "",
                "Corruption upward counts:",
                "",
                markdown_table(
                    ["corruption", "conf before", "conf after", "status before", "status after", "unit before", "unit after", "ref before", "ref after", "status-after changed scenes", "ref-after changed scenes"],
                    [
                        [
                            kind,
                            row["before_upward_counts"]["confidence_score"],
                            row["after_upward_counts"]["confidence_score"],
                            row["before_upward_counts"]["status"],
                            row["after_upward_counts"]["status"],
                            row["before_upward_counts"]["unit_status"],
                            row["after_upward_counts"]["unit_status"],
                            row["before_upward_counts"]["reference_status"],
                            row["after_upward_counts"]["reference_status"],
                            row["changed_after_field_scene_counts"]["status_after"],
                            row["changed_after_field_scene_counts"]["reference_status_after"],
                        ]
                        for kind, row in cohort["corruptions"].items()
                    ],
                ),
                "",
                f"- common unperturbed numeric fields: {normal['common_numeric_field_count']}",
                f"- common numeric digest equal: {int(normal['common_numeric_digest_equal'])}",
                "- 재현 기술: 휘발 필드(runtime·타임스탬프) 제외 수치 전 필드 동일 (공통 정상 평가 수치 필드 범위).",
                "",
            ]
        )
    lines.extend(
        [
            "## 원본 무수정 및 산출물 SHA",
            "",
            f"- source manifest mismatch count: {source_manifest['mismatch_count']}",
            f"- source manifest before digest: `{source_manifest['before']['digest']}`",
            f"- source manifest after digest: `{source_manifest['after']['digest']}`",
            *[f"- {name}: `{digest}`" for name, digest in sorted(artifact_hashes.items())],
            "",
            "## 미해결",
            "",
            "- ratio 신호가 없는 GRID/reference-only anchor가 교란인지 진짜 증거인지 단일 관측만으로 식별하는 문제는 남는다.",
            "- 300종 속성 시험은 두 200-IR 코호트와 여섯 교란 생성 family, 고정 seed 범위의 전수 결과다. 가능한 모든 입력 변형에 대한 형식 증명은 아니다.",
            "- 이 보고서는 수치와 변화량만 기록하며 평가 게이트 판정을 출력하지 않는다.",
            "",
            "LOOP_COMPLETE: L1c",
        ]
    )
    return "\n".join(lines) + "\n"


def verify_final_artifacts(seals: Mapping[str, Any]) -> dict[str, Any]:
    required_names = {
        "prereg.json",
        "feyerabend_c1_v2.py",
        "loop_l1c.py",
        "c1v4_results.json",
        "replay_delta.json",
        "evidence.xlsx",
        "REPORT.md",
    }
    actual_names = {path.name for path in CELL_DIR.iterdir() if path.is_file()}
    if actual_names != required_names:
        raise RuntimeError(
            f"artifact names mismatch: missing={sorted(required_names-actual_names)} extra={sorted(actual_names-required_names)}"
        )
    if any(not (CELL_DIR / name).is_file() or (CELL_DIR / name).stat().st_size <= 0 for name in required_names):
        raise RuntimeError("one or more required artifacts are missing or empty")
    if sha256_file(PREREG_PATH) != seals["prereg_json_sha256"]:
        raise RuntimeError("prereg.json changed after repair")
    if sha256_file(EVIDENCE_PATH) != seals["evidence_xlsx_sha256"]:
        raise RuntimeError("evidence.xlsx changed after repair")
    report_final_line = REPORT_PATH.read_text(encoding="utf-8").rstrip("\n").splitlines()[-1]
    if report_final_line != "LOOP_COMPLETE: L1c":
        raise RuntimeError(f"unexpected report final line: {report_final_line!r}")
    c1v4 = json.loads(C1V4_RESULTS_PATH.read_text(encoding="utf-8"))
    delta = json.loads(REPLAY_DELTA_PATH.read_text(encoding="utf-8"))
    if len(c1v4["scenes"]) != 200 or len(c1v4["v1_replay"]["scenes"]) != 200:
        raise RuntimeError("replay scene counts are not 200 + 200")
    if c1v4["selftest"]["passed"] != 19 or c1v4["selftest"]["total"] != 19:
        raise RuntimeError("selftest count is not 19/19")
    if delta["source_readonly_manifest"]["mismatch_count"] != 0:
        raise RuntimeError("read-only source mismatch")
    return {
        "artifact_count": len(required_names),
        "l1b_scene_count": len(c1v4["scenes"]),
        "v1_scene_count": len(c1v4["v1_replay"]["scenes"]),
        "selftest_passed": c1v4["selftest"]["passed"],
        "selftest_total": c1v4["selftest"]["total"],
        "property_case_count": c1v4["selftest"]["monotonicity_property"]["case_count"],
        "property_upward_count": sum(c1v4["selftest"]["monotonicity_property"]["upward_counts"].values()),
        "source_manifest_mismatch_count": delta["source_readonly_manifest"]["mismatch_count"],
        "report_final_line": report_final_line,
    }


def execute_full() -> int:
    started_wall = time.perf_counter()
    started_cpu = time.process_time()
    seals = verify_dual_seal()
    source_before = readonly_manifest()

    transcript_stream = io.StringIO()
    selftests = run_packet_selftests(transcript_stream)
    transcript = transcript_stream.getvalue()

    l1b_rows, l1b_aggregates = replay_cohort(L1B_SCENE_DIR)
    v1_rows, v1_aggregates = replay_cohort(V1_SCENE_DIR)
    l1b_baseline = json.loads(L1B_BASELINE_PATH.read_text(encoding="utf-8"))
    v1_baseline = json.loads(V1_BASELINE_PATH.read_text(encoding="utf-8"))

    deltas = {
        "schema": "ariadne.e2.loop_l1c.replay_delta.v1",
        "cohorts": {
            "l1b_200": cohort_delta(
                "l1b_200", l1b_baseline, l1b_rows, l1b_aggregates
            ),
            "c1_v1_200": cohort_delta(
                "c1_v1_200", v1_baseline, v1_rows, v1_aggregates
            ),
        },
        "counterexample_regression": selftests["counterexample"],
        "monotonicity_property": selftests["monotonicity_property"],
    }
    source_after = readonly_manifest()
    source_manifest = compare_manifests(source_before, source_after)
    if source_manifest["mismatch_count"] != 0:
        raise RuntimeError(f"read-only source mismatch: {source_manifest['mismatches']}")
    deltas["source_readonly_manifest"] = source_manifest

    runtime = {
        "wall_seconds": time.perf_counter() - started_wall,
        "cpu_seconds": time.process_time() - started_cpu,
        "python": platform.python_version(),
        "platform": platform.platform(),
        "volatile_fields": ["runtime.wall_seconds", "runtime.cpu_seconds"],
    }
    c1v4 = {
        "schema": "ariadne.e2.loop_l1c.c1v4_results.v1",
        "contract": {
            "packet": str(PACKET_PATH),
            "write_root": str(CELL_DIR),
            "repo_mutation_allowed": False,
            "cad_access_allowed": False,
            "test_access_allowed": False,
            "gate_verdict_emitted": False,
        },
        "dual_seal": {key: value for key, value in seals.items() if key != "payload"},
        "sealed_configuration": seals["payload"],
        "estimator": {
            "path": str(ESTIMATOR_PATH),
            "sha256": sha256_file(ESTIMATOR_PATH),
            "original_path": str(ARTIFACT_C1_PATH),
            "original_sha256": sha256_file(ARTIFACT_C1_PATH),
            "repair_revision": "c1v4_ratio_consistency",
            "repair_principle": "ratio-bearing DIM/TEXT reference support requires selected ratio-mode inlier membership",
        },
        "selftest": {**selftests, "transcript": transcript},
        "inputs": {
            "l1b_scene_dir": directory_record(L1B_SCENE_DIR),
            "v1_scene_dir": directory_record(V1_SCENE_DIR),
            "l1b_baseline": file_record(L1B_BASELINE_PATH),
            "v1_baseline": file_record(V1_BASELINE_PATH),
        },
        "aggregates": l1b_aggregates,
        "scenes": l1b_rows,
        "v1_replay": {"aggregates": v1_aggregates, "scenes": v1_rows},
        "replay_delta": deltas["cohorts"],
        "source_readonly_manifest": source_manifest,
        "runtime": runtime,
    }

    write_json_atomic(C1V4_RESULTS_PATH, c1v4)
    write_json_atomic(REPLAY_DELTA_PATH, deltas)
    artifact_hashes = {
        "c1v4_results.json": sha256_file(C1V4_RESULTS_PATH),
        "evidence.xlsx": sha256_file(EVIDENCE_PATH),
        "feyerabend_c1_v2.py": sha256_file(ESTIMATOR_PATH),
        "loop_l1c.py": sha256_file(SCRIPT_PATH),
        "prereg.json": sha256_file(PREREG_PATH),
        "replay_delta.json": sha256_file(REPLAY_DELTA_PATH),
    }
    write_text_atomic(
        REPORT_PATH,
        render_report(seals, selftests, deltas, source_manifest, artifact_hashes),
    )
    verification = verify_final_artifacts(seals)
    print(json.dumps(verification, ensure_ascii=False, sort_keys=True))
    return 0


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--selftest",
        action="store_true",
        help="run the inherited 17 checks, live regression, and 300 seeded perturbations without writes",
    )
    arguments = parser.parse_args(argv)
    if arguments.selftest:
        stream = io.StringIO()
        results = run_packet_selftests(stream)
        print(stream.getvalue(), end="")
        return 0 if results["passed"] == results["total"] else 1
    return execute_full()


if __name__ == "__main__":
    raise SystemExit(main())
