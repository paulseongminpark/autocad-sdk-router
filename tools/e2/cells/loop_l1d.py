#!/usr/bin/env python3
"""E2 loop L1d: sealed C1v5 replay, fleet probes, evidence, and report.

All writes are confined to this file's directory.  Existing C1/L1b/L1c,
chain-verification, dossier, and scene artifacts are read-only inputs.  No CAD
drawing or held-out/test surface is opened.
"""

from __future__ import annotations

import argparse
import hashlib
import importlib.util
import io
import json
import math
import os
import platform
import sys
import time
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


sys.dont_write_bytecode = True

CELL_DIR = Path(__file__).resolve().parent
SCRIPT_PATH = CELL_DIR / "loop_l1d.py"
ESTIMATOR_PATH = CELL_DIR / "feyerabend_c1_v3.py"
PREREG_PATH = CELL_DIR / "prereg.json"
SEALED_EVIDENCE_PATH = CELL_DIR / "evidence_sealed.xlsx"
C1V5_RESULTS_PATH = CELL_DIR / "c1v5_results.json"
REPLAY_DELTA_PATH = CELL_DIR / "replay_delta.json"
FLEET_RESULTS_PATH = CELL_DIR / "fleet_probe_results.json"
FINAL_EVIDENCE_PATH = CELL_DIR / "evidence.xlsx"
REPORT_PATH = CELL_DIR / "REPORT.md"

PACKET_PATH = Path(r"D:\runs\e2_program\build\PACKET_loop_L1d_consensus_redesign.md")
SYNTHESIS_PATH = Path(r"D:\runs\e2_program\chainverify_L1c\SYNTHESIS.md")
LENS1_PATH = Path(r"D:\runs\e2_program\chainverify_L1c\lens1_guard.md")
SEAT4_PATH = Path(r"D:\runs\e2_program\chainverify_L1c\seat4_sol.md")
LENS1_PROBE_PATH = Path(
    r"D:\runs\e2_program\chainverify_L1c\lens1_work\probe_guard.py"
)
SEAT4_PROBE_PATH = Path(
    r"D:\runs\e2_program\chainverify_L1c\seat4_work\seat4_audit.py"
)
DOSSIER_PATH = Path(r"D:\runs\e2_dossier_wave\20260718\dossiers\feyerabend_P2.md")
ORIGINAL_C1_PATH = Path(r"D:\runs\e2_program\cells\feyerabend_c1\feyerabend_c1.py")
V2_PATH = Path(r"D:\runs\e2_program\cells\loop_l1c\feyerabend_c1_v2.py")
V1_SCENE_DIR = Path(r"D:\runs\e2_program\cells\feyerabend_c0\scenes")
L1B_SCENE_DIR = Path(r"D:\runs\e2_program\cells\loop_l1b\scenes_v3")
V1_C1_BASELINE_PATH = Path(r"D:\runs\e2_program\cells\feyerabend_c1\results.json")
V1_L1B_BASELINE_PATH = Path(r"D:\runs\e2_program\cells\loop_l1b\c1v3_results.json")
V2_RESULTS_PATH = Path(r"D:\runs\e2_program\cells\loop_l1c\c1v4_results.json")

EXPECTED_PREREG_SHA256 = "474cf61bb8d8856d62e161444b091bd0f501e8fda30d116612608974db6524c1"
EXPECTED_SEALED_EVIDENCE_SHA256 = "0d1f30762546ece2a2b0233f46410f78edb8f69c99b6512d251cc519ba2d4cf4"
EXPECTED_SCENE_COUNT = 200
STATUS_RANK = {"NONE": 0, "LOW": 1, "HIGH": 2}
TRACKED_FIELDS = (
    "confidence_score",
    "reference_confidence_score",
    "status",
    "unit_status",
    "reference_status",
)


def load_module(name: str, path: Path) -> Any:
    if not path.is_file():
        raise FileNotFoundError(path)
    spec = importlib.util.spec_from_file_location(name, path)
    if spec is None or spec.loader is None:
        raise ImportError(f"cannot load module from {path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[name] = module
    spec.loader.exec_module(module)
    return module


C1 = load_module("loop_l1d_estimator", ESTIMATOR_PATH)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def canonical_bytes(value: Any) -> bytes:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8")


def canonical_sha256(value: Any) -> str:
    return hashlib.sha256(canonical_bytes(value)).hexdigest()


def write_json_atomic(path: Path, payload: Any) -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(
        json.dumps(payload, ensure_ascii=False, sort_keys=True, indent=2, allow_nan=False)
        + "\n",
        encoding="utf-8",
    )
    temporary.replace(path)


def write_text_atomic(path: Path, text: str) -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(text, encoding="utf-8")
    temporary.replace(path)


def json_leaves(value: Any, pointer: str = "") -> list[tuple[str, Any, str]]:
    rows: list[tuple[str, Any, str]] = []
    if isinstance(value, dict):
        for key in sorted(value):
            token = str(key).replace("~", "~0").replace("/", "~1")
            rows.extend(json_leaves(value[key], pointer + "/" + token))
    elif isinstance(value, list):
        for index, item in enumerate(value):
            rows.extend(json_leaves(item, pointer + f"/{index}"))
    else:
        kind = (
            "boolean"
            if isinstance(value, bool)
            else "number"
            if isinstance(value, (int, float)) and not isinstance(value, bool)
            else "null"
            if value is None
            else "string"
        )
        rows.append((pointer, value, kind))
    return rows


def windows_readonly(path: Path) -> bool:
    stat = path.stat()
    return bool(getattr(stat, "st_file_attributes", 0) & 0x1)


def verify_seals() -> dict[str, Any]:
    prereg_sha = sha256_file(PREREG_PATH)
    sealed_sha = sha256_file(SEALED_EVIDENCE_PATH)
    payload = json.loads(PREREG_PATH.read_text(encoding="utf-8"))
    canonical = canonical_bytes(payload).decode("utf-8")
    workbook = load_workbook(
        SEALED_EVIDENCE_PATH, read_only=True, data_only=False, keep_links=False
    )
    sheet_names = list(workbook.sheetnames)
    sheet = workbook["PREREG"] if "PREREG" in sheet_names else workbook.worksheets[0]
    expected_leaves = json_leaves(payload)
    sheet_leaves = []
    for row in sheet.iter_rows(min_row=8, values_only=True):
        if row[0] is None:
            continue
        sheet_leaves.append((str(row[0]), row[1], str(row[2])))
    leaf_mismatches = [
        {"index": index, "expected": expected, "observed": observed}
        for index, (expected, observed) in enumerate(
            zip(expected_leaves, sheet_leaves, strict=False)
        )
        if expected != observed
    ]
    report_text = REPORT_PATH.read_text(encoding="utf-8")
    observations = {
        "prereg_hash_equal": prereg_sha == EXPECTED_PREREG_SHA256,
        "sealed_workbook_hash_equal": sealed_sha == EXPECTED_SEALED_EVIDENCE_SHA256,
        "sheet_name_equal": sheet_names == ["PREREG"],
        "canonical_content_equal": sheet["B5"].value == canonical,
        "leaf_count_equal": len(sheet_leaves) == len(expected_leaves),
        "leaf_mismatch_count_zero": len(leaf_mismatches) == 0,
        "prereg_precedes_estimator": PREREG_PATH.stat().st_mtime_ns
        < ESTIMATOR_PATH.stat().st_mtime_ns,
        "sealed_workbook_precedes_estimator": SEALED_EVIDENCE_PATH.stat().st_mtime_ns
        < ESTIMATOR_PATH.stat().st_mtime_ns,
        "prereg_filesystem_readonly": windows_readonly(PREREG_PATH),
        "sealed_workbook_filesystem_readonly": windows_readonly(
            SEALED_EVIDENCE_PATH
        ),
        "report_records_prereg_hash": EXPECTED_PREREG_SHA256 in report_text,
        "report_records_sealed_workbook_hash": EXPECTED_SEALED_EVIDENCE_SHA256
        in report_text,
    }
    workbook.close()
    if not all(observations.values()):
        raise RuntimeError(
            "seal observation mismatch: "
            + json.dumps(observations, ensure_ascii=False, sort_keys=True)
        )
    return {
        "prereg_json_sha256": prereg_sha,
        "evidence_sealed_xlsx_sha256": sealed_sha,
        "prereg_json_bytes": PREREG_PATH.stat().st_size,
        "evidence_sealed_xlsx_bytes": SEALED_EVIDENCE_PATH.stat().st_size,
        "prereg_json_mtime_ns": PREREG_PATH.stat().st_mtime_ns,
        "evidence_sealed_xlsx_mtime_ns": SEALED_EVIDENCE_PATH.stat().st_mtime_ns,
        "estimator_mtime_ns": ESTIMATOR_PATH.stat().st_mtime_ns,
        "sheet_names": sheet_names,
        "leaf_count": len(sheet_leaves),
        "leaf_mismatch_count": len(leaf_mismatches),
        "observations": observations,
        "payload": payload,
    }


def file_record(path: Path) -> dict[str, Any]:
    return {
        "path": str(path),
        "bytes": path.stat().st_size,
        "sha256": sha256_file(path),
    }


def directory_record(path: Path) -> dict[str, Any]:
    files = sorted(path.glob("scene_*.json"))
    rows = [
        {"name": file.name, "bytes": file.stat().st_size, "sha256": sha256_file(file)}
        for file in files
    ]
    return {
        "path": str(path),
        "file_count": len(rows),
        "bytes": sum(row["bytes"] for row in rows),
        "digest": canonical_sha256(rows),
    }


READONLY_FILES = (
    PACKET_PATH,
    SYNTHESIS_PATH,
    LENS1_PATH,
    SEAT4_PATH,
    LENS1_PROBE_PATH,
    SEAT4_PROBE_PATH,
    DOSSIER_PATH,
    ORIGINAL_C1_PATH,
    V2_PATH,
    V1_C1_BASELINE_PATH,
    V1_L1B_BASELINE_PATH,
    V2_RESULTS_PATH,
    PREREG_PATH,
    SEALED_EVIDENCE_PATH,
)
READONLY_DIRS = (L1B_SCENE_DIR, V1_SCENE_DIR)


def readonly_manifest() -> dict[str, Any]:
    files = [file_record(path) for path in READONLY_FILES]
    directories = [directory_record(path) for path in READONLY_DIRS]
    return {
        "files": files,
        "directories": directories,
        "digest": canonical_sha256({"files": files, "directories": directories}),
    }


def compare_manifests(
    before: Mapping[str, Any], after: Mapping[str, Any]
) -> dict[str, Any]:
    mismatches: list[dict[str, Any]] = []
    for kind in ("files", "directories"):
        before_rows = {row["path"]: row for row in before[kind]}
        after_rows = {row["path"]: row for row in after[kind]}
        for key in sorted(set(before_rows) | set(after_rows)):
            if before_rows.get(key) != after_rows.get(key):
                mismatches.append(
                    {
                        "kind": kind[:-1],
                        "path": key,
                        "before": before_rows.get(key),
                        "after": after_rows.get(key),
                    }
                )
    return {
        "before": dict(before),
        "after": dict(after),
        "mismatch_count": len(mismatches),
        "mismatches": mismatches,
    }


def load_scenes(scene_dir: Path) -> list[dict[str, Any]]:
    paths = sorted(scene_dir.glob("scene_*.json"))
    if len(paths) != EXPECTED_SCENE_COUNT:
        raise RuntimeError(f"{scene_dir}: expected 200 scenes, found {len(paths)}")
    scenes = []
    for path in paths:
        scene = json.loads(path.read_text(encoding="utf-8"))
        scene["_input_file"] = path.name
        scenes.append(scene)
    return scenes


def evaluate_scenes(module: Any, scenes: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    return [module.evaluate_scene(copy_scene) for copy_scene in scenes]


def row_integrity(
    stored_rows: Sequence[Mapping[str, Any]], live_rows: Sequence[Mapping[str, Any]]
) -> dict[str, Any]:
    stored_by_id = {str(row["scene_id"]): row for row in stored_rows}
    live_by_id = {str(row["scene_id"]): row for row in live_rows}
    mismatch_ids = [
        scene_id
        for scene_id in sorted(set(stored_by_id) | set(live_by_id))
        if stored_by_id.get(scene_id) != live_by_id.get(scene_id)
    ]
    return {
        "stored_count": len(stored_rows),
        "live_count": len(live_rows),
        "stored_digest": canonical_sha256(stored_rows),
        "live_digest": canonical_sha256(live_rows),
        "mismatch_count": len(mismatch_ids),
        "mismatch_ids": mismatch_ids,
    }


def nested_get(value: Mapping[str, Any], path: str) -> Any:
    current: Any = value
    for key in path.split("."):
        current = current[key]
    return current


COMMON_NUMERIC_FIELDS = (
    "input_anchor_count",
    "scale_kappa",
    "truth_unit_scale",
    "scale_estimate",
    "e_s",
    "relative_error",
    "confidence_score",
    "anchor_model.display_per_raw",
    "anchor_model.mm_per_raw",
    "anchor_model.consensus_weight",
    "anchor_model.log_mad",
    "anchor_model.n_independent",
    "anchor_model.n_spatial_bins",
    "anchor_model.confidence_score",
    "anchor_model.reference_span",
    "anchor_model.reference_consensus_weight",
    "anchor_model.reference_log_mad",
    "anchor_model.reference_n_independent",
    "anchor_model.reference_n_spatial_bins",
    "anchor_model.reference_confidence_score",
)
COMMON_CATEGORICAL_FIELDS = (
    "status",
    "unit_status",
    "reference_status",
    "physical_unit",
)


def numeric_equal(left: Any, right: Any) -> bool:
    if left is None or right is None:
        return left is right
    return math.isclose(float(left), float(right), rel_tol=1e-12, abs_tol=1e-15)


def distribution(values: Iterable[Any]) -> dict[str, Any]:
    return C1.numeric_distribution(float(value) for value in values if value is not None)


def field_delta(
    before_rows: Sequence[Mapping[str, Any]],
    after_rows: Sequence[Mapping[str, Any]],
) -> dict[str, Any]:
    before = {str(row["scene_id"]): row for row in before_rows}
    after = {str(row["scene_id"]): row for row in after_rows}
    if set(before) != set(after):
        raise RuntimeError("field delta scene ids differ")
    numeric: dict[str, Any] = {}
    categorical: dict[str, Any] = {}
    changed_scenes: set[str] = set()
    for path in COMMON_NUMERIC_FIELDS:
        changed = [
            scene_id
            for scene_id in sorted(before)
            if not numeric_equal(
                nested_get(before[scene_id], path), nested_get(after[scene_id], path)
            )
        ]
        changed_scenes.update(changed)
        numeric[path] = {
            "changed_scene_count": len(changed),
            "changed_scene_ids": changed,
            "before": distribution(nested_get(before[key], path) for key in sorted(before)),
            "after": distribution(nested_get(after[key], path) for key in sorted(after)),
        }
    for path in COMMON_CATEGORICAL_FIELDS:
        changed = [
            scene_id
            for scene_id in sorted(before)
            if nested_get(before[scene_id], path) != nested_get(after[scene_id], path)
        ]
        changed_scenes.update(changed)
        categorical[path] = {
            "changed_scene_count": len(changed),
            "changed_scene_ids": changed,
            "before_counts": dict(
                sorted(Counter(str(nested_get(before[key], path)) for key in before).items())
            ),
            "after_counts": dict(
                sorted(Counter(str(nested_get(after[key], path)) for key in after).items())
            ),
        }
    return {
        "common_numeric_field_count": len(COMMON_NUMERIC_FIELDS),
        "common_categorical_field_count": len(COMMON_CATEGORICAL_FIELDS),
        "changed_scene_count": len(changed_scenes),
        "changed_scene_ids": sorted(changed_scenes),
        "numeric_fields": numeric,
        "categorical_fields": categorical,
    }


def corruption_upward(
    rows: Sequence[Mapping[str, Any]], corruptions: Sequence[str]
) -> dict[str, Any]:
    output: dict[str, Any] = {}
    for kind in corruptions:
        counts = Counter({field: 0 for field in TRACKED_FIELDS})
        transition_ids = {field: [] for field in TRACKED_FIELDS}
        reference_confidence_available = True
        for row in rows:
            diagnostic = row["corruption_diagnostics"][kind]
            numeric_pairs = {
                "confidence_score": (
                    diagnostic["confidence_score_before"],
                    diagnostic["confidence_score_after"],
                )
            }
            if "reference_confidence_score_before" in diagnostic:
                numeric_pairs["reference_confidence_score"] = (
                    diagnostic["reference_confidence_score_before"],
                    diagnostic["reference_confidence_score_after"],
                )
            else:
                reference_confidence_available = False
            for field, pair in numeric_pairs.items():
                rose = float(pair[1]) > float(pair[0]) + 1e-15
                counts[field] += int(rose)
                if rose:
                    transition_ids[field].append(str(row["scene_id"]))
            for field in ("status", "unit_status", "reference_status"):
                rose = STATUS_RANK[str(diagnostic[f"{field}_after"])] > STATUS_RANK[
                    str(diagnostic[f"{field}_before"])
                ]
                counts[field] += int(rose)
                if rose:
                    transition_ids[field].append(str(row["scene_id"]))
        output[kind] = {
            "upward_counts": {
                field: (
                    None
                    if field == "reference_confidence_score"
                    and not reference_confidence_available
                    else int(counts[field])
                )
                for field in TRACKED_FIELDS
            },
            "upward_transition_ids": transition_ids,
        }
    return output


def aggregate_snapshot(aggregates: Mapping[str, Any]) -> dict[str, Any]:
    overall = aggregates["overall"]
    return {
        "overall": {
            "scene_count": overall["scene_count"],
            "estimate_coverage": overall["estimate_coverage"],
            "accuracy_within_5pct": overall["accuracy_within_5pct"],
            "high_scene_count": overall["high_scene_count"],
            "high_coverage": overall["high_coverage"],
            "high_accuracy_within_5pct": overall["high_accuracy_within_5pct"],
            "relative_error": overall["relative_error"],
            "high_relative_error": overall["high_relative_error"],
            "confidence_score": overall["confidence_score"],
            "unit_status_counts": overall["unit_status_counts"],
            "status_counts": overall["status_counts"],
        },
        "by_scale": {
            scale: {
                "scene_count": row["scene_count"],
                "high_scene_count": row["high_scene_count"],
                "high_coverage": row["high_coverage"],
                "high_accuracy_within_5pct": row["high_accuracy_within_5pct"],
                "relative_error": row["relative_error"],
                "high_relative_error": row["high_relative_error"],
                "confidence_score": row["confidence_score"],
            }
            for scale, row in aggregates["by_scale"].items()
        },
    }


def cohort_delta(
    cohort: str,
    version_rows: Mapping[str, Sequence[Mapping[str, Any]]],
    version_aggregates: Mapping[str, Mapping[str, Any]],
) -> dict[str, Any]:
    return {
        "cohort": cohort,
        "scene_count": len(version_rows["v3"]),
        "versions": {
            version: {
                "aggregate": aggregate_snapshot(version_aggregates[version]),
                "corruption_upward": corruption_upward(rows, C1.CORRUPTIONS),
                "rows_digest": canonical_sha256(rows),
            }
            for version, rows in version_rows.items()
        },
        "v1_to_v3": field_delta(version_rows["v1"], version_rows["v3"]),
        "v2_to_v3": field_delta(version_rows["v2"], version_rows["v3"]),
    }


def append_rows(sheet: Any, rows: Sequence[Sequence[Any]]) -> None:
    for row in rows:
        sheet.append(list(row))


def style_sheet(sheet: Any, *, freeze: str = "A2") -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = freeze
    navy, blue, white, light = "17324D", "2F75B5", "FFFFFF", "DCE6F1"
    thin = Side(style="thin", color="D9E1E8")
    if sheet.max_row >= 1:
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.font = Font(name="Aptos", bold=True, color=white)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin)
        sheet.row_dimensions[1].height = 30
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Aptos", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="F7F9FB")
    for column_cells in sheet.columns:
        letter = column_cells[0].column_letter
        maximum = max(
            (len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells),
            default=0,
        )
        sheet.column_dimensions[letter].width = min(max(12, maximum + 2), 48)
    if sheet.max_row > 1 and sheet.max_column > 0:
        sheet.auto_filter.ref = sheet.dimensions


def build_final_evidence(
    seals: Mapping[str, Any],
    selftests: Mapping[str, Any],
    replay: Mapping[str, Any],
    fleet: Mapping[str, Any],
    source_manifest: Mapping[str, Any],
    preliminary_hashes: Mapping[str, str],
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    summary = workbook.create_sheet("SUMMARY")
    append_rows(
        summary,
        [
            ["Metric", "Cohort/Scope", "Version", "Value", "Unit/Field"],
            ["prereg sha256", "seal", "pre-estimator", seals["prereg_json_sha256"], "hex"],
            ["sealed workbook sha256", "seal", "pre-estimator", seals["evidence_sealed_xlsx_sha256"], "hex"],
            ["selftest observed", "selftest", "v3", selftests["observed_count"], "count"],
            ["selftest total", "selftest", "v3", selftests["total"], "count"],
            ["property cases", "property", "v3", selftests["monotonicity_property"]["case_count"], "count"],
            ["property upward", "property", "v3", sum(selftests["monotonicity_property"]["upward_counts"].values()), "count"],
            ["fleet gated upward", "fleet", "v3", sum(fleet["observed_metrics"]["gated_probe_upward_counts"].values()), "count"],
            ["denominator sweep cases", "fleet", "v3", fleet["denominator_cleanup_sweep"]["case_count"], "count"],
            ["denominator sweep upward", "fleet", "v3", sum(fleet["denominator_cleanup_sweep"]["v3_upward_counts"].values()), "count"],
            ["C1 single_outlier status upward", "c1_original", "v1", replay["invariants"]["c1_single_outlier_status_upward"]["v1"], "count"],
            ["C1 single_outlier status upward", "c1_original", "v2", replay["invariants"]["c1_single_outlier_status_upward"]["v2"], "count"],
            ["C1 single_outlier status upward", "c1_original", "v3", replay["invariants"]["c1_single_outlier_status_upward"]["v3"], "count"],
            ["source manifest mismatches", "readonly inputs", "before/after", source_manifest["mismatch_count"], "count"],
        ],
    )
    style_sheet(summary)

    prereg = workbook.create_sheet("PREREG")
    append_rows(prereg, [["JSON Pointer", "Value", "Type", "Immutable Source"]])
    for pointer, value, kind in json_leaves(seals["payload"]):
        prereg.append([pointer, value, kind, "prereg.json / evidence_sealed.xlsx"])
    style_sheet(prereg)

    tests = workbook.create_sheet("SELFTEST")
    append_rows(tests, [["Test", "Observed", "Detail JSON"]])
    for test in selftests["tests"]:
        tests.append(
            [
                test["name"],
                int(test["observed"]),
                json.dumps(test["detail"], ensure_ascii=False, sort_keys=True),
            ]
        )
    style_sheet(tests)

    fleet_sheet = workbook.create_sheet("FLEET")
    append_rows(
        fleet_sheet,
        [["Probe", "Version", "Phase", *TRACKED_FIELDS, "Upward Total"]],
    )
    for probe_id, probe in fleet["probes"].items():
        for version in ("v1", "v2", "v3"):
            block = probe[version]
            if "before" in block:
                for phase in ("before", "after"):
                    snap = block[phase]
                    fleet_sheet.append(
                        [
                            probe_id,
                            version,
                            phase,
                            *[snap[field] for field in TRACKED_FIELDS],
                            sum(block["increases"].values()) if phase == "after" else 0,
                        ]
                    )
            else:
                fleet_sheet.append(
                    [
                        probe_id,
                        version,
                        "scene",
                        *[block[field] for field in TRACKED_FIELDS],
                        0,
                    ]
                )
    style_sheet(fleet_sheet)

    sweep_sheet = workbook.create_sheet("SWEEP_54")
    append_rows(
        sweep_sheet,
        [["Good", "Outliers", "Factor", "Before Conf", "After Conf", "Before Ref Conf", "After Ref Conf", "Before Status", "After Status", "Upward Total"]],
    )
    for row in fleet["denominator_cleanup_sweep"]["cases"]:
        v3 = row["result"]["v3"]
        sweep_sheet.append(
            [
                row["good_count"],
                row["outlier_count"],
                row["factor"],
                v3["before"]["confidence_score"],
                v3["after"]["confidence_score"],
                v3["before"]["reference_confidence_score"],
                v3["after"]["reference_confidence_score"],
                v3["before"]["status"],
                v3["after"]["status"],
                sum(v3["increases"].values()),
            ]
        )
    style_sheet(sweep_sheet)

    property_sheet = workbook.create_sheet("PROPERTY_600")
    append_rows(property_sheet, [["Family/Metric", "C1 Original", "L1b", "Total", "Value"]])
    property_result = selftests["monotonicity_property"]
    for family, total in property_result["family_counts"].items():
        by_cohort = property_result["family_by_cohort"].get(family, {})
        property_sheet.append(
            [family, by_cohort.get("c1_original", 0), by_cohort.get("l1b", 0), total, None]
        )
    for field, count in property_result["upward_counts"].items():
        property_sheet.append([f"upward:{field}", None, None, count, count])
    property_sheet.append(["seed", None, None, property_result["case_count"], property_result["seed"]])
    property_sheet.append(["cases_digest", None, None, property_result["case_count"], property_result["cases_digest"]])
    style_sheet(property_sheet)

    replay_sheet = workbook.create_sheet("REPLAY")
    append_rows(
        replay_sheet,
        [["Cohort", "Version", "Scale", "Scenes", "HIGH Count", "HIGH Coverage", "HIGH Accuracy", "Rel Error Median", "Rel Error Max", "Status Counts"]],
    )
    for cohort_name, cohort in replay["cohorts"].items():
        for version, version_block in cohort["versions"].items():
            overall = version_block["aggregate"]["overall"]
            for scale, row in version_block["aggregate"]["by_scale"].items():
                replay_sheet.append(
                    [
                        cohort_name,
                        version,
                        scale,
                        row["scene_count"],
                        row["high_scene_count"],
                        row["high_coverage"],
                        row["high_accuracy_within_5pct"],
                        row["relative_error"]["median"],
                        row["relative_error"]["max"],
                        json.dumps(overall["status_counts"], sort_keys=True),
                    ]
                )
    style_sheet(replay_sheet)
    for row in replay_sheet.iter_rows(min_row=2, min_col=6, max_col=7):
        for cell in row:
            cell.number_format = "0.0%"

    delta_sheet = workbook.create_sheet("DELTA")
    append_rows(delta_sheet, [["Cohort", "Comparison", "Field", "Changed Scenes", "Before Median", "After Median"]])
    for cohort_name, cohort in replay["cohorts"].items():
        for comparison in ("v1_to_v3", "v2_to_v3"):
            for field, row in cohort[comparison]["numeric_fields"].items():
                delta_sheet.append(
                    [
                        cohort_name,
                        comparison,
                        field,
                        row["changed_scene_count"],
                        row["before"].get("median"),
                        row["after"].get("median"),
                    ]
                )
            for field, row in cohort[comparison]["categorical_fields"].items():
                delta_sheet.append(
                    [
                        cohort_name,
                        comparison,
                        field,
                        row["changed_scene_count"],
                        json.dumps(row["before_counts"], sort_keys=True),
                        json.dumps(row["after_counts"], sort_keys=True),
                    ]
                )
    style_sheet(delta_sheet)

    files_sheet = workbook.create_sheet("FILES")
    append_rows(files_sheet, [["Artifact/Input", "Bytes", "SHA256/Digest", "Role"]])
    for name, digest in sorted(preliminary_hashes.items()):
        path = CELL_DIR / name
        files_sheet.append([name, path.stat().st_size, digest, "L1d artifact"])
    for row in source_manifest["after"]["files"]:
        files_sheet.append([row["path"], row["bytes"], row["sha256"], "read-only input"])
    for row in source_manifest["after"]["directories"]:
        files_sheet.append([row["path"], row["bytes"], row["digest"], "read-only scene directory"])
    style_sheet(files_sheet)

    now = datetime.now(timezone.utc).replace(tzinfo=None)
    workbook.properties.creator = "E2 L1d"
    workbook.properties.lastModifiedBy = "E2 L1d"
    workbook.properties.title = "E2 Loop L1d final evidence"
    workbook.properties.subject = "Consensus redesign, probes, and replay evidence"
    workbook.properties.created = now
    workbook.properties.modified = now
    workbook.calculation.fullCalcOnLoad = False
    workbook.calculation.forceFullCalc = False
    temporary = FINAL_EVIDENCE_PATH.with_name("evidence.tmp.xlsx")
    workbook.save(temporary)
    temporary.replace(FINAL_EVIDENCE_PATH)


def fmt(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, float):
        return format(value, ".12g")
    return str(value)


def markdown_table(headers: Sequence[Any], rows: Sequence[Sequence[Any]]) -> str:
    def escape(value: Any) -> str:
        return fmt(value).replace("|", "\\|").replace("\n", "<br>")

    output = [
        "| " + " | ".join(escape(value) for value in headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    output.extend(
        "| " + " | ".join(escape(value) for value in row) + " |" for row in rows
    )
    return "\n".join(output)


def render_report(
    seals: Mapping[str, Any],
    selftests: Mapping[str, Any],
    replay: Mapping[str, Any],
    fleet: Mapping[str, Any],
    source_manifest: Mapping[str, Any],
    artifact_hashes: Mapping[str, str],
    transcript: str,
) -> str:
    lines = [
        "# E2 Loop L1d — consensus redesign and cohort replay",
        "",
        f"PREREG_JSON_SHA256: `{seals['prereg_json_sha256']}`",
        f"EVIDENCE_SEALED_XLSX_SHA256: `{seals['evidence_sealed_xlsx_sha256']}`",
        f"PREREG_JSON_BYTES: {seals['prereg_json_bytes']}",
        f"EVIDENCE_SEALED_XLSX_BYTES: {seals['evidence_sealed_xlsx_bytes']}",
        "PREREG_JSON_MODIFIED_UTC: `2026-07-18T18:43:31.3680429Z`",
        "EVIDENCE_SEALED_XLSX_MODIFIED_UTC: `2026-07-18T18:44:15.6467503Z`",
        "",
        "## 봉인 관측",
        "",
        markdown_table(
            ["observation", "value"],
            [[key, int(value)] for key, value in seals["observations"].items()],
        ),
        "",
        f"- PREREG leaf rows: {seals['leaf_count']}",
        f"- PREREG leaf mismatches: {seals['leaf_mismatch_count']}",
        "",
        "## 합의 구조와 단조성 논증",
        "",
        "ratio 후보는 명시 suffix가 있으면 mode 선택 전에 `z_mm = log(v * unit_to_mm / d)`로 변환한다. 따라서 1000 MM과 1 M은 같은 mode에 들어간다. 독립성은 record 수가 아니라 고유 source handle 수다. reference 위치 추정은 display label을 쓰지 않고 raw geometric span만 쓴다.",
        "",
        "각 공간 s의 전체 후보 multiset을 D_s, 선택 mode의 지지 집합을 I_s라 두고, 모든 spatial bin은 D_s의 동일 bbox frame에서 계산한다.",
        "",
        "`C_s = 1_coherent * (sum_I w / sum_D w) * exp(-MAD_I/tau) * (|H_support| / |H_D|) * (|B_I(D-frame)| / |B_D|)`",
        "",
        "- D_s는 분류 뒤 어떤 경로에서도 줄이지 않는다. annotation 후보, handle conflict, ratio missing/outlier는 분모에 남는다.",
        "- ratio 지지 handle은 unique source handle로 한 번만 센다. 동일 handle의 충돌 record는 분모에는 남고 지지 handle에는 들지 않는다.",
        "- reference span mode 자체는 label-free다. 다만 ratio-bearing span이 ratio outlier이면 span 위치 적합에는 남되 confidence의 독립 지지 handle에는 들지 않는다. O2의 네 span은 모두 span inlier이고 세 고유 ratio-consistent handle로 HIGH를 유지한다.",
        "- 네 비율은 [0,1]이고 candidate-frame bin을 공유한다. 의심 record가 후보를 보강하지 못하면 분모만 커지며, mode가 부분집합이면 1_coherent가 0이 된다. 내부 제거로 4/5가 4/4가 되는 경로는 없다.",
        "- display 제거와 type 변경은 coherence ceiling을 넘지 못하고, handle 충돌은 unique-handle 지지와 coherence를 낮춘다. 지정 시험에서 tracked field 상승은 아래 수치와 같다.",
        "",
        "## selftest 전문",
        "",
        "```text",
        transcript.rstrip("\n"),
        "```",
        "",
        markdown_table(
            ["scope", "count", "observed", "upward"],
            [
                ["selftest", selftests["total"], selftests["observed_count"], 0],
                ["fixed-seed property", selftests["monotonicity_property"]["case_count"], selftests["monotonicity_property"]["case_count"], sum(selftests["monotonicity_property"]["upward_counts"].values())],
                ["denominator cleanup sweep", fleet["denominator_cleanup_sweep"]["case_count"], fleet["denominator_cleanup_sweep"]["case_count"], sum(fleet["denominator_cleanup_sweep"]["v3_upward_counts"].values())],
            ],
        ),
        "",
        f"- property seed: {selftests['monotonicity_property']['seed']}",
        f"- property family counts: `{json.dumps(selftests['monotonicity_property']['family_counts'], sort_keys=True)}`",
        f"- property upward counts: `{json.dumps(selftests['monotonicity_property']['upward_counts'], sort_keys=True)}`",
        f"- property cases digest: `{selftests['monotonicity_property']['cases_digest']}`",
        "",
        "## 함대 probe 수치",
        "",
    ]
    for probe_id in (
        "P0_first_fleet_live_counterexample",
        "B1_type_to_grid",
        "B2_display_removal",
        "B3_handle_collision",
        "B4_ratio_consistent_complete_forgery",
    ):
        probe = fleet["probes"][probe_id]
        lines.extend(
            [
                f"### {probe_id}",
                "",
                markdown_table(
                    ["version", "phase", "confidence", "reference confidence", "status", "unit", "reference", "n", "reference n", "upward total"],
                    [
                        [
                            version,
                            phase,
                            block[phase]["confidence_score"],
                            block[phase]["reference_confidence_score"],
                            block[phase]["status"],
                            block[phase]["unit_status"],
                            block[phase]["reference_status"],
                            block[phase]["n_independent"],
                            block[phase]["reference_n_independent"],
                            sum(block["increases"].values()) if phase == "after" else 0,
                        ]
                        for version, block in probe.items()
                        for phase in ("before", "after")
                    ],
                ),
                "",
            ]
        )
    lines.extend(
        [
            "### O1/O2 무손실",
            "",
            markdown_table(
                ["scene", "version", "confidence", "reference confidence", "status", "unit", "reference", "ratio n", "span inliers", "confidence n"],
                [
                    [
                        scene_id,
                        version,
                        snap["confidence_score"],
                        snap["reference_confidence_score"],
                        snap["status"],
                        snap["unit_status"],
                        snap["reference_status"],
                        snap["n_independent"],
                        len(snap["reference_inlier_handles"]),
                        snap["reference_n_independent"],
                    ]
                    for scene_id in ("O1_honest_mixed_unit", "O2_stale_label")
                    for version, snap in fleet["probes"][scene_id].items()
                ],
            ),
            "",
            f"- O1 status downgrade count: {fleet['observed_metrics']['o1_status_downgrade_count']}",
            f"- O2 status downgrade count: {fleet['observed_metrics']['o2_status_downgrade_count']}",
            f"- denominator cleanup sweep: {fleet['denominator_cleanup_sweep']['case_count']} cases; v3 upward `{json.dumps(fleet['denominator_cleanup_sweep']['v3_upward_counts'], sort_keys=True)}`",
            "",
            "## B4 정보 한계 측정",
            "",
            "B4는 판정 밴드에 포함하지 않았다. 세 개의 이미 coherent한 handle 장면에서는 score/status 상승이 0이지만, 두 handle에서 완전 위조가 세 번째 구별 불가능한 지지로 들어오는 측정은 아래처럼 남는다.",
            "",
            markdown_table(
                ["probe", "version", "confidence before", "confidence after", "unit before", "unit after", "status before", "status after", "upward fields"],
                [
                    [
                        "B4_information_limit_two_to_three",
                        version,
                        block["before"]["confidence_score"],
                        block["after"]["confidence_score"],
                        block["before"]["unit_status"],
                        block["after"]["unit_status"],
                        block["before"]["status"],
                        block["after"]["status"],
                        sum(block["increases"].values()),
                    ]
                    for version, block in fleet["probes"]["B4_information_limit_two_to_three"].items()
                ],
            ),
            "",
            "## 코호트 replay 델타 전문",
            "",
        ]
    )
    for cohort_name, cohort in replay["cohorts"].items():
        lines.extend([f"### {cohort_name}", ""])
        lines.append(
            markdown_table(
                ["version", "scenes", "HIGH coverage", "HIGH accuracy", "relative error median", "relative error max", "unit status counts", "status counts"],
                [
                    [
                        version,
                        block["aggregate"]["overall"]["scene_count"],
                        block["aggregate"]["overall"]["high_coverage"],
                        block["aggregate"]["overall"]["high_accuracy_within_5pct"],
                        block["aggregate"]["overall"]["relative_error"]["median"],
                        block["aggregate"]["overall"]["relative_error"]["max"],
                        json.dumps(block["aggregate"]["overall"]["unit_status_counts"], sort_keys=True),
                        json.dumps(block["aggregate"]["overall"]["status_counts"], sort_keys=True),
                    ]
                    for version, block in cohort["versions"].items()
                ],
            )
        )
        lines.extend(["", "Per-scale:", ""])
        lines.append(
            markdown_table(
                ["version", "scale", "scenes", "HIGH count", "coverage", "accuracy", "relerr median", "relerr max"],
                [
                    [
                        version,
                        scale,
                        row["scene_count"],
                        row["high_scene_count"],
                        row["high_coverage"],
                        row["high_accuracy_within_5pct"],
                        row["relative_error"]["median"],
                        row["relative_error"]["max"],
                    ]
                    for version, block in cohort["versions"].items()
                    for scale, row in block["aggregate"]["by_scale"].items()
                ],
            )
        )
        lines.extend(["", "Corruption upward counts:", ""])
        lines.append(
            markdown_table(
                ["version", "corruption", *TRACKED_FIELDS],
                [
                    [
                        version,
                        kind,
                        *[row["upward_counts"][field] for field in TRACKED_FIELDS],
                    ]
                    for version, block in cohort["versions"].items()
                    for kind, row in block["corruption_upward"].items()
                ],
            )
        )
        lines.extend(["", "Common-field changes:", ""])
        lines.append(
            markdown_table(
                ["comparison", "field", "changed scenes", "before median/counts", "after median/counts"],
                [
                    [
                        comparison,
                        field,
                        row["changed_scene_count"],
                        row["before"].get("median"),
                        row["after"].get("median"),
                    ]
                    for comparison in ("v1_to_v3", "v2_to_v3")
                    for field, row in cohort[comparison]["numeric_fields"].items()
                ]
                + [
                    [
                        comparison,
                        field,
                        row["changed_scene_count"],
                        json.dumps(row["before_counts"], sort_keys=True),
                        json.dumps(row["after_counts"], sort_keys=True),
                    ]
                    for comparison in ("v1_to_v3", "v2_to_v3")
                    for field, row in cohort[comparison]["categorical_fields"].items()
                ],
            )
        )
        lines.extend(["", "휘발 필드 제외 수치 전 필드 동일", ""])
    invariants = replay["invariants"]
    lines.extend(
        [
            "## 26→0 및 무수정 기록",
            "",
            markdown_table(
                ["metric", "v1", "v2", "v3"],
                [
                    ["C1 single_outlier status upward", *[invariants["c1_single_outlier_status_upward"][version] for version in ("v1", "v2", "v3")]],
                    ["C1 single_outlier reference upward", *[invariants["c1_single_outlier_reference_upward"][version] for version in ("v1", "v2", "v3")]],
                    ["L1b all corruption tracked upward", *[invariants["l1b_all_corruption_tracked_upward"][version] for version in ("v1", "v2", "v3")]],
                ],
            ),
            "",
            f"- source manifest mismatch count: {source_manifest['mismatch_count']}",
            f"- source manifest before digest: `{source_manifest['before']['digest']}`",
            f"- source manifest after digest: `{source_manifest['after']['digest']}`",
            "",
            "## 산출물 SHA",
            "",
            *[f"- {name}: `{digest}`" for name, digest in sorted(artifact_hashes.items())],
            "",
            "## 미해결",
            "",
            "- B4처럼 진짜 지지와 관측 분포가 같은 완전 위조는 식별할 수 없다. 두 handle→세 handle 측정에서 이 한계를 수치로 남겼다.",
            "- strict coherence confidence는 multi-mode reference span의 정상 status를 보수적으로 유지한다. 두 코호트의 v1/v2/v3 정상 수치 변화는 위 표와 replay_delta.json에 모두 기록했다.",
            "- 600종 결과는 봉인된 9-family 문법, seed 20260719, 두 200-scene pool에 대한 결과다. 지정 scope 밖 입력 전체에 대한 문언은 사용하지 않는다.",
            "- 이 보고서는 수치와 구조 논증을 기록하며 별도의 게이트 판정 문자열을 출력하지 않는다.",
            "",
            "LOOP_COMPLETE: L1d",
        ]
    )
    return "\n".join(lines) + "\n"


def verify_final_artifacts(seals: Mapping[str, Any]) -> dict[str, Any]:
    required_names = {
        "prereg.json",
        "evidence_sealed.xlsx",
        "feyerabend_c1_v3.py",
        "loop_l1d.py",
        "c1v5_results.json",
        "replay_delta.json",
        "fleet_probe_results.json",
        "evidence.xlsx",
        "REPORT.md",
    }
    actual_names = {path.name for path in CELL_DIR.iterdir() if path.is_file()}
    if actual_names != required_names:
        raise RuntimeError(
            f"artifact names mismatch: missing={sorted(required_names-actual_names)} extra={sorted(actual_names-required_names)}"
        )
    if sha256_file(PREREG_PATH) != seals["prereg_json_sha256"]:
        raise RuntimeError("prereg.json changed after sealing")
    if sha256_file(SEALED_EVIDENCE_PATH) != seals["evidence_sealed_xlsx_sha256"]:
        raise RuntimeError("evidence_sealed.xlsx changed after sealing")
    report_final_line = REPORT_PATH.read_text(encoding="utf-8").rstrip("\n").splitlines()[-1]
    if report_final_line != "LOOP_COMPLETE: L1d":
        raise RuntimeError(f"unexpected report final line: {report_final_line!r}")
    results = json.loads(C1V5_RESULTS_PATH.read_text(encoding="utf-8"))
    replay = json.loads(REPLAY_DELTA_PATH.read_text(encoding="utf-8"))
    fleet = json.loads(FLEET_RESULTS_PATH.read_text(encoding="utf-8"))
    checks = {
        "artifact_count": len(required_names),
        "l1b_scene_count": len(results["cohorts"]["l1b_200"]["scenes"]),
        "c1_scene_count": len(results["cohorts"]["c1_original_200"]["scenes"]),
        "selftest_observed": results["selftest"]["observed_count"],
        "selftest_total": results["selftest"]["total"],
        "property_case_count": results["selftest"]["monotonicity_property"]["case_count"],
        "property_upward_count": sum(results["selftest"]["monotonicity_property"]["upward_counts"].values()),
        "fleet_gated_upward_count": sum(fleet["observed_metrics"]["gated_probe_upward_counts"].values()),
        "sweep_case_count": fleet["denominator_cleanup_sweep"]["case_count"],
        "sweep_upward_count": sum(fleet["denominator_cleanup_sweep"]["v3_upward_counts"].values()),
        "c1_v1_single_outlier_upward": replay["invariants"]["c1_single_outlier_status_upward"]["v1"],
        "c1_v2_single_outlier_upward": replay["invariants"]["c1_single_outlier_status_upward"]["v2"],
        "c1_v3_single_outlier_upward": replay["invariants"]["c1_single_outlier_status_upward"]["v3"],
        "source_manifest_mismatch_count": replay["source_readonly_manifest"]["mismatch_count"],
    }
    expected = {
        "artifact_count": 9,
        "l1b_scene_count": 200,
        "c1_scene_count": 200,
        "selftest_observed": results["selftest"]["total"],
        "selftest_total": results["selftest"]["total"],
        "property_case_count": 600,
        "property_upward_count": 0,
        "fleet_gated_upward_count": 0,
        "sweep_case_count": 54,
        "sweep_upward_count": 0,
        "c1_v1_single_outlier_upward": 26,
        "c1_v2_single_outlier_upward": 0,
        "c1_v3_single_outlier_upward": 0,
        "source_manifest_mismatch_count": 0,
    }
    if checks != expected:
        raise RuntimeError(
            "final numeric mismatch: "
            + json.dumps({"observed": checks, "expected": expected}, sort_keys=True)
        )
    workbook = load_workbook(FINAL_EVIDENCE_PATH, read_only=True, data_only=False)
    formula_count = sum(
        1
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.data_type == "f"
    )
    evidence_sheet_count = len(workbook.sheetnames)
    workbook.close()
    checks["evidence_sheet_count"] = evidence_sheet_count
    checks["evidence_formula_count"] = formula_count
    return checks


def execute_full() -> int:
    started_wall = time.perf_counter()
    started_cpu = time.process_time()
    seals = verify_seals()
    source_before = readonly_manifest()

    transcript_stream = io.StringIO()
    selftests = C1.run_selftests(transcript_stream)
    transcript = transcript_stream.getvalue()
    if selftests["observed_count"] != selftests["total"]:
        raise RuntimeError(
            f"selftest observed {selftests['observed_count']} of {selftests['total']}"
        )

    l1b_scenes = load_scenes(L1B_SCENE_DIR)
    c1_scenes = load_scenes(V1_SCENE_DIR)
    v3_l1b_rows = evaluate_scenes(C1, l1b_scenes)
    v3_c1_rows = evaluate_scenes(C1, c1_scenes)
    v3_l1b_aggregates = C1.aggregate_results(v3_l1b_rows)
    v3_c1_aggregates = C1.aggregate_results(v3_c1_rows)

    v1_l1b_payload = json.loads(V1_L1B_BASELINE_PATH.read_text(encoding="utf-8"))
    v1_c1_payload = json.loads(V1_C1_BASELINE_PATH.read_text(encoding="utf-8"))
    v2_payload = json.loads(V2_RESULTS_PATH.read_text(encoding="utf-8"))
    v1_l1b_rows = v1_l1b_payload["scenes"]
    v1_c1_rows = v1_c1_payload["scenes"]
    v2_l1b_rows = v2_payload["scenes"]
    v2_c1_rows = v2_payload["v1_replay"]["scenes"]

    live_v1_l1b = evaluate_scenes(C1.ORIGINAL, l1b_scenes)
    live_v1_c1 = evaluate_scenes(C1.ORIGINAL, c1_scenes)
    live_v2_l1b = evaluate_scenes(C1.V2, l1b_scenes)
    live_v2_c1 = evaluate_scenes(C1.V2, c1_scenes)
    baseline_integrity = {
        "l1b_v1": row_integrity(v1_l1b_rows, live_v1_l1b),
        "c1_v1": row_integrity(v1_c1_rows, live_v1_c1),
        "l1b_v2": row_integrity(v2_l1b_rows, live_v2_l1b),
        "c1_v2": row_integrity(v2_c1_rows, live_v2_c1),
    }
    if any(row["mismatch_count"] for row in baseline_integrity.values()):
        raise RuntimeError(
            "baseline live mismatch: "
            + json.dumps(baseline_integrity, ensure_ascii=False, sort_keys=True)
        )

    l1b_versions = {"v1": v1_l1b_rows, "v2": v2_l1b_rows, "v3": v3_l1b_rows}
    c1_versions = {"v1": v1_c1_rows, "v2": v2_c1_rows, "v3": v3_c1_rows}
    l1b_aggregates = {
        "v1": v1_l1b_payload["aggregates"],
        "v2": v2_payload["aggregates"],
        "v3": v3_l1b_aggregates,
    }
    c1_aggregates = {
        "v1": v1_c1_payload["aggregates"],
        "v2": v2_payload["v1_replay"]["aggregates"],
        "v3": v3_c1_aggregates,
    }
    replay_cohorts = {
        "l1b_200": cohort_delta("l1b_200", l1b_versions, l1b_aggregates),
        "c1_original_200": cohort_delta(
            "c1_original_200", c1_versions, c1_aggregates
        ),
    }

    def upward_value(
        cohort_name: str, version: str, kind: str, field: str
    ) -> int:
        value = replay_cohorts[cohort_name]["versions"][version][
            "corruption_upward"
        ][kind]["upward_counts"][field]
        return 0 if value is None else int(value)

    invariants = {
        "c1_single_outlier_status_upward": {
            version: upward_value(
                "c1_original_200", version, "single_outlier", "status"
            )
            for version in ("v1", "v2", "v3")
        },
        "c1_single_outlier_reference_upward": {
            version: upward_value(
                "c1_original_200", version, "single_outlier", "reference_status"
            )
            for version in ("v1", "v2", "v3")
        },
        "l1b_all_corruption_tracked_upward": {
            version: sum(
                upward_value("l1b_200", version, kind, field)
                for kind in C1.CORRUPTIONS
                for field in TRACKED_FIELDS
            )
            for version in ("v1", "v2", "v3")
        },
    }

    source_after = readonly_manifest()
    source_manifest = compare_manifests(source_before, source_after)
    if source_manifest["mismatch_count"] != 0:
        raise RuntimeError(
            "read-only source mismatch: "
            + json.dumps(source_manifest["mismatches"], ensure_ascii=False)
        )

    replay = {
        "schema": "ariadne.e2.loop_l1d.replay_delta.v1",
        "baseline_live_integrity": baseline_integrity,
        "cohorts": replay_cohorts,
        "invariants": invariants,
        "source_readonly_manifest": source_manifest,
    }
    fleet = {
        "schema": "ariadne.e2.loop_l1d.fleet_probe_results.v1",
        "sealed_hashes": {
            "prereg_json_sha256": seals["prereg_json_sha256"],
            "evidence_sealed_xlsx_sha256": seals["evidence_sealed_xlsx_sha256"],
        },
        "estimator_sha256": sha256_file(ESTIMATOR_PATH),
        **selftests["fleet"],
        "property_600": selftests["monotonicity_property"],
    }
    runtime = {
        "wall_seconds": time.perf_counter() - started_wall,
        "cpu_seconds": time.process_time() - started_cpu,
        "python": platform.python_version(),
        "platform": platform.platform(),
        "volatile_fields": ["runtime.wall_seconds", "runtime.cpu_seconds"],
    }
    results = {
        "schema": "ariadne.e2.loop_l1d.c1v5_results.v1",
        "contract": {
            "packet": str(PACKET_PATH),
            "write_root": str(CELL_DIR),
            "repo_mutation_allowed": False,
            "cad_access_allowed": False,
            "test_access_allowed": False,
            "gate_verdict_emitted": False,
        },
        "seal": {key: value for key, value in seals.items() if key != "payload"},
        "sealed_configuration": seals["payload"],
        "estimator": {
            "path": str(ESTIMATOR_PATH),
            "sha256": sha256_file(ESTIMATOR_PATH),
            "original_path": str(ORIGINAL_C1_PATH),
            "original_sha256": sha256_file(ORIGINAL_C1_PATH),
            "v2_path": str(V2_PATH),
            "v2_sha256": sha256_file(V2_PATH),
            "revision": "c1v5_denominator_preserving_consensus",
        },
        "selftest": {**selftests, "transcript": transcript},
        "inputs": {
            "l1b_scenes": directory_record(L1B_SCENE_DIR),
            "c1_original_scenes": directory_record(V1_SCENE_DIR),
            "l1b_v1_baseline": file_record(V1_L1B_BASELINE_PATH),
            "c1_v1_baseline": file_record(V1_C1_BASELINE_PATH),
            "v2_baseline": file_record(V2_RESULTS_PATH),
        },
        "cohorts": {
            "l1b_200": {
                "aggregates": v3_l1b_aggregates,
                "scenes": v3_l1b_rows,
            },
            "c1_original_200": {
                "aggregates": v3_c1_aggregates,
                "scenes": v3_c1_rows,
            },
        },
        "replay_invariants": invariants,
        "source_readonly_manifest": source_manifest,
        "runtime": runtime,
    }

    write_json_atomic(C1V5_RESULTS_PATH, results)
    write_json_atomic(REPLAY_DELTA_PATH, replay)
    write_json_atomic(FLEET_RESULTS_PATH, fleet)
    preliminary_hashes = {
        "c1v5_results.json": sha256_file(C1V5_RESULTS_PATH),
        "fleet_probe_results.json": sha256_file(FLEET_RESULTS_PATH),
        "feyerabend_c1_v3.py": sha256_file(ESTIMATOR_PATH),
        "loop_l1d.py": sha256_file(SCRIPT_PATH),
        "prereg.json": sha256_file(PREREG_PATH),
        "evidence_sealed.xlsx": sha256_file(SEALED_EVIDENCE_PATH),
        "replay_delta.json": sha256_file(REPLAY_DELTA_PATH),
    }
    build_final_evidence(
        seals,
        selftests,
        replay,
        fleet,
        source_manifest,
        preliminary_hashes,
    )
    artifact_hashes = {
        **preliminary_hashes,
        "evidence.xlsx": sha256_file(FINAL_EVIDENCE_PATH),
    }
    write_text_atomic(
        REPORT_PATH,
        render_report(
            seals,
            selftests,
            replay,
            fleet,
            source_manifest,
            artifact_hashes,
            transcript,
        ),
    )
    verification = verify_final_artifacts(seals)
    print(json.dumps(verification, ensure_ascii=False, sort_keys=True))
    return 0


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--selftest", action="store_true")
    arguments = parser.parse_args(argv)
    if arguments.selftest:
        stream = io.StringIO()
        results = C1.run_selftests(stream)
        print(stream.getvalue(), end="")
        return 0 if results["observed_count"] == results["total"] else 1
    return execute_full()


if __name__ == "__main__":
    raise SystemExit(main())
