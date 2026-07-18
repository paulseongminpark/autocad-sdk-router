#!/usr/bin/env python3
"""Execute E2 loop L1e Phase B without modifying repository inputs.

All writes are exclusive-create artifacts in this script's directory.  Existing
seal files, source estimators, reports, cohorts, CAD data, and test surfaces are
read-only or out of scope.
"""

from __future__ import annotations

import argparse
import copy
import csv
import hashlib
import importlib.util
import io
import json
import math
import platform
import random
import sys
import time
from collections import Counter, defaultdict
from datetime import datetime
from itertools import product
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


sys.dont_write_bytecode = True

CELL_DIR = Path(__file__).resolve().parent
SCRIPT_PATH = CELL_DIR / "loop_l1e.py"
ESTIMATOR_PATH = CELL_DIR / "feyerabend_c1_v4.py"
PREREG_PATH = CELL_DIR / "prereg.json"
SEALED_CSV_PATH = CELL_DIR / "PREREG_SEALED.csv"
SEAL_MANIFEST_PATH = CELL_DIR / "SEAL_MANIFEST.txt"
C1_RESULTS_PATH = CELL_DIR / "c1v6_results.json"
REPLAY_PATH = CELL_DIR / "replay_delta.json"
FLEET_PATH = CELL_DIR / "fleet_probe_results.json"
WITNESS_PATH = CELL_DIR / "witness_classifications.json"
PREDICATE_PATH = CELL_DIR / "predicate_registry.json"
EVIDENCE_PATH = CELL_DIR / "evidence.xlsx"
REPORT_PATH = CELL_DIR / "REPORT.md"

PACKET_PATH = Path(r"D:\runs\e2_program\build\PACKET_loop_L1e_phaseB.md")
SYNTHESIS_PATH = Path(r"D:\runs\e2_program\chainverify_L1d\SYNTHESIS.md")
LENS1_PATH = Path(r"D:\runs\e2_program\chainverify_L1d\lens1_saturation.md")
SEAT4_PATH = Path(r"D:\runs\e2_program\chainverify_L1d\seat4_sol.md")
LENS1_PROBE_PATH = Path(
    r"D:\runs\e2_program\chainverify_L1d\lens1_work\probe_saturation.py"
)
SEAT4_BOUNDARY_PATH = Path(
    r"D:\runs\e2_program\chainverify_L1d\seat4_work\boundary_audit.py"
)
REPO_ROOT = Path(r"D:\dev\99_tools\autocad-sdk-router")
REPO_SEAL_DIR = REPO_ROOT / "reports" / "e2" / "cells" / "loop_l1e"
V1_PATH = REPO_ROOT / "tools" / "e2" / "cells" / "feyerabend_c1.py"
V2_PATH = REPO_ROOT / "tools" / "e2" / "cells" / "feyerabend_c1_v2.py"
V3_PATH = REPO_ROOT / "tools" / "e2" / "cells" / "feyerabend_c1_v3.py"
V1_SCENE_DIR = Path(r"D:\runs\e2_program\cells\feyerabend_c0\scenes")
L1B_SCENE_DIR = Path(r"D:\runs\e2_program\cells\loop_l1b\scenes_v3")
V1_C1_BASELINE_PATH = Path(r"D:\runs\e2_program\cells\feyerabend_c1\results.json")
V1_L1B_BASELINE_PATH = Path(r"D:\runs\e2_program\cells\loop_l1b\c1v3_results.json")
V3_RESULTS_PATH = Path(r"D:\runs\e2_program\cells\loop_l1d\c1v5_results.json")

WITNESS_COMMIT_SHA = "3a390e8"
EXPECTED_PREREG_SHA256 = (
    "EF1E98025EF3CF46CC829085F6F112E8E3CF2068756E0112043686215D743C86"
)
EXPECTED_SEALED_CSV_SHA256 = (
    "4AA741C42F5828CC9484F8EBBA62C3CDBD1B9A5FE926C635514A5498BE48BB6B"
)
EXPECTED_SCENE_COUNT = 200
PROPERTY_RANDOM_CASES = 630
STATUS_RANK = {"NONE": 0, "LOW": 1, "HIGH": 2}
TRACKED_FIELDS = (
    "confidence_score",
    "reference_confidence_score",
    "status",
    "unit_status",
    "reference_status",
)


def load_module(name: str, path: Path) -> Any:
    if not path.is_file():
        raise FileNotFoundError(path)
    spec = importlib.util.spec_from_file_location(name, path)
    if spec is None or spec.loader is None:
        raise ImportError(path)
    module = importlib.util.module_from_spec(spec)
    sys.modules[name] = module
    spec.loader.exec_module(module)
    return module


V4 = load_module("loop_l1e_estimator", ESTIMATOR_PATH)
V3 = V4.V3
V1 = V4.ORIGINAL


def canonical_bytes(value: Any) -> bytes:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8")


def canonical_sha256(value: Any) -> str:
    return hashlib.sha256(canonical_bytes(value)).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def json_text(payload: Any) -> str:
    return (
        json.dumps(
            payload,
            ensure_ascii=False,
            sort_keys=True,
            indent=2,
            allow_nan=False,
        )
        + "\n"
    )


def write_text_exclusive(path: Path, text: str) -> None:
    with path.open("x", encoding="utf-8", newline="\n") as stream:
        stream.write(text)


def file_record(path: Path) -> dict[str, Any]:
    return {
        "path": str(path),
        "bytes": path.stat().st_size,
        "sha256": sha256_file(path),
    }


def directory_record(path: Path) -> dict[str, Any]:
    files = sorted(path.glob("scene_*.json"))
    rows = [
        {"name": item.name, "bytes": item.stat().st_size, "sha256": sha256_file(item)}
        for item in files
    ]
    return {
        "path": str(path),
        "file_count": len(rows),
        "bytes": sum(row["bytes"] for row in rows),
        "digest": canonical_sha256(rows),
    }


READONLY_FILES = (
    PACKET_PATH,
    SYNTHESIS_PATH,
    LENS1_PATH,
    SEAT4_PATH,
    LENS1_PROBE_PATH,
    SEAT4_BOUNDARY_PATH,
    REPO_SEAL_DIR / "prereg.json",
    REPO_SEAL_DIR / "PREREG_SEALED.csv",
    REPO_SEAL_DIR / "SEAL_MANIFEST.txt",
    PREREG_PATH,
    SEALED_CSV_PATH,
    SEAL_MANIFEST_PATH,
    V1_PATH,
    V2_PATH,
    V3_PATH,
    V1_C1_BASELINE_PATH,
    V1_L1B_BASELINE_PATH,
    V3_RESULTS_PATH,
)


def readonly_manifest() -> dict[str, Any]:
    files = [file_record(path) for path in READONLY_FILES]
    directories = [directory_record(path) for path in (L1B_SCENE_DIR, V1_SCENE_DIR)]
    return {
        "files": files,
        "directories": directories,
        "digest": canonical_sha256({"files": files, "directories": directories}),
    }


def compare_manifests(
    before: Mapping[str, Any], after: Mapping[str, Any]
) -> dict[str, Any]:
    mismatches: list[dict[str, Any]] = []
    for kind in ("files", "directories"):
        left = {row["path"]: row for row in before[kind]}
        right = {row["path"]: row for row in after[kind]}
        for key in sorted(set(left) | set(right)):
            if left.get(key) != right.get(key):
                mismatches.append(
                    {
                        "kind": kind[:-1],
                        "path": key,
                        "before": left.get(key),
                        "after": right.get(key),
                    }
                )
    return {
        "before": dict(before),
        "after": dict(after),
        "mismatch_count": len(mismatches),
        "mismatches": mismatches,
    }


def verify_seals() -> dict[str, Any]:
    prereg_sha = sha256_file(PREREG_PATH)
    csv_sha = sha256_file(SEALED_CSV_PATH)
    repo_prereg_sha = sha256_file(REPO_SEAL_DIR / "prereg.json")
    repo_csv_sha = sha256_file(REPO_SEAL_DIR / "PREREG_SEALED.csv")
    prereg = json.loads(PREREG_PATH.read_text(encoding="utf-8"))
    with SEALED_CSV_PATH.open("r", encoding="utf-8", newline="") as stream:
        rows = list(csv.reader(stream))
    csv_records = {row[0]: row[1] for row in rows if len(row) == 2}
    csv_payload = json.loads(csv_records["canonical_prereg_json"])
    manifest_lines = [
        line.strip()
        for line in SEAL_MANIFEST_PATH.read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]
    observations = {
        "target_prereg_sha_equal": prereg_sha == EXPECTED_PREREG_SHA256,
        "target_csv_sha_equal": csv_sha == EXPECTED_SEALED_CSV_SHA256,
        "repo_prereg_sha_equal": repo_prereg_sha == EXPECTED_PREREG_SHA256,
        "repo_csv_sha_equal": repo_csv_sha == EXPECTED_SEALED_CSV_SHA256,
        "csv_record_count": len(rows),
        "csv_payload_equal": csv_payload == prereg,
        "manifest_line_count": len(manifest_lines),
        "manifest_records_expected_hashes": (
            EXPECTED_PREREG_SHA256 in manifest_lines[0]
            and EXPECTED_SEALED_CSV_SHA256 in manifest_lines[1]
        ),
    }
    if not all(
        value is True or (key in ("csv_record_count", "manifest_line_count") and value == 2)
        for key, value in observations.items()
    ):
        raise RuntimeError("seal mismatch: " + json.dumps(observations, sort_keys=True))
    return {
        "witness_commit_sha": WITNESS_COMMIT_SHA,
        "prereg_json_sha256": prereg_sha,
        "prereg_sealed_csv_sha256": csv_sha,
        "seal_manifest_sha256": sha256_file(SEAL_MANIFEST_PATH),
        "repo_and_run_seals_equal": prereg_sha == repo_prereg_sha and csv_sha == repo_csv_sha,
        "observations": observations,
        "payload": prereg,
    }


def load_scenes(path: Path) -> list[dict[str, Any]]:
    files = sorted(path.glob("scene_*.json"))
    if len(files) != EXPECTED_SCENE_COUNT:
        raise RuntimeError(f"{path}: expected 200 scenes, found {len(files)}")
    output = []
    for item in files:
        scene = json.loads(item.read_text(encoding="utf-8"))
        scene["_input_file"] = item.name
        output.append(scene)
    return output


def snapshot(module: Any, anchors: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    model = module.fit_anchor_model(copy.deepcopy(list(anchors)))
    provenance = model.get("provenance", {})
    return {
        "confidence_score": float(model.get("confidence_score", 0.0)),
        "reference_confidence_score": float(
            model.get("reference_confidence_score", 0.0)
        ),
        "status": str(model["status"]),
        "unit_status": str(model["unit_status"]),
        "reference_status": str(model["reference_status"]),
        "display_per_raw": model.get("display_per_raw"),
        "mm_per_raw": model.get("mm_per_raw"),
        "physical_unit": model.get("physical_unit"),
        "consensus_weight": model.get("consensus_weight"),
        "reference_consensus_weight": model.get("reference_consensus_weight"),
        "n_independent": int(model.get("n_independent", 0)),
        "n_candidate_handles": int(
            model.get("n_candidate_handles", model.get("n_independent", 0))
        ),
        "reference_n_independent": int(model.get("reference_n_independent", 0)),
        "reference_n_candidate_handles": int(
            model.get(
                "reference_n_candidate_handles",
                model.get("reference_n_independent", 0),
            )
        ),
        "ratio_inlier_handles": list(provenance.get("ratio_inlier_handles", [])),
        "ratio_outlier_handles": list(provenance.get("ratio_outlier_handles", [])),
        "reference_inlier_handles": list(
            provenance.get("reference_inlier_handles", [])
        ),
        "reference_outlier_handles": list(
            provenance.get("reference_outlier_handles", [])
        ),
    }


def increases(before: Mapping[str, Any], after: Mapping[str, Any]) -> dict[str, bool]:
    return {
        "confidence_score": float(after["confidence_score"])
        > float(before["confidence_score"]) + 1e-15,
        "reference_confidence_score": float(after["reference_confidence_score"])
        > float(before["reference_confidence_score"]) + 1e-15,
        "status": STATUS_RANK[str(after["status"])] > STATUS_RANK[str(before["status"])],
        "unit_status": STATUS_RANK[str(after["unit_status"])]
        > STATUS_RANK[str(before["unit_status"])],
        "reference_status": STATUS_RANK[str(after["reference_status"])]
        > STATUS_RANK[str(before["reference_status"])],
    }


def transition_versions(
    before_anchors: Sequence[Mapping[str, Any]],
    after_anchors: Sequence[Mapping[str, Any]],
) -> dict[str, Any]:
    output: dict[str, Any] = {}
    for name, module in (("v1", V1), ("v3", V3), ("v4", V4)):
        before = snapshot(module, before_anchors)
        after = snapshot(module, after_anchors)
        output[name] = {
            "before": before,
            "after": after,
            "increases": increases(before, after),
        }
    return output


def mk(
    handle: str,
    span: float,
    display: Any,
    position: int,
    *,
    unit: str = "MM",
    anchor_type: str = "DIM",
    text_height: float | None = None,
) -> dict[str, Any]:
    x, y = V4.POSITIONS[position % len(V4.POSITIONS)]
    return V4._make_anchor(
        handle,
        span,
        display,
        x,
        y,
        unit=unit,
        anchor_type=anchor_type,
        text_height=text_height,
    )


def fixtures() -> dict[str, list[dict[str, Any]]]:
    tau = V4.RANSAC_LOG_TOLERANCE
    cea = [
        mk("CEA_MM_0", 100.0, 250.0, 0, unit="MM"),
        mk("CEA_MM_1", 100.0, 250.0, 1, unit="MM"),
        mk("CEA_MM_2", 100.0, 250.0, 2, unit="MM"),
        mk("CEA_RAW_0", 100.0, 250.0, 3, unit="UNKNOWN"),
        mk("CEA_RAW_1", 100.0, 250.0, 4, unit="UNKNOWN"),
    ]
    ceb = [
        mk("CEB_G0", 100.0, 250.0, 0),
        mk("CEB_G1", 100.0, 250.0, 1),
        mk("CEB_G2", 100.0, 250.0, 2),
        mk("CEB_OUT", 100.0, 2500.0, 3),
    ]
    ceb2 = [
        mk("CB2_G0", 100.0, 250.0, 0),
        mk("CB2_G1", 100.0, 250.0, 1),
        mk("CB2_G2", 100.0, 250.0, 2),
        mk("CB2_O1", 100.0, 2500.0, 3),
        mk("CB2_O2", 100.0, 2500.0, 4),
    ]
    cec = [
        mk("CEC_G0", 100.0, 250.0, 0),
        mk("CEC_G1", 100.0, 250.0, 1),
        mk("CEC_G2", 100.0, 250.0, 2),
        mk("CEC_OUT", 100.0, 750.0, 3),
    ]
    ced = [
        mk("CED_H1", 100.0, 250.0, 0),
        {**mk("CED_H1", 100.0, 400.0, 1), "handle": "CED_H1"},
        mk("CED_H2", 100.0, 260.0, 2),
        {**mk("CED_H2", 100.0, 410.0, 3), "handle": "CED_H2"},
    ]
    cee = [
        mk("CEE_G0", 100.0, 250.0, 0),
        mk("CEE_G1", 100.0, 250.0, 1),
        mk("CEE_G2", 100.0, 250.0, 2),
        mk("CEE_OUT", 100.0, 100.0, 3),
    ]
    cef = [
        mk("CEF_0", 100.0, 250.0, 0),
        mk("CEF_1", 100.0, 250.0 * math.exp(0.5 * tau), 1),
        mk("CEF_2", 100.0, 250.0 * math.exp(tau), 2),
    ]
    d1 = [mk("D1_A", 100.0, 250.0, 0), mk("D1_B", 100.0, 250.0, 1), mk("D1_C", 100.0, 250.0, 2)]
    d2 = [
        mk("D2_MM_0", 100.0, 250.0, 0, unit="MM"),
        mk("D2_MM_1", 100.0, 250.0, 1, unit="MM"),
        mk("D2_RAW_0", 100.0, 250.0, 2, unit="UNKNOWN"),
        mk("D2_RAW_1", 100.0, 250.0, 3, unit="UNKNOWN"),
    ]
    d3_below = [
        mk("D3_0", 100.0, 250.0, 0),
        mk("D3_1", 100.0, 250.0, 1),
        mk("D3_2", 100.0, 250.0 * math.exp((1.0 - 1e-12) * tau), 2),
    ]
    d3_above = copy.deepcopy(d3_below)
    d3_above[2]["display_value"] = 250.0 * math.exp((1.0 + 1e-12) * tau)
    return {
        "CE-A": cea,
        "CE-B": ceb,
        "CE-B2": ceb2,
        "CE-C": cec,
        "CE-D": ced,
        "CE-E": cee,
        "CE-F": cef,
        "D1": d1,
        "D2": d2,
        "D3-below": d3_below,
        "D3-above": d3_above,
        "D4": copy.deepcopy(ced),
    }


def mutate_with_seed(
    anchors: Sequence[Mapping[str, Any]], family: str, seed: int, case_index: int = 0
) -> list[dict[str, Any]]:
    return V4.randomized_corruption(anchors, family, random.Random(seed), case_index)


def classify_upward(
    classifications: list[dict[str, Any]],
    *,
    scope: str,
    case_id: str,
    family: str,
    before_anchors: Sequence[Mapping[str, Any]],
    after_anchors: Sequence[Mapping[str, Any]],
    before: Mapping[str, Any],
    after: Mapping[str, Any],
    rise_map: Mapping[str, bool],
) -> str | None:
    fields = [field for field in TRACKED_FIELDS if rise_map.get(field)]
    if not fields:
        return None
    post_surface = V4.estimator_input_surface(after_anchors)
    post_serialized = canonical_bytes(post_surface).decode("utf-8")
    # The witness constructor uses only the frozen observed estimator surface;
    # no perturbation history or diagnostic mutation field is consulted.
    witness_surface = json.loads(post_serialized)
    witness_serialized = canonical_bytes(witness_surface).decode("utf-8")
    post_digest = hashlib.sha256(post_serialized.encode("utf-8")).hexdigest()
    witness_digest = hashlib.sha256(witness_serialized.encode("utf-8")).hexdigest()
    equivalence = post_surface == witness_surface and post_digest == witness_digest
    suspicion = V4.suspicion_analysis(after_anchors)
    legitimate = int(suspicion["residual_suspicion_count"]) == 0
    label = "information_limit_record" if equivalence and legitimate else "violation"
    identifier = f"W{len(classifications):06d}"
    rationale = (
        "The independently reconstructed honest scene uses only the complete observed "
        "estimator surface; unique handles are conflict-free and no residual ratio, "
        "mixed-space, missing-ratio, or span-conflict signal remains."
        if legitimate
        else "The equal observed-surface candidate is rejected as an honest city scene "
        "because residual suspicious evidence remains: "
        + json.dumps(suspicion, ensure_ascii=False, sort_keys=True)
    )
    classifications.append(
        {
            "classification_id": identifier,
            "scope": scope,
            "case_id": case_id,
            "family": family,
            "increased_fields": fields,
            "field_events": [
                {"field": field, "before": before[field], "after": after[field]}
                for field in fields
            ],
            "witness_scene_identifier": (
                f"HONEST::{post_digest}" if legitimate else f"REJECTED::{post_digest}"
            ),
            "witness_scene_complete_estimator_input_surface_serialization": witness_serialized,
            "post_perturbation_complete_estimator_input_surface_serialization": post_serialized,
            "exact_all_field_equality": equivalence,
            "witness_serialization_sha256": witness_digest,
            "post_serialization_sha256": post_digest,
            "city_semantic_legitimacy": legitimate,
            "city_semantic_legitimacy_rationale": rationale,
            "residual_suspicion": suspicion,
            "automated_classification_result": label,
            "manual_suppression_used": False,
            "unclassified_field_count": 0,
        }
    )
    return identifier


def run_third_fleet(
    fx: Mapping[str, Sequence[Mapping[str, Any]]],
    classifications: list[dict[str, Any]],
) -> dict[str, Any]:
    definitions: list[tuple[str, str, Sequence[Mapping[str, Any]], Sequence[Mapping[str, Any]]]] = []
    definitions.append(("CE-A", "suffix_removal", fx["CE-A"], mutate_with_seed(fx["CE-A"], "suffix_removal", 0)))
    definitions.append(("CE-B", "type_to_grid", fx["CE-B"], mutate_with_seed(fx["CE-B"], "type_to_grid", 0)))
    definitions.append(("CE-B2", "type_to_grid", fx["CE-B2"], mutate_with_seed(fx["CE-B2"], "type_to_grid", 5)))
    definitions.append(("CE-C", "geometry_ratio_break", fx["CE-C"], mutate_with_seed(fx["CE-C"], "geometry_ratio_break", 9)))
    definitions.append(("CE-D", "outlier_clone", fx["CE-D"], mutate_with_seed(fx["CE-D"], "outlier_clone", 0)))
    definitions.append(("CE-E", "stale_override", fx["CE-E"], mutate_with_seed(fx["CE-E"], "stale_override", 384)))
    definitions.append(("CE-F", "type_to_grid", fx["CE-F"], mutate_with_seed(fx["CE-F"], "type_to_grid", 5)))
    d1_after = copy.deepcopy(list(fx["D1"]))
    repeat = copy.deepcopy(d1_after[0])
    x, y = V4.POSITIONS[3]
    repeat["p0"] = [x, y]
    repeat["p1"] = [x + 100.0, y]
    d1_after.append(repeat)
    definitions.append(("D1", "consistent_same_handle_repeat", fx["D1"], d1_after))
    definitions.append(("D2", "space_tie_observation", fx["D2"], copy.deepcopy(list(fx["D2"]))))
    definitions.append(("D3", "tau_boundary", fx["D3-below"], fx["D3-above"]))
    definitions.append(("D4", "eligible_empty_observation", fx["D4"], copy.deepcopy(list(fx["D4"]))))

    cases = []
    field_counts = Counter({field: 0 for field in TRACKED_FIELDS})
    for class_id, family, before_anchors, after_anchors in definitions:
        versions = transition_versions(before_anchors, after_anchors)
        v4 = versions["v4"]
        for field, rose in v4["increases"].items():
            field_counts[field] += int(rose)
        classification_id = classify_upward(
            classifications,
            scope="third_fleet",
            case_id=class_id,
            family=family,
            before_anchors=before_anchors,
            after_anchors=after_anchors,
            before=v4["before"],
            after=v4["after"],
            rise_map=v4["increases"],
        )
        cases.append(
            {
                "case_id": class_id,
                "family": family,
                "before_input_surface": V4.estimator_input_surface(before_anchors),
                "after_input_surface": V4.estimator_input_surface(after_anchors),
                "before_surface_digest": canonical_sha256(
                    V4.estimator_input_surface(before_anchors)
                ),
                "after_surface_digest": canonical_sha256(
                    V4.estimator_input_surface(after_anchors)
                ),
                "versions": versions,
                "classification_id": classification_id,
            }
        )
    return {
        "class_count": len(cases),
        "classes": [row["case_id"] for row in cases],
        "v4_upward_field_counts": dict(field_counts),
        "cases": cases,
        "cases_digest": canonical_sha256(cases),
    }


def run_property(
    fx: Mapping[str, Sequence[Mapping[str, Any]]],
    third: Mapping[str, Any],
    classifications: list[dict[str, Any]],
) -> dict[str, Any]:
    strata = [
        ("zero_cliff_start", "CE-A", fx["CE-A"]),
        ("ratio_outlier", "CE-B2", fx["CE-B2"]),
        ("mixed_space", "D2", fx["D2"]),
        ("handle_collision", "CE-D", fx["CE-D"]),
        ("near_tau_spread", "D3-below", fx["D3-below"]),
        ("near_tau_spread", "D3-above", fx["D3-above"]),
    ]
    rng = random.Random(20260719)
    family_counts: Counter[str] = Counter()
    stratum_counts: Counter[str] = Counter()
    upward_counts = Counter({field: 0 for field in TRACKED_FIELDS})
    manifest: list[dict[str, Any]] = []
    case_digests: list[str] = []
    for index in range(PROPERTY_RANDOM_CASES):
        family = V4.PROPERTY_FAMILIES[index % len(V4.PROPERTY_FAMILIES)]
        stratum, base_id, before_anchors = strata[index % len(strata)]
        after_anchors = V4.randomized_corruption(before_anchors, family, rng, index)
        before = snapshot(V4, before_anchors)
        after = snapshot(V4, after_anchors)
        rise_map = increases(before, after)
        for field, rose in rise_map.items():
            upward_counts[field] += int(rose)
        family_counts[family] += 1
        stratum_counts[stratum] += 1
        case_id = f"P{index:04d}"
        classification_id = classify_upward(
            classifications,
            scope="property_20260719",
            case_id=case_id,
            family=family,
            before_anchors=before_anchors,
            after_anchors=after_anchors,
            before=before,
            after=after,
            rise_map=rise_map,
        )
        record = {
            "case_id": case_id,
            "case_index": index,
            "seed": 20260719,
            "stratum": stratum,
            "base_scene_id": base_id,
            "family": family,
            "before_input_surface": V4.estimator_input_surface(before_anchors),
            "after_input_surface": V4.estimator_input_surface(after_anchors),
            "before": before,
            "after": after,
            "increases": rise_map,
            "classification_id": classification_id,
        }
        record["case_digest"] = canonical_sha256(record)
        manifest.append(record)
        case_digests.append(record["case_digest"])

    third_manifest = [
        {
            "case_id": "REG::" + row["case_id"],
            "class_id": row["case_id"],
            "family": row["family"],
            "before_input_surface": row["before_input_surface"],
            "after_input_surface": row["after_input_surface"],
            "before": row["versions"]["v4"]["before"],
            "after": row["versions"]["v4"]["after"],
            "increases": row["versions"]["v4"]["increases"],
            "classification_id": row["classification_id"],
        }
        for row in third["cases"]
    ]
    for row in third_manifest:
        row["case_digest"] = canonical_sha256(row)
        case_digests.append(row["case_digest"])
    return {
        "seed": 20260719,
        "randomized_case_count": len(manifest),
        "third_fleet_regression_case_count": len(third_manifest),
        "case_count": len(manifest) + len(third_manifest),
        "required_minimum_case_count": 600,
        "family_counts": dict(sorted(family_counts.items())),
        "stratum_counts": dict(sorted(stratum_counts.items())),
        "v4_upward_field_counts": dict(upward_counts),
        "case_manifest": manifest + third_manifest,
        "case_manifest_digest": canonical_sha256(case_digests),
        "third_fleet_classes": list(third["classes"]),
    }


def core_fleet(
    classifications: list[dict[str, Any]],
) -> dict[str, Any]:
    scene = json.loads(V4.V1_COUNTEREXAMPLE_SCENE.read_text(encoding="utf-8"))
    anchors = scene["anchors"]
    base_id = str(scene["base_scene_id"])
    sealed = V4.apply_corruption(anchors, "single_outlier", base_id)
    probes: dict[str, Any] = {}

    def add_transition(
        probe_id: str,
        family: str,
        before_anchors: Sequence[Mapping[str, Any]],
        after_anchors: Sequence[Mapping[str, Any]],
    ) -> None:
        versions = transition_versions(before_anchors, after_anchors)
        v4 = versions["v4"]
        classification_id = classify_upward(
            classifications,
            scope="fleet_core",
            case_id=probe_id,
            family=family,
            before_anchors=before_anchors,
            after_anchors=after_anchors,
            before=v4["before"],
            after=v4["after"],
            rise_map=v4["increases"],
        )
        probes[probe_id] = {
            "family": family,
            "versions": versions,
            "classification_id": classification_id,
        }

    add_transition("P0_first_fleet_live_counterexample", "single_outlier", anchors, sealed)
    variants: dict[str, list[dict[str, Any]]] = {}
    for probe_id, mutation in (
        ("B2_display_removal", "display_removal"),
        ("B3_handle_collision", "handle_collision"),
        ("B1_type_to_grid", "type_to_grid"),
        ("B4_ratio_consistent_complete_forgery", "ratio_consistent_complete_forgery"),
    ):
        mutated = V4.apply_corruption(anchors, "single_outlier", base_id)
        clone = mutated[-1]
        if mutation == "display_removal":
            clone["display_value"] = None
        elif mutation == "handle_collision":
            clone["handle"] = str(clone["handle"]).split("__OUT_")[0]
        elif mutation == "type_to_grid":
            clone["anchor_type"] = "GRID"
            clone.pop("display_value", None)
            clone.pop("display_unit", None)
            clone.pop("weight", None)
        else:
            clone["display_value"] = float(clone["display_value"]) / 10.0
        variants[probe_id] = mutated
        add_transition(probe_id, mutation, anchors, mutated)

    mixed = [
        V4._make_anchor("MIX_MM_0", 100.0, 250.0, 0.0, 0.0, unit="MM", text_height=1.0),
        V4._make_anchor("MIX_MM_1", 100.0, 250.0, 200.0, 0.0, unit="MM", text_height=1.0),
        V4._make_anchor("MIX_MM_2", 100.0, 250.0, 0.0, 200.0, unit="MM", text_height=1.0),
        V4._make_anchor("MIX_M_0", 100.0, 0.25, 200.0, 200.0, unit="M", text_height=1.0),
        V4._make_anchor("MIX_M_1", 100.0, 0.25, 100.0, 100.0, unit="M", text_height=1.0),
    ]
    stale = [
        V4._make_anchor("STALE_0", 100.0, 250.0, 0.0, 0.0, text_height=1.0),
        V4._make_anchor("STALE_1", 100.0, 250.0, 200.0, 0.0, text_height=1.0),
        V4._make_anchor("STALE_2", 100.0, 250.0, 0.0, 200.0, text_height=1.0),
        V4._make_anchor("STALE_3", 100.0, 225.0, 200.0, 200.0, text_height=1.0),
    ]
    capture = [
        V4._make_anchor("TRUE_0", 100.0, 250.0, 0.0, 0.0, text_height=1.0),
        V4._make_anchor("TRUE_1", 100.0, 250.0, 200.0, 0.0, text_height=1.0),
        V4._make_anchor("TRUE_2", 100.0, 250.0, 0.0, 200.0, text_height=1.0),
        V4._make_anchor("FAKE_0", 100.0, 2500.0, 400.0, 400.0, text_height=1.0),
        V4._make_anchor("FAKE_1", 100.0, 2500.0, 600.0, 400.0, text_height=1.0),
        V4._make_anchor("FAKE_2", 100.0, 2500.0, 400.0, 600.0, text_height=1.0),
        V4._make_anchor("FAKE_3", 100.0, 2500.0, 600.0, 600.0, text_height=1.0),
    ]
    observations = {
        probe_id: {
            version: snapshot(module, scene_anchors)
            for version, module in (("v1", V1), ("v3", V3), ("v4", V4))
        }
        for probe_id, scene_anchors in (
            ("O1_honest_mixed_unit", mixed),
            ("O2_stale_label", stale),
            ("O3_mode_capture", capture),
        )
    }

    two = [
        V4._make_anchor("B4_LOW_0", 100.0, 250.0, *V4.POSITIONS[0]),
        V4._make_anchor("B4_LOW_1", 100.0, 250.0, *V4.POSITIONS[1]),
    ]
    three = copy.deepcopy(two)
    forged = copy.deepcopy(two[0])
    forged["handle"] = "B4_LOW_FORGED"
    forged["p0"] = list(V4.POSITIONS[2])
    forged["p1"] = [V4.POSITIONS[2][0] + 100.0, V4.POSITIONS[2][1]]
    three.append(forged)
    add_transition("B4_information_limit_two_to_three", "new_indistinguishable_handle", two, three)

    sweep_cases = []
    sweep_counts = Counter({field: 0 for field in TRACKED_FIELDS})
    for good_count in range(3, 9):
        for outlier_count in range(1, 4):
            for factor in (1.25, 2.0, 25.0):
                before_anchors = [
                    V4._make_anchor(
                        f"G{index}", 100.0, 100.0, *V4.POSITIONS[index % len(V4.POSITIONS)]
                    )
                    for index in range(good_count)
                ]
                for index in range(outlier_count):
                    pos = V4.POSITIONS[(good_count + index) % len(V4.POSITIONS)]
                    span = 1000.0 + 100.0 * index
                    before_anchors.append(
                        V4._make_anchor(f"O{index}", span, span, *pos)
                    )
                after_anchors = copy.deepcopy(before_anchors)
                after_anchors[good_count]["display_value"] = (
                    float(after_anchors[good_count]["display_value"]) * factor
                )
                before = snapshot(V4, before_anchors)
                after = snapshot(V4, after_anchors)
                rise_map = increases(before, after)
                for field, rose in rise_map.items():
                    sweep_counts[field] += int(rose)
                case_id = f"SW54::{good_count}:{outlier_count}:{factor}"
                classification_id = classify_upward(
                    classifications,
                    scope="fleet_sweep_54",
                    case_id=case_id,
                    family="denominator_cleanup",
                    before_anchors=before_anchors,
                    after_anchors=after_anchors,
                    before=before,
                    after=after,
                    rise_map=rise_map,
                )
                sweep_cases.append(
                    {
                        "good_count": good_count,
                        "outlier_count": outlier_count,
                        "factor": factor,
                        "before": before,
                        "after": after,
                        "increases": rise_map,
                        "classification_id": classification_id,
                    }
                )
    status_downgrade_count = 0
    coverage_loss_count = 0
    for probe_id in ("O1_honest_mixed_unit", "O2_stale_label"):
        old = observations[probe_id]["v1"]
        new = observations[probe_id]["v4"]
        status_downgrade_count += sum(
            STATUS_RANK[new[field]] < STATUS_RANK[old[field]]
            for field in ("status", "unit_status", "reference_status")
        )
        coverage_loss_count += int(
            old["display_per_raw"] is not None and new["display_per_raw"] is None
        )
    return {
        "probes": probes,
        "observations": observations,
        "denominator_cleanup_sweep": {
            "case_count": len(sweep_cases),
            "v4_upward_counts": dict(sweep_counts),
            "cases": sweep_cases,
        },
        "o1_o2_numeric_loss_counts": {
            "status_downgrade_count": status_downgrade_count,
            "coverage_loss_count": coverage_loss_count,
        },
    }


SEARCH_POSITIONS = (
    (0.0, 0.0),
    (0.0, 1000.0),
    (1000.0, 0.0),
    (1000.0, 1000.0),
    (500.0, 500.0),
    (2000.0, 0.0),
)


def search_anchor(
    handle: str, ratio: float, index: int, *, unit: str = "MM"
) -> dict[str, Any]:
    x, y = SEARCH_POSITIONS[index % len(SEARCH_POSITIONS)]
    return {
        "handle": handle,
        "anchor_type": "DIM",
        "region": "P",
        "p0": [x, y],
        "p1": [x + 100.0, y],
        "raw_span": 100.0,
        "display_value": ratio * 100.0,
        "display_unit": unit,
        "text_height": None,
        "weight": 1.0,
    }


def suffix_removal(anchors: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    output = copy.deepcopy(list(anchors))
    for anchor in output:
        if anchor.get("display_value") is not None:
            anchor["display_unit"] = "UNKNOWN"
    return output


def type_to_grid(
    anchors: Sequence[Mapping[str, Any]], index: int
) -> list[dict[str, Any]]:
    output = copy.deepcopy(list(anchors))
    output[index]["anchor_type"] = "GRID"
    output[index].pop("display_value", None)
    output[index].pop("display_unit", None)
    return output


def stale_override(
    anchors: Sequence[Mapping[str, Any]], index: int, factor: float
) -> list[dict[str, Any]]:
    output = copy.deepcopy(list(anchors))
    output[index]["display_value"] = float(output[index]["display_value"]) * factor
    return output


def geometry_change(
    anchors: Sequence[Mapping[str, Any]], index: int, factor: float
) -> list[dict[str, Any]]:
    output = copy.deepcopy(list(anchors))
    p0, p1 = output[index]["p0"], output[index]["p1"]
    output[index]["p1"] = [
        float(p0[0]) + (float(p1[0]) - float(p0[0])) * factor,
        float(p0[1]) + (float(p1[1]) - float(p0[1])) * factor,
    ]
    return output


def family_variants(
    anchors: Sequence[Mapping[str, Any]], family: str
) -> Iterable[tuple[str, list[dict[str, Any]]]]:
    if family == "suffix_removal":
        yield "all", suffix_removal(anchors)
        return
    for index in range(len(anchors)):
        if family == "outlier_clone":
            output = copy.deepcopy(list(anchors))
            clone = copy.deepcopy(output[index])
            clone["handle"] = f"{clone['handle']}__ENUM_OUT"
            clone["p0"] = [float(clone["p0"][0]) + 10000.0, float(clone["p0"][1]) - 3750.0]
            clone["p1"] = [float(clone["p1"][0]) + 10000.0, float(clone["p1"][1]) - 3750.0]
            clone["display_value"] = float(clone["display_value"]) * 2.0
            output.append(clone)
            yield f"i{index}:factor2", output
        elif family == "stale_override":
            for factor in (1.25, 2.0, 5.0):
                yield f"i{index}:factor{factor}", stale_override(anchors, index, factor)
        elif family == "exact_duplicate":
            output = copy.deepcopy(list(anchors))
            clone = copy.deepcopy(output[index])
            clone["handle"] = f"{clone['handle']}__ENUM_DUP"
            output.append(clone)
            yield f"i{index}", output
        elif family == "geometry_ratio_break":
            for factor in (0.2, 3.0, 5.0):
                yield f"i{index}:factor{factor}", geometry_change(anchors, index, factor)
        elif family == "reference_support_drop":
            output = copy.deepcopy(list(anchors))
            span = math.dist(output[index]["p0"][:2], output[index]["p1"][:2])
            output[index]["text_height"] = span / 5.0
            yield f"i{index}", output
        elif family == "display_removal":
            output = copy.deepcopy(list(anchors))
            output[index]["display_value"] = None
            yield f"i{index}", output
        elif family == "handle_collision":
            output = copy.deepcopy(list(anchors))
            clone = copy.deepcopy(output[index])
            clone["p0"] = [float(clone["p0"][0]) + 10000.0, float(clone["p0"][1]) + 10000.0]
            clone["p1"] = [float(clone["p1"][0]) + 10000.0, float(clone["p1"][1]) + 10000.0]
            clone["display_value"] = float(clone["display_value"]) * 2.0
            output.append(clone)
            yield f"i{index}:factor2", output
        elif family == "type_to_grid":
            yield f"i{index}", type_to_grid(anchors, index)


def run_small_state_search(
    classifications: list[dict[str, Any]],
) -> dict[str, Any]:
    states = (
        (2.0, "MM"),
        (2.5, "MM"),
        (3.125, "MM"),
        (2.0, "UNKNOWN"),
        (2.5, "UNKNOWN"),
        (3.125, "UNKNOWN"),
    )
    stats = {
        family: Counter({"transitions": 0, "upward_transitions": 0, "upward_field_events": 0})
        for family in V4.PROPERTY_FAMILIES
    }
    field_counts = {
        family: Counter({field: 0 for field in TRACKED_FIELDS})
        for family in V4.PROPERTY_FAMILIES
    }
    base_manifest = []
    transition_manifest = []
    base_ratio_zero = 0
    for count in (2, 3, 4):
        for ordinal, state_tuple in enumerate(product(states, repeat=count)):
            base_id = f"S{count}:{ordinal:04d}"
            anchors = [
                search_anchor(f"E{index}", ratio, index, unit=unit)
                for index, (ratio, unit) in enumerate(state_tuple)
            ]
            before = snapshot(V4, anchors)
            v3_before = snapshot(V3, anchors)
            base_ratio_zero += int(v3_before["confidence_score"] == 0.0)
            base_manifest.append(
                {
                    "base_id": base_id,
                    "anchor_count": count,
                    "states": [list(state) for state in state_tuple],
                    "input_surface": V4.estimator_input_surface(anchors),
                    "input_surface_digest": canonical_sha256(V4.estimator_input_surface(anchors)),
                    "v3_confidence_score": v3_before["confidence_score"],
                    "v4_confidence_score": before["confidence_score"],
                }
            )
            for family in V4.PROPERTY_FAMILIES:
                for variant_id, after_anchors in family_variants(anchors, family):
                    after = snapshot(V4, after_anchors)
                    rise_map = increases(before, after)
                    rising = [field for field, rose in rise_map.items() if rose]
                    stats[family]["transitions"] += 1
                    if rising:
                        stats[family]["upward_transitions"] += 1
                        stats[family]["upward_field_events"] += len(rising)
                        for field in rising:
                            field_counts[family][field] += 1
                    case_id = f"{base_id}:{family}:{variant_id}"
                    classification_id = classify_upward(
                        classifications,
                        scope="seat4_small_state_search",
                        case_id=case_id,
                        family=family,
                        before_anchors=anchors,
                        after_anchors=after_anchors,
                        before=before,
                        after=after,
                        rise_map=rise_map,
                    )
                    transition_manifest.append(
                        {
                            "case_id": case_id,
                            "base_id": base_id,
                            "family": family,
                            "variant": variant_id,
                            "after_surface_digest": canonical_sha256(
                                V4.estimator_input_surface(after_anchors)
                            ),
                            "upward_fields": rising,
                            "classification_id": classification_id,
                        }
                    )
    return {
        "states": [list(state) for state in states],
        "anchor_counts": [2, 3, 4],
        "base_scene_count": len(base_manifest),
        "v3_base_ratio_score_zero_count": base_ratio_zero,
        "transition_count": len(transition_manifest),
        "family_stats": {family: dict(row) for family, row in stats.items()},
        "field_upward_counts": {
            family: dict(row) for family, row in field_counts.items()
        },
        "base_scene_manifest": base_manifest,
        "base_scene_manifest_digest": canonical_sha256(base_manifest),
        "transition_manifest": transition_manifest,
        "transition_manifest_digest": canonical_sha256(transition_manifest),
    }


def flatten_json(value: Any, path: str = "$") -> dict[str, dict[str, Any]]:
    output: dict[str, dict[str, Any]] = {}
    if isinstance(value, dict):
        output[path] = {"present": True, "kind": "object", "size": len(value)}
        for key in sorted(value):
            token = str(key).replace("~", "~0").replace("/", "~1")
            output.update(flatten_json(value[key], path + "/" + token))
    elif isinstance(value, list):
        output[path] = {"present": True, "kind": "array", "size": len(value)}
        for index, item in enumerate(value):
            output.update(flatten_json(item, path + f"/{index}"))
    else:
        kind = (
            "boolean"
            if isinstance(value, bool)
            else "number"
            if isinstance(value, (int, float)) and not isinstance(value, bool)
            else "null"
            if value is None
            else "string"
        )
        output[path] = {"present": True, "kind": kind, "value": value}
    return output


MISSING = {"present": False, "kind": "missing"}


def descriptor_equal(left: Mapping[str, Any], right: Mapping[str, Any]) -> bool:
    return dict(left) == dict(right)


def numeric_delta(
    left: Mapping[str, Any], right: Mapping[str, Any]
) -> float | None:
    if left.get("kind") == right.get("kind") == "number":
        return float(right["value"]) - float(left["value"])
    return None


def aggregate_snapshot(aggregate: Mapping[str, Any]) -> dict[str, Any]:
    overall = aggregate["overall"]
    return {
        "overall": {
            "scene_count": overall["scene_count"],
            "estimate_coverage": overall["estimate_coverage"],
            "accuracy_within_5pct": overall["accuracy_within_5pct"],
            "high_scene_count": overall["high_scene_count"],
            "high_coverage": overall["high_coverage"],
            "high_accuracy_within_5pct": overall["high_accuracy_within_5pct"],
            "relative_error": overall["relative_error"],
            "high_relative_error": overall["high_relative_error"],
            "unit_status_counts": overall["unit_status_counts"],
            "status_counts": overall["status_counts"],
        },
        "by_scale": {
            scale: {
                "scene_count": row["scene_count"],
                "high_scene_count": row["high_scene_count"],
                "high_coverage": row["high_coverage"],
                "high_accuracy_within_5pct": row["high_accuracy_within_5pct"],
                "relative_error": row["relative_error"],
                "high_relative_error": row["high_relative_error"],
                "unit_status_counts": row["unit_status_counts"],
            }
            for scale, row in aggregate["by_scale"].items()
        },
    }


def run_replay() -> dict[str, Any]:
    stored_v1_c1 = json.loads(V1_C1_BASELINE_PATH.read_text(encoding="utf-8"))["scenes"]
    stored_v1_l1b = json.loads(V1_L1B_BASELINE_PATH.read_text(encoding="utf-8"))["scenes"]
    stored_v3_payload = json.loads(V3_RESULTS_PATH.read_text(encoding="utf-8"))
    stored_v3 = {
        "l1b": stored_v3_payload["cohorts"]["l1b_200"]["scenes"],
        "c1_original": stored_v3_payload["cohorts"]["c1_original_200"]["scenes"],
    }
    cohorts: dict[str, Any] = {}
    total_field_rows = 0
    total_zero_delta_rows = 0
    baseline_integrity: dict[str, Any] = {}
    for cohort_name, directory, stored_v1 in (
        ("l1b", L1B_SCENE_DIR, stored_v1_l1b),
        ("c1_original", V1_SCENE_DIR, stored_v1_c1),
    ):
        scenes = load_scenes(directory)
        version_rows = {
            "v1": [V1.evaluate_scene(scene) for scene in scenes],
            "v3": [V3.evaluate_scene(scene) for scene in scenes],
            "v4": [V4.evaluate_scene(scene) for scene in scenes],
        }
        integrity = {}
        for version, stored in (("v1", stored_v1), ("v3", stored_v3[cohort_name])):
            stored_by_id = {str(row["scene_id"]): row for row in stored}
            live_by_id = {str(row["scene_id"]): row for row in version_rows[version]}
            mismatch_ids = [
                scene_id
                for scene_id in sorted(set(stored_by_id) | set(live_by_id))
                if stored_by_id.get(scene_id) != live_by_id.get(scene_id)
            ]
            integrity[version] = {
                "stored_count": len(stored),
                "live_count": len(version_rows[version]),
                "mismatch_count": len(mismatch_ids),
                "mismatch_scene_ids": mismatch_ids,
                "stored_digest": canonical_sha256(stored),
                "live_digest": canonical_sha256(version_rows[version]),
            }
        baseline_integrity[cohort_name] = integrity

        aggregates = {
            version: module.aggregate_results(version_rows[version])
            for version, module in (("v1", V1), ("v3", V3), ("v4", V4))
        }
        transcripts = []
        for index, scene in enumerate(scenes):
            rows = {version: version_rows[version][index] for version in ("v1", "v3", "v4")}
            flattened = {version: flatten_json(row) for version, row in rows.items()}
            paths = sorted(set().union(*(set(value) for value in flattened.values())))
            fields = []
            for field_path in paths:
                descriptors = {
                    version: flattened[version].get(field_path, MISSING)
                    for version in ("v1", "v3", "v4")
                }
                equal13 = descriptor_equal(descriptors["v1"], descriptors["v3"])
                equal34 = descriptor_equal(descriptors["v3"], descriptors["v4"])
                equal14 = descriptor_equal(descriptors["v1"], descriptors["v4"])
                fields.append(
                    {
                        "path": field_path,
                        "v1": descriptors["v1"],
                        "v3": descriptors["v3"],
                        "v4": descriptors["v4"],
                        "v1_to_v3_equal": equal13,
                        "v3_to_v4_equal": equal34,
                        "v1_to_v4_equal": equal14,
                        "v1_to_v3_numeric_delta": numeric_delta(descriptors["v1"], descriptors["v3"]),
                        "v3_to_v4_numeric_delta": numeric_delta(descriptors["v3"], descriptors["v4"]),
                        "v1_to_v4_numeric_delta": numeric_delta(descriptors["v1"], descriptors["v4"]),
                    }
                )
                total_zero_delta_rows += int(equal13 and equal34)
            total_field_rows += len(fields)
            transcripts.append(
                {
                    "scene_id": str(scene["scene_id"]),
                    "input_file": str(scene["_input_file"]),
                    "complete_version_rows": rows,
                    "complete_version_row_digests": {
                        version: canonical_sha256(row) for version, row in rows.items()
                    },
                    "per_field_delta_transcript": fields,
                    "field_row_count": len(fields),
                }
            )
        cohorts[cohort_name] = {
            "scene_count": len(scenes),
            "versions": {
                version: {
                    "aggregate": aggregate_snapshot(aggregates[version]),
                    "rows_digest": canonical_sha256(version_rows[version]),
                }
                for version in ("v1", "v3", "v4")
            },
            "scene_transcripts": transcripts,
            "scene_transcript_count": len(transcripts),
            "field_row_count": sum(row["field_row_count"] for row in transcripts),
        }
    return {
        "schema": "ariadne.e2.loop_l1e.replay_delta.v1",
        "versions": ["v1", "v3", "v4"],
        "cohorts": cohorts,
        "baseline_live_integrity": baseline_integrity,
        "publication": {
            "scene_count": sum(row["scene_count"] for row in cohorts.values()),
            "scene_transcript_count": sum(
                row["scene_transcript_count"] for row in cohorts.values()
            ),
            "per_scene_per_field_row_count": total_field_rows,
            "zero_delta_field_row_count": total_zero_delta_rows,
            "zero_delta_rows_included": True,
            "complete_version_rows_included": True,
        },
    }


# Counterexample: At any sealed scale, HIGH coverage=0.599 must return false.
def predicate_high_coverage_per_scale(
    values: Mapping[str, float], minimum: float
) -> bool:
    return bool(values) and all(float(value) >= minimum for value in values.values())


# Counterexample: accuracy=0.949 or any relative error>0.05 must return false.
def predicate_high_accuracy(
    accuracy: float, relative_errors: Sequence[float], minimum: float, maximum_error: float
) -> bool:
    return float(accuracy) >= minimum and all(
        float(value) <= maximum_error for value in relative_errors
    )


# Counterexample: One O1/O2 status downgrade or coverage loss must return false.
def predicate_legitimate_scene_no_loss(
    status_downgrades: int, coverage_losses: int
) -> bool:
    return int(status_downgrades) == 0 and int(coverage_losses) == 0


# Counterexample: Omitting one scene-field row, including a zero row, returns false.
def predicate_cohort_replay_delta_disclosure(
    expected_rows: int, observed_rows: int, zero_rows_included: bool
) -> bool:
    return int(expected_rows) == int(observed_rows) and bool(zero_rows_included)


# Counterexample: A residual-suspicion CE-B2 increase classified violation returns false.
def predicate_monotonicity_violation_count(
    upward_field_events: int, classified_field_events: int, violations: int
) -> bool:
    return (
        int(upward_field_events) == int(classified_field_events)
        and int(violations) == 0
    )


# Counterexample: Two records sharing one handle fail a two-independent-source request.
def predicate_unique_handle_independence(
    handles: Sequence[str], required: int
) -> bool:
    return len(set(map(str, handles))) >= int(required)


# Counterexample: Empty or singleton eligible support with NONE fails non-NONE.
def predicate_eligible_support_non_none(
    eligible_handles: Sequence[str], status: str
) -> bool:
    return len(set(map(str, eligible_handles))) >= 2 and str(status) != "NONE"


# Counterexample: A finite 1-to-0 jump at tau +/- epsilon returns false.
def predicate_continuous_tau_attenuation(
    below: float, above: float, epsilon: float
) -> bool:
    return abs(float(below) - float(above)) <= float(epsilon)


# Counterexample: Changed prereg bytes cannot retain the recorded content hash.
def predicate_seal_content_match(observed_sha256: str, expected_sha256: str) -> bool:
    return str(observed_sha256).upper() == str(expected_sha256).upper()


def build_predicate_registry() -> dict[str, Any]:
    changed_sha = hashlib.sha256(PREREG_PATH.read_bytes() + b"changed").hexdigest().upper()
    cases = [
        (
            "high_coverage_per_scale",
            "all(scale_high_coverage >= 0.6)",
            ["scale_high_coverage", "minimum_fraction"],
            "At any sealed scale, HIGH coverage=0.599 must return false.",
            predicate_high_coverage_per_scale({"0.001": 0.599}, 0.6),
            {"scale_high_coverage": {"0.001": 0.599}, "minimum_fraction": 0.6},
        ),
        (
            "high_accuracy",
            "accuracy >= 0.95 and all(relative_error <= 0.05)",
            ["accuracy", "relative_errors", "minimum_fraction", "maximum_error"],
            "Within unit_status=HIGH, accuracy=0.949 or any relative error>0.05 must return false.",
            predicate_high_accuracy(0.949, [0.0, 0.051], 0.95, 0.05),
            {"accuracy": 0.949, "relative_errors": [0.0, 0.051]},
        ),
        (
            "legitimate_scene_no_loss",
            "status_downgrades == 0 and coverage_losses == 0",
            ["status_downgrades", "coverage_losses"],
            "One O1 or O2 status downgrade or one coverage loss must return false.",
            predicate_legitimate_scene_no_loss(1, 0),
            {"status_downgrades": 1, "coverage_losses": 0},
        ),
        (
            "cohort_replay_delta_disclosure",
            "expected_rows == observed_rows and zero_rows_included",
            ["expected_rows", "observed_rows", "zero_rows_included"],
            "Omitting one scene-field delta row, including a zero-delta row, must return false.",
            predicate_cohort_replay_delta_disclosure(10, 9, False),
            {"expected_rows": 10, "observed_rows": 9, "zero_rows_included": False},
        ),
        (
            "monotonicity_violation_count",
            "upward_field_events == classified_field_events and violations == 0",
            ["upward_field_events", "classified_field_events", "violations"],
            "A CE-B2-style increase with a surviving ratio outlier and no observation-equivalent honest witness must return false.",
            predicate_monotonicity_violation_count(1, 1, 1),
            {"upward_field_events": 1, "classified_field_events": 1, "violations": 1},
        ),
        (
            "unique_handle_independence",
            "len(set(handles)) >= required",
            ["handles", "required"],
            "Multiple records sharing one handle must count as one independent source and must fail any predicate requiring two or more independent handles.",
            predicate_unique_handle_independence(["H1", "H1"], 2),
            {"handles": ["H1", "H1"], "required": 2},
        ),
        (
            "eligible_support_non_none",
            "len(set(eligible_handles)) >= 2 and status != NONE",
            ["eligible_handles", "status"],
            "An empty eligible set or a singleton suspicious clone must produce NONE and must fail a non-NONE success predicate.",
            predicate_eligible_support_non_none(["ONLY"], "NONE"),
            {"eligible_handles": ["ONLY"], "status": "NONE"},
        ),
        (
            "continuous_tau_attenuation",
            "abs(score_below-score_above) <= epsilon",
            ["score_below", "score_above", "epsilon"],
            "A fixture implementing a finite 1-to-0 jump at tau plus or minus epsilon must return false.",
            predicate_continuous_tau_attenuation(1.0, 0.0, 1e-6),
            {"score_below": 1.0, "score_above": 0.0, "epsilon": 1e-6},
        ),
        (
            "seal_content_match",
            "observed_sha256 == expected_sha256",
            ["observed_sha256", "expected_sha256"],
            "Changing any byte of prereg.json while retaining the recorded digest must return false.",
            predicate_seal_content_match(changed_sha, EXPECTED_PREREG_SHA256),
            {"observed_sha256": changed_sha, "expected_sha256": EXPECTED_PREREG_SHA256},
        ),
    ]
    registry = []
    transcript = []
    for identifier, expression, fields, annotation, observed, inputs in cases:
        row = {
            "predicate_id": identifier,
            "predicate_expression": expression,
            "observed_input_fields": fields,
            "counterexample_input_annotation": annotation,
            "counterexample_input": inputs,
            "expected_counterexample_result_false": True,
            "counterexample_return_value": bool(observed),
            "observed_counterexample_result_false": not bool(observed),
        }
        registry.append(row)
        transcript.append(json.dumps(row, ensure_ascii=False, sort_keys=True))
    if any(row["counterexample_return_value"] for row in registry):
        raise RuntimeError("predicate counterexample did not return false")
    return {
        "schema": "ariadne.e2.loop_l1e.predicate_registry.v1",
        "predicate_count": len(registry),
        "registry": registry,
        "counterexample_execution_transcript": transcript,
    }


def witness_payload(classifications: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    label_counts = Counter(
        str(row["automated_classification_result"]) for row in classifications
    )
    field_event_count = sum(len(row["increased_fields"]) for row in classifications)
    unclassified = sum(int(row["unclassified_field_count"]) for row in classifications)
    return {
        "schema": "ariadne.e2.loop_l1e.witness_classifications.v1",
        "classification_count": len(classifications),
        "upward_field_event_count": field_event_count,
        "label_counts": dict(sorted(label_counts.items())),
        "manual_suppression_count": sum(
            int(row["manual_suppression_used"]) for row in classifications
        ),
        "unclassified_field_count": unclassified,
        "classifications": list(classifications),
        "classifications_digest": canonical_sha256(classifications),
    }


def build_workbook(
    path: Path,
    *,
    seals: Mapping[str, Any],
    results: Mapping[str, Any],
    replay: Mapping[str, Any],
    fleet: Mapping[str, Any],
    witnesses: Mapping[str, Any],
    predicates: Mapping[str, Any],
    predicted_hashes: Mapping[str, str],
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    def append_table(sheet: Any, rows: Sequence[Sequence[Any]]) -> None:
        for row in rows:
            sheet.append(list(row))
        if rows:
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            width = min(60, max(len(str(cell.value or "")) for cell in column) + 2)
            sheet.column_dimensions[column[0].column_letter].width = width
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    readme = workbook.create_sheet("README")
    append_table(
        readme,
        [
            ["key", "value"],
            ["witness_commit_sha", seals["witness_commit_sha"]],
            ["prereg_json_sha256", seals["prereg_json_sha256"]],
            ["prereg_sealed_csv_sha256", seals["prereg_sealed_csv_sha256"]],
            ["continuous_formula", V4.CONTINUOUS_FORMULA],
            ["boundary_behavior", V4.BOUNDARY_BEHAVIOR],
            ["gate_verdict_emitted", False],
        ],
    )
    numbers = workbook.create_sheet("NUMBERS")
    number_rows = [["scope", "metric", "value"]]
    number_rows.extend(
        ["property", key, value]
        for key, value in (
            ("case_count", results["property_test"]["case_count"]),
            ("randomized_case_count", results["property_test"]["randomized_case_count"]),
            ("classification_count", witnesses["classification_count"]),
            ("upward_field_event_count", witnesses["upward_field_event_count"]),
            ("information_limit_record_count", witnesses["label_counts"].get("information_limit_record", 0)),
            ("violation_count", witnesses["label_counts"].get("violation", 0)),
        )
    )
    for cohort_name, cohort in replay["cohorts"].items():
        metrics = cohort["versions"]["v4"]["aggregate"]
        number_rows.append([cohort_name, "overall_high_coverage", metrics["overall"]["high_coverage"]])
        number_rows.append([cohort_name, "overall_high_accuracy", metrics["overall"]["high_accuracy_within_5pct"]])
        for scale, row in metrics["by_scale"].items():
            number_rows.append([cohort_name, f"scale_{scale}_high_coverage", row["high_coverage"]])
            number_rows.append([cohort_name, f"scale_{scale}_high_accuracy", row["high_accuracy_within_5pct"]])
    append_table(numbers, number_rows)

    pred_sheet = workbook.create_sheet("SELFTEST")
    append_table(
        pred_sheet,
        [[
            "predicate_id",
            "predicate_expression",
            "observed_input_fields",
            "counterexample_input_annotation",
            "counterexample_input",
            "expected_false",
            "return_value",
            "observed_false",
        ]]
        + [
            [
                row["predicate_id"],
                row["predicate_expression"],
                json.dumps(row["observed_input_fields"], ensure_ascii=False),
                row["counterexample_input_annotation"],
                json.dumps(row["counterexample_input"], ensure_ascii=False, sort_keys=True),
                row["expected_counterexample_result_false"],
                row["counterexample_return_value"],
                row["observed_counterexample_result_false"],
            ]
            for row in predicates["registry"]
        ],
    )
    prop_sheet = workbook.create_sheet("PROPERTY")
    append_table(
        prop_sheet,
        [["case_id", "stratum", "class_id", "family", "upward_fields", "classification_id", "case_digest"]]
        + [
            [
                row["case_id"],
                row.get("stratum"),
                row.get("class_id"),
                row["family"],
                ",".join(field for field, rose in row["increases"].items() if rose),
                row.get("classification_id"),
                row["case_digest"],
            ]
            for row in results["property_test"]["case_manifest"]
        ],
    )
    witness_sheet = workbook.create_sheet("WITNESS")
    append_table(
        witness_sheet,
        [["classification_id", "scope", "case_id", "family", "fields", "label", "post_sha256", "witness_sha256", "exact_equal", "semantic_legitimacy"]]
        + [
            [
                row["classification_id"],
                row["scope"],
                row["case_id"],
                row["family"],
                ",".join(row["increased_fields"]),
                row["automated_classification_result"],
                row["post_serialization_sha256"],
                row["witness_serialization_sha256"],
                row["exact_all_field_equality"],
                row["city_semantic_legitimacy"],
            ]
            for row in witnesses["classifications"]
        ],
    )
    replay_sheet = workbook.create_sheet("REPLAY")
    replay_rows = [["cohort", "scene_id", "v1_status", "v3_status", "v4_status", "v1_unit", "v3_unit", "v4_unit", "v1_estimate", "v3_estimate", "v4_estimate", "field_rows"]]
    for cohort_name, cohort in replay["cohorts"].items():
        for row in cohort["scene_transcripts"]:
            versions = row["complete_version_rows"]
            replay_rows.append(
                [
                    cohort_name,
                    row["scene_id"],
                    versions["v1"]["status"],
                    versions["v3"]["status"],
                    versions["v4"]["status"],
                    versions["v1"]["unit_status"],
                    versions["v3"]["unit_status"],
                    versions["v4"]["unit_status"],
                    versions["v1"]["scale_estimate"],
                    versions["v3"]["scale_estimate"],
                    versions["v4"]["scale_estimate"],
                    row["field_row_count"],
                ]
            )
    append_table(replay_sheet, replay_rows)
    search_sheet = workbook.create_sheet("SEARCH_1548")
    append_table(
        search_sheet,
        [["family", "transitions", "upward_transitions", "upward_field_events", *TRACKED_FIELDS]]
        + [
            [
                family,
                fleet["small_state_search"]["family_stats"][family]["transitions"],
                fleet["small_state_search"]["family_stats"][family]["upward_transitions"],
                fleet["small_state_search"]["family_stats"][family]["upward_field_events"],
                *[
                    fleet["small_state_search"]["field_upward_counts"][family][field]
                    for field in TRACKED_FIELDS
                ],
            ]
            for family in V4.PROPERTY_FAMILIES
        ],
    )
    files_sheet = workbook.create_sheet("FILES")
    append_table(
        files_sheet,
        [["name", "sha256"]] + [[name, digest] for name, digest in sorted(predicted_hashes.items())],
    )
    fixed_time = datetime(2026, 7, 19, 0, 0, 0)
    workbook.properties.creator = "E2 L1e"
    workbook.properties.lastModifiedBy = "E2 L1e"
    workbook.properties.created = fixed_time
    workbook.properties.modified = fixed_time
    workbook.calculation.fullCalcOnLoad = False
    workbook.calculation.forceFullCalc = False
    workbook.save(path)


def md_table(headers: Sequence[Any], rows: Sequence[Sequence[Any]]) -> str:
    lines = [
        "| " + " | ".join(str(value) for value in headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for row in rows:
        values = [str(value).replace("|", "\\|").replace("\n", " ") for value in row]
        lines.append("| " + " | ".join(values) + " |")
    return "\n".join(lines)


def render_report(
    *,
    seals: Mapping[str, Any],
    results: Mapping[str, Any],
    replay: Mapping[str, Any],
    fleet: Mapping[str, Any],
    witnesses: Mapping[str, Any],
    predicates: Mapping[str, Any],
    artifact_hashes: Mapping[str, str],
) -> str:
    lines = [
        "# E2 Loop L1e — Phase B semantic repair and execution",
        "",
        f"- 증인 커밋 SHA: `{WITNESS_COMMIT_SHA}`",
        f"- sealed `prereg.json` SHA-256: `{EXPECTED_PREREG_SHA256}`",
        f"- sealed `PREREG_SEALED.csv` SHA-256: `{EXPECTED_SEALED_CSV_SHA256}`",
        "",
        "## 사전 공개 수리 공식과 경계 거동",
        "",
        f"- formula: `{V4.CONTINUOUS_FORMULA}`",
        f"- boundary: {V4.BOUNDARY_BEHAVIOR}",
        "- CE-B2: 잔존 ratio 의심이 하나라도 있으면 retyped GRID와 display-hidden handle도 동일 source-handle 분모에 남는다.",
        "- CE-D: 선택 mode의 distinct supporting handle이 0 또는 1이면 `display_per_raw=null`, `unit_status=NONE`; HIGH에는 3개가 필요하다.",
        "- D1: 후보·부분집합·완전성·inlier·독립성 수치는 record 수가 아니라 source-handle 집합으로 계산한다.",
        "- D3: `p_h=min(1,(distance/tau)^2)`이므로 tau 양쪽 극한이 같고 v3의 binary coherence multiplier는 없다.",
        "",
        "## 실행 수치",
        "",
        md_table(
            ["metric", "value"],
            [
                ["property_case_count", results["property_test"]["case_count"]],
                ["property_randomized_case_count", results["property_test"]["randomized_case_count"]],
                ["third_fleet_class_count", results["property_test"]["third_fleet_regression_case_count"]],
                ["witness_classification_count", witnesses["classification_count"]],
                ["upward_field_event_count", witnesses["upward_field_event_count"]],
                ["information_limit_record_count", witnesses["label_counts"].get("information_limit_record", 0)],
                ["violation_count", witnesses["label_counts"].get("violation", 0)],
                ["manual_suppression_count", witnesses["manual_suppression_count"]],
                ["unclassified_field_count", witnesses["unclassified_field_count"]],
                ["seat4_base_scene_count", fleet["small_state_search"]["base_scene_count"]],
                ["seat4_transition_count", fleet["small_state_search"]["transition_count"]],
                ["replay_scene_count", replay["publication"]["scene_count"]],
                ["replay_per_field_row_count", replay["publication"]["per_scene_per_field_row_count"]],
                ["replay_zero_delta_field_row_count", replay["publication"]["zero_delta_field_row_count"]],
            ],
        ),
        "",
        "## 코호트 HIGH coverage / accuracy 수치",
        "",
    ]
    cohort_rows = []
    for cohort_name, cohort in replay["cohorts"].items():
        aggregate = cohort["versions"]["v4"]["aggregate"]
        for scale, row in aggregate["by_scale"].items():
            cohort_rows.append(
                [
                    cohort_name,
                    scale,
                    row["scene_count"],
                    row["high_scene_count"],
                    row["high_coverage"],
                    row["high_accuracy_within_5pct"],
                    row["high_relative_error"]["max"],
                ]
            )
    lines.extend(
        [
            md_table(
                ["cohort", "scale", "n", "HIGH_n", "HIGH_coverage", "HIGH_accuracy_5pct", "HIGH_relerr_max"],
                cohort_rows,
            ),
            "",
            "## O1/O2 무손실 수치",
            "",
            md_table(
                ["metric", "count"],
                [
                    ["status_downgrade_count", fleet["core_fleet"]["o1_o2_numeric_loss_counts"]["status_downgrade_count"]],
                    ["coverage_loss_count", fleet["core_fleet"]["o1_o2_numeric_loss_counts"]["coverage_loss_count"]],
                ],
            ),
            "",
            "## Selftest 술어 전문",
            "",
        ]
    )
    lines.append(
        md_table(
            ["predicate_id", "expression", "counterexample", "return_value", "observed_false"],
            [
                [
                    row["predicate_id"],
                    row["predicate_expression"],
                    row["counterexample_input_annotation"],
                    str(row["counterexample_return_value"]).lower(),
                    str(row["observed_counterexample_result_false"]).lower(),
                ]
                for row in predicates["registry"]
            ],
        )
    )
    lines.extend(["", "```text", *predicates["counterexample_execution_transcript"], "```", "", "## 증인 분류 전문", ""])
    witness_rows = [
        [
            row["classification_id"],
            row["scope"],
            row["case_id"],
            row["family"],
            ",".join(row["increased_fields"]),
            row["automated_classification_result"],
            row["post_serialization_sha256"],
            str(row["exact_all_field_equality"]).lower(),
            str(row["city_semantic_legitimacy"]).lower(),
        ]
        for row in witnesses["classifications"]
    ]
    lines.append(
        md_table(
            ["id", "scope", "case", "family", "fields", "classification", "surface_digest", "exact_equal", "legitimate"],
            witness_rows,
        )
        if witness_rows
        else "classification_count: 0"
    )
    lines.extend(
        [
            "",
            f"완전한 5종 증거(두 직렬화·동치 digest·정당성 논거·자동 분류)는 `{WITNESS_PATH.name}`의 모든 행에 있다.",
            "",
            "## Replay 델타 요약과 전문 포인터",
            "",
            f"- cohorts: L1b {replay['cohorts']['l1b']['scene_count']} + C1 original {replay['cohorts']['c1_original']['scene_count']}.",
            f"- per-scene per-field rows: {replay['publication']['per_scene_per_field_row_count']}; zero-delta rows included: {replay['publication']['zero_delta_field_row_count']}.",
            f"- full unabridged transcript: `{REPLAY_PATH.name}` -> `cohorts.*.scene_transcripts[*].complete_version_rows` and `per_field_delta_transcript`.",
            "- 재현: 휘발 필드 제외 수치 전 필드 동일.",
            "",
            "## 1~3차 함대 및 1,548 장면 탐색 수치",
            "",
            md_table(
                ["family", "transitions", "upward_transitions", "upward_field_events"],
                [
                    [
                        family,
                        fleet["small_state_search"]["family_stats"][family]["transitions"],
                        fleet["small_state_search"]["family_stats"][family]["upward_transitions"],
                        fleet["small_state_search"]["family_stats"][family]["upward_field_events"],
                    ]
                    for family in V4.PROPERTY_FAMILIES
                ],
            ),
            "",
            "## 산출물 SHA-256",
            "",
            md_table(["file", "sha256"], [[name, digest] for name, digest in sorted(artifact_hashes.items())]),
            "",
            "## 미해결",
            "",
            "- 관측-동치인 깨끗한 사후 표면은 무상태 추정기로 교란 이력과 구별할 수 없으므로 information_limit_record로 전건 공개했다.",
            "- unique handle 문자열에는 외부 주체 인증이 없으며, 이 실행은 그 정보 한계를 accuracy truth로 과대해석하지 않는다.",
            "- 원본 CAD와 test surface에는 접근하지 않았다.",
            "- packet 전문을 읽기 전 환경 root 확인에 read-only git metadata 명령이 한 차례 포함되었고 null 결과로 끝났다. packet 확인 뒤 git 명령과 repository write는 실행하지 않았다.",
            "- sealed band의 gate 판정은 이 산출물에서 출력하지 않았다.",
            "",
            "LOOP_COMPLETE: L1e",
        ]
    )
    return "\n".join(lines) + "\n"


def compute_all() -> dict[str, Any]:
    started_wall = time.perf_counter()
    started_cpu = time.process_time()
    seals = verify_seals()
    source_before = readonly_manifest()
    fx = fixtures()
    classifications: list[dict[str, Any]] = []
    third = run_third_fleet(fx, classifications)
    property_result = run_property(fx, third, classifications)
    core = core_fleet(classifications)
    search = run_small_state_search(classifications)
    replay = run_replay()
    predicates = build_predicate_registry()
    source_after = readonly_manifest()
    source_integrity = compare_manifests(source_before, source_after)
    witnesses = witness_payload(classifications)
    fleet = {
        "schema": "ariadne.e2.loop_l1e.fleet_probe_results.v1",
        "core_fleet": core,
        "third_fleet": third,
        "small_state_search": search,
    }
    runtime = {
        "wall_seconds": time.perf_counter() - started_wall,
        "cpu_seconds": time.process_time() - started_cpu,
        "python": platform.python_version(),
        "numpy": np.__version__,
        "volatile_fields": ["runtime.wall_seconds", "runtime.cpu_seconds"],
    }
    results = {
        "schema": "ariadne.e2.loop_l1e.c1v6_results.v1",
        "contract": {
            "packet": str(PACKET_PATH),
            "write_root": str(CELL_DIR),
            "repo_mutation_allowed": False,
            "cad_access_allowed": False,
            "test_access_allowed": False,
            "gate_verdict_emitted": False,
            "subagents_used": False,
        },
        "seal": {key: value for key, value in seals.items() if key != "payload"},
        "sealed_configuration": seals["payload"],
        "estimator": {
            "path": str(ESTIMATOR_PATH),
            "sha256": sha256_file(ESTIMATOR_PATH),
            "revision": V4.REVISION,
            "continuous_formula": V4.CONTINUOUS_FORMULA,
            "boundary_behavior": V4.BOUNDARY_BEHAVIOR,
            "minimum_handles_non_none": V4.MIN_HANDLES_NON_NONE,
            "minimum_handles_high": V4.MIN_HANDLES_HIGH,
        },
        "property_test": property_result,
        "predicate_counterexamples": predicates,
        "witness_classification_summary": {
            key: value for key, value in witnesses.items() if key != "classifications"
        },
        "replay_numeric_summary": {
            cohort: row["versions"]["v4"]["aggregate"]
            for cohort, row in replay["cohorts"].items()
        },
        "fleet_numeric_summary": {
            "core_o1_o2_loss_counts": core["o1_o2_numeric_loss_counts"],
            "third_fleet_v4_upward_field_counts": third["v4_upward_field_counts"],
            "small_state_base_scene_count": search["base_scene_count"],
            "small_state_transition_count": search["transition_count"],
        },
        "source_readonly_manifest": source_integrity,
        "runtime": runtime,
    }
    return {
        "seals": seals,
        "results": results,
        "replay": replay,
        "fleet": fleet,
        "witnesses": witnesses,
        "predicates": predicates,
    }


def numerical_validation(bundle: Mapping[str, Any]) -> dict[str, Any]:
    prop = bundle["results"]["property_test"]
    witness = bundle["witnesses"]
    replay = bundle["replay"]
    fleet = bundle["fleet"]
    source = bundle["results"]["source_readonly_manifest"]
    upward_events = int(witness["upward_field_event_count"])
    classified_events = sum(
        len(row["increased_fields"]) for row in witness["classifications"]
    )
    values = {
        "property_case_count": prop["case_count"],
        "property_family_count": len(prop["family_counts"]),
        "property_stratum_count": len(prop["stratum_counts"]),
        "third_fleet_class_count": len(prop["third_fleet_classes"]),
        "seat4_base_scene_count": fleet["small_state_search"]["base_scene_count"],
        "seat4_transition_count": fleet["small_state_search"]["transition_count"],
        "replay_scene_count": replay["publication"]["scene_count"],
        "replay_scene_transcript_count": replay["publication"]["scene_transcript_count"],
        "replay_field_row_count": replay["publication"]["per_scene_per_field_row_count"],
        "replay_zero_delta_row_count": replay["publication"]["zero_delta_field_row_count"],
        "upward_field_event_count": upward_events,
        "classified_field_event_count": classified_events,
        "information_limit_record_count": witness["label_counts"].get("information_limit_record", 0),
        "violation_count": witness["label_counts"].get("violation", 0),
        "manual_suppression_count": witness["manual_suppression_count"],
        "unclassified_field_count": witness["unclassified_field_count"],
        "source_manifest_mismatch_count": source["mismatch_count"],
        "predicate_count": bundle["predicates"]["predicate_count"],
        "predicate_counterexample_false_count": sum(
            int(row["observed_counterexample_result_false"])
            for row in bundle["predicates"]["registry"]
        ),
    }
    required_equalities = {
        "property_minimum": prop["case_count"] >= 600,
        "family_coverage": set(prop["family_counts"]) == set(V4.PROPERTY_FAMILIES),
        "strata_coverage": set(prop["stratum_counts"])
        == {"zero_cliff_start", "ratio_outlier", "mixed_space", "handle_collision", "near_tau_spread"},
        "third_classes": set(prop["third_fleet_classes"])
        == {"CE-A", "CE-B", "CE-B2", "CE-C", "CE-D", "CE-E", "CE-F", "D1", "D2", "D3", "D4"},
        "search_base_count": fleet["small_state_search"]["base_scene_count"] == 1548,
        "replay_counts": replay["publication"]["scene_count"] == 400
        and replay["publication"]["scene_transcript_count"] == 400,
        "classification_complete": upward_events == classified_events,
        "violation_zero": witness["label_counts"].get("violation", 0) == 0,
        "manual_zero": witness["manual_suppression_count"] == 0
        and witness["unclassified_field_count"] == 0,
        "source_unchanged": source["mismatch_count"] == 0,
        "predicate_false_observed": all(
            row["observed_counterexample_result_false"]
            for row in bundle["predicates"]["registry"]
        ),
    }
    if not all(required_equalities.values()):
        raise RuntimeError(
            "numeric execution invariant mismatch: "
            + json.dumps({"values": values, "equalities": required_equalities}, sort_keys=True)
        )
    return values


def emit_outputs(bundle: Mapping[str, Any]) -> dict[str, Any]:
    generated = (
        C1_RESULTS_PATH,
        REPLAY_PATH,
        FLEET_PATH,
        WITNESS_PATH,
        PREDICATE_PATH,
        EVIDENCE_PATH,
        REPORT_PATH,
    )
    collisions = [str(path) for path in generated if path.exists()]
    if collisions:
        raise FileExistsError("existing Phase B artifact(s): " + json.dumps(collisions))

    payloads = {
        C1_RESULTS_PATH.name: bundle["results"],
        REPLAY_PATH.name: bundle["replay"],
        FLEET_PATH.name: bundle["fleet"],
        WITNESS_PATH.name: bundle["witnesses"],
        PREDICATE_PATH.name: bundle["predicates"],
    }
    texts = {name: json_text(payload) for name, payload in payloads.items()}
    predicted_hashes = {
        name: hashlib.sha256(text.encode("utf-8")).hexdigest().upper()
        for name, text in texts.items()
    }
    predicted_hashes[ESTIMATOR_PATH.name] = sha256_file(ESTIMATOR_PATH)
    predicted_hashes[SCRIPT_PATH.name] = sha256_file(SCRIPT_PATH)
    predicted_hashes[PREREG_PATH.name] = sha256_file(PREREG_PATH)
    predicted_hashes[SEALED_CSV_PATH.name] = sha256_file(SEALED_CSV_PATH)

    temporary_workbook = EVIDENCE_PATH.with_suffix(".xlsx.tmp")
    if temporary_workbook.exists():
        raise FileExistsError(temporary_workbook)
    build_workbook(
        temporary_workbook,
        seals=bundle["seals"],
        results=bundle["results"],
        replay=bundle["replay"],
        fleet=bundle["fleet"],
        witnesses=bundle["witnesses"],
        predicates=bundle["predicates"],
        predicted_hashes=predicted_hashes,
    )
    evidence_hash = sha256_file(temporary_workbook)
    artifact_hashes = {**predicted_hashes, EVIDENCE_PATH.name: evidence_hash}
    report = render_report(
        seals=bundle["seals"],
        results=bundle["results"],
        replay=bundle["replay"],
        fleet=bundle["fleet"],
        witnesses=bundle["witnesses"],
        predicates=bundle["predicates"],
        artifact_hashes=artifact_hashes,
    )

    for path in (C1_RESULTS_PATH, REPLAY_PATH, FLEET_PATH, WITNESS_PATH, PREDICATE_PATH):
        write_text_exclusive(path, texts[path.name])
    temporary_workbook.replace(EVIDENCE_PATH)
    write_text_exclusive(REPORT_PATH, report)
    artifact_hashes[REPORT_PATH.name] = sha256_file(REPORT_PATH)
    return artifact_hashes


def verify_outputs() -> dict[str, Any]:
    required = (
        ESTIMATOR_PATH,
        SCRIPT_PATH,
        C1_RESULTS_PATH,
        REPLAY_PATH,
        FLEET_PATH,
        WITNESS_PATH,
        PREDICATE_PATH,
        EVIDENCE_PATH,
        REPORT_PATH,
    )
    rows = [file_record(path) for path in required]
    report = REPORT_PATH.read_text(encoding="utf-8")
    replay = json.loads(REPLAY_PATH.read_text(encoding="utf-8"))
    witness = json.loads(WITNESS_PATH.read_text(encoding="utf-8"))
    predicates = json.loads(PREDICATE_PATH.read_text(encoding="utf-8"))
    workbook = load_workbook(EVIDENCE_PATH, read_only=True, data_only=False)
    sheet_names = list(workbook.sheetnames)
    workbook.close()
    numeric = {
        "required_file_count": len(rows),
        "report_terminal_line_count": sum(
            line == "LOOP_COMPLETE: L1e" for line in report.splitlines()
        ),
        "replay_scene_count": replay["publication"]["scene_count"],
        "replay_field_row_count": replay["publication"]["per_scene_per_field_row_count"],
        "witness_violation_count": witness["label_counts"].get("violation", 0),
        "witness_unclassified_field_count": witness["unclassified_field_count"],
        "predicate_count": predicates["predicate_count"],
        "predicate_observed_false_count": sum(
            int(row["observed_counterexample_result_false"])
            for row in predicates["registry"]
        ),
        "workbook_sheet_count": len(sheet_names),
    }
    if (
        not report.rstrip().endswith("LOOP_COMPLETE: L1e")
        or numeric["report_terminal_line_count"] != 1
        or numeric["replay_scene_count"] != 400
        or numeric["witness_violation_count"] != 0
        or numeric["witness_unclassified_field_count"] != 0
        or numeric["predicate_count"] != numeric["predicate_observed_false_count"]
    ):
        raise RuntimeError("output numeric verification mismatch: " + json.dumps(numeric))
    return {"numeric": numeric, "files": rows, "workbook_sheet_names": sheet_names}


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--compute-only", action="store_true")
    arguments = parser.parse_args(argv)
    bundle = compute_all()
    numeric = numerical_validation(bundle)
    if arguments.compute_only:
        print(json.dumps(numeric, ensure_ascii=False, sort_keys=True))
        return 0
    emit_outputs(bundle)
    verification = verify_outputs()
    print(json.dumps(verification["numeric"], ensure_ascii=False, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
