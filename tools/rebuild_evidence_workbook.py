#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Evidence workbook (xlsx) generator for a full-roundtrip capstone run.

Paul reviews each rebuilt drawing against the original BY EYE, with the
question "what evidence did the agent rebuild this from?" open next to it.
This tool folds one run directory (census IR + regen patch + deferred +
interior diff + dim semantic gate) into the 7-sheet workbook shape first
hand-built for R4h/R4l (reports/interior100/*_rebuild_evidence.xlsx) so
every future run regenerates the workbook identically instead of ad-hoc.

Usage:
    python tools/rebuild_evidence_workbook.py \
        --run-dir runs/e2e_1dwg_R4m_wipeout_20260709 \
        --out reports/interior100/R4m_rebuild_evidence.xlsx \
        --title R4m
"""
from __future__ import annotations

import argparse
import json
import os
from collections import Counter
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import Font

_JSON_ENCODING = "utf-8-sig"


def _load(path: str) -> Optional[Dict[str, Any]]:
    if not os.path.isfile(path):
        return None
    with open(path, encoding=_JSON_ENCODING) as fh:
        return json.load(fh)


def _sheet(wb, title: str, headers: List[str]):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    return ws


def _kind_histogram(entities: List[Dict[str, Any]]) -> str:
    counts = Counter(((e.get("geometry") or {}).get("kind") or "?") for e in entities)
    return ", ".join(f"{k}:{v}" for k, v in counts.most_common())


def build_workbook(run_dir: str, out_path: str, title: str) -> Dict[str, Any]:
    census = _load(os.path.join(run_dir, "census", "dwg_graph_ir.json")) or {}
    summary = _load(os.path.join(run_dir, "summary.json")) or {}
    deferred_doc = _load(os.path.join(run_dir, "deferred.json"))
    interior = _load(os.path.join(run_dir, "interior_diff.json")) or {}
    dim_gate = _load(os.path.join(run_dir, "dim_semantic_gate.json")) or {}
    patch = _load(os.path.join(run_dir, "regen", "patch.json")) or {}

    entities = [e for e in (census.get("entities") or []) if isinstance(e, dict)]
    block_defs = [b for b in (census.get("block_definitions") or []) if isinstance(b, dict)]
    operations = [op for op in (patch.get("operations") or []) if isinstance(op, dict)]
    deferred = (deferred_doc if isinstance(deferred_doc, list)
                else (deferred_doc or {}).get("deferred", []) or [])

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 1. 개요
    ws = _sheet(wb, "개요", ["항목", "값"])
    staged = summary.get("staged") or {}
    regen = summary.get("regen") or {}
    gate = (regen.get("interior_gate") or {})
    totals = (interior.get("totals") or {})
    excluded = totals.get("derived_cache_excluded") or {}
    rows = [
        ("런", title),
        ("run_dir", os.path.abspath(run_dir)),
        ("원본 DWG", staged.get("original_path")),
        ("원본 sha256", staged.get("original_sha256")),
        ("원본 크기(B)", staged.get("original_byte_size")),
        ("모델스페이스 엔티티", summary.get("filtered_entity_count")),
        ("블록 정의 수", len(block_defs)),
        ("재생 op 수", len(operations)),
        ("유예 수", len(deferred)),
        ("interior fraction", totals.get("interior_diff0_fraction")),
        ("interior diff0/분모", f"{totals.get('diff0_total')}/{totals.get('a_entity_total')}"),
        ("*D 파생캐시 제외", (f"a={excluded.get('a_def_count')}defs/{excluded.get('a_entity_total')}ents, "
                          f"b={excluded.get('b_def_count')}defs/{excluded.get('b_entity_total')}ents")
         if excluded else "(레거시 전수 비교)"),
        ("interior gate", f"{gate.get('status')} (baseline {gate.get('baseline')})"),
        ("L5 치수 게이트", dim_gate.get("status")),
        ("전체 status", summary.get("status")),
    ]
    for row in rows:
        ws.append(list(row))

    # 2. 모델스페이스
    ws = _sheet(wb, f"모델스페이스_{len(entities)}", ["handle", "class", "layer", "kind"])
    for e in entities:
        ws.append([e.get("handle"), e.get("class"), e.get("layer"),
                   (e.get("geometry") or {}).get("kind")])

    # 3. 블록정의
    ws = _sheet(wb, "블록정의", ["블록명", "내부 엔티티", "kind 분포", "익명"])
    for b in block_defs:
        des = [e for e in (b.get("def_entities") or []) if isinstance(e, dict)]
        name = b.get("name") or ""
        ws.append([name, len(des), _kind_histogram(des),
                   "Y" if name.startswith("*") else ""])

    # 4. 재생 op
    ws = _sheet(wb, f"재생op_{len(operations)}", ["#", "operation", "block", "kind", "layer"])
    for i, op in enumerate(operations):
        args = op.get("args") or {}
        ent = args.get("entity") or {}
        ws.append([i, op.get("operation"),
                   args.get("block_name") or args.get("name"),
                   ent.get("kind") if isinstance(ent, dict) else None,
                   args.get("layer")])

    # 5. 유예
    ws = _sheet(wb, f"유예_{len(deferred)}", ["블록명", "idx", "handle", "kind", "사유"])
    for d in deferred:
        ws.append([d.get("block_name"), d.get("def_entity_index"), d.get("handle"),
                   d.get("kind"), d.get("reason")])

    # 6. 잔여분석 (diff0 미만 def만)
    ws = _sheet(wb, "잔여분석",
                ["def", "a_total", "diff0", "removed", "added", "modified", "가족"])
    for row in sorted((interior.get("per_def") or []),
                      key=lambda r: -(int(r.get("a_total", 0)) - int(r.get("diff0", 0)))):
        if int(row.get("a_total", 0)) == int(row.get("diff0", 0)) and not row.get("missing_side"):
            continue
        name = row.get("name") or ""
        family = ("*D 파생캐시" if name.startswith("*D")
                  else "익명" if name.startswith("*") else "명명")
        ws.append([name, row.get("a_total"), row.get("diff0"), row.get("removed"),
                   row.get("added"), row.get("modified"), family])

    # 7. 치수 의미게이트
    dims = [e for e in entities if (e.get("geometry") or {}).get("kind") == "dimension"]
    ws = _sheet(wb, f"치수_{len(dims)}_의미게이트",
                ["handle", "layer", "측정값", "회전(rad)", "L5 판정"])
    verdict = dim_gate.get("status") or "?"
    for e in dims:
        g = e.get("geometry") or {}
        ws.append([e.get("handle"), e.get("layer"), g.get("measurement"),
                   g.get("rotation"), verdict])

    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb.save(out_path)
    return {
        "out": os.path.abspath(out_path),
        "sheets": [ws.title for ws in wb.worksheets],
        "modelspace": len(entities),
        "block_defs": len(block_defs),
        "operations": len(operations),
        "deferred": len(deferred),
    }


def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--run-dir", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--title", required=True)
    args = ap.parse_args(argv)
    result = build_workbook(args.run_dir, args.out, args.title)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
