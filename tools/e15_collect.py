#!/usr/bin/env python3
"""E1.5 collector: validate judge outputs, compute the preregistered metrics.

Reads reports/e1/annot_v1/raw/<judge>/shard_NN.json (JSON array per shard, schema v1)
plus the v0 ornith baseline, joins by unit_id, and emits calibration_v1.{json,md}
and e15_evidence.xlsx. All metrics here are the ones preregistered in
reports/e1/annot_v1/prereg_e15.json - nothing added post hoc without an
"exploratory" tag.
"""
from __future__ import annotations

import json
import math
from collections import Counter, defaultdict
from itertools import combinations
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
V1_DIR = ROOT / "reports" / "e1" / "annot_v1"
RAW_DIR = V1_DIR / "raw"
ORNITH_V0 = ROOT / "reports" / "e1" / "ornith_annot_v0.jsonl"

ROLES = {"평면 부분도", "심볼", "치수캐시", "가구", "기타"}
LIVE_JUDGES = ["opus48_max", "fable5_high", "sol56_xhigh", "sonnet5_xhigh", "grok45_xhigh"]
TOP_TIER = ["opus48_max", "fable5_high", "sol56_xhigh"]
RATIONALE_FIELDS = ("evidence", "rule", "counterfactual", "uncertainty")


def _rationale_complete(ans: dict) -> bool:
    r = ans.get("rationale")
    if not isinstance(r, dict):
        return False
    for f in RATIONALE_FIELDS:
        v = r.get(f)
        if not isinstance(v, str) or not v.strip():
            return False
    ev = r.get("evidence", "")
    return isinstance(ev, str) and len(ev.strip()) >= 20


def _valid_answer(ans: dict) -> tuple[bool, str]:
    if not isinstance(ans, dict):
        return False, "not an object"
    if not isinstance(ans.get("unit_id"), str) or not ans["unit_id"]:
        return False, "missing unit_id"
    if ans.get("role") not in ROLES:
        return False, f"bad role: {ans.get('role')!r}"
    wl = ans.get("wall_likelihood")
    if not isinstance(wl, (int, float)) or isinstance(wl, bool) or not (0.0 <= float(wl) <= 1.0):
        return False, "bad wall_likelihood"
    if not isinstance(ans.get("wall_line_handles"), list):
        return False, "bad wall_line_handles"
    if not _rationale_complete(ans):
        return False, "rationale incomplete"
    return True, ""


def load_judge(judge: str) -> tuple[dict[str, dict], dict]:
    answers: dict[str, dict] = {}
    stats = {"shards": 0, "answers": 0, "valid": 0, "errors": []}
    jdir = RAW_DIR / judge
    for shard in sorted(jdir.glob("shard_*.json")):
        stats["shards"] += 1
        try:
            arr = json.loads(shard.read_text(encoding="utf-8-sig"))
        except Exception as exc:  # noqa: BLE001 - report, never crash the collector
            stats["errors"].append(f"{shard.name}: parse failure {exc}")
            continue
        if not isinstance(arr, list):
            stats["errors"].append(f"{shard.name}: not a JSON array")
            continue
        for ans in arr:
            stats["answers"] += 1
            ok, why = _valid_answer(ans)
            if ok:
                stats["valid"] += 1
                answers[ans["unit_id"]] = ans
            else:
                uid = ans.get("unit_id") if isinstance(ans, dict) else "?"
                stats["errors"].append(f"{shard.name}:{uid}: {why}")
    return answers, stats


def load_ornith_v0() -> dict[str, dict]:
    answers: dict[str, dict] = {}
    with open(ORNITH_V0, "r", encoding="utf-8-sig") as handle:
        for line in handle:
            if not line.strip():
                continue
            row = json.loads(line)
            parsed = row.get("parsed") or {}
            uid = row.get("unit_id")
            if not uid or parsed.get("role") is None:
                continue
            answers[uid] = {
                "unit_id": uid,
                "def": parsed.get("def"),
                "role": parsed.get("role"),
                "wall_likelihood": parsed.get("wall_likelihood"),
                "wall_line_handles": parsed.get("wall_line_handles") or [],
                "notes": parsed.get("notes"),
            }
    return answers


def _bucket(x: float) -> str:
    if x <= 0.3:
        return "low"
    if x >= 0.7:
        return "high"
    return "mid"


def _pearson(xs: list[float], ys: list[float]) -> float | None:
    if len(xs) < 2:
        return None
    mx, my = sum(xs) / len(xs), sum(ys) / len(ys)
    cov = sum((a - mx) * (b - my) for a, b in zip(xs, ys))
    vx = sum((a - mx) ** 2 for a in xs)
    vy = sum((b - my) ** 2 for b in ys)
    if not vx or not vy:
        return None
    return cov / math.sqrt(vx * vy)


def _handles(ans: dict) -> set[str]:
    out = set()
    for h in ans.get("wall_line_handles") or []:
        v = h.get("handle") if isinstance(h, dict) else h
        if isinstance(v, str) and v.strip():
            out.add(v.strip())
    return out


def _jaccard(a: set[str], b: set[str]) -> float | None:
    u = a | b
    if not u:
        return None
    return len(a & b) / len(u)


def pairwise(a: dict[str, dict], b: dict[str, dict]) -> dict:
    common = sorted(set(a) & set(b))
    role_agree = sum(1 for u in common if a[u]["role"] == b[u]["role"])
    xs, ys = [], []
    bucket_agree = 0
    jac = []
    for u in common:
        wa, wb = a[u].get("wall_likelihood"), b[u].get("wall_likelihood")
        if isinstance(wa, (int, float)) and isinstance(wb, (int, float)):
            xs.append(float(wa))
            ys.append(float(wb))
            if _bucket(float(wa)) == _bucket(float(wb)):
                bucket_agree += 1
        j = _jaccard(_handles(a[u]), _handles(b[u]))
        if j is not None:
            jac.append(j)
    return {
        "n_common": len(common),
        "role_agreement": role_agree / len(common) if common else None,
        "likelihood_pearson": _pearson(xs, ys),
        "bucket_agreement": bucket_agree / len(xs) if xs else None,
        "handle_jaccard_mean": sum(jac) / len(jac) if jac else None,
        "handle_jaccard_n": len(jac),
    }


def fleiss_kappa(judgesets: dict[str, dict[str, dict]]) -> float | None:
    units = set.intersection(*(set(v) for v in judgesets.values())) if judgesets else set()
    if not units:
        return None
    cats = sorted(ROLES)
    n_j = len(judgesets)
    rows = []
    for u in sorted(units):
        counts = Counter(judgesets[j][u]["role"] for j in judgesets)
        rows.append([counts.get(c, 0) for c in cats])
    n_u = len(rows)
    p_i = [(sum(c * c for c in row) - n_j) / (n_j * (n_j - 1)) for row in rows]
    p_bar = sum(p_i) / n_u
    totals = [sum(row[k] for row in rows) for k in range(len(cats))]
    p_j = [t / (n_u * n_j) for t in totals]
    p_e = sum(p * p for p in p_j)
    if p_e >= 1.0:
        return None
    return (p_bar - p_e) / (1 - p_e)


def main() -> None:
    judgesets: dict[str, dict[str, dict]] = {}
    validity: dict[str, dict] = {}
    for judge in LIVE_JUDGES:
        if (RAW_DIR / judge).is_dir():
            answers, stats = load_judge(judge)
            judgesets[judge] = answers
            rate = stats["valid"] / stats["answers"] if stats["answers"] else 0.0
            validity[judge] = {**stats, "valid_rate": round(rate, 4),
                               "B3_valid": rate >= 0.95, "errors": stats["errors"][:40]}
    ornith = load_ornith_v0()
    all_sets = {**judgesets, "ornith35b_v0": ornith}

    pairs = {}
    for a, b in combinations(sorted(all_sets), 2):
        pairs[f"{a}|{b}"] = pairwise(all_sets[a], all_sets[b])

    def mean_over(names: list[str], key: str) -> float | None:
        vals = []
        for a, b in combinations(sorted(names), 2):
            v = pairs.get(f"{a}|{b}", {}).get(key)
            if v is not None:
                vals.append(v)
        return sum(vals) / len(vals) if vals else None

    live_present = [j for j in LIVE_JUDGES if j in judgesets and judgesets[j]]
    top_present = [j for j in TOP_TIER if j in judgesets and judgesets[j]]
    top_role = mean_over(top_present, "role_agreement") if len(top_present) >= 2 else None
    top_pearson = mean_over(top_present, "likelihood_pearson") if len(top_present) >= 2 else None

    def band_b1(x):
        if x is None:
            return "NOT_COMPUTABLE"
        return "well_posed" if x >= 0.70 else ("marginal" if x >= 0.60 else "ambiguity_dominates")

    result = {
        "schema": "ariadne.e15_calibration.v1",
        "prereg": "reports/e1/annot_v1/prereg_e15.json",
        "judges_present": sorted(all_sets),
        "validity": validity,
        "pairwise": pairs,
        "fleiss_kappa_role_live": fleiss_kappa({j: judgesets[j] for j in live_present}) if len(live_present) >= 2 else None,
        "top_tier_mean_role_agreement": top_role,
        "top_tier_mean_likelihood_pearson": top_pearson,
        "verdicts": {
            "B1_task_well_posed": band_b1(top_role),
            "B2_ladder_visible": (None if top_role is None else bool(top_role - 0.5491 >= 0.10)),
            "B4_likelihood_usable": (None if top_pearson is None else
                                     ("silver_ok" if top_pearson >= 0.70 else
                                      ("gray" if top_pearson >= 0.50 else "unreliable"))),
        },
    }
    out_json = V1_DIR / "calibration_v1.json"
    out_json.write_text(json.dumps(result, ensure_ascii=False, indent=1), encoding="utf-8", newline="\n")

    lines = ["# E1.5 calibration v1", "", f"Judges: {', '.join(sorted(all_sets))}", ""]
    lines.append("| pair | n | role agree | pearson | bucket | jaccard |")
    lines.append("| --- | ---: | ---: | ---: | ---: | ---: |")
    for k in sorted(pairs):
        p = pairs[k]
        fmt = lambda v: "-" if v is None else f"{v:.4f}"  # noqa: E731
        lines.append(f"| {k} | {p['n_common']} | {fmt(p['role_agreement'])} | "
                     f"{fmt(p['likelihood_pearson'])} | {fmt(p['bucket_agreement'])} | {fmt(p['handle_jaccard_mean'])} |")
    lines += ["", f"Fleiss kappa (live judges, role): {result['fleiss_kappa_role_live']}",
              f"Top-tier mean role agreement: {top_role}",
              f"Top-tier mean likelihood pearson: {top_pearson}",
              f"Verdicts: {json.dumps(result['verdicts'], ensure_ascii=False)}", ""]
    (V1_DIR / "calibration_v1.md").write_text("\n".join(lines), encoding="utf-8", newline="\n")

    try:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "pairwise"
        ws.append(["pair", "n_common", "role_agreement", "likelihood_pearson", "bucket_agreement", "handle_jaccard_mean"])
        for k in sorted(pairs):
            p = pairs[k]
            ws.append([k, p["n_common"], p["role_agreement"], p["likelihood_pearson"], p["bucket_agreement"], p["handle_jaccard_mean"]])
        ws2 = wb.create_sheet("validity")
        ws2.append(["judge", "shards", "answers", "valid", "valid_rate", "B3_valid"])
        for j, s in validity.items():
            ws2.append([j, s["shards"], s["answers"], s["valid"], s["valid_rate"], s["B3_valid"]])
        ws3 = wb.create_sheet("answers")
        ws3.append(["judge", "unit_id", "def", "role", "wall_likelihood", "n_handles", "rationale_complete"])
        for j in sorted(all_sets):
            for u in sorted(all_sets[j]):
                a = all_sets[j][u]
                ws3.append([j, u, a.get("def"), a.get("role"), a.get("wall_likelihood"),
                            len(a.get("wall_line_handles") or []), _rationale_complete(a)])
        wb.save(V1_DIR / "e15_evidence.xlsx")
        xlsx = "written"
    except Exception as exc:  # noqa: BLE001
        xlsx = f"SKIPPED: {exc}"
    print(json.dumps({"judges": sorted(all_sets), "validity": {j: v["valid_rate"] for j, v in validity.items()},
                      "verdicts": result["verdicts"], "xlsx": xlsx}, ensure_ascii=False))


if __name__ == "__main__":
    main()
